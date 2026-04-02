"""
Segment Analysis Module for XBRL to Excel Conversion

セグメント情報の分析シートを生成するモジュール
"""

import unicodedata


def add_segment_analysis_sheets(workbook, segment_sheets_info, debug_log=None):
    """
    セグメント情報シートから分析シートを生成

    Args:
        workbook: openpyxlワークブック
        segment_sheets_info: セグメントシート情報のリスト
            各要素は辞書で以下のキーを持つ:
            - sheet_name: シート名
            - ordered_keys: プレゼンテーションツリーの要素リスト [(full_path, pref_label), ...]
            - all_years_data: 全期間のデータ {role: {full_path: {(std, dim, period): value}}}
            - role: ロールURI
            - sorted_role_cols: ソート済みカラム情報のリスト
            - role_columns: ロールのカラム情報セット
            - current_standard: 現在の会計基準
            - segment_dict: セグメント用語辞書
            - common_dict: 共通用語辞書
            - labels_map: ラベルマップ
            - used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    debug_log(f"[Segment Analysis] Starting segment analysis sheet generation for {len(segment_sheets_info)} sheets")

    for info in segment_sheets_info:
        _create_segment_analysis_sheet(
            workbook=workbook,
            sheet_name=info['sheet_name'],
            ordered_keys=info['ordered_keys'],
            all_years_data=info['all_years_data'],
            role=info['role'],
            sorted_role_cols=info['sorted_role_cols'],
            role_columns=info['role_columns'],
            current_standard=info['current_standard'],
            segment_dict=info['segment_dict'],
            common_dict=info['common_dict'],
            labels_map=info['labels_map'],
            used_sheet_names=info['used_sheet_names'],
            global_element_period_values=info.get('global_element_period_values', {}),
            debug_log=debug_log
        )

        # 元の注記シートの末尾にセグメント別研究開発費を追加
        _append_rd_to_notes_sheet(
            workbook=workbook,
            sheet_name=info['sheet_name'],
            sorted_role_cols=info['sorted_role_cols'],
            global_element_period_values=info.get('global_element_period_values', {}),
            debug_log=debug_log
        )

        # Create PPM analysis sheet for Japanese GAAP only
        if '日本基準' in info['sheet_name']:
            # Use the same sheet name transformation as _create_segment_analysis_sheet
            # (replacing "注記" with "連結")
            base_name = info['sheet_name'].replace("注記", "連結")
            analysis_sheet_name = base_name + "_分析"
            if len(analysis_sheet_name) > 31:
                analysis_sheet_name = base_name[:28] + "_分析"

            _create_ppm_analysis_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                filing_pairs=info.get('filing_pairs', []),
                debug_log=debug_log
            )

            # PPM後に作成することで、PPMが分析シートに追加した集計列も反映される
            _create_data_acquisition_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                filing_pairs=info.get('filing_pairs', []),
                debug_log=debug_log
            )

            _create_composition_ratio_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_yoy_growth_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_ebitda_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_sales_ratio_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_employee_ratio_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

        elif 'IFRS' in info['sheet_name']:
            base_name = info['sheet_name'].replace("注記", "連結")
            analysis_sheet_name = base_name + "_分析"
            if len(analysis_sheet_name) > 31:
                analysis_sheet_name = base_name[:28] + "_分析"

            _create_ppm_analysis_sheet_ifrs(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                filing_pairs=info.get('filing_pairs', []),
                debug_log=debug_log
            )

            # PPM後に作成することで、PPMが分析シートに追加した集計列も反映される
            _create_data_acquisition_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                filing_pairs=info.get('filing_pairs', []),
                debug_log=debug_log
            )

            _create_ebitda_sheet_ifrs(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_sales_ratio_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_employee_ratio_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_composition_ratio_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )

            _create_yoy_growth_sheet(
                workbook=workbook,
                analysis_sheet_name=analysis_sheet_name,
                used_sheet_names=info['used_sheet_names'],
                debug_log=debug_log
            )


def _create_segment_analysis_sheet(workbook, sheet_name, ordered_keys, all_years_data,
                                   role, sorted_role_cols, role_columns, current_standard,
                                   segment_dict, common_dict, labels_map, used_sheet_names,
                                   global_element_period_values, debug_log):
    """
    セグメント分析シートを作成（内部関数）

    Args:
        workbook: openpyxlワークブック
        sheet_name: 元のシート名
        ordered_keys: プレゼンテーションツリーの要素リスト
        all_years_data: 全期間のデータ
        role: ロールURI
        sorted_role_cols: ソート済みカラム情報
        role_columns: ロールのカラム情報
        current_standard: 現在の会計基準
        segment_dict: セグメント用語辞書
        common_dict: 共通用語辞書
        labels_map: ラベルマップ
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    # 1. 「注記」を「連結」に置換
    base_name = sheet_name.replace("注記", "連結")

    # 2. 分析シート名を生成
    analysis_sheet_name = base_name + "_分析"

    # 3. Excelのシート名制限（31文字）をチェック
    if len(analysis_sheet_name) > 31:
        # 31文字を超える場合は、置換後の名前をカットして "_分析" を結合
        # (31 - 3 = 28文字分を本体から取得)
        analysis_sheet_name = base_name[:28] + "_分析"

    debug_log(f"[Segment Analysis] Creating analysis sheet: {analysis_sheet_name}")

    # 分析シートを作成
    aws = workbook.create_sheet(title=analysis_sheet_name)
    used_sheet_names.add(analysis_sheet_name)

    # セグメントを横軸に配置（ユニークなディメンション）
    unique_dims = []
    for c in sorted_role_cols:
        # c is (std, dim, period)
        d = c[1] if len(c) == 3 else c[0]
        if d not in unique_dims:
            unique_dims.append(d)

    # 「報告セグメント及びその他の合計」がある場合は「報告セグメント」（単独列）を除去する。
    # 単独の「報告セグメント」列はデータが空になることが多く冗長なため。
    # PPM分析用の chart_end_col も自動的に「報告セグメント及びその他の合計」を参照するようになる。
    _HAS_TOTAL_WITH_OTHER = any(
        d == "報告セグメント及びその他の合計" for d in unique_dims
    )
    if _HAS_TOTAL_WITH_OTHER:
        unique_dims = [d for d in unique_dims if d != "報告セグメント"]

    # IFRS: 報告セグメント合計列が存在しない場合は合成列を追加
    synthetic_total_dim = None
    reporting_dims_for_total = []

    if 'IFRS' in sheet_name:
        has_total = any(
            "報告セグメント" in str(d) and "以外" not in str(d)
            for d in unique_dims
        )
        if not has_total:
            ikai_idx = next(
                (i for i, d in enumerate(unique_dims) if "以外" in str(d)),
                None
            )
            if ikai_idx is not None and ikai_idx > 0:
                reporting_dims_for_total = unique_dims[:ikai_idx]
                unique_dims.insert(ikai_idx, "報告セグメント合計")
                synthetic_total_dim = "報告セグメント合計"
            elif ikai_idx is None and unique_dims:
                reporting_dims_for_total = list(unique_dims)
                unique_dims.append("報告セグメント合計")
                synthetic_total_dim = "報告セグメント合計"

    # 日本基準: 「共通」ラベルのdimが存在する場合、列を再構成
    # 列順: [全報告セグメント, 報告セグメント合計, 共通+その他個別, 報告セグメント及びその他の合計, 全社・消去等]
    # 「報告セグメント及びその他の合計」がXBRLになければ合成列として追加する。
    # PPMグラフ対象は「報告セグメント合計」列以左のみ（_ADJUSTMENT_KEYWORDS の「合計」フィルタで除外）。
    synthetic_goukei_dim = None        # 合成「報告セグメント及びその他の合計」dim名
    goukei_source_dims   = []          # synthetic_goukei_dim の合計元dims
    if synthetic_total_dim is None:
        _kyotsu_exists = any(str(d) == '共通' for d in unique_dims)
        if _kyotsu_exists:
            # 集計・調整系を識別するキーワード（これに該当しない個別dimは報告 or 非報告個別）
            _AGG_KW = ['報告セグメント', '全社', '消去', '調整項目', '全体', '連結', '単体']
            # 「その他」系は報告セグメントではなく共通と同グループに置く
            _OTHER_KW = ['その他']
            # 純粋な報告セグメント（集計・調整・共通・その他を除く個別セグメント）
            _report_segs = [
                d for d in unique_dims
                if str(d) != '共通'
                and not any(kw in str(d) for kw in _AGG_KW)
                and not any(kw in str(d) for kw in _OTHER_KW)
            ]
            # 「その他」等の非報告個別セグメント（共通を除く）
            _other_individual = [
                d for d in unique_dims
                if str(d) != '共通'
                and not any(kw in str(d) for kw in _AGG_KW)
                and any(kw in str(d) for kw in _OTHER_KW)
            ]
            # 集計・調整系（「報告セグメント及びその他の合計」含む）
            _agg_dims = [
                d for d in unique_dims
                if str(d) != '共通'
                and any(kw in str(d) for kw in _AGG_KW)
            ]
            # 「報告セグメント及びその他の合計」がXBRLに存在するか確認
            _goukei_existing = next(
                (d for d in _agg_dims if '報告セグメント及びその他' in str(d)), None
            )
            _agg_others = [d for d in _agg_dims if d != _goukei_existing]

            if _report_segs:
                # 非報告個別dims（共通 + その他系）
                _non_report_indiv = ['共通'] + _other_individual
                # 「報告セグメント及びその他の合計」は常に合成値で計算する
                # （XBRLの既存値は共通を含まないケースがあるため上書き）
                if _goukei_existing:
                    # XBRLの既存列を除外し、合成列として作り直す
                    _agg_others = [d for d in _agg_others if d != _goukei_existing]
                # 列順: [報告セグメント群, 報告セグメント合計, 共通+その他, 報告セグメント及びその他の合計, 残集計]
                unique_dims = (_report_segs + ['報告セグメント合計']
                               + _non_report_indiv + ['報告セグメント及びその他の合計'] + _agg_others)
                synthetic_goukei_dim = '報告セグメント及びその他の合計'
                goukei_source_dims   = _report_segs + _non_report_indiv
                reporting_dims_for_total = _report_segs
                synthetic_total_dim = "報告セグメント合計"

    # 報告セグメント合計列のSUM式用: 列位置を事前計算
    # unique_dims 内の synthetic_total_dim の位置 → シート上の列番号 (A=1, B=2, C=3...)
    synthetic_total_col_idx = None   # シート列番号（1-based）
    synthetic_sum_first_col = None   # SUM範囲の先頭列レター
    synthetic_sum_last_col  = None   # SUM範囲の末尾列レター
    if synthetic_total_dim:
        from openpyxl.utils import get_column_letter as _gcl
        tot_pos = unique_dims.index(synthetic_total_dim)   # 0-based in unique_dims
        synthetic_total_col_idx = 3 + tot_pos              # C=3 が unique_dims[0]
        # reporting_dims_for_total は unique_dims の 0 ~ tot_pos-1
        synthetic_sum_first_col = _gcl(3)                  # C
        synthetic_sum_last_col  = _gcl(3 + tot_pos - 1)   # tot_pos 列分の最後
        del _gcl

    # 合成「報告セグメント及びその他の合計」列の列インデックスを事前計算
    synthetic_goukei_col_idx = None  # シート列番号（1-based）
    if synthetic_goukei_dim:
        _gk_pos = unique_dims.index(synthetic_goukei_dim)  # 0-based in unique_dims
        synthetic_goukei_col_idx = 3 + _gk_pos             # C=3 が unique_dims[0]

    # All available years for this role (ascending)
    unique_periods = sorted(list(set(c[2] if len(c) == 3 else c[1] for c in role_columns)))

    debug_log(f"[Segment Analysis] Found {len(unique_dims)} segments and {len(unique_periods)} periods")

    # ヘッダー行を作成
    aws.append(["勘定科目", "年度"] + unique_dims)

    # 1. 有効な年度の特定と、勘定科目データのアグリゲーション（同じラベルの項目をマージ）
    valid_periods = set()
    label_to_data = {}   # display_label -> {period -> {dim -> value}}
    label_info = {}      # display_label -> {full_path, depth, pref_label, el}
    label_order = []     # 登場順を保持

    for full_path_data in ordered_keys:
        full_path, pref_label = full_path_data
        el = full_path.split('::')[-1]
        if '|' in el:
            el = el.split('|')[0]

        # 不要な要素タイプをスキップ
        if el.endswith(("TextBlock", "Abstract", "Axis", "Member", "Table")):
            continue

        parts = el.split('_')
        base_name = parts[-1] if len(parts) > 1 else el

        # ラベルを取得
        if base_name in segment_dict:
            label = segment_dict[base_name]
        elif base_name in common_dict:
            label = common_dict[base_name]
        else:
            label = labels_map.get(el)
            if label:
                label = label.replace(' [メンバー]', '').replace(' [要素]', '').replace(' [区分]', '').strip()

        if not label:
            label = _convert_camel_case_to_title(base_name)

        # ラベルから不要な接尾辞を削除
        display_label = label
        display_label = display_label.replace(' [目次項目]', '').replace(' [タイトル項目]', '')
        display_label = display_label.replace('（IFRS）', '').replace('(IFRS)', '')
        display_label = display_label.replace('、経営指標等', '')
        display_label = display_label.replace('、流動資産', '').replace('、非流動資産', '')
        display_label = display_label.replace('、流動負債', '').replace('、非流動負債', '')
        display_label = display_label.strip()

        depth = len(full_path.split('::')) - 1
        
        if display_label not in label_to_data:
            label_to_data[display_label] = {}
            label_order.append(display_label)
        
        # メタデータを更新（後から出現したものを優先）
        label_info[display_label] = {
            'full_path': full_path,
            'depth': depth,
            'pref_label': pref_label,
            'el': el
        }

        is_sales_item = (display_label == "外部顧客への売上高") or ("売上高" in display_label and "計" in display_label)

        for period in unique_periods:
            for dim in unique_dims:
                found_v = None
                stds_to_check = [current_standard] if current_standard != 'JP_ALL' else ['IFRS', 'JP', 'US', 'JMIS']
                for s in stds_to_check:
                    v = all_years_data[role][full_path].get((s, dim, period))
                    if v is not None:
                        found_v = v
                        break
                
                if found_v is not None and found_v != "":
                    if period not in label_to_data[display_label]:
                        label_to_data[display_label][period] = {}
                    
                    # データをマージ（後から出現したパスのデータで上書き）
                    label_to_data[display_label][period][dim] = found_v
                    
                    # 数値チェックと売上高判定による有効年度追加
                    val_clean = unicodedata.normalize('NFKC', str(found_v)).replace(',', '').strip()
                    try:
                        if val_clean and not any(c.isalpha() for c in val_clean):
                            if is_sales_item:
                                valid_periods.add(period)
                    except:
                        pass

    # 全期間通して一度もデータ（数値・非数値問わず）がない項目を削除
    final_label_order = [l for l in label_order if label_to_data[l]]

    # 有効年度を確定（売上高がない場合は全年度を使用）
    sorted_valid_periods = sorted(list(valid_periods))
    if not sorted_valid_periods:
        sorted_valid_periods = unique_periods

    debug_log(f"[Segment Analysis] Aligned to {len(sorted_valid_periods)} valid periods by merging duplicate labels")

    # セグメント情報等シートのみインデントなし（他のシートはdepthベースのインデントを保持）
    _no_indent = 'セグメント情報等' in sheet_name

    # 3. データ行を作成
    seen_rows_analysis = set()
    for d_label in final_label_order:
        info = label_info[d_label]
        it_depth = info['depth']
        it_el = info['el']
        it_pref_label = info['pref_label']
        it_full_path = info['full_path'] # CF判定用
        it_fp_data = (it_full_path, it_pref_label)
        indent_prefix = "" if _no_indent else "　" * it_depth

        for period in sorted_valid_periods:
            row_data_analysis = [indent_prefix + d_label, period]
            
            # マージ済みのデータを取得
            period_data = label_to_data[d_label].get(period, {})

            for dim in unique_dims:
                if dim == synthetic_total_dim or dim == synthetic_goukei_dim:
                    # 合成合計列: 対象dims の合計を計算
                    _src_dims = reporting_dims_for_total if dim == synthetic_total_dim else goukei_source_dims
                    total_val = 0.0
                    has_any_val = False
                    for rd in _src_dims:
                        rv = period_data.get(rd, "")
                        if rv:
                            rv_clean = unicodedata.normalize('NFKC', str(rv)).replace(',', '').strip()
                            try:
                                if rv_clean and not any(c.isalpha() for c in rv_clean):
                                    total_val += float(rv_clean)
                                    has_any_val = True
                            except:
                                pass
                    val = total_val if has_any_val else ""
                else:
                    val = period_data.get(dim, "")

                    if val:
                        val_clean = unicodedata.normalize('NFKC', str(val)).replace(',', '').strip()
                        try:
                            if val_clean and not any(c.isalpha() for c in val_clean):
                                val = float(val_clean)
                        except:
                            pass

                row_data_analysis.append(val)

            # 数値データが1つもない行（テキストのみ）はスキップ
            has_numeric = any(isinstance(v, (int, float)) for v in row_data_analysis[2:])
            if not has_numeric:
                continue

            # 重複チェックはラベルと年度の組み合わせで行う
            row_key = (d_label, period)
            if row_key in seen_rows_analysis:
                continue
            seen_rows_analysis.add(row_key)

            from openpyxl.utils import get_column_letter as _gcl2
            next_row = aws.max_row + 1

            # 報告セグメント合計列をSUM式で上書き
            if synthetic_total_dim and synthetic_sum_first_col and synthetic_sum_last_col:
                row_data_analysis[synthetic_total_col_idx - 1] = (
                    f"=SUM({synthetic_sum_first_col}{next_row}:{synthetic_sum_last_col}{next_row})"
                )

            # 報告セグメント及びその他の合計列をSUM式で上書き（合成列の場合）
            if synthetic_goukei_dim and synthetic_goukei_col_idx:
                # goukei_source_dims のシート列を特定してSUM式を生成
                _gk_cols = [
                    _gcl2(3 + unique_dims.index(d))
                    for d in goukei_source_dims
                    if d in unique_dims
                ]
                if _gk_cols:
                    _gk_formula = "=SUM(" + ",".join(f"{cl}{next_row}" for cl in _gk_cols) + ")"
                    row_data_analysis[synthetic_goukei_col_idx - 1] = _gk_formula

            del _gcl2
            aws.append(row_data_analysis)

        # キャッシュ・フロー計算書の場合、特定の要素で停止
        if 'キャッシュ・フロー' in sheet_name and 'CashAndCashEquivalents' in it_el:
            if it_pref_label and it_pref_label.endswith(('periodEndLabel', 'totalLabel')):
                # Check if this is at natural end of hierarchy
                try:
                    current_idx = ordered_keys.index(it_fp_data)
                    if current_idx >= len(ordered_keys) - 1:
                        break
                except ValueError:
                    # Should not happen if it's in final_label_order
                    pass

                # Check if there are more CashAndCash items ahead
                has_more_cash_items = False
                for next_idx in range(current_idx + 1, len(ordered_keys)):
                    next_fp, _ = ordered_keys[next_idx]
                    next_el_name = next_fp.split('::')[-1]
                    if '|' in next_el_name:
                        next_el_name = next_el_name.split('|')[0]
                    if next_el_name.endswith(("Abstract", "TextBlock", "Table", "Axis", "Member")):
                        continue
                    # If we find another CashAndCash item, don't break yet
                    if 'CashAndCashEquivalents' in next_el_name:
                        has_more_cash_items = True
                        break
                    break

                if has_more_cash_items:
                    continue  # Don't break, process the next CashAndCash item

                # Check depth of next items if no more CashAndCash items
                current_depth = len(full_path.split('::')) - 1
                is_at_end = True
                for next_idx in range(current_idx + 1, len(ordered_keys)):
                    next_fp, _ = ordered_keys[next_idx]
                    next_el_name = next_fp.split('::')[-1]
                    if '|' in next_el_name:
                        next_el_name = next_el_name.split('|')[0]
                    if next_el_name.endswith(("Abstract", "TextBlock", "Table", "Axis", "Member")):
                        continue
                    next_depth = len(next_fp.split('::')) - 1
                    if next_depth > current_depth:
                        is_at_end = False
                    break
                if is_at_end:
                    break

    # プレゼンテーションツリーに含まれない追加項目をシート末尾に追加
    # (研究開発費、設備投資額、従業員数など)
    _EXTRA_SEGMENT_ELEMENTS = [
        ('jpcrp_cor_ResearchAndDevelopmentExpensesResearchAndDevelopmentActivities', '研究開発費'),
        ('jpcrp_cor_CapitalExpendituresOverviewOfCapitalExpendituresEtc', '設備投資額'),
        ('jpcrp_cor_NumberOfEmployees', '従業員数'),
        ('jpcrp_cor_AverageNumberOfTemporaryWorkers', '平均臨時雇用人員'),
    ]
    fact_std = current_standard if current_standard not in ('JP_ALL',) else 'JP'

    # 「報告セグメント及びその他の合計」列の位置を特定（構成比の分母として使う）
    _TOTAL_SUFFIXES = ('合計', '全体', '全社', '消去', '調整', '連結財務諸表')
    _reporting_total_dim = next(
        (d for d in unique_dims if '報告セグメント' in str(d) and '以外' not in str(d) and '合計' in str(d)),
        None
    )
    # 合計列を構成する個別セグメント次元（合計・全体・全社・消去・調整を除く）
    _summable_dims = [
        d for d in unique_dims
        if d != synthetic_total_dim and not any(s in str(d) for s in _TOTAL_SUFFIXES)
    ]

    for extra_el, extra_label in _EXTRA_SEGMENT_ELEMENTS:
        extra_vals = global_element_period_values.get(extra_el, {})
        if not extra_vals:
            continue

        # セグメント次元にデータがあるか確認
        has_segment_data = False
        for period in sorted_valid_periods:
            for dim in unique_dims:
                if dim == synthetic_total_dim:
                    continue
                if extra_vals.get((fact_std, dim, period)) is not None:
                    has_segment_data = True
                    break
            if has_segment_data:
                break
        if not has_segment_data:
            continue

        debug_log(f"[Segment Analysis] Appending '{extra_label}' rows")
        for period in sorted_valid_periods:
            row_data = [extra_label, period]
            # 各dim列の値を収集（後で合計列の補完に使う）
            dim_values = {}
            for dim in unique_dims:
                if dim == synthetic_total_dim:
                    total = 0.0
                    has_any = False
                    for rd in reporting_dims_for_total:
                        v = extra_vals.get((fact_std, rd, period))
                        if v is not None:
                            vc = unicodedata.normalize('NFKC', str(v)).replace(',', '').strip()
                            try:
                                total += float(vc)
                                has_any = True
                            except ValueError:
                                pass
                    dim_values[dim] = total if has_any else ""
                else:
                    v = extra_vals.get((fact_std, dim, period))
                    if v is not None:
                        vc = unicodedata.normalize('NFKC', str(v)).replace(',', '').strip()
                        try:
                            dim_values[dim] = float(vc)
                        except ValueError:
                            dim_values[dim] = v
                    else:
                        dim_values[dim] = ""

            # 「報告セグメント及びその他の合計」列が空の場合、個別セグメントの合計で補完
            if _reporting_total_dim and dim_values.get(_reporting_total_dim) == "":
                seg_sum = 0.0
                has_any_seg = False
                for sd in _summable_dims:
                    sv = dim_values.get(sd, "")
                    if isinstance(sv, (int, float)):
                        seg_sum += sv
                        has_any_seg = True
                if has_any_seg:
                    dim_values[_reporting_total_dim] = seg_sum

            for dim in unique_dims:
                row_data.append(dim_values[dim])
            aws.append(row_data)

    # セル書式を適用
    for row in aws.iter_rows(min_row=2, max_row=aws.max_row, min_col=3, max_col=aws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = r'#,##0_ ;[Red]\-#,##0 '

    # ウィンドウ枠の固定 (B2で固定: A列と1行目を固定)
    aws.freeze_panes = 'B2'

    # 列幅の設定
    # A列: 31
    aws.column_dimensions['A'].width = 31
    # B列以降: 12
    from openpyxl.utils import get_column_letter
    for col_idx in range(2, aws.max_column + 1):
        aws.column_dimensions[get_column_letter(col_idx)].width = 12

    debug_log(f"[Segment Analysis] Completed analysis sheet: {analysis_sheet_name} with {aws.max_row - 1} data rows")


def _append_rd_to_notes_sheet(workbook, sheet_name, sorted_role_cols,
                              global_element_period_values, debug_log):
    """元の注記シートの末尾にセグメント別に追加項目（研究開発費等）の行を追加する"""
    if sheet_name not in workbook.sheetnames:
        return

    _EXTRA_SEGMENT_ELEMENTS = [
        ('jpcrp_cor_ResearchAndDevelopmentExpensesResearchAndDevelopmentActivities', '研究開発費'),
        ('jpcrp_cor_CapitalExpendituresOverviewOfCapitalExpendituresEtc', '設備投資額'),
        ('jpcrp_cor_NumberOfEmployees', '従業員数'),
        ('jpcrp_cor_AverageNumberOfTemporaryWorkers', '平均臨時雇用人員'),
    ]

    non_segment_dims = ('全体', '単体', '連結', '全社', '連結財務諸表計上額')
    ws = workbook[sheet_name]
    appended = False

    for extra_el, extra_label in _EXTRA_SEGMENT_ELEMENTS:
        extra_vals = global_element_period_values.get(extra_el, {})
        if not extra_vals:
            continue

        has_segment_data = any(
            extra_vals.get(col_key) is not None
            for col_key in sorted_role_cols
            if len(col_key) == 3 and col_key[1] not in non_segment_dims
        )
        if not has_segment_data:
            continue

        debug_log(f"[Segment Analysis] Appending '{extra_label}' row to notes sheet: {sheet_name}")

        row_data = [extra_label, extra_el]
        for col_key in sorted_role_cols:
            v = extra_vals.get(col_key)
            if v is not None:
                vc = unicodedata.normalize('NFKC', str(v)).replace(',', '').strip()
                try:
                    row_data.append(float(vc))
                except ValueError:
                    row_data.append(v)
            else:
                row_data.append("")

        ws.append(row_data)
        appended = True

    if not appended:
        return

    # 追加した行に書式を適用
    last_row = ws.max_row
    start_row = last_row - sum(
        1 for el, _ in _EXTRA_SEGMENT_ELEMENTS
        if global_element_period_values.get(el) and any(
            global_element_period_values[el].get(c) is not None
            for c in sorted_role_cols
            if len(c) == 3 and c[1] not in non_segment_dims
        )
    ) + 1
    for r in range(start_row, last_row + 1):
        for cell in ws[r][2:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = r'#,##0_ ;[Red]\-#,##0 '


def _convert_camel_case_to_title(name):
    """
    キャメルケースをタイトルケースに変換

    Args:
        name: 変換する文字列

    Returns:
        変換後の文字列
    """
    import re
    # Insert space before uppercase letters
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1 \2', name)
    # Insert space before uppercase letters that follow lowercase letters or numbers
    return re.sub('([a-z0-9])([A-Z])', r'\1 \2', s1)


# ============================================================================
# 共有ヘルパー関数
# （_create_data_acquisition_sheet / _create_ppm_analysis_sheet /
#   _create_ppm_analysis_sheet_ifrs で共通利用）
# ============================================================================

def _to_period_str(pv):
    """period セルの値を 'YYYY-MM-DD' 文字列に正規化する。"""
    if pv is None:
        return None
    if hasattr(pv, 'strftime'):
        return pv.strftime('%Y-%m-%d')
    return str(pv).strip()


def _read_numeric(ws, row, col):
    """セルの実数値を返す。SUM式セルは個別列を合計して代替する。"""
    if row is None:
        return None
    val = ws.cell(row, col).value
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str) and val.startswith('=SUM('):
        total = 0.0
        has_val = False
        for c in range(3, col):
            v = ws.cell(row, c).value
            if isinstance(v, (int, float)):
                total += v
                has_val = True
        return total if has_val else None
    return None


def _build_period_lookup(analysis_ws):
    """analysis_ws から (label, period_str) -> 行番号 の辞書と
    登場順ラベルリストを構築して返す。

    Returns:
        (period_lookup, unique_labels_ordered)
    """
    unique_labels_ordered = []
    period_lookup = {}
    for r in range(2, analysis_ws.max_row + 1):
        label_val = analysis_ws.cell(r, 1).value
        period_val = analysis_ws.cell(r, 2).value
        if not label_val:
            continue
        norm_label = str(label_val).strip()
        if not unique_labels_ordered or unique_labels_ordered[-1] != norm_label:
            unique_labels_ordered.append(norm_label)
        ps = _to_period_str(period_val)
        if ps:
            period_lookup[(norm_label, ps)] = r
    return period_lookup, unique_labels_ordered


def _build_col_info(analysis_ws, max_col):
    """analysis_ws のヘッダー行からセグメント列情報を構築して返す。

    Returns:
        (col_to_dim, hokoku_col, igai_col, goukei_col,
         hokoku_is_synthesized, goukei_is_synthesized)
    """
    col_to_dim = {}
    for c in range(3, max_col + 1):
        hv = analysis_ws.cell(1, c).value
        if hv is not None:
            col_to_dim[c] = str(hv)

    hokoku_col = next((c for c, d in col_to_dim.items() if d == '報告セグメント合計'), None)
    igai_col   = next((c for c, d in col_to_dim.items()
                       if '以外' in d and '報告セグメント' in d), None)
    goukei_col = next((c for c, d in col_to_dim.items()
                       if '報告セグメント及びその他の合計' in d), None)
    hokoku_is_synthesized = (col_to_dim.get(hokoku_col, '') == '報告セグメント合計')
    goukei_is_synthesized = (goukei_col is not None and
                              col_to_dim.get(goukei_col, '') == '報告セグメント及びその他の合計')
    return col_to_dim, hokoku_col, igai_col, goukei_col, hokoku_is_synthesized, goukei_is_synthesized


def _make_get_val_for_filing(analysis_ws, col_to_dim,
                              hokoku_col, igai_col, goukei_col,
                              hokoku_is_synthesized, goukei_is_synthesized):
    """_get_val_for_filing クロージャを生成して返す。

    戻り値の関数シグネチャ: (src_row, c, fp, is_current) -> float | None
    analysis_ws 列 c の値を filing_pair の dims でフィルタして返す。
    """
    def _get_val_for_filing(src_row, c, fp, is_current):
        if src_row is None:
            return None
        allowed = fp.get('current_dims', set()) if is_current else fp.get('prior_dims', set())

        if c == hokoku_col and hokoku_is_synthesized:
            total = 0.0
            has_val = False
            for cc in range(3, hokoku_col):
                if cc in (igai_col, goukei_col):
                    continue
                dim = col_to_dim.get(cc, '')
                if allowed and dim not in allowed:
                    continue
                v = _read_numeric(analysis_ws, src_row, cc)
                if v is not None:
                    total += v
                    has_val = True
            return total if has_val else None

        if c == goukei_col and goukei_is_synthesized:
            # 報告セグメント合計（hokoku）＋ hokoku〜goukei 間の個別列（共通・その他）
            # dims フィルタを適用して合算する
            v_h = (_get_val_for_filing(src_row, hokoku_col, fp, is_current)
                   if (hokoku_col and hokoku_col != goukei_col) else None)
            # hokoku_col と goukei_col の間にある非集計個別列を合算
            v_between = 0.0
            has_between = False
            if hokoku_col and goukei_col:
                for _bc in range(hokoku_col + 1, goukei_col):
                    if _bc == igai_col:
                        continue
                    _bdim = col_to_dim.get(_bc, '')
                    if allowed and _bdim not in allowed:
                        continue
                    _bv = _read_numeric(analysis_ws, src_row, _bc)
                    if _bv is not None:
                        v_between += _bv
                        has_between = True
            # igai_col も加算（従来通り）
            v_i = None
            if igai_col:
                igai_dim = col_to_dim.get(igai_col, '')
                if not allowed or igai_dim in allowed:
                    v_i = _read_numeric(analysis_ws, src_row, igai_col)
            if v_h is None and not has_between and v_i is None:
                return None
            return (v_h or 0.0) + v_between + (v_i or 0.0)

        dim = col_to_dim.get(c, '')
        if allowed and dim not in allowed:
            return None
        return _read_numeric(analysis_ws, src_row, c)

    return _get_val_for_filing


def _create_data_acquisition_sheet(workbook, analysis_sheet_name, used_sheet_names,
                                   filing_pairs, debug_log):
    """
    「データ取得」シートを作成（内部関数）

    分析シートに含まれる全勘定科目を filing_pairs ベースで縦持ちに展開する。
    勘定科目の並び順は分析シートと同じ順序を保持する。

    列構造: 勘定科目 | 報告年度 | 前期・当期 | 会計年度 | セグメント列...
    各報告年度につき前期・当期の2行を出力する。

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        filing_pairs: [{current: 'YYYY-MM-DD', prior: 'YYYY-MM-DD'|None,
                        current_dims: set, prior_dims: set}, ...] 昇順
        debug_log: デバッグログ関数
    """
    from openpyxl.utils import get_column_letter

    # --- シート名を生成 ---
    acq_sheet_name = analysis_sheet_name.replace("_分析", "_データ取得")
    if len(acq_sheet_name) > 31:
        acq_sheet_name = analysis_sheet_name[:22] + "_データ取得"

    if acq_sheet_name in workbook.sheetnames:
        debug_log(f"[Data Acquisition] Sheet '{acq_sheet_name}' already exists, skipping")
        return

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[Data Acquisition] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    if not filing_pairs:
        debug_log("[Data Acquisition] No filing_pairs provided, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    max_col = analysis_ws.max_column

    # -----------------------------------------------------------------------
    # 1. セグメント列情報・dims フィルタ付き読み取り関数を構築
    # -----------------------------------------------------------------------
    seg_headers = [analysis_ws.cell(1, c).value or "" for c in range(3, max_col + 1)]
    col_to_dim, hokoku_col, igai_col, goukei_col, hokoku_is_synthesized, goukei_is_synthesized = \
        _build_col_info(analysis_ws, max_col)
    _get_val_for_filing = _make_get_val_for_filing(
        analysis_ws, col_to_dim, hokoku_col, igai_col, goukei_col,
        hokoku_is_synthesized, goukei_is_synthesized)

    # -----------------------------------------------------------------------
    # 2. period_lookup と unique_labels_ordered を構築
    # -----------------------------------------------------------------------
    period_lookup, unique_labels_ordered = _build_period_lookup(analysis_ws)

    if not unique_labels_ordered:
        debug_log("[Data Acquisition] Analysis sheet has no data rows, skipping")
        return

    # -----------------------------------------------------------------------
    # 2b. ラベルグループ化
    #     片方が他方のプレフィックスである場合を同一概念の変形とみなす。
    #     例: 「事業利益」と「事業利益（△は損失）」→ canonical =「事業利益」
    # -----------------------------------------------------------------------
    canonical_to_aliases = {}  # canonical_label -> [alias, ...]  (canonical含む)
    label_to_canonical = {}    # 全ラベル -> canonical

    for lbl in unique_labels_ordered:
        if lbl in label_to_canonical:
            continue
        matched_canonical = None
        for canon in list(canonical_to_aliases.keys()):
            if lbl.startswith(canon) or canon.startswith(lbl):
                matched_canonical = canon
                break
        if matched_canonical:
            canonical_to_aliases[matched_canonical].append(lbl)
            label_to_canonical[lbl] = matched_canonical
        else:
            canonical_to_aliases[lbl] = [lbl]
            label_to_canonical[lbl] = lbl

    unique_canonicals_ordered = []
    _seen_canon = set()
    for lbl in unique_labels_ordered:
        canon = label_to_canonical[lbl]
        if canon not in _seen_canon:
            unique_canonicals_ordered.append(canon)
            _seen_canon.add(canon)

    merged = {c: aliases for c, aliases in canonical_to_aliases.items() if len(aliases) > 1}
    if merged:
        debug_log(f"[Data Acquisition] Merged label groups: {merged}")

    # -----------------------------------------------------------------------
    # 3. シートを作成してヘッダー行を出力
    # -----------------------------------------------------------------------
    acq_ws = workbook.create_sheet(title=acq_sheet_name)
    used_sheet_names.add(acq_sheet_name)
    debug_log(f"[Data Acquisition] Creating sheet: {acq_sheet_name}")

    acq_ws.append(["勘定科目", "報告年度", "前期・当期", "会計年度"] + seg_headers)

    # -----------------------------------------------------------------------
    # 4. 各勘定科目 × 各 filing_pair につき前期・当期の2行を出力
    # -----------------------------------------------------------------------
    row_count = 0
    for label in unique_canonicals_ordered:
        aliases = canonical_to_aliases.get(label, [label])
        for fp in filing_pairs:
            cur_p = _to_period_str(fp.get('current'))
            pri_p = _to_period_str(fp.get('prior'))
            if not cur_p:
                continue

            for flag_label, period_str in [("前期", pri_p), ("当期", cur_p)]:
                row_data = [label, cur_p, flag_label, period_str if period_str else ""]
                is_current = (flag_label == "当期")

                # エイリアス全候補からデータがある行を探す
                src_row = None
                if period_str:
                    for alias in aliases:
                        r = period_lookup.get((alias, period_str))
                        if r is not None:
                            src_row = r
                            break

                for c in range(3, max_col + 1):
                    if src_row is None:
                        row_data.append("")
                    else:
                        v = _get_val_for_filing(src_row, c, fp, is_current)
                        row_data.append(v if v is not None else "")

                acq_ws.append(row_data)
                row_count += 1

        # 勘定科目間に空行を挟む
        acq_ws.append([])

    # -----------------------------------------------------------------------
    # 5. 書式・列幅・ウィンドウ枠
    # -----------------------------------------------------------------------
    for row in acq_ws.iter_rows(min_row=2, max_row=acq_ws.max_row, min_col=5,
                                 max_col=acq_ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = r'#,##0_ ;[Red]\-#,##0 '

    acq_ws.freeze_panes = 'E2'

    acq_ws.column_dimensions['A'].width = 31
    acq_ws.column_dimensions['B'].width = 12
    acq_ws.column_dimensions['C'].width = 8
    acq_ws.column_dimensions['D'].width = 12
    for col_idx in range(5, acq_ws.max_column + 1):
        acq_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    debug_log(f"[Data Acquisition] Completed sheet: {acq_sheet_name} with {row_count} data rows")


def _create_ppm_analysis_sheet(workbook, analysis_sheet_name, used_sheet_names, filing_pairs, debug_log):
    """
    PPM分析シートを作成（内部関数）

    各有報の当期・前期ペアを使ってPPM分析用シートを生成する。
    列構造: 勘定科目 | 報告年度 | 前期・当期 | 会計年度 | セグメント列...
    各報告年度につき前期・当期の2行を出力し、同一有報内のペアで成長率を計算する。

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        filing_pairs: [{current: 'YYYY-MM-DD', prior: 'YYYY-MM-DD'|None}, ...] 昇順
        debug_log: デバッグログ関数
    """
    import re
    import math
    import datetime
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BubbleChart, Reference, Series

    # --- PPM分析シート名を生成 ---
    ppm_sheet_name = analysis_sheet_name + "_PPM分析用"
    if len(ppm_sheet_name) > 31:
        ppm_sheet_name = analysis_sheet_name[:18] + "_PPM分析用"

    debug_log(f"[PPM Analysis] Creating PPM analysis sheet: {ppm_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[PPM Analysis] Analysis sheet '{analysis_sheet_name}' not found, skipping PPM sheet")
        return

    if not filing_pairs:
        debug_log("[PPM Analysis] No filing_pairs provided, skipping PPM sheet")
        return

    analysis_ws = workbook[analysis_sheet_name]
    ppm_ws = workbook.create_sheet(title=ppm_sheet_name)
    used_sheet_names.add(ppm_sheet_name)

    max_col = analysis_ws.max_column

    # -----------------------------------------------------------------------
    # 1. analysis_ws を走査して (ラベル, period_str) -> 行番号 のルックアップを構築
    # -----------------------------------------------------------------------
    period_lookup, unique_labels_ordered = _build_period_lookup(analysis_ws)

    # -----------------------------------------------------------------------
    # 2. 売上・利益ラベルを検出
    #    売上: 「計」
    #    利益: 「利益」「損失」「純益」を含むラベルの全候補
    # -----------------------------------------------------------------------
    target_sales_label = None
    profit_label_candidates = []
    for label in unique_labels_ordered:
        if target_sales_label is None and label == "計":
            target_sales_label = label
        if "利益" in label or "損失" in label or "純益" in label:
            profit_label_candidates.append(label)

    # Financial industry detection: 連結粗利益 / 業務粗利益 → sales label
    #                                連結業務純益 / 業務純益  → profit label (top priority)
    _financial_sales_kws  = ["連結粗利益", "業務粗利益"]
    _financial_profit_kws = ["連結業務純益", "業務純益"]
    _financial_sales_label  = None
    _financial_profit_label = None
    for _kw in _financial_sales_kws:
        for _lbl in unique_labels_ordered:
            if _lbl == _kw or _kw in _lbl:
                _financial_sales_label = _lbl
                break
        if _financial_sales_label:
            break
    if _financial_sales_label:
        for _kw in _financial_profit_kws:
            for _lbl in unique_labels_ordered:
                if _lbl == _kw or _kw in _lbl:
                    _financial_profit_label = _lbl
                    break
            if _financial_profit_label:
                break
        target_sales_label = _financial_sales_label
        # Prepend financial profit label so it takes top priority
        if _financial_profit_label and _financial_profit_label not in profit_label_candidates:
            profit_label_candidates.insert(0, _financial_profit_label)
        elif _financial_profit_label:
            profit_label_candidates.remove(_financial_profit_label)
            profit_label_candidates.insert(0, _financial_profit_label)
        debug_log(f"[PPM Analysis] Financial industry detected: sales='{target_sales_label}', profit='{_financial_profit_label}'")

    # 「計」がない場合（かつ金融業でもない場合）のフォールバック
    if target_sales_label is None:
        _sales_fallback_keywords = ["外部顧客への売上収益", "売上収益", "営業収益", "売上高"]
        for _kw in _sales_fallback_keywords:
            for _lbl in unique_labels_ordered:
                if _kw in _lbl:
                    target_sales_label = _lbl
                    debug_log(f"[PPM Analysis] Sales label fallback: '{target_sales_label}' (matched keyword '{_kw}')")
                    break
            if target_sales_label is not None:
                break

    # セグメント利益を営業利益より優先するよう並び替え（金融業判定済み後）
    # 優先順: セグメント利益含むもの → その他利益/損失ラベル
    def _profit_sort_key(lbl):
        if lbl == profit_label_candidates[0] if profit_label_candidates else False:
            return 0   # already-pinned top candidate keeps its position
        if "セグメント利益" in lbl:
            return 1
        return 2
    # Only re-sort non-pinned entries when no financial label was pinned
    if not _financial_profit_label:
        profit_label_candidates.sort(key=lambda lbl: (0 if "セグメント利益" in lbl else 1))

    # Derive display labels
    # Sales: replace "計" with "売上"; all other labels shown as-is
    sales_display_label = "　売上" if target_sales_label == "計" else ("　" + target_sales_label if target_sales_label else "　売上")

    debug_log(f"[PPM Analysis] Sales label='{target_sales_label}', Profit candidates={profit_label_candidates}")

    # -----------------------------------------------------------------------
    # 3b. 列位置の検出（報告セグメント / 以外 / 及びその他の合計）
    # -----------------------------------------------------------------------
    hokoku_col = None
    igai_col   = None
    goukei_col = None
    for _ci in range(3, max_col + 1):
        _hv = analysis_ws.cell(1, _ci).value
        if not _hv:
            continue
        _hv_str = str(_hv)
        if "報告セグメント" not in _hv_str:
            continue
        if "以外" in _hv_str:
            igai_col = _ci
        elif "及びその他" in _hv_str:
            goukei_col = _ci
        elif hokoku_col is None:
            hokoku_col = _ci

    if hokoku_col is None:
        if igai_col and igai_col > 3:
            _sum_start_letter = get_column_letter(3)
            _sum_end_letter   = get_column_letter(igai_col - 1)
            analysis_ws.insert_cols(igai_col)
            analysis_ws.cell(1, igai_col).value = "報告セグメント合計"
            for _ri in range(2, analysis_ws.max_row + 1):
                if any(isinstance(analysis_ws.cell(_ri, c).value, (int, float))
                       for c in range(3, igai_col)):
                    analysis_ws.cell(_ri, igai_col).value = (
                        f"=SUM({_sum_start_letter}{_ri}:{_sum_end_letter}{_ri})"
                    )
            hokoku_col = igai_col
            igai_col  += 1
            if goukei_col is not None:
                goukei_col += 1
            max_col += 1
            debug_log(f"[PPM Analysis] Inserted '報告セグメント合計' column at col {hokoku_col}")
        else:
            hokoku_col = max_col

    if igai_col and goukei_col is None:
        new_col = max_col + 1
        analysis_ws.cell(1, new_col).value = "報告セグメント及びその他の合計"
        _h_letter = get_column_letter(hokoku_col)
        _i_letter = get_column_letter(igai_col)
        for _ri in range(2, analysis_ws.max_row + 1):
            if any(isinstance(analysis_ws.cell(_ri, c).value, (int, float))
                   for c in range(3, max_col + 1)):
                analysis_ws.cell(_ri, new_col).value = (
                    f"=SUM({_h_letter}{_ri},{_i_letter}{_ri})"
                )
        goukei_col = new_col
        max_col = new_col
        debug_log(f"[PPM Analysis] Added '報告セグメント及びその他の合計' column at col {new_col}")

    chart_end_col = goukei_col or hokoku_col

    # -----------------------------------------------------------------------
    # 4. valid_pairs の構築（analysis_ws にデータがある報告年度のみ、最新11件）
    # -----------------------------------------------------------------------
    # analysis_ws のセグメント列（col 3以降）に実数値があるか確認するヘルパー
    def _has_data_row(row):
        if row is None:
            return False
        for c in range(3, max_col + 1):
            v = analysis_ws.cell(row, c).value
            if isinstance(v, (int, float)):
                return True
        return False

    def _get_profit_src(period_str):
        for candidate in profit_label_candidates:
            row = period_lookup.get((candidate, period_str))
            if row is not None and _has_data_row(row):
                return candidate, row
        return None, None

    valid_pairs = []
    for fp in filing_pairs:
        cur_p = _to_period_str(fp.get('current'))
        pri_p = _to_period_str(fp.get('prior'))
        if not cur_p:
            continue
        cur_s = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
        _, cur_p_src = _get_profit_src(cur_p)
        if _has_data_row(cur_s) or _has_data_row(cur_p_src):
            valid_pairs.append({
                'current': cur_p,
                'prior': pri_p,
                'current_dims': fp.get('current_dims', set()),
                'prior_dims':   fp.get('prior_dims',   set()),
            })

    # 報告年度（current）で昇順ソート
    valid_pairs.sort(key=lambda fp: fp['current'] or '')

    MAX_FILINGS = 11
    if len(valid_pairs) > MAX_FILINGS:
        valid_pairs = valid_pairs[-MAX_FILINGS:]
    N = len(valid_pairs)

    if N == 0:
        debug_log("[PPM Analysis] No valid filing pairs found, skipping PPM sheet")
        workbook.remove(ppm_ws)
        used_sheet_names.discard(ppm_sheet_name)
        return

    debug_log(f"[PPM Analysis] Building PPM sheet for {N} filing years")

    # -----------------------------------------------------------------------
    # SEG_OFFSET: ppm_ws のセグメント列は analysis_ws 列 +2 にオフセット
    #   analysis_ws col 3 → ppm_ws col 5 (E)
    # -----------------------------------------------------------------------
    SEG_OFFSET = 2
    ppm_max_col = max_col + SEG_OFFSET   # ppm_ws の最大列

    # -----------------------------------------------------------------------
    # col_to_dim / _get_val_for_filing（dims フィルタ付き読み取り）
    # -----------------------------------------------------------------------
    # hokoku_col / igai_col / goukei_col は 3b で確定済みのため _build_col_info で
    # 上書きしない。col_to_dim と is_synthesized フラグだけを構築する。
    col_to_dim = {c: str(analysis_ws.cell(1, c).value)
                  for c in range(3, max_col + 1)
                  if analysis_ws.cell(1, c).value is not None}
    hokoku_is_synthesized = (col_to_dim.get(hokoku_col, '') == '報告セグメント合計')
    goukei_is_synthesized = (goukei_col is not None and
                              col_to_dim.get(goukei_col, '') == '報告セグメント及びその他の合計')
    _get_val_for_filing = _make_get_val_for_filing(
        analysis_ws, col_to_dim, hokoku_col, igai_col, goukei_col,
        hokoku_is_synthesized, goukei_is_synthesized)

    # -----------------------------------------------------------------------
    # 5. ヘッダー行（1行）
    #    A=勘定科目, B=報告年度, C=前期・当期, D=会計年度, E+=セグメント名
    # -----------------------------------------------------------------------
    header = ["勘定科目", "報告年度", "前期・当期", "会計年度"]
    for c in range(3, max_col + 1):
        hv = analysis_ws.cell(1, c).value
        header.append(hv if hv is not None else "")
    ppm_ws.append(header)

    # -----------------------------------------------------------------------
    # 6. 売上セクション（2*N 行）
    # -----------------------------------------------------------------------
    sales_start_row = ppm_ws.max_row + 1

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        pri_p = fp['prior']
        pri_src = period_lookup.get((target_sales_label, pri_p)) if (target_sales_label and pri_p) else None
        cur_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None

        sales_lbl = "　" + (target_sales_label or "売上")
        # 前期 row
        pri_row = [sales_lbl, cur_p, "前期", pri_p if pri_p else ""]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(pri_src, c, fp, False)
            pri_row.append(v if v is not None else "")
        ppm_ws.append(pri_row)

        # 当期 row
        cur_row = [sales_lbl, cur_p, "当期", cur_p]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(cur_src, c, fp, True)
            cur_row.append(v if v is not None else "")
        ppm_ws.append(cur_row)

    # 空行
    ppm_ws.append([])

    # -----------------------------------------------------------------------
    # 7. セグメント利益セクション（2*N 行）
    # -----------------------------------------------------------------------
    profit_start_row = ppm_ws.max_row + 1

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        pri_p = fp.get('prior')
        pri_profit_lbl, pri_src_default = _get_profit_src(pri_p) if pri_p else (None, None)
        cur_profit_lbl, cur_src = _get_profit_src(cur_p)

        # 前期・当期で同一勘定科目ラベルを使用する
        consistent_lbl = cur_profit_lbl or pri_profit_lbl or "セグメント利益"
        if pri_p and cur_profit_lbl:
            pri_src_consistent = period_lookup.get((cur_profit_lbl, pri_p))
            pri_src = pri_src_consistent if (pri_src_consistent is not None and _has_data_row(pri_src_consistent)) else pri_src_default
        else:
            pri_src = pri_src_default

        # 前期 row
        pri_row = ["　" + consistent_lbl, cur_p, "前期", pri_p if pri_p else ""]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(pri_src, c, fp, False)
            pri_row.append(v if v is not None else "")
        ppm_ws.append(pri_row)

        # 当期 row
        cur_row = ["　" + consistent_lbl, cur_p, "当期", cur_p]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(cur_src, c, fp, True)
            cur_row.append(v if v is not None else "")
        ppm_ws.append(cur_row)

    profit_end_row = ppm_ws.max_row

    # 空行
    ppm_ws.append([])

    # -----------------------------------------------------------------------
    # 8. 売上高対前年増加率（N+1 行）
    #    行0: 最古の前期（ベース年、数式なし）
    #    行1..N: 各報告年度（当期/前期 の比）
    # -----------------------------------------------------------------------
    growth_start_row = ppm_ws.max_row + 1
    # growth_rates[i]: dict {analysis_col -> float}  i=0 はベース年（空）
    growth_rates = [{}]   # ベース年分として空dict

    # ベース年行
    base_prior = valid_pairs[0]['prior']
    base_row = ["売上高対前年増加率", base_prior if base_prior else "", "", base_prior if base_prior else ""]
    base_row += [""] * (max_col - 2)
    ppm_ws.append(base_row)

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        # ppm_ws での前期・当期行の行番号（売上セクション内）
        pri_sales_row = sales_start_row + 2 * i        # 前期 row
        cur_sales_row = sales_start_row + 2 * i + 1   # 当期 row

        g_row = ["売上高対前年増加率", cur_p, "", cur_p]
        year_growth = {}

        for c in range(3, max_col + 1):
            ppm_col = get_column_letter(c + SEG_OFFSET)
            if fp['prior'] is None:
                g_row.append("")
            else:
                formula = (f"=IF(OR({ppm_col}{pri_sales_row}=\"\","
                           f"{ppm_col}{cur_sales_row}=\"\"),"
                           f"\"\",{ppm_col}{cur_sales_row}/{ppm_col}{pri_sales_row}-1)")
                g_row.append(formula)
            # 実値（軸計算用）
            pri_src = period_lookup.get((target_sales_label, fp['prior'])) if (target_sales_label and fp['prior']) else None
            cur_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
            pri_val = _get_val_for_filing(pri_src, c, fp, False)
            cur_val = _get_val_for_filing(cur_src, c, fp, True)
            if pri_val is not None and cur_val is not None and pri_val != 0:
                year_growth[c] = cur_val / pri_val - 1

        growth_rates.append(year_growth)
        ppm_ws.append(g_row)

    growth_end_row = ppm_ws.max_row

    # 空行
    ppm_ws.append([])

    # -----------------------------------------------------------------------
    # 9. 売上高利益率（N+1 行）
    #    行0: ベース年（最古の前期の利益率）
    #    行1..N: 各報告年度の当期利益率
    # -----------------------------------------------------------------------
    margin_start_row = ppm_ws.max_row + 1
    # profit_margins[i]: dict {analysis_col -> float}  i=0 はベース年
    profit_margins = []

    # ベース年行（最古の報告年度の前期データ）
    base_sales_ppm_row   = sales_start_row           # 最古の前期 売上行
    base_profit_ppm_row  = profit_start_row          # 最古の前期 利益行
    base_margin_row_data = ["売上高利益率", base_prior if base_prior else "", "", base_prior if base_prior else ""]
    base_margin = {}
    _fp0 = valid_pairs[0]
    for c in range(3, max_col + 1):
        ppm_col = get_column_letter(c + SEG_OFFSET)
        formula = (f"=IF(OR({ppm_col}{base_profit_ppm_row}=\"\","
                   f"{ppm_col}{base_sales_ppm_row}=\"\"),"
                   f"\"\",{ppm_col}{base_profit_ppm_row}/{ppm_col}{base_sales_ppm_row})")
        base_margin_row_data.append(formula)
        # 実値（最古のfiling pairの前期データでフィルタ）
        pri_s_src = period_lookup.get((target_sales_label, base_prior)) if (target_sales_label and base_prior) else None
        _, pri_p_src = _get_profit_src(base_prior) if base_prior else (None, None)
        s_val = _get_val_for_filing(pri_s_src, c, _fp0, False)
        p_val = _get_val_for_filing(pri_p_src, c, _fp0, False)
        if s_val is not None and p_val is not None and s_val != 0:
            base_margin[c] = p_val / s_val
    ppm_ws.append(base_margin_row_data)
    profit_margins.append(base_margin)

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        cur_sales_ppm_row  = sales_start_row  + 2 * i + 1   # 当期 売上行
        cur_profit_ppm_row = profit_start_row + 2 * i + 2   # 当期 利益行

        m_row = ["売上高利益率", cur_p, "", cur_p]
        year_margin = {}
        for c in range(3, max_col + 1):
            ppm_col = get_column_letter(c + SEG_OFFSET)
            formula = (f"=IF(OR({ppm_col}{cur_profit_ppm_row}=\"\","
                       f"{ppm_col}{cur_sales_ppm_row}=\"\"),"
                       f"\"\",{ppm_col}{cur_profit_ppm_row}/{ppm_col}{cur_sales_ppm_row})")
            m_row.append(formula)
            # 実値（当期データでフィルタ）
            cur_s_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
            _, cur_p_src = _get_profit_src(cur_p)
            s_val = _get_val_for_filing(cur_s_src, c, fp, True)
            p_val = _get_val_for_filing(cur_p_src, c, fp, True)
            if s_val is not None and p_val is not None and s_val != 0:
                year_margin[c] = p_val / s_val
        profit_margins.append(year_margin)
        ppm_ws.append(m_row)

    margin_end_row = ppm_ws.max_row

    # 売上実値（チャートの完全性チェック用）: valid_pairs[i] の当期データ
    sales_values = []
    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        cur_s_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
        sv = {}
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(cur_s_src, c, fp, True)
            if v is not None:
                sv[c] = v
        sales_values.append(sv)

    # -----------------------------------------------------------------------
    # 10. 書式設定
    # -----------------------------------------------------------------------
    ppm_ws.freeze_panes = 'B2'
    ppm_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, ppm_max_col + 1):
        ppm_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # 売上・利益セクションは数値書式
    for row_idx in range(sales_start_row, profit_end_row + 1):
        for col_idx in range(5, ppm_max_col + 1):
            ppm_ws.cell(row_idx, col_idx).number_format = r'#,##0_ ;[Red]\-#,##0 '

    # 成長率・利益率セクションはパーセント書式
    for row_idx in list(range(growth_start_row, growth_end_row + 1)) + list(range(margin_start_row, margin_end_row + 1)):
        for col_idx in range(5, ppm_max_col + 1):
            ppm_ws.cell(row_idx, col_idx).number_format = '0%'

    # -----------------------------------------------------------------------
    # 11. チャート用集約データエリア
    # -----------------------------------------------------------------------
    # ppm_ws.append([]) は max_row を更新しないため、sec_start を外部から渡す方式に変更。
    # C列（col=3）: 年度・ラベル、D列以降（col=4+）: セグメントデータ
    LATEST_IDX    = N - 1          # valid_pairs の最新インデックス
    FIVE_AGO_IDX  = N - 6          # 5期前インデックス（N>=6 のとき有効）
    FIVE_AGO_OFFSET = 5
    COL_YEAR = 4   # D列: 年度・ラベル
    COL_DATA = 5   # E列: セグメントデータ開始

    def _fmt_year_str(yv):
        if yv is None:
            return ""
        if hasattr(yv, 'strftime'):
            return yv.strftime('%Y/%m')
        s = str(yv)
        if '-' in s:
            try:
                return datetime.datetime.strptime(s, '%Y-%m-%d').strftime('%Y/%m')
            except ValueError:
                return s[:7].replace('-', '/')
        return s[:4]

    # growth_rates[0]=ベース, growth_rates[i+1]=valid_pairs[i]
    # profit_margins[0]=ベース, profit_margins[i+1]=valid_pairs[i]
    _ADJUSTMENT_KEYWORDS = ('合計', '全体', '全社', '消去', '調整', '連結財務諸表', 'その他')

    def _valid_cols(filing_idx):
        """3指標すべてが揃っている analysis_ws 列インデックスのリストを返す"""
        mi = filing_idx + 1   # profit_margins / growth_rates のインデックス（0 はベース年）
        result = []
        for c in range(3, chart_end_col + 1):
            # 調整項目等の非セグメント列は hokoku_col / goukei_col 以外では除外
            if c not in (hokoku_col, goukei_col):
                dim_name = col_to_dim.get(c, '')
                if any(s in dim_name for s in _ADJUSTMENT_KEYWORDS):
                    continue
            if (growth_rates[mi].get(c)  is not None and
                profit_margins[mi].get(c) is not None and
                sales_values[filing_idx].get(c) is not None):
                result.append(c)
        return result

    def _append_data_section(filing_idx, sec_start):
        """集約データ4行を指定行・C列起点で ppm_ws.cell() 書き込み"""
        vcols = _valid_cols(filing_idx)
        vcols_chart = [ci for ci in vcols if ci <= hokoku_col]

        cur_p = valid_pairs[filing_idx]['current']
        cur_sales_ppm  = sales_start_row  + 2 * filing_idx + 1
        # growth_start_row / margin_start_row は直前の append([]) により max_row が更新されないため
        # 実際のデータ行は start_row + 1 にある。filing_idx 分のオフセット (+1) を加えて +2 とする。
        growth_ppm_row = growth_start_row + filing_idx + 2
        margin_ppm_row = margin_start_row + filing_idx + 2

        # ヘッダ行 (sec_start): C=年度, D+=セグメント名
        ppm_ws.cell(sec_start, COL_YEAR).value = cur_p
        for k, ci in enumerate(vcols):
            ppm_ws.cell(sec_start, COL_DATA + k).value = f"={get_column_letter(ci + SEG_OFFSET)}1"
        if vcols_chart and vcols_chart[-1] == hokoku_col:
            hok_k = vcols.index(hokoku_col)
            ppm_ws.cell(sec_start, COL_DATA + hok_k).value = "計"

        # 利益率行 (sec_start+1): C=ラベル参照, D+=値参照
        ppm_ws.cell(sec_start + 1, COL_YEAR).value = f"=A{margin_ppm_row}"
        for k, ci in enumerate(vcols):
            ppm_ws.cell(sec_start + 1, COL_DATA + k).value = (
                f"={get_column_letter(ci + SEG_OFFSET)}{margin_ppm_row}")

        # 成長率行 (sec_start+2): C=ラベル参照, D+=値参照
        ppm_ws.cell(sec_start + 2, COL_YEAR).value = f"=A{growth_ppm_row}"
        for k, ci in enumerate(vcols):
            ppm_ws.cell(sec_start + 2, COL_DATA + k).value = (
                f"={get_column_letter(ci + SEG_OFFSET)}{growth_ppm_row}")

        # 売上行 (sec_start+3): C="売上", D+=値参照（hokoku_col は *1% でバブルサイズ調整、goukei_col は通常表示）
        ppm_ws.cell(sec_start + 3, COL_YEAR).value = target_sales_label or "売上"
        for k, ci in enumerate(vcols):
            cl = get_column_letter(ci + SEG_OFFSET)
            ppm_ws.cell(sec_start + 3, COL_DATA + k).value = (
                f"={cl}{cur_sales_ppm}*1%" if ci == hokoku_col else f"={cl}{cur_sales_ppm}")

        sec_end = sec_start + 3
        n_vcols  = len(vcols)
        n_vchart = len(vcols_chart)
        sec_max_col       = (COL_DATA + n_vcols  - 1) if n_vcols  else COL_YEAR
        chart_sec_max_col = (COL_DATA + n_vchart - 1) if n_vchart else COL_YEAR

        for k in range(n_vcols):
            cl = get_column_letter(COL_DATA + k)
            ppm_ws[f'{cl}{sec_start + 1}'].number_format = '0%'
            ppm_ws[f'{cl}{sec_start + 2}'].number_format = '0%'
            ppm_ws[f'{cl}{sec_start + 3}'].number_format = r'#,##0_);[Red](#,##0)'

        return sec_start, sec_end, sec_max_col, vcols, chart_sec_max_col

    # -----------------------------------------------------------------------
    # 12. 軸範囲計算
    # -----------------------------------------------------------------------
    def _axis_values(metric_list, filing_indices):
        """metric_list: list of dict (col->float), filing_indices: valid_pairs インデックス"""
        vals = []
        for fi in filing_indices:
            mi = fi + 1
            if 0 <= mi < len(metric_list):
                for c, v in metric_list[mi].items():
                    if 3 <= c <= hokoku_col and v is not None:
                        if c != hokoku_col:
                            dim_name = col_to_dim.get(c, '')
                            if any(s in dim_name for s in _ADJUSTMENT_KEYWORDS):
                                continue
                        vals.append(v)
        return vals

    def _rounded_range(vals):
        if not vals:
            return -0.05, 0.40
        mn, mx = min(vals), max(vals)
        return math.floor(mn / 0.05) * 0.05, math.ceil(mx / 0.05) * 0.05

    rel_filings = [LATEST_IDX]
    if N > FIVE_AGO_OFFSET:
        rel_filings.append(FIVE_AGO_IDX)

    common_x_min, common_x_max = _rounded_range(_axis_values(profit_margins, rel_filings))
    common_y_min, common_y_max = _rounded_range(_axis_values(growth_rates,   rel_filings))

    # -----------------------------------------------------------------------
    # 13. バブルチャート作成ヘルパー
    # -----------------------------------------------------------------------
    def _make_chart(title_str):
        chart = BubbleChart()
        chart.style = 2
        chart.height, chart.width = 15, 15
        chart.title = f"PPM分析 {title_str}"
        chart.x_axis.title = "売上高利益率"
        chart.y_axis.title = "売上高対前年増加率"
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.scaling.min = common_x_min
        chart.x_axis.scaling.max = common_x_max
        chart.y_axis.scaling.min = common_y_min
        chart.y_axis.scaling.max = common_y_max
        chart.legend = None
        return chart

    def _add_series(chart, ws, sec_start, sec_max_col):
        xv = Reference(ws, min_col=COL_DATA, min_row=sec_start + 1, max_col=sec_max_col, max_row=sec_start + 1)
        yv = Reference(ws, min_col=COL_DATA, min_row=sec_start + 2, max_col=sec_max_col, max_row=sec_start + 2)
        sz = Reference(ws, min_col=COL_DATA, min_row=sec_start + 3, max_col=sec_max_col, max_row=sec_start + 3)
        chart.series.append(Series(values=yv, xvalues=xv, zvalues=sz, title=""))

    # 最新期セクション: margin_end_row の直下に1行空けてから開始
    lat_sec_start = margin_end_row + 2
    lat_start, lat_end, lat_max_col, _, lat_chart_max_col = _append_data_section(LATEST_IDX, lat_sec_start)
    chart_latest = _make_chart(_fmt_year_str(valid_pairs[LATEST_IDX]['current']))
    _add_series(chart_latest, ppm_ws, lat_start, lat_chart_max_col)

    # -----------------------------------------------------------------------
    # 14. 5年前データセクション（N >= 6 の場合）
    # -----------------------------------------------------------------------
    chart_5y = None
    if N > FIVE_AGO_OFFSET:
        # 最新期セクション末尾から1行空けて5年前セクション開始
        five_sec_start = lat_end + 2
        five_start, five_end, five_max_col, _, five_chart_max_col = _append_data_section(FIVE_AGO_IDX, five_sec_start)
        chart_5y = _make_chart(_fmt_year_str(valid_pairs[FIVE_AGO_IDX]['current']))
        _add_series(chart_5y, ppm_ws, five_start, five_chart_max_col)
        debug_log("[PPM Analysis] Added 5-year comparison section")

    # -----------------------------------------------------------------------
    # 15. チャートをシートに配置
    # -----------------------------------------------------------------------
    chart_row = ppm_ws.max_row + 2
    ppm_ws.add_chart(chart_latest, f'B{chart_row}')
    if chart_5y:
        ppm_ws.add_chart(chart_5y, f'I{chart_row}')

    debug_log(f"[PPM Analysis] Completed PPM analysis sheet: {ppm_sheet_name}")


def _create_ppm_analysis_sheet_ifrs(workbook, analysis_sheet_name, used_sheet_names, filing_pairs, debug_log):
    """
    IFRS用PPM分析シートを作成（内部関数）

    各有報の当期・前期ペアを使ってPPM分析用シートを生成する。
    列構造: 勘定科目 | 報告年度 | 前期・当期 | 会計年度 | セグメント列...
    各報告年度につき前期・当期の2行を出力し、同一有報内のペアで成長率を計算する。
    """
    import re
    import math
    import datetime
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BubbleChart, Reference, Series

    # --- PPM分析シート名を生成 ---
    ppm_sheet_name = analysis_sheet_name + "_PPM分析用"
    if len(ppm_sheet_name) > 31:
        ppm_sheet_name = analysis_sheet_name[:18] + "_PPM分析用"

    debug_log(f"[PPM IFRS] Creating PPM analysis sheet: {ppm_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[PPM IFRS] Analysis sheet '{analysis_sheet_name}' not found, skipping PPM sheet")
        return

    if not filing_pairs:
        debug_log("[PPM IFRS] No filing_pairs provided, skipping PPM sheet")
        return

    analysis_ws = workbook[analysis_sheet_name]
    ppm_ws = workbook.create_sheet(title=ppm_sheet_name)
    used_sheet_names.add(ppm_sheet_name)

    max_col = analysis_ws.max_column

    # -----------------------------------------------------------------------
    # 1. analysis_ws を走査して (ラベル, period_str) -> 行番号 のルックアップを構築
    # -----------------------------------------------------------------------
    period_lookup, unique_labels_ordered = _build_period_lookup(analysis_ws)

    # -----------------------------------------------------------------------
    # 2. IFRSラベル検出
    #    売上収益: 「収益」or「売上」を含み「外部顧客」「セグメント間」を除く
    #    利益: 「利益」を含むラベルの全候補
    # -----------------------------------------------------------------------
    target_sales_label = None
    profit_label_candidates = []
    for label in unique_labels_ordered:
        if target_sales_label is None:
            if (("収益" in label or "売上" in label)
                    and "外部顧客" not in label
                    and "セグメント間" not in label):
                target_sales_label = label
        if "利益" in label:
            profit_label_candidates.append(label)

    # Fallback: if no sales label found (e.g. no aggregate row), try "外部顧客への売上収益" etc.
    if target_sales_label is None:
        _sales_fallback_keywords = ["外部顧客への売上収益", "売上収益", "営業収益", "売上高"]
        for _kw in _sales_fallback_keywords:
            for _lbl in unique_labels_ordered:
                if _kw in _lbl:
                    target_sales_label = _lbl
                    debug_log(f"[PPM IFRS] Sales label fallback: '{target_sales_label}' (matched keyword '{_kw}')")
                    break
            if target_sales_label is not None:
                break

    # Fallback: セグメント利益が見つからない場合、「当期損益」「当期利益」「当期損失」を探す
    if not profit_label_candidates:
        _profit_fallback_keywords = ["当期損益", "当期利益", "当期損失"]
        for _kw in _profit_fallback_keywords:
            for _lbl in unique_labels_ordered:
                if _kw in _lbl and _lbl not in profit_label_candidates:
                    profit_label_candidates.append(_lbl)
        if profit_label_candidates:
            debug_log(f"[PPM IFRS] Profit fallback used: {profit_label_candidates}")

    # Prioritize segment profit labels over operating profit labels
    def _profit_sort_key(lbl):
        if "セグメント利益" in lbl:
            return 0
        return 1
    profit_label_candidates.sort(key=_profit_sort_key)

    debug_log(f"[PPM IFRS] Sales label='{target_sales_label}', Profit candidates={profit_label_candidates}")

    # -----------------------------------------------------------------------
    # 3b. 列位置の検出（報告セグメント / 以外 / 及びその他の合計）
    # -----------------------------------------------------------------------
    hokoku_col = None
    igai_col   = None
    goukei_col = None
    for _ci in range(3, max_col + 1):
        _hv = analysis_ws.cell(1, _ci).value
        if not _hv:
            continue
        _hv_str = str(_hv)
        if "報告セグメント" not in _hv_str:
            continue
        if "以外" in _hv_str:
            igai_col = _ci
        elif "及びその他" in _hv_str:
            goukei_col = _ci
        elif hokoku_col is None:
            hokoku_col = _ci

    if hokoku_col is None:
        if igai_col and igai_col > 3:
            _sum_start_letter = get_column_letter(3)
            _sum_end_letter   = get_column_letter(igai_col - 1)
            analysis_ws.insert_cols(igai_col)
            analysis_ws.cell(1, igai_col).value = "報告セグメント合計"
            for _ri in range(2, analysis_ws.max_row + 1):
                if any(isinstance(analysis_ws.cell(_ri, c).value, (int, float))
                       for c in range(3, igai_col)):
                    analysis_ws.cell(_ri, igai_col).value = (
                        f"=SUM({_sum_start_letter}{_ri}:{_sum_end_letter}{_ri})"
                    )
            hokoku_col = igai_col
            igai_col  += 1
            if goukei_col is not None:
                goukei_col += 1
            max_col += 1
            debug_log(f"[PPM IFRS] Inserted '報告セグメント合計' column at col {hokoku_col}")
        else:
            hokoku_col = max_col

    if igai_col and goukei_col is None:
        new_col = max_col + 1
        analysis_ws.cell(1, new_col).value = "報告セグメント及びその他の合計"
        _h_letter = get_column_letter(hokoku_col)
        _i_letter = get_column_letter(igai_col)
        for _ri in range(2, analysis_ws.max_row + 1):
            if any(isinstance(analysis_ws.cell(_ri, c).value, (int, float))
                   for c in range(3, max_col + 1)):
                analysis_ws.cell(_ri, new_col).value = (
                    f"=SUM({_h_letter}{_ri},{_i_letter}{_ri})"
                )
        goukei_col = new_col
        max_col = new_col
        debug_log(f"[PPM IFRS] Added '報告セグメント及びその他の合計' column at col {new_col}")

    chart_end_col = goukei_col or hokoku_col

    # -----------------------------------------------------------------------
    # 4. valid_pairs の構築（analysis_ws にデータがある報告年度のみ、最新11件）
    # -----------------------------------------------------------------------
    def _has_data_row(row):
        if row is None:
            return False
        for c in range(3, max_col + 1):
            v = analysis_ws.cell(row, c).value
            if isinstance(v, (int, float)):
                return True
        return False

    def _get_profit_src(period_str):
        for candidate in profit_label_candidates:
            row = period_lookup.get((candidate, period_str))
            if row is not None and _has_data_row(row):
                return candidate, row
        return None, None

    valid_pairs = []
    for fp in filing_pairs:
        cur_p = _to_period_str(fp.get('current'))
        pri_p = _to_period_str(fp.get('prior'))
        if not cur_p:
            continue
        cur_s = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
        _, cur_p_src = _get_profit_src(cur_p)
        if _has_data_row(cur_s) or _has_data_row(cur_p_src):
            valid_pairs.append({
                'current': cur_p,
                'prior': pri_p,
                'current_dims': fp.get('current_dims', set()),
                'prior_dims':   fp.get('prior_dims',   set()),
            })

    # 報告年度（current）で昇順ソート
    valid_pairs.sort(key=lambda fp: fp['current'] or '')

    MAX_FILINGS = 11
    if len(valid_pairs) > MAX_FILINGS:
        valid_pairs = valid_pairs[-MAX_FILINGS:]
    N = len(valid_pairs)

    if N == 0:
        debug_log("[PPM IFRS] No valid filing pairs found, skipping PPM sheet")
        workbook.remove(ppm_ws)
        used_sheet_names.discard(ppm_sheet_name)
        return

    debug_log(f"[PPM IFRS] Building PPM sheet for {N} filing years")

    SEG_OFFSET = 2
    ppm_max_col = max_col + SEG_OFFSET

    # -----------------------------------------------------------------------
    # col_to_dim / _get_val_for_filing（dims フィルタ付き読み取り）
    # -----------------------------------------------------------------------
    # hokoku_col / igai_col / goukei_col は 3b で確定済みのため _build_col_info で
    # 上書きしない。col_to_dim と is_synthesized フラグだけを構築する。
    col_to_dim = {c: str(analysis_ws.cell(1, c).value)
                  for c in range(3, max_col + 1)
                  if analysis_ws.cell(1, c).value is not None}
    hokoku_is_synthesized = (col_to_dim.get(hokoku_col, '') == '報告セグメント合計')
    goukei_is_synthesized = (goukei_col is not None and
                              col_to_dim.get(goukei_col, '') == '報告セグメント及びその他の合計')
    _get_val_for_filing = _make_get_val_for_filing(
        analysis_ws, col_to_dim, hokoku_col, igai_col, goukei_col,
        hokoku_is_synthesized, goukei_is_synthesized)

    # -----------------------------------------------------------------------
    # 5. ヘッダー行
    # -----------------------------------------------------------------------
    header = ["勘定科目", "報告年度", "前期・当期", "会計年度"]
    for c in range(3, max_col + 1):
        hv = analysis_ws.cell(1, c).value
        header.append(hv if hv is not None else "")
    ppm_ws.append(header)

    # -----------------------------------------------------------------------
    # 6. 売上収益セクション（2*N 行）
    # -----------------------------------------------------------------------
    sales_start_row = ppm_ws.max_row + 1

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        pri_p = fp['prior']
        pri_src = period_lookup.get((target_sales_label, pri_p)) if (target_sales_label and pri_p) else None
        cur_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None

        sales_lbl = "　" + (target_sales_label or "売上収益")
        pri_row = [sales_lbl, cur_p, "前期", pri_p if pri_p else ""]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(pri_src, c, fp, False)
            pri_row.append(v if v is not None else "")
        ppm_ws.append(pri_row)

        cur_row = [sales_lbl, cur_p, "当期", cur_p]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(cur_src, c, fp, True)
            cur_row.append(v if v is not None else "")
        ppm_ws.append(cur_row)

    # 空行
    ppm_ws.append([])

    # -----------------------------------------------------------------------
    # 7. セグメント利益セクション（2*N 行）
    # -----------------------------------------------------------------------
    profit_start_row = ppm_ws.max_row + 1

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        pri_p = fp['prior']
        pri_profit_lbl, pri_src_default = _get_profit_src(pri_p) if pri_p else (None, None)
        cur_profit_lbl, cur_src = _get_profit_src(cur_p)

        # 前期・当期で同一勘定科目ラベルを使用する。
        # 当期ラベル（cur_profit_lbl）を優先し、前期でも同ラベルのデータを探す。
        # 同ラベルが前期に存在しない場合は前期のデフォルトラベルのデータを使用。
        consistent_lbl = cur_profit_lbl or pri_profit_lbl or "セグメント利益"
        if pri_p and cur_profit_lbl:
            pri_src_consistent = period_lookup.get((cur_profit_lbl, pri_p))
            pri_src = pri_src_consistent if (pri_src_consistent is not None and _has_data_row(pri_src_consistent)) else pri_src_default
        else:
            pri_src = pri_src_default

        pri_row = ["　" + consistent_lbl, cur_p, "前期", pri_p if pri_p else ""]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(pri_src, c, fp, False)
            pri_row.append(v if v is not None else "")
        ppm_ws.append(pri_row)

        cur_row = ["　" + consistent_lbl, cur_p, "当期", cur_p]
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(cur_src, c, fp, True)
            cur_row.append(v if v is not None else "")
        ppm_ws.append(cur_row)

    profit_end_row = ppm_ws.max_row

    # 空行
    ppm_ws.append([])

    # -----------------------------------------------------------------------
    # 8. 売上収益対前年増加率（N+1 行）
    # -----------------------------------------------------------------------
    growth_start_row = ppm_ws.max_row + 1
    growth_rates = [{}]   # ベース年分として空dict

    base_prior = valid_pairs[0]['prior']
    base_row = ["売上収益対前年増加率", base_prior if base_prior else "", "", base_prior if base_prior else ""]
    base_row += [""] * (max_col - 2)
    ppm_ws.append(base_row)

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        pri_sales_row = sales_start_row + 2 * i
        cur_sales_row = sales_start_row + 2 * i + 1

        g_row = ["売上収益対前年増加率", cur_p, "", cur_p]
        year_growth = {}

        for c in range(3, max_col + 1):
            ppm_col = get_column_letter(c + SEG_OFFSET)
            if fp['prior'] is None:
                g_row.append("")
            else:
                formula = (f"=IF(OR({ppm_col}{pri_sales_row}=\"\","
                           f"{ppm_col}{cur_sales_row}=\"\"),"
                           f"\"\",{ppm_col}{cur_sales_row}/{ppm_col}{pri_sales_row}-1)")
                g_row.append(formula)
            pri_src = period_lookup.get((target_sales_label, fp['prior'])) if (target_sales_label and fp['prior']) else None
            cur_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
            pri_val = _get_val_for_filing(pri_src, c, fp, False)
            cur_val = _get_val_for_filing(cur_src, c, fp, True)
            if pri_val is not None and cur_val is not None and pri_val != 0:
                year_growth[c] = cur_val / pri_val - 1

        growth_rates.append(year_growth)
        ppm_ws.append(g_row)

    growth_end_row = ppm_ws.max_row

    # 空行
    ppm_ws.append([])

    # -----------------------------------------------------------------------
    # 9. 売上高利益率（N+1 行）
    # -----------------------------------------------------------------------
    margin_start_row = ppm_ws.max_row + 1
    profit_margins = []

    base_sales_ppm_row  = sales_start_row
    base_profit_ppm_row = profit_start_row
    base_margin_row_data = ["売上高利益率", base_prior if base_prior else "", "", base_prior if base_prior else ""]
    base_margin = {}
    _fp0 = valid_pairs[0]
    for c in range(3, max_col + 1):
        ppm_col = get_column_letter(c + SEG_OFFSET)
        formula = (f"=IF(OR({ppm_col}{base_profit_ppm_row}=\"\","
                   f"{ppm_col}{base_sales_ppm_row}=\"\"),"
                   f"\"\",{ppm_col}{base_profit_ppm_row}/{ppm_col}{base_sales_ppm_row})")
        base_margin_row_data.append(formula)
        pri_s_src = period_lookup.get((target_sales_label, base_prior)) if (target_sales_label and base_prior) else None
        _, pri_p_src = _get_profit_src(base_prior) if base_prior else (None, None)
        s_val = _get_val_for_filing(pri_s_src, c, _fp0, False)
        p_val = _get_val_for_filing(pri_p_src, c, _fp0, False)
        if s_val is not None and p_val is not None and s_val != 0:
            base_margin[c] = p_val / s_val
    ppm_ws.append(base_margin_row_data)
    profit_margins.append(base_margin)

    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        cur_sales_ppm_row  = sales_start_row  + 2 * i + 1
        cur_profit_ppm_row = profit_start_row + 2 * i + 2   # 当期 利益行

        m_row = ["売上高利益率", cur_p, "", cur_p]
        year_margin = {}
        for c in range(3, max_col + 1):
            ppm_col = get_column_letter(c + SEG_OFFSET)
            formula = (f"=IF(OR({ppm_col}{cur_profit_ppm_row}=\"\","
                       f"{ppm_col}{cur_sales_ppm_row}=\"\"),"
                       f"\"\",{ppm_col}{cur_profit_ppm_row}/{ppm_col}{cur_sales_ppm_row})")
            m_row.append(formula)
            cur_s_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
            _, cur_p_src = _get_profit_src(cur_p)
            s_val = _get_val_for_filing(cur_s_src, c, fp, True)
            p_val = _get_val_for_filing(cur_p_src, c, fp, True)
            if s_val is not None and p_val is not None and s_val != 0:
                year_margin[c] = p_val / s_val
        profit_margins.append(year_margin)
        ppm_ws.append(m_row)

    margin_end_row = ppm_ws.max_row

    sales_values = []
    for i, fp in enumerate(valid_pairs):
        cur_p = fp['current']
        cur_s_src = period_lookup.get((target_sales_label, cur_p)) if target_sales_label else None
        sv = {}
        for c in range(3, max_col + 1):
            v = _get_val_for_filing(cur_s_src, c, fp, True)
            if v is not None:
                sv[c] = v
        sales_values.append(sv)

    # -----------------------------------------------------------------------
    # 10. 書式設定
    # -----------------------------------------------------------------------
    ppm_ws.freeze_panes = 'B2'
    ppm_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, ppm_max_col + 1):
        ppm_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    for row_idx in range(sales_start_row, profit_end_row + 1):
        for col_idx in range(5, ppm_max_col + 1):
            ppm_ws.cell(row_idx, col_idx).number_format = r'#,##0_ ;[Red]\-#,##0 '

    for row_idx in list(range(growth_start_row, growth_end_row + 1)) + list(range(margin_start_row, margin_end_row + 1)):
        for col_idx in range(5, ppm_max_col + 1):
            ppm_ws.cell(row_idx, col_idx).number_format = '0%'

    # -----------------------------------------------------------------------
    # 11. チャート用集約データエリア
    # -----------------------------------------------------------------------
    LATEST_IDX    = N - 1
    FIVE_AGO_IDX  = N - 6
    FIVE_AGO_OFFSET = 5
    COL_YEAR = 4   # D列: 年度・ラベル
    COL_DATA = 5   # E列: セグメントデータ開始

    def _fmt_year_str(yv):
        if yv is None:
            return ""
        if hasattr(yv, 'strftime'):
            return yv.strftime('%Y/%m')
        s = str(yv)
        if '-' in s:
            try:
                return datetime.datetime.strptime(s, '%Y-%m-%d').strftime('%Y/%m')
            except ValueError:
                return s[:7].replace('-', '/')
        return s[:4]

    _ADJUSTMENT_KEYWORDS = ('合計', '全体', '全社', '消去', '調整', '連結財務諸表', 'その他')

    def _valid_cols(filing_idx):
        mi = filing_idx + 1
        result = []
        for c in range(3, chart_end_col + 1):
            # 調整項目等の非セグメント列は hokoku_col / goukei_col 以外では除外
            if c not in (hokoku_col, goukei_col):
                dim_name = col_to_dim.get(c, '')
                if any(s in dim_name for s in _ADJUSTMENT_KEYWORDS):
                    continue
            if (growth_rates[mi].get(c)  is not None and
                profit_margins[mi].get(c) is not None and
                sales_values[filing_idx].get(c) is not None):
                result.append(c)
        return result

    def _append_data_section(filing_idx, sec_start):
        """集約データ4行を指定行・C列起点で ppm_ws.cell() 書き込み"""
        vcols = _valid_cols(filing_idx)
        vcols_chart = [ci for ci in vcols if ci <= hokoku_col]

        cur_p = valid_pairs[filing_idx]['current']
        cur_sales_ppm  = sales_start_row  + 2 * filing_idx + 1
        # growth_start_row / margin_start_row は直前の append([]) により max_row が更新されないため
        # 実際のデータ行は start_row + 1 にある。filing_idx 分のオフセット (+1) を加えて +2 とする。
        growth_ppm_row = growth_start_row + filing_idx + 2
        margin_ppm_row = margin_start_row + filing_idx + 2

        # ヘッダ行 (sec_start): C=年度, D+=セグメント名
        ppm_ws.cell(sec_start, COL_YEAR).value = cur_p
        for k, ci in enumerate(vcols):
            ppm_ws.cell(sec_start, COL_DATA + k).value = f"={get_column_letter(ci + SEG_OFFSET)}1"
        if vcols_chart and vcols_chart[-1] == hokoku_col:
            hok_k = vcols.index(hokoku_col)
            ppm_ws.cell(sec_start, COL_DATA + hok_k).value = "計"

        # 利益率行 (sec_start+1)
        ppm_ws.cell(sec_start + 1, COL_YEAR).value = f"=A{margin_ppm_row}"
        for k, ci in enumerate(vcols):
            ppm_ws.cell(sec_start + 1, COL_DATA + k).value = (
                f"={get_column_letter(ci + SEG_OFFSET)}{margin_ppm_row}")

        # 成長率行 (sec_start+2)
        ppm_ws.cell(sec_start + 2, COL_YEAR).value = f"=A{growth_ppm_row}"
        for k, ci in enumerate(vcols):
            ppm_ws.cell(sec_start + 2, COL_DATA + k).value = (
                f"={get_column_letter(ci + SEG_OFFSET)}{growth_ppm_row}")

        # 売上行 (sec_start+3): hokoku_col は *1% でバブルサイズ調整、goukei_col は通常表示
        ppm_ws.cell(sec_start + 3, COL_YEAR).value = target_sales_label or "売上収益"
        for k, ci in enumerate(vcols):
            cl = get_column_letter(ci + SEG_OFFSET)
            ppm_ws.cell(sec_start + 3, COL_DATA + k).value = (
                f"={cl}{cur_sales_ppm}*1%" if ci == hokoku_col else f"={cl}{cur_sales_ppm}")

        sec_end = sec_start + 3
        n_vcols  = len(vcols)
        n_vchart = len(vcols_chart)
        sec_max_col       = (COL_DATA + n_vcols  - 1) if n_vcols  else COL_YEAR
        chart_sec_max_col = (COL_DATA + n_vchart - 1) if n_vchart else COL_YEAR

        for k in range(n_vcols):
            cl = get_column_letter(COL_DATA + k)
            ppm_ws[f'{cl}{sec_start + 1}'].number_format = '0%'
            ppm_ws[f'{cl}{sec_start + 2}'].number_format = '0%'
            ppm_ws[f'{cl}{sec_start + 3}'].number_format = r'#,##0_);[Red](#,##0)'

        return sec_start, sec_end, sec_max_col, vcols, chart_sec_max_col

    # -----------------------------------------------------------------------
    # 12. 軸範囲計算
    # -----------------------------------------------------------------------
    def _axis_values(metric_list, filing_indices):
        vals = []
        for fi in filing_indices:
            mi = fi + 1
            if 0 <= mi < len(metric_list):
                for c, v in metric_list[mi].items():
                    if 3 <= c <= hokoku_col and v is not None:
                        if c != hokoku_col:
                            dim_name = col_to_dim.get(c, '')
                            if any(s in dim_name for s in _ADJUSTMENT_KEYWORDS):
                                continue
                        vals.append(v)
        return vals

    def _rounded_range(vals):
        if not vals:
            return -0.05, 0.40
        mn, mx = min(vals), max(vals)
        return math.floor(mn / 0.05) * 0.05, math.ceil(mx / 0.05) * 0.05

    rel_filings = [LATEST_IDX]
    if N > FIVE_AGO_OFFSET:
        rel_filings.append(FIVE_AGO_IDX)

    common_x_min, common_x_max = _rounded_range(_axis_values(profit_margins, rel_filings))
    common_y_min, common_y_max = _rounded_range(_axis_values(growth_rates,   rel_filings))

    # -----------------------------------------------------------------------
    # 13. バブルチャート作成ヘルパー
    # -----------------------------------------------------------------------
    def _make_chart(title_str):
        chart = BubbleChart()
        chart.style = 2
        chart.height, chart.width = 15, 15
        chart.title = f"PPM分析 {title_str}"
        chart.x_axis.title = "売上高利益率"
        chart.y_axis.title = "売上収益対前年増加率"
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.scaling.min = common_x_min
        chart.x_axis.scaling.max = common_x_max
        chart.y_axis.scaling.min = common_y_min
        chart.y_axis.scaling.max = common_y_max
        chart.legend = None
        return chart

    def _add_series(chart, ws, sec_start, sec_max_col):
        xv = Reference(ws, min_col=COL_DATA, min_row=sec_start + 1, max_col=sec_max_col, max_row=sec_start + 1)
        yv = Reference(ws, min_col=COL_DATA, min_row=sec_start + 2, max_col=sec_max_col, max_row=sec_start + 2)
        sz = Reference(ws, min_col=COL_DATA, min_row=sec_start + 3, max_col=sec_max_col, max_row=sec_start + 3)
        chart.series.append(Series(values=yv, xvalues=xv, zvalues=sz, title=""))

    # 最新期セクション: margin_end_row の直下に1行空けてから開始
    lat_sec_start = margin_end_row + 2
    lat_start, lat_end, lat_max_col, _, lat_chart_max_col = _append_data_section(LATEST_IDX, lat_sec_start)
    chart_latest = _make_chart(_fmt_year_str(valid_pairs[LATEST_IDX]['current']))
    _add_series(chart_latest, ppm_ws, lat_start, lat_chart_max_col)

    # -----------------------------------------------------------------------
    # 14. 5年前データセクション（N >= 6 の場合）
    # -----------------------------------------------------------------------
    chart_5y = None
    if N > FIVE_AGO_OFFSET:
        # 最新期セクション末尾から1行空けて5年前セクション開始
        five_sec_start = lat_end + 2
        five_start, five_end, five_max_col, _, five_chart_max_col = _append_data_section(FIVE_AGO_IDX, five_sec_start)
        chart_5y = _make_chart(_fmt_year_str(valid_pairs[FIVE_AGO_IDX]['current']))
        _add_series(chart_5y, ppm_ws, five_start, five_chart_max_col)
        debug_log("[PPM IFRS] Added 5-year comparison section")

    # -----------------------------------------------------------------------
    # 15. チャートをシートに配置
    # -----------------------------------------------------------------------
    last_data_row = five_end if chart_5y else lat_end
    chart_row = last_data_row + 2
    ppm_ws.add_chart(chart_latest, f'B{chart_row}')
    if chart_5y:
        ppm_ws.add_chart(chart_5y, f'I{chart_row}')

    debug_log(f"[PPM IFRS] Completed PPM analysis sheet: {ppm_sheet_name}")


def _create_composition_ratio_sheet(workbook, analysis_sheet_name, used_sheet_names, debug_log):
    """
    構成比シートを作成（内部関数）

    分析シートの各セルを「報告セグメント」列で割った構成比を表示するシートを生成する。
    数式: =IF(OR('分析'!C2="", '分析'!$H2=""), "", '分析'!C2/'分析'!$H2)

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    from openpyxl.utils import get_column_letter

    comp_sheet_name = analysis_sheet_name + "_構成比"
    if len(comp_sheet_name) > 31:
        comp_sheet_name = analysis_sheet_name[:25] + "_構成比"

    debug_log(f"[Composition] Creating composition ratio sheet: {comp_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[Composition] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    comp_ws = workbook.create_sheet(title=comp_sheet_name)
    used_sheet_names.add(comp_sheet_name)

    escaped = analysis_sheet_name.replace("'", "''")
    max_col = analysis_ws.max_column
    max_row = analysis_ws.max_row

    # 「報告セグメント」列を動的に検出
    denom_col = max_col
    for col_idx in range(3, max_col + 1):
        hv = analysis_ws.cell(1, col_idx).value
        if hv and "報告セグメント" in str(hv) and "以外" not in str(hv):
            denom_col = col_idx
            break

    denom_cl = get_column_letter(denom_col)

    # ヘッダ行（行1）: 勘定科目・年度・各セグメント列をそのままコピー参照（報告セグメントまで）
    header = ["勘定科目", "年度"]
    for col_idx in range(3, denom_col + 1):
        cl = get_column_letter(col_idx)
        header.append(f"=IF('{escaped}'!{cl}1=\"\",\"\",'{escaped}'!{cl}1)")
    comp_ws.append(header)

    # データ行（行2以降）
    for row in range(2, max_row + 1):
        row_data = []
        # A列: 勘定科目
        row_data.append(f"='{escaped}'!A{row}")
        # B列: 年度
        row_data.append(f"='{escaped}'!B{row}")
        # C列以降: 構成比（報告セグメント列まで）
        for col_idx in range(3, denom_col + 1):
            cl = get_column_letter(col_idx)
            formula = (
                f"=IF(OR('{escaped}'!{cl}{row}=\"\","
                f"'{escaped}'!${denom_cl}{row}=\"\"),"
                f"\"\","
                f"'{escaped}'!{cl}{row}/'{escaped}'!${denom_cl}{row})"
            )
            row_data.append(formula)
        comp_ws.append(row_data)

    # 書式設定
    comp_ws.freeze_panes = 'B2'
    comp_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, denom_col + 1):
        comp_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # 数値セル（C列以降、2行目以降）に % 書式を設定
    for row in comp_ws.iter_rows(min_row=2, min_col=3, max_col=denom_col):
        for cell in row:
            cell.number_format = '0%'

    debug_log(f"[Composition] Completed composition ratio sheet: {comp_sheet_name}")


def _create_yoy_growth_sheet(workbook, analysis_sheet_name, used_sheet_names, debug_log):
    """
    対前年増加率シートを作成（内部関数）

    分析シートの各セルについて、当年/前年 - 1 の対前年増加率を表示するシートを生成する。
    数式: =IF(OR('分析'!C2="", '分析'!C3=""), "", '分析'!C2/'分析'!C3-1)
    全カラム（max_col まで）を出力する。

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    from openpyxl.utils import get_column_letter

    yoy_sheet_name = analysis_sheet_name + "_対前年増加率"
    if len(yoy_sheet_name) > 31:
        yoy_sheet_name = analysis_sheet_name[:22] + "_対前年増加率"

    debug_log(f"[YoY] Creating YoY growth sheet: {yoy_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[YoY] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    yoy_ws = workbook.create_sheet(title=yoy_sheet_name)
    used_sheet_names.add(yoy_sheet_name)

    escaped = analysis_sheet_name.replace("'", "''")
    max_col = analysis_ws.max_column
    max_row = analysis_ws.max_row

    # 各勘定科目グループの先頭行（一番古い年度）を検出
    # A列の値が変わる行 = 新しい勘定科目の先頭
    first_rows_of_group = set()
    prev_label = object()  # sentinel
    for row in range(2, max_row + 1):
        label = analysis_ws.cell(row, 1).value
        if label != prev_label:
            first_rows_of_group.add(row)
            prev_label = label

    # ヘッダ行（行1）
    header = ["勘定科目", "年度"]
    for col_idx in range(3, max_col + 1):
        cl = get_column_letter(col_idx)
        header.append(f"=IF('{escaped}'!{cl}1=\"\",\"\",'{escaped}'!{cl}1)")
    yoy_ws.append(header)

    # 行2以降: 各勘定の先頭行はデータ列を空白、それ以外は対前年増加率
    for row in range(2, max_row + 1):
        row_data = [f"='{escaped}'!A{row}", f"='{escaped}'!B{row}"]
        if row in first_rows_of_group:
            # 一番古い年度: 前年データなしのため空白
            row_data += [""] * (max_col - 2)
        else:
            for col_idx in range(3, max_col + 1):
                cl = get_column_letter(col_idx)
                formula = (
                    f"=IF(OR('{escaped}'!{cl}{row - 1}=\"\","
                    f"'{escaped}'!{cl}{row}=\"\"),"
                    f"\"\","
                    f"'{escaped}'!{cl}{row - 1}/'{escaped}'!{cl}{row}-1)"
                )
                row_data.append(formula)
        yoy_ws.append(row_data)

    # 書式設定
    yoy_ws.freeze_panes = 'B2'
    yoy_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, max_col + 1):
        yoy_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # 数値セル（C列以降、3行目以降）に % 書式を設定
    for row in yoy_ws.iter_rows(min_row=3, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = '0%'

    debug_log(f"[YoY] Completed YoY growth sheet: {yoy_sheet_name}")


def _create_ebitda_sheet(workbook, analysis_sheet_name, used_sheet_names, debug_log):
    """
    EBITDA分析シートを作成（内部関数）

    分析シートから売上高（「計」）・セグメント利益（最初の「利益」or「損失」）・
    償却費/償却額を含む全勘定・減損損失を動的に検出し、EBITDAを計算するシートを生成する。

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    import re
    from openpyxl.utils import get_column_letter

    ebitda_sheet_name = analysis_sheet_name + "_EBITDA"
    if len(ebitda_sheet_name) > 31:
        ebitda_sheet_name = analysis_sheet_name[:24] + "_EBITDA"

    debug_log(f"[EBITDA] Creating EBITDA sheet: {ebitda_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[EBITDA] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    ebitda_ws = workbook.create_sheet(title=ebitda_sheet_name)
    used_sheet_names.add(ebitda_sheet_name)

    escaped = analysis_sheet_name.replace("'", "''")
    max_col = analysis_ws.max_column

    # -----------------------------------------------------------------------
    # 1. analysis_ws を走査してラベルと年度のルックアップを構築
    # -----------------------------------------------------------------------
    unique_labels_ordered = []
    lookup = {}
    max_year = -1

    def _extract_year(period_val):
        if period_val is None:
            return None
        if hasattr(period_val, 'year'):
            return period_val.year
        s = str(period_val)
        if '-' in s:
            try:
                return int(s.split('-')[0])
            except ValueError:
                pass
        m = re.search(r'(\d{4})', s)
        return int(m.group(1)) if m else None

    for r in range(2, analysis_ws.max_row + 1):
        label_val = analysis_ws.cell(r, 1).value
        period_val = analysis_ws.cell(r, 2).value
        if not label_val:
            continue
        norm_label = str(label_val).strip()
        if not unique_labels_ordered or unique_labels_ordered[-1] != norm_label:
            unique_labels_ordered.append(norm_label)
        year = _extract_year(period_val)
        if year:
            lookup[(norm_label, year)] = r
            if year > max_year:
                max_year = year

    if max_year == -1:
        debug_log("[EBITDA] No years found in analysis sheet, skipping")
        return

    # -----------------------------------------------------------------------
    # 2. 対象ラベルを検索
    # -----------------------------------------------------------------------
    target_sales_label = None
    target_profit_label = None
    amortization_labels = []
    impairment_label = None

    for label in unique_labels_ordered:
        if target_sales_label is None and label == "計":
            target_sales_label = label
        if target_profit_label is None and ("利益" in label or "損失" in label):
            target_profit_label = label
        if ("償却費" in label or "償却額" in label) and label not in amortization_labels:
            amortization_labels.append(label)
        if impairment_label is None and "減損損失" in label:
            impairment_label = label

    # EBITDA構成要素: セグメント利益 + 償却費/償却額 + 減損損失
    ebitda_items = []
    if target_profit_label:
        ebitda_items.append(target_profit_label)
    ebitda_items.extend(amortization_labels)
    if impairment_label and impairment_label not in ebitda_items:
        ebitda_items.append(impairment_label)

    debug_log(f"[EBITDA] max_year={max_year}, Sales='{target_sales_label}', "
              f"EBITDA items={ebitda_items}")

    # -----------------------------------------------------------------------
    # 3. 11年分の年度リスト（昇順）
    # -----------------------------------------------------------------------
    NUM_YEARS = 11
    target_years = list(range(max_year - 10, max_year + 1))

    # -----------------------------------------------------------------------
    # 4. シートを構築
    # -----------------------------------------------------------------------

    # --- ヘッダー行 ---
    header_row = []
    for col_idx in range(1, max_col + 1):
        cl = get_column_letter(col_idx)
        header_row.append(f"=IF('{escaped}'!{cl}1=\"\",\"\",'{escaped}'!{cl}1)")
    ebitda_ws.append(header_row)

    def _write_data_block(label, src_rows_list, display_label=None):
        """11年分データブロックを書き込み、ブロック開始行を返す"""
        block_start = ebitda_ws.max_row + 1
        lbl = display_label if display_label is not None else label
        for idx, src_row in enumerate(src_rows_list):
            row_data = [lbl]
            for col_idx in range(2, max_col + 1):
                cl = get_column_letter(col_idx)
                if col_idx == 2:
                    formula = f"='{escaped}'!B{src_row}" if src_row else ""
                else:
                    formula = (
                        f"=IF('{escaped}'!{cl}{src_row}=\"\",\"\",'{escaped}'!{cl}{src_row})"
                        if src_row else ""
                    )
                row_data.append(formula)
            ebitda_ws.append(row_data)
        return block_start

    # 売上高ブロック
    sales_src_rows = [
        lookup.get((target_sales_label, y)) if target_sales_label else None
        for y in target_years
    ]
    sales_block_start = _write_data_block(
        target_sales_label or "売上高", sales_src_rows, "　売上高"
    )

    # EBITDAアイテムのブロック
    item_block_starts = []
    item_is_negative = []  # 負ののれんを含む場合はTrue
    for item_label in ebitda_items:
        item_src_rows = [lookup.get((item_label, y)) for y in target_years]
        block_start = _write_data_block(item_label, item_src_rows)
        item_block_starts.append(block_start)
        item_is_negative.append("負ののれん" in item_label)

    # --- 空行 ---
    ebitda_ws.append([""] * max_col)

    # --- EBITDAブロック ---
    ebitda_block_start = ebitda_ws.max_row + 1
    for idx in range(NUM_YEARS):
        row_data = ["　EBITDA"]
        row_data.append(f"=B{sales_block_start + idx}")
        for col_idx in range(3, max_col + 1):
            cl = get_column_letter(col_idx)
            if item_block_starts:
                refs = ",".join(
                    f"{cl}{start + idx}"
                    for start, is_neg in zip(item_block_starts, item_is_negative)
                    if not is_neg
                )
                refs_minus = ",".join(
                    f"{cl}{start + idx}"
                    for start, is_neg in zip(item_block_starts, item_is_negative)
                    if is_neg
                )
                if not refs_minus:
                    formula = f"=IF(COUNT({refs})=0,\"\",SUM({refs}))" if refs else ""
                else:
                    count_refs = refs if refs else refs_minus
                    sum_part = f"SUM({refs})-SUM({refs_minus})" if refs else f"-SUM({refs_minus})"
                    formula = f"=IF(COUNT({count_refs})=0,\"\",{sum_part})"
            else:
                formula = ""
            row_data.append(formula)
        ebitda_ws.append(row_data)
    ebitda_block_end = ebitda_ws.max_row

    # --- 空行 ---
    ebitda_ws.append([""] * max_col)

    # --- EBITDA/売上高ブロック ---
    ratio_block_start = ebitda_ws.max_row + 1
    for idx in range(NUM_YEARS):
        row_data = ["EBITDA/売上高"]
        row_data.append(f"=B{sales_block_start + idx}")
        for col_idx in range(3, max_col + 1):
            cl = get_column_letter(col_idx)
            ebitda_row = ebitda_block_start + idx
            sales_row = sales_block_start + idx
            formula = (
                f"=IF(OR({cl}{ebitda_row}=\"\",{cl}{sales_row}=\"\"),"
                f"\"\",{cl}{ebitda_row}/{cl}{sales_row})"
            )
            row_data.append(formula)
        ebitda_ws.append(row_data)

    # --- 空行 ---
    ebitda_ws.append([""] * max_col)

    # --- EBITDA対前年増加率ブロック ---
    for idx in range(NUM_YEARS):
        row_data = ["EBITDA対前年増加率"]
        row_data.append(f"=B{sales_block_start + idx}")
        if idx == 0:
            # 最古年度は前年データなしのため空白
            row_data += [""] * (max_col - 2)
        else:
            for col_idx in range(3, max_col + 1):
                cl = get_column_letter(col_idx)
                cur_row = ebitda_block_start + idx
                prev_row = ebitda_block_start + idx - 1
                formula = (
                    f"=IF(OR({cl}{cur_row}=\"\",{cl}{prev_row}=\"\"),"
                    f"\"\",{cl}{cur_row}/{cl}{prev_row}-1)"
                )
                row_data.append(formula)
        ebitda_ws.append(row_data)

    # -----------------------------------------------------------------------
    # 5. 書式設定
    # -----------------------------------------------------------------------
    # 数値書式（ヘッダー以降〜EBITDAブロック末: 金額）
    for row in ebitda_ws.iter_rows(min_row=2, max_row=ebitda_block_end, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = r'#,##0_ ;[Red]\-#,##0 '

    # % 書式（EBITDA/売上高、EBITDA対前年増加率）
    for row in ebitda_ws.iter_rows(min_row=ratio_block_start, max_row=ebitda_ws.max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = '0%'

    # ウィンドウ枠固定
    ebitda_ws.freeze_panes = 'B2'

    # 列幅
    ebitda_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, max_col + 1):
        ebitda_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    debug_log(f"[EBITDA] Completed EBITDA sheet: {ebitda_sheet_name}")


def _create_ebitda_sheet_ifrs(workbook, analysis_sheet_name, used_sheet_names, debug_log):
    """
    IFRS用EBITDA分析シートを作成（内部関数）

    IFRSセグメント分析シートから売上収益・セグメント利益・償却費等を検出し、
    EBITDAを計算するシートを生成する。

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名（例: 連結_セグメント情報等(IFRS)_分析）
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    import re
    from openpyxl.utils import get_column_letter

    ebitda_sheet_name = analysis_sheet_name + "_EBITDA"
    if len(ebitda_sheet_name) > 31:
        ebitda_sheet_name = analysis_sheet_name[:24] + "_EBITDA"

    debug_log(f"[EBITDA IFRS] Creating EBITDA sheet: {ebitda_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[EBITDA IFRS] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    ebitda_ws = workbook.create_sheet(title=ebitda_sheet_name)
    used_sheet_names.add(ebitda_sheet_name)

    escaped = analysis_sheet_name.replace("'", "''")
    max_col = analysis_ws.max_column

    # -----------------------------------------------------------------------
    # 1. analysis_ws を走査してラベルと年度のルックアップを構築
    # -----------------------------------------------------------------------
    unique_labels_ordered = []
    lookup = {}
    max_year = -1

    def _extract_year(period_val):
        if period_val is None:
            return None
        if hasattr(period_val, 'year'):
            return period_val.year
        s = str(period_val)
        if '-' in s:
            try:
                return int(s.split('-')[0])
            except ValueError:
                pass
        m = re.search(r'(\d{4})', s)
        return int(m.group(1)) if m else None

    for r in range(2, analysis_ws.max_row + 1):
        label_val = analysis_ws.cell(r, 1).value
        period_val = analysis_ws.cell(r, 2).value
        if not label_val:
            continue
        norm_label = str(label_val).strip()
        if not unique_labels_ordered or unique_labels_ordered[-1] != norm_label:
            unique_labels_ordered.append(norm_label)
        year = _extract_year(period_val)
        if year:
            lookup[(norm_label, year)] = r
            if year > max_year:
                max_year = year

    if max_year == -1:
        debug_log("[EBITDA IFRS] No years found in analysis sheet, skipping")
        return

    # -----------------------------------------------------------------------
    # 2. 対象ラベルを検索
    # -----------------------------------------------------------------------
    target_sales_label = None    # 売上収益・営業収益・収益・売上高
    profit_label_candidates = [] # 全「利益」ヒット（複数年度でラベル変更に対応）
    refs_labels = []             # 償却費・償却額・減損損失（戻入除く）
    refs_minus_labels = []       # 減損の戻入・負ののれん

    for label in unique_labels_ordered:
        # 売上高: 「収益」or「売上」を含み、「外部顧客」「セグメント間」を除く
        if target_sales_label is None:
            if (("収益" in label or "売上" in label)
                    and "外部顧客" not in label
                    and "セグメント間" not in label):
                target_sales_label = label

        # セグメント利益: 「利益」を含む全候補を収集
        if "利益" in label:
            profit_label_candidates.append(label)

        # 減損の戻入（「戻入」or「戻し入」を含む）・負ののれん → refs_minus
        if (("減損" in label and ("戻入" in label or "戻し入" in label))
                or "負ののれん" in label):
            if label not in refs_minus_labels:
                refs_minus_labels.append(label)
        # 償却費・償却額・減損損失（戻入なし）→ refs
        elif (("償却費" in label or "償却額" in label or "減損損失" in label)
              and "戻入" not in label and "戻し入" not in label):
            if label not in refs_labels:
                refs_labels.append(label)

    debug_log(f"[EBITDA IFRS] max_year={max_year}, Sales='{target_sales_label}', "
              f"Profit candidates={profit_label_candidates}, refs={refs_labels}, refs_minus={refs_minus_labels}")

    # -----------------------------------------------------------------------
    # 3. 11年分の年度リスト（昇順）
    # -----------------------------------------------------------------------
    NUM_YEARS = 11
    target_years = list(range(max_year - 10, max_year + 1))

    # -----------------------------------------------------------------------
    # 4. シートを構築
    # -----------------------------------------------------------------------

    # --- ヘッダー行 ---
    header_row = []
    for col_idx in range(1, max_col + 1):
        cl = get_column_letter(col_idx)
        header_row.append(f"=IF('{escaped}'!{cl}1=\"\",\"\",'{escaped}'!{cl}1)")
    ebitda_ws.append(header_row)

    def _first_valid_row(label):
        """ラベル表示用に最新年から最初の有効な行番号を返す"""
        for y in reversed(target_years):
            r = lookup.get((label, y))
            if r is not None:
                return r
        return None

    def _row_has_data(row):
        """analysis_ws の指定行（列3以降）に数値データが1つ以上あるか確認する。"""
        if row is None:
            return False
        for c in range(3, max_col + 1):
            v = analysis_ws.cell(row, c).value
            if isinstance(v, (int, float)):
                return True
        return False

    def _write_data_block_ifrs(label, src_rows_list, year_fallbacks=None):
        """セル参照ラベルでデータブロックを書き込み、ブロック開始行を返す。
        year_fallbacks: src_rowがNoneのとき年度列に表示するリテラル値のリスト（Noneなら空欄）"""
        block_start = ebitda_ws.max_row + 1
        label_row = _first_valid_row(label)
        for idx, src_row in enumerate(src_rows_list):
            label_formula = f"='{escaped}'!A{label_row}" if label_row is not None else label
            row_data = [label_formula]
            for col_idx in range(2, max_col + 1):
                cl = get_column_letter(col_idx)
                if col_idx == 2:
                    if src_row:
                        formula = f"='{escaped}'!B{src_row}"
                    elif year_fallbacks is not None:
                        formula = year_fallbacks[idx]
                    else:
                        formula = ""
                else:
                    formula = (
                        f"=IF('{escaped}'!{cl}{src_row}=\"\",\"\",'{escaped}'!{cl}{src_row})"
                        if src_row else ""
                    )
                row_data.append(formula)
            ebitda_ws.append(row_data)
        return block_start

    # 売上高ブロック
    sales_src_rows = [
        lookup.get((target_sales_label, y)) if target_sales_label else None
        for y in target_years
    ]
    sales_block_start = _write_data_block_ifrs(
        target_sales_label or "売上高", sales_src_rows, year_fallbacks=target_years
    )

    # -----------------------------------------------------------------------
    # セグメント利益ブロック（複数ラベル対応）
    # 年度ごとに最初にデータがある候補が担当する。
    # 各候補を別ブロックとして出力し、担当外の年はデータ空（年度は表示）。
    # -----------------------------------------------------------------------
    # 年度ごとの担当ラベルと行番号を決定
    profit_assignment = {}  # year -> (label, row)
    for y in target_years:
        for candidate in profit_label_candidates:
            row = lookup.get((candidate, y))
            if row is not None and _row_has_data(row):
                profit_assignment[y] = (candidate, row)
                break

    # 利益候補ごとにブロックを作成（担当年のみ実データ行を渡す）
    profit_block_starts = []
    for candidate in profit_label_candidates:
        candidate_src_rows = []
        has_any = False
        for y in target_years:
            assignment = profit_assignment.get(y)
            if assignment is not None and assignment[0] == candidate:
                candidate_src_rows.append(assignment[1])
                has_any = True
            else:
                candidate_src_rows.append(None)
        if has_any:
            profit_block_starts.append(
                _write_data_block_ifrs(candidate, candidate_src_rows, year_fallbacks=target_years)
            )

    if not profit_block_starts:
        # 利益ラベルが全くない場合はダミー行を1ブロック追加
        dummy_rows = [None] * NUM_YEARS
        profit_block_starts.append(
            _write_data_block_ifrs("セグメント利益", dummy_rows, year_fallbacks=target_years)
        )

    # refsブロック
    refs_block_starts = []
    for lbl in refs_labels:
        src_rows = [lookup.get((lbl, y)) for y in target_years]
        refs_block_starts.append(_write_data_block_ifrs(lbl, src_rows, year_fallbacks=target_years))

    # refs_minusブロック
    refs_minus_block_starts = []
    for lbl in refs_minus_labels:
        src_rows = [lookup.get((lbl, y)) for y in target_years]
        refs_minus_block_starts.append(_write_data_block_ifrs(lbl, src_rows, year_fallbacks=target_years))

    # --- 空行 ---
    ebitda_ws.append([""] * max_col)

    # --- EBITDAブロック ---
    # 複数利益ブロックは年度ごとに排他的（担当年のみデータ）なので合算してよい。
    # profit_sum = IF(p1="",0,p1)+IF(p2="",0,p2)+...
    # any_profit_cond = OR(p1<>"", p2<>"", ...)
    ebitda_block_start = ebitda_ws.max_row + 1
    for idx in range(NUM_YEARS):
        row_data = ["EBITDA"]
        row_data.append(f"=B{sales_block_start + idx}")
        for col_idx in range(3, max_col + 1):
            cl = get_column_letter(col_idx)
            profit_cells = [f"{cl}{bs + idx}" for bs in profit_block_starts]
            refs_cells = [f"{cl}{bs + idx}" for bs in refs_block_starts]
            refs_minus_cells = [f"{cl}{bs + idx}" for bs in refs_minus_block_starts]

            # 利益の合計式（空白を0として加算）
            if len(profit_cells) == 1:
                profit_sum = profit_cells[0]
                any_profit_cond = f"{profit_cells[0]}<>\"\""
            else:
                profit_sum = "+".join(f"IF({p}=\"\",0,{p})" for p in profit_cells)
                any_profit_cond = "OR(" + ",".join(f"{p}<>\"\"" for p in profit_cells) + ")"

            if not refs_cells and not refs_minus_cells:
                formula = f"=IF(NOT({any_profit_cond}),\"\",{profit_sum})"
            elif not refs_minus_cells:
                refs_str = ",".join(refs_cells)
                formula = (
                    f"=IF(OR(NOT({any_profit_cond}),COUNT({refs_str})=0),"
                    f"\"\",{profit_sum}+SUM({refs_str}))"
                )
            elif not refs_cells:
                refs_minus_str = ",".join(refs_minus_cells)
                formula = (
                    f"=IF(OR(NOT({any_profit_cond}),COUNT({refs_minus_str})=0),"
                    f"\"\",{profit_sum}-SUM({refs_minus_str}))"
                )
            else:
                refs_str = ",".join(refs_cells)
                refs_minus_str = ",".join(refs_minus_cells)
                formula = (
                    f"=IF(OR(NOT({any_profit_cond}),COUNT({refs_str})=0),"
                    f"\"\",{profit_sum}+SUM({refs_str})-SUM({refs_minus_str}))"
                )
            row_data.append(formula)
        ebitda_ws.append(row_data)
    ebitda_block_end = ebitda_ws.max_row

    # --- 空行 ---
    ebitda_ws.append([""] * max_col)

    # --- EBITDA/売上高ブロック ---
    ratio_block_start = ebitda_ws.max_row + 1
    for idx in range(NUM_YEARS):
        row_data = ["EBITDA/売上高"]
        row_data.append(f"=B{sales_block_start + idx}")
        for col_idx in range(3, max_col + 1):
            cl = get_column_letter(col_idx)
            ebitda_row = ebitda_block_start + idx
            sales_row = sales_block_start + idx
            formula = (
                f"=IF(OR({cl}{ebitda_row}=\"\",{cl}{sales_row}=\"\"),"
                f"\"\",{cl}{ebitda_row}/{cl}{sales_row})"
            )
            row_data.append(formula)
        ebitda_ws.append(row_data)

    # --- 空行 ---
    ebitda_ws.append([""] * max_col)

    # --- EBITDA対前年増加率ブロック ---
    for idx in range(NUM_YEARS):
        row_data = ["EBITDA対前年増加率"]
        row_data.append(f"=B{sales_block_start + idx}")
        if idx == 0:
            row_data += [""] * (max_col - 2)
        else:
            for col_idx in range(3, max_col + 1):
                cl = get_column_letter(col_idx)
                cur_row = ebitda_block_start + idx
                prev_row = ebitda_block_start + idx - 1
                formula = (
                    f"=IF(OR({cl}{cur_row}=\"\",{cl}{prev_row}=\"\"),"
                    f"\"\",{cl}{cur_row}/{cl}{prev_row}-1)"
                )
                row_data.append(formula)
        ebitda_ws.append(row_data)

    # -----------------------------------------------------------------------
    # 5. 書式設定
    # -----------------------------------------------------------------------
    for row in ebitda_ws.iter_rows(min_row=2, max_row=ebitda_block_end, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = r'#,##0_ ;[Red]\-#,##0 '

    for row in ebitda_ws.iter_rows(min_row=ratio_block_start, max_row=ebitda_ws.max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = '0%'

    ebitda_ws.freeze_panes = 'B2'
    ebitda_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, max_col + 1):
        ebitda_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    debug_log(f"[EBITDA IFRS] Completed EBITDA sheet: {ebitda_sheet_name}")


def _create_sales_ratio_sheet(workbook, analysis_sheet_name, used_sheet_names, debug_log):
    """
    売上高比率シートを作成（内部関数）

    分析シートの各セルを同一期間の売上高（「計」or「売上収益」、
    見つからない場合は利益行の直前の勘定）で割った比率を表示するシートを生成する。
    数式例: =IF(OR('分析'!C2="", '分析'!C$24=""), "", '分析'!C2/'分析'!C$24)

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    from openpyxl.utils import get_column_letter

    ratio_sheet_name = analysis_sheet_name + "_売上高比率"
    if len(ratio_sheet_name) > 31:
        ratio_sheet_name = analysis_sheet_name[:22] + "_売上高比率"

    debug_log(f"[SalesRatio] Creating sales ratio sheet: {ratio_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[SalesRatio] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    max_col = analysis_ws.max_column
    max_row = analysis_ws.max_row

    # -----------------------------------------------------------------------
    # 1. 各行のラベルと期間を収集し、ラベル→行番号・期間→行番号リストを構築
    # -----------------------------------------------------------------------
    # label_rows: label -> [row_index, ...]  (出現順)
    # period_label_row: (label, period_str) -> row_index
    label_first_occurrence = {}   # label -> first row index
    all_labels_ordered = []       # 重複なし・出現順
    period_label_row = {}         # (period_str, label) -> row_index

    def _to_period_str(v):
        if v is None:
            return None
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v).strip()

    for r in range(2, max_row + 1):
        lv = analysis_ws.cell(r, 1).value
        pv = analysis_ws.cell(r, 2).value
        if not lv:
            continue
        label = str(lv).strip()
        ps = _to_period_str(pv)
        if label not in label_first_occurrence:
            label_first_occurrence[label] = r
            all_labels_ordered.append(label)
        if ps:
            period_label_row[(ps, label)] = r

    if not all_labels_ordered:
        debug_log(f"[SalesRatio] No data rows in analysis sheet, skipping")
        return

    # -----------------------------------------------------------------------
    # 2. 売上高ラベルを検出
    #    優先順: 完全一致"計" -> 完全一致"売上収益" -> 部分一致"売上収益" ->
    #           最初の利益/損失ラベルの直前のラベル
    #    ※ラベルは全角スペースを含む場合があるためstrip()して比較
    # -----------------------------------------------------------------------
    sales_label = None

    # 優先1: stripped後に完全一致 "計"
    for lbl in all_labels_ordered:
        if lbl.strip() == "計":
            sales_label = lbl
            break

    # 優先2: stripped後に完全一致 "売上収益" (IFRSで多い)
    if sales_label is None:
        for lbl in all_labels_ordered:
            if lbl.strip() == "売上収益":
                sales_label = lbl
                break

    # 優先3: "売上収益" を含む (部分一致フォールバック)
    if sales_label is None:
        for lbl in all_labels_ordered:
            if "売上収益" in lbl:
                sales_label = lbl
                break

    # 優先4: 最初の利益/損失ラベルの直前のラベル
    if sales_label is None:
        for i, lbl in enumerate(all_labels_ordered):
            if "利益" in lbl or "損失" in lbl:
                if i > 0:
                    sales_label = all_labels_ordered[i - 1]
                break

    if sales_label is None:
        debug_log(f"[SalesRatio] Could not detect sales label, skipping")
        return

    debug_log(f"[SalesRatio] Sales label detected: '{sales_label}'")

    # -----------------------------------------------------------------------
    # 3. 期間ごとの売上高行番号を構築
    #    period_str -> row_in_analysis_ws
    # -----------------------------------------------------------------------
    period_sales_row = {}  # period_str -> row index in analysis_ws
    for (ps, lbl), row_idx in period_label_row.items():
        if lbl == sales_label:
            period_sales_row[ps] = row_idx

    if not period_sales_row:
        debug_log(f"[SalesRatio] No period data found for sales label '{sales_label}', skipping")
        return

    # -----------------------------------------------------------------------
    # 4. シートを作成しヘッダー行を出力
    # -----------------------------------------------------------------------
    ratio_ws = workbook.create_sheet(title=ratio_sheet_name)
    used_sheet_names.add(ratio_sheet_name)

    escaped = analysis_sheet_name.replace("'", "''")

    # ヘッダー行: 勘定科目・年度・各セグメント列をそのままコピー参照
    header = ["勘定科目", "年度"]
    for col_idx in range(3, max_col + 1):
        cl = get_column_letter(col_idx)
        header.append(f"=IF('{escaped}'!{cl}1=\"\",\"\",'{escaped}'!{cl}1)")
    ratio_ws.append(header)

    # -----------------------------------------------------------------------
    # 5. データ行: 各行 × 各列に売上高比率の数式を出力
    # -----------------------------------------------------------------------
    for row in range(2, max_row + 1):
        pv = analysis_ws.cell(row, 2).value
        ps = _to_period_str(pv)
        sales_row = period_sales_row.get(ps) if ps else None

        row_data = [
            f"='{escaped}'!A{row}",
            f"='{escaped}'!B{row}",
        ]
        for col_idx in range(3, max_col + 1):
            cl = get_column_letter(col_idx)
            if sales_row is None:
                row_data.append("")
            else:
                formula = (
                    f"=IF(OR('{escaped}'!{cl}{row}=\"\","
                    f"'{escaped}'!{cl}${sales_row}=\"\"),"
                    f"\"\","
                    f"'{escaped}'!{cl}{row}/'{escaped}'!{cl}${sales_row})"
                )
                row_data.append(formula)
        ratio_ws.append(row_data)

    # -----------------------------------------------------------------------
    # 6. 書式設定・列幅・ウィンドウ枠
    # -----------------------------------------------------------------------
    for row in ratio_ws.iter_rows(min_row=2, max_row=ratio_ws.max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = '0.0%'

    ratio_ws.freeze_panes = 'B2'
    ratio_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, max_col + 1):
        ratio_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    debug_log(f"[SalesRatio] Completed sales ratio sheet: {ratio_sheet_name}")


def _create_employee_ratio_sheet(workbook, analysis_sheet_name, used_sheet_names, debug_log):
    """
    従業員比率シートを作成（内部関数）

    分析シートの各セルを同一期間の従業員数（下から2番目のラベル）で割った比率を
    表示するシートを生成する。
    数式例: =IF(OR('分析'!C2="", '分析'!C$24=""), "", '分析'!C2/'分析'!C$24)

    Args:
        workbook: openpyxlワークブック
        analysis_sheet_name: 参照元の分析シート名
        used_sheet_names: 使用済みシート名のセット
        debug_log: デバッグログ関数
    """
    from openpyxl.utils import get_column_letter

    ratio_sheet_name = analysis_sheet_name + "_従業員比率"
    if len(ratio_sheet_name) > 31:
        ratio_sheet_name = analysis_sheet_name[:22] + "_従業員比率"

    debug_log(f"[EmpRatio] Creating employee ratio sheet: {ratio_sheet_name}")

    if analysis_sheet_name not in workbook.sheetnames:
        debug_log(f"[EmpRatio] Analysis sheet '{analysis_sheet_name}' not found, skipping")
        return

    analysis_ws = workbook[analysis_sheet_name]
    max_col = analysis_ws.max_column
    max_row = analysis_ws.max_row

    # -----------------------------------------------------------------------
    # 1. 各行のラベルと期間を収集
    # -----------------------------------------------------------------------
    all_labels_ordered = []   # 重複なし・出現順
    period_label_row = {}     # (period_str, label) -> row_index

    def _to_period_str(v):
        if v is None:
            return None
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v).strip()

    seen_labels = set()
    for r in range(2, max_row + 1):
        lv = analysis_ws.cell(r, 1).value
        pv = analysis_ws.cell(r, 2).value
        if not lv:
            continue
        label = str(lv).strip()
        ps = _to_period_str(pv)
        if label not in seen_labels:
            seen_labels.add(label)
            all_labels_ordered.append(label)
        if ps:
            period_label_row[(ps, label)] = r

    if len(all_labels_ordered) < 2:
        debug_log(f"[EmpRatio] Not enough labels in analysis sheet, skipping")
        return

    # -----------------------------------------------------------------------
    # 2. 従業員数ラベルを末尾から検索（「従業員数」を含む最後のラベル）
    # -----------------------------------------------------------------------
    employee_label = None
    for lbl in reversed(all_labels_ordered):
        if "従業員数" in lbl:
            employee_label = lbl
            break

    if employee_label is None:
        debug_log(f"[EmpRatio] No employee count label found, skipping")
        return

    debug_log(f"[EmpRatio] Employee label detected: '{employee_label}'")

    # -----------------------------------------------------------------------
    # 3. 期間ごとの従業員数行番号を構築
    # -----------------------------------------------------------------------
    period_employee_row = {}  # period_str -> row index in analysis_ws
    for (ps, lbl), row_idx in period_label_row.items():
        if lbl == employee_label:
            period_employee_row[ps] = row_idx

    if not period_employee_row:
        debug_log(f"[EmpRatio] No period data found for employee label '{employee_label}', skipping")
        return

    # -----------------------------------------------------------------------
    # 4. シートを作成しヘッダー行を出力
    # -----------------------------------------------------------------------
    ratio_ws = workbook.create_sheet(title=ratio_sheet_name)
    used_sheet_names.add(ratio_sheet_name)

    escaped = analysis_sheet_name.replace("'", "''")

    header = ["勘定科目", "年度"]
    for col_idx in range(3, max_col + 1):
        cl = get_column_letter(col_idx)
        header.append(f"=IF('{escaped}'!{cl}1=\"\",\"\",'{escaped}'!{cl}1)")
    ratio_ws.append(header)

    # -----------------------------------------------------------------------
    # 5. データ行: 各行 × 各列に従業員比率の数式を出力
    # -----------------------------------------------------------------------
    for row in range(2, max_row + 1):
        pv = analysis_ws.cell(row, 2).value
        ps = _to_period_str(pv)
        emp_row = period_employee_row.get(ps) if ps else None

        row_data = [
            f"='{escaped}'!A{row}",
            f"='{escaped}'!B{row}",
        ]
        for col_idx in range(3, max_col + 1):
            cl = get_column_letter(col_idx)
            if emp_row is None:
                row_data.append("")
            else:
                formula = (
                    f"=IF(OR('{escaped}'!{cl}{row}=\"\","
                    f"'{escaped}'!{cl}${emp_row}=\"\"),"
                    f"\"\","
                    f"'{escaped}'!{cl}{row}/'{escaped}'!{cl}${emp_row})"
                )
                row_data.append(formula)
        ratio_ws.append(row_data)

    # -----------------------------------------------------------------------
    # 6. 書式設定・列幅・ウィンドウ枠
    # -----------------------------------------------------------------------
    for row in ratio_ws.iter_rows(min_row=2, max_row=ratio_ws.max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = r'#,##0_ ;[Red]\-#,##0 '

    ratio_ws.freeze_panes = 'B2'
    ratio_ws.column_dimensions['A'].width = 31
    for col_idx in range(2, max_col + 1):
        ratio_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    debug_log(f"[EmpRatio] Completed employee ratio sheet: {ratio_sheet_name}")
