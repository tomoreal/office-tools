#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XBRL to Excel Converter

このスクリプトは、EDINETからダウンロードしたXBRLファイルをExcelに変換します。

【現在の状態】
- 総行数: 3680行（コメント追加後）
- 機能: 完成・動作確認済み
- 分割: 不要（将来必要になった時のみ実施）

【プログラム構成】
このファイルは以下の5つの層で構成されています：

1. INFRASTRUCTURE LAYER (61-380行)
   - ログ管理、ファイル操作、セキュリティチェック
   - 将来の分割先: infra/logging.py, infra/file.py

2. TAXONOMY LAYER (108-905行)
   - タクソノミ管理・更新・解析
   - 将来の分割先: taxonomy/repository.py, parser.py, updater.py

3. XBRL LAYER (383-1412行)
   - XBRL解析・階層構築
   - 将来の分割先: xbrl/loader.py, parser.py, context.py

4. CORE LAYER (1413-3650行)
   - メイン処理パイプライン（process_xbrl_zips）
   - 将来の分割先: core/processor.py, pipeline.py, model/xbrl_data.py

5. CLI LAYER (3652-3680行)
   - コマンドライン引数処理
   - 将来の分割先: cli.py

【将来の分割について】
分割が必要になる条件（以下のいずれか）:
- 保守が困難になった時（バグ修正に3日以上）
- 複数人で開発する時（gitコンフリクト頻発）
- 大規模な機能追加（PDF出力、API化など）
- テスト自動化が必要になった時

詳細は「将来の分割案.md」を参照してください。
"""

import os
import sys
import zipfile
import tempfile
import shutil
import glob
import time
import json
import urllib.request
import re
import gzip
import logging
import subprocess
from threading import Lock
from contextlib import contextmanager

try:
    import fcntl
    HAS_FCNTL = True
except ImportError:
    HAS_FCNTL = False

try:
    from lxml import etree
    HAS_LXML = True
except ImportError:
    import xml.etree.ElementTree as etree
    HAS_LXML = False

# Import financial analysis module
from financial_analysis import add_financial_analysis_sheets

# Import segment analysis module
from segment_analysis import add_segment_analysis_sheets

# Import ccc analysis module
from ccc_analysis import add_ccc_analysis_sheets

# Import diversity analysis module
from diversity_analysis import add_diversity_sheet, add_human_capital_sheet

# Import EDINET Taxonomy Dictionary
from edinet_taxonomy_dict import common_dict as EDINET_COMMON_DICT

# Base directory for the script and caching
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Delay loading heavy libraries until needed (helps CGI performance)
HAS_PANDAS = False
HAS_OPENPYXL = False

# Control verbose logging via environment variable (default: enabled for debugging)
VERBOSE_LOGGING = os.environ.get('XBRL_VERBOSE', '1') == '1'

# Thread lock for taxonomy cache operations to prevent race conditions
# when multiple workers try to download/extract/write taxonomy cache simultaneously
_TAXONOMY_LOCK = Lock()

# Pre-compiled regular expressions (performance optimization)
# These patterns are used frequently in loops, so pre-compiling avoids repeated compilation
_RE_CAMEL_CASE_1 = re.compile(r'(.)([A-Z][a-z]+)')
_RE_CAMEL_CASE_2 = re.compile(r'([a-z0-9])([A-Z])')
_RE_TAXONOMY_YEAR = re.compile(r'http://disclosure\.edinet-fsa\.go\.jp/taxonomy/[a-z]+(?:_[a-z]+)?/(\d{4})-\d{2}-\d{2}')
_RE_SEGMENT_SUFFIX = re.compile(r'-(\d+)$')
_RE_TAXONOMY_INDEX = re.compile(r'<a href="(/search/\d+\.html)">(\d{4})年版EDINETタクソノミ</a>')
_RE_TAXONOMY_ZIP = re.compile(r'(/search/\d+/1c_Taxonomy\.zip)')

# Configure logging using Python's standard logging module
# This provides better performance (buffering) and thread safety compared to manual file I/O
_LOG_FILE = os.path.join(SCRIPT_DIR, 'convert_xbrl_debug.log')

# ============================================================================
# INFRASTRUCTURE LAYER - Logging & File Operations
# ============================================================================
# 【将来の分割先】infra/logging.py, infra/file.py
#
# このセクションには以下が含まれます:
# - TimestampFormatter (61-87行): ログフォーマッタ
# - rotate_logs_manually (184-222行): ログローテーション
# - debug_log (224-238行): デバッグログ出力
# - vprint (342-347行): 詳細ログ出力
# - validate_zip_path (240-248行): ZIPパス検証
# - check_zip_bomb (249-264行): ZIP爆弾チェック
# - file_lock (266-316行): ファイルロック
# ============================================================================

# Custom formatter that includes timestamp
class TimestampFormatter(logging.Formatter):
    def formatTime(self, record, datefmt=None):
        return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.created))

# Set up the logger
_logger = logging.getLogger('xbrl_converter')
_logger.setLevel(logging.DEBUG if VERBOSE_LOGGING else logging.INFO)

# File handler with buffering
_file_handler = logging.FileHandler(_LOG_FILE, mode='a', encoding='utf-8')
_file_handler.setLevel(logging.DEBUG)
_file_formatter = TimestampFormatter('%(asctime)s %(message)s')
_file_handler.setFormatter(_file_formatter)
_logger.addHandler(_file_handler)

# Console handler (stderr) for server log visibility
_console_handler = logging.StreamHandler(sys.stderr)
_console_handler.setLevel(logging.INFO)
_console_formatter = logging.Formatter('%(message)s')
_console_handler.setFormatter(_console_formatter)
_logger.addHandler(_console_handler)

# Prevent propagation to root logger
_logger.propagate = False

# Flag to ensure log rotation is only checked once per session
_log_rotation_checked = False

# Flag to ensure EDINET taxonomy dict update is checked at most once per session
# (keyed by the highest taxonomy_year seen so far)
_taxonomy_dict_last_checked_year = None

# ============================================================================
# TAXONOMY LAYER - Taxonomy Management & Updates
# ============================================================================
# 【将来の分割先】taxonomy/repository.py, taxonomy/parser.py, taxonomy/updater.py
#
# このセクションには以下が含まれます:
# - get_edinet_taxonomy_dict_year (93-116行): タクソノミ年度取得
# - check_and_update_edinet_taxonomy (118-182行): タクソノミ更新チェック
# - fetch_taxonomy_url (479-534行): タクソノミURL取得
# - get_standard_labels (536-730行): タクソノミラベル取得（メイン）
# - parse_labels_file (732-850行): ラベルファイル解析
# - build_suffix_index (318-340行): サフィックスインデックス構築
#
# 分割時の注意:
# - repository.py: I/O処理のみ（キャッシュ・取得）
# - parser.py: 純粋な解析ロジック
# - updater.py: 更新処理（副作用）
# ============================================================================

def get_edinet_taxonomy_dict_year():
    """Read the taxonomy year embedded in edinet_taxonomy_dict.py's docstring.

    The file header contains a line like:
        Generated: 2026-03-18 23:48:57
    We treat the calendar year of generation (e.g. 2026) as the proxy for
    which taxonomy version the dict was built from.

    Returns:
        str | None: Four-digit year string (e.g. '2026'), or None on failure.
    """
    dict_path = os.path.join(SCRIPT_DIR, 'edinet_taxonomy_dict.py')
    if not os.path.exists(dict_path):
        return None
    try:
        with open(dict_path, 'r', encoding='utf-8') as f:
            header = f.read(512)  # Only need the first few lines
        m = re.search(r'Generated:\s*(\d{4})', header)
        if m:
            return m.group(1)
    except Exception:
        pass
    return None


def check_and_update_edinet_taxonomy(taxonomy_year):
    """Run update_edinet_taxonomy.py when the XBRL references a newer taxonomy year.

    This is triggered once per session for the highest taxonomy_year encountered.
    It compares the year embedded in edinet_taxonomy_dict.py against the year
    found in the XBRL _pre.xml file.  If the XBRL references a newer year, the
    update script is called as a subprocess so the dict stays current.

    Args:
        taxonomy_year (str): Four-digit year extracted from the XBRL file
                             (e.g. '2025').
    """
    global _taxonomy_dict_last_checked_year

    if not taxonomy_year:
        return

    # Skip if we've already checked for this year (or a newer one) this session
    if _taxonomy_dict_last_checked_year and _taxonomy_dict_last_checked_year >= taxonomy_year:
        return

    _taxonomy_dict_last_checked_year = taxonomy_year

    dict_year = get_edinet_taxonomy_dict_year()
    debug_log(f"[TaxonomyDict] XBRL year={taxonomy_year}, dict_year={dict_year}")

    # Trigger update when the XBRL references a newer year than the current dict
    if dict_year is None or taxonomy_year > dict_year:
        debug_log(f"[TaxonomyDict] Newer taxonomy detected ({taxonomy_year} > {dict_year}). Running update_edinet_taxonomy.py...")
        update_script = os.path.join(SCRIPT_DIR, 'update_edinet_taxonomy.py')
        if not os.path.exists(update_script):
            debug_log(f"[TaxonomyDict] update_edinet_taxonomy.py not found at {update_script}, skipping.")
            return
        try:
            result = subprocess.run(
                [sys.executable, update_script],
                capture_output=True,
                text=True,
                timeout=120,
                cwd=SCRIPT_DIR,
            )
            if result.returncode == 0:
                debug_log(f"[TaxonomyDict] update_edinet_taxonomy.py completed successfully.")
                # Reload the updated module so this session uses the new dict
                import importlib
                try:
                    import edinet_taxonomy_dict
                    importlib.reload(edinet_taxonomy_dict)
                    # Re-bind the global EDINET_COMMON_DICT to the refreshed module
                    global EDINET_COMMON_DICT
                    EDINET_COMMON_DICT = edinet_taxonomy_dict.common_dict
                    debug_log(f"[TaxonomyDict] edinet_taxonomy_dict reloaded ({len(EDINET_COMMON_DICT)} items).")
                except Exception as e:
                    debug_log(f"[TaxonomyDict] WARNING: Could not reload edinet_taxonomy_dict: {e}")
            else:
                debug_log(f"[TaxonomyDict] update_edinet_taxonomy.py exited with code {result.returncode}.")
                if result.stderr:
                    debug_log(f"[TaxonomyDict] stderr: {result.stderr[:500]}")
        except subprocess.TimeoutExpired:
            debug_log("[TaxonomyDict] update_edinet_taxonomy.py timed out after 120s.")
        except Exception as e:
            debug_log(f"[TaxonomyDict] ERROR running update_edinet_taxonomy.py: {e}")
    else:
        debug_log(f"[TaxonomyDict] Dict is up-to-date (year {dict_year} >= {taxonomy_year}), skipping update.")


def rotate_logs_manually(log_file):
    """cronが使えない環境向け：プログラム実行時にログをチェックし、1週間単位でローテーション・圧縮を行う"""
    if not os.path.exists(log_file):
        return

    # 1週間（7日）の秒数
    WEEK_SECONDS = 7 * 24 * 3600
    
    # ログファイルの最終更新時間を取得
    file_mtime = os.path.getmtime(log_file)
    if time.time() - file_mtime < WEEK_SECONDS:
        return # まだ1週間経っていないので何もしない

    try:
        import gzip
        import shutil
        # ローテーション処理 (4回分＝1ヶ月分保持)
        # 1. 一番古いファイルを削除し、順にずらす
        for i in range(4, 1, -1):
            old_file = f"{log_file}.{i}.gz"
            prev_file = f"{log_file}.{i-1}.gz"
            if i == 4 and os.path.exists(old_file):
                os.remove(old_file)
            if os.path.exists(prev_file):
                os.rename(prev_file, f"{log_file}.{i}.gz")

        # 2. 直近の非圧縮ログ (.1) を圧縮して .2.gz にする
        log_1 = f"{log_file}.1"
        if os.path.exists(log_1):
            with open(log_1, 'rb') as f_in:
                with gzip.open(f"{log_file}.2.gz", 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(log_1)

        # 3. 現在のログを .1 にリネーム
        os.rename(log_file, log_1)
    except Exception as e:
        # ログローテーション自体のエラーはstderrにのみ出力（デッドロック回避）
        print(f"Log rotation error: {e}", file=sys.stderr)

def debug_log(message):
    """Write message to a persistent debug log file for user visibility.

    Now uses Python's logging module for better performance (buffering)
    and thread safety instead of manual file I/O.
    """
    global _log_rotation_checked

    # Check for manual rotation (only once per session for performance)
    if not _log_rotation_checked:
        rotate_logs_manually(_LOG_FILE)
        _log_rotation_checked = True

    # Use logging module which handles buffering and thread safety
    _logger.info(message)

def validate_zip_path(target_path, base_dir):
    """Ensure the target path is within the base directory to prevent Zip Slip."""
    abs_target = os.path.abspath(target_path)
    abs_base = os.path.abspath(base_dir)
    # Use commonpath to ensure target is actually within base_dir
    # (prevents prefix match bypass where '/tmp/base_dir_extra' starts with '/tmp/base_dir')
    if os.path.commonpath([abs_target, abs_base]) != abs_base:
        raise Exception(f"Zip Slip detected: {abs_target} is outside of {abs_base}")

def check_zip_bomb(zip_ref, max_size=500 * 1024 * 1024):
    """Check for ZIP bomb attacks by validating total uncompressed size.

    Args:
        zip_ref: zipfile.ZipFile object
        max_size: Maximum allowed total uncompressed size (default: 500MB)

    Raises:
        Exception: If total uncompressed size exceeds max_size
    """
    total_size = 0
    for info in zip_ref.infolist():
        total_size += info.file_size
        if total_size > max_size:
            raise Exception(f"ZIP bomb detected: total uncompressed size ({total_size:,} bytes) exceeds limit ({max_size:,} bytes)")

@contextmanager
def file_lock(lock_path, timeout=60):
    """Cross-process file lock using fcntl (Unix) with fallback to threading.Lock.

    This prevents race conditions when multiple processes (e.g., CGI, Gunicorn, uWSGI)
    try to download/extract the same taxonomy simultaneously.

    Args:
        lock_path: Path to the lock file
        timeout: Maximum time to wait for lock (seconds)

    Yields:
        None (lock is held during context)
    """
    lock_file = None
    try:
        # Create lock file directory if it doesn't exist
        lock_dir = os.path.dirname(lock_path)
        if lock_dir:
            os.makedirs(lock_dir, exist_ok=True)

        if HAS_FCNTL:
            # Use fcntl for cross-process locking (Unix/Linux)
            lock_file = open(lock_path, 'w')
            start_time = time.time()

            while True:
                try:
                    fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                    debug_log(f"Acquired file lock: {lock_path}")
                    break
                except BlockingIOError:
                    if time.time() - start_time > timeout:
                        raise TimeoutError(f"Failed to acquire lock after {timeout}s: {lock_path}")
                    time.sleep(0.1)

            yield

            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
            debug_log(f"Released file lock: {lock_path}")
        else:
            # Fallback to threading.Lock for Windows (process-local only)
            # Note: This doesn't protect against multi-process race conditions
            debug_log(f"WARNING: fcntl not available, using threading.Lock (not cross-process safe)")
            with _TAXONOMY_LOCK:
                yield
    finally:
        if lock_file:
            try:
                lock_file.close()
            except Exception:
                pass

def http_request_with_retry(url, max_retries=3, timeout=10, delay=1.0):
    """HTTP request with automatic retry on failure.

    Retries the request up to max_retries times with exponential backoff.
    This handles transient network errors, timeouts, and temporary server issues.

    Args:
        url: URL to fetch
        max_retries: Maximum number of retry attempts (default: 3)
        timeout: Request timeout in seconds (default: 10)
        delay: Initial delay between retries in seconds (default: 1.0)

    Returns:
        response object from urllib.request.urlopen

    Raises:
        Exception: If all retry attempts fail
    """
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            debug_log(f"HTTP request to {url} (attempt {attempt}/{max_retries})")
            response = urllib.request.urlopen(url, timeout=timeout)
            if attempt > 1:
                debug_log(f"SUCCESS: HTTP request succeeded on attempt {attempt}")
            return response
        except Exception as e:
            last_error = e
            error_msg = str(e)
            debug_log(f"ERROR: HTTP request failed (attempt {attempt}/{max_retries}): {error_msg}")

            if attempt < max_retries:
                # Exponential backoff: 1s, 2s, 4s, ...
                wait_time = delay * (2 ** (attempt - 1))
                debug_log(f"Retrying in {wait_time:.1f} seconds...")
                time.sleep(wait_time)
            else:
                debug_log(f"FAILED: All {max_retries} retry attempts exhausted for {url}")

    # All retries failed
    raise Exception(f"Failed to fetch {url} after {max_retries} attempts: {last_error}")

def http_retrieve_with_retry(url, filepath, max_retries=3, timeout=30, delay=1.0):
    """Download file with automatic retry on failure.

    Similar to http_request_with_retry but for file downloads using urlretrieve.

    Args:
        url: URL to download
        filepath: Local path to save the file
        max_retries: Maximum number of retry attempts (default: 3)
        timeout: Request timeout in seconds (default: 30)
        delay: Initial delay between retries in seconds (default: 1.0)

    Raises:
        Exception: If all retry attempts fail
    """
    import socket

    # Set global timeout for urlretrieve (it doesn't accept timeout parameter)
    original_timeout = socket.getdefaulttimeout()
    socket.setdefaulttimeout(timeout)

    last_error = None

    try:
        for attempt in range(1, max_retries + 1):
            try:
                debug_log(f"Downloading {url} to {filepath} (attempt {attempt}/{max_retries})")
                urllib.request.urlretrieve(url, filepath)
                if attempt > 1:
                    debug_log(f"SUCCESS: Download succeeded on attempt {attempt}")
                return
            except Exception as e:
                last_error = e
                error_msg = str(e)
                debug_log(f"ERROR: Download failed (attempt {attempt}/{max_retries}): {error_msg}")

                # Clean up partial download
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                        debug_log(f"Removed partial download: {filepath}")
                    except Exception:
                        pass

                if attempt < max_retries:
                    wait_time = delay * (2 ** (attempt - 1))
                    debug_log(f"Retrying in {wait_time:.1f} seconds...")
                    time.sleep(wait_time)
                else:
                    debug_log(f"FAILED: All {max_retries} retry attempts exhausted for {url}")

        # All retries failed
        raise Exception(f"Failed to download {url} after {max_retries} attempts: {last_error}")
    finally:
        # Restore original timeout
        socket.setdefaulttimeout(original_timeout)

def build_suffix_index(labels_map):
    """Build a suffix index for O(1) label lookups.

    Converts O(N) suffix searches to O(1) hash lookups.
    For a labels_map with 20,000+ entries, this significantly improves performance
    when searching for labels by suffix (e.g., '_OperatingRevenue').

    Args:
        labels_map: dict mapping element names to labels

    Returns:
        dict: suffix -> (full_key, label) mapping
            Keys are element suffixes after the last '_' (e.g., 'OperatingRevenue')
    """
    suffix_index = {}
    for full_key, label in labels_map.items():
        if '_' in full_key:
            # Extract suffix after last underscore
            suffix = full_key.split('_')[-1]
            # Only keep the first match for each suffix (priority)
            if suffix not in suffix_index:
                suffix_index[suffix] = (full_key, label)
    return suffix_index

def vprint(*args, **kwargs):
    """Verbose print - only prints if VERBOSE_LOGGING is enabled."""
    if VERBOSE_LOGGING:
       msg = " ".join(map(str, args))
       _logger.debug(f"[VERBOSE] {msg}")


# ============================================================================
# XBRL LAYER - XBRL Parsing & Data Extraction
# ============================================================================
# 【将来の分割先】xbrl/loader.py, xbrl/parser.py, xbrl/context.py
#
# このセクションには以下が含まれます:
# - safe_xpath (349-430行): XPath実行ヘルパー
# - find_xbrl_files (432-478行): XBRLファイル検出
# - clean_label (852-874行): ラベルクリーンアップ
# - convert_camel_case_to_title (876-880行): キャメルケース変換
# - parse_presentation_linkbase (882-1004行): プレゼンテーション解析
# - parse_instance_contexts_and_units (1006-1142行): コンテキスト・単位解析
# - parse_ixbrl_facts (1144-1294行): iXBRL事実値抽出
# - create_hierarchy (1296-1344行): 階層構築
# - merge_sequences (1346-1355行): シーケンスマージ
#
# 分割時の注意:
# - loader.py: ZIP展開、ファイル検出
# - parser.py: XBRLパースのメインロジック（500行程度は許容）
# - context.py: Context/Unit処理（XBRLの鬼門）
# ============================================================================

def safe_xpath(tree_or_elem, query, namespaces=None):
    """Safe XPath helper that works with both lxml and standard xml.etree.ElementTree.
    Note: ElementTree supports only a subset of XPath.
    """
    if HAS_LXML:
        return tree_or_elem.xpath(query, namespaces=namespaces)
    else:
        # Fallback for standard ElementTree (basic namespace handling)
        # Note: ET uses {url}prefix syntax for namespaces in findall
        # We try to convert basic queries but complex XPath might skip.
        try:
            # If it's a Tree object, get root first
            root = tree_or_elem.getroot() if hasattr(tree_or_elem, 'getroot') else tree_or_elem
            if namespaces:
                # Basic conversion for link:loc -> {http://...}loc if query is simple
                # For now, we return empty or try simple findall if query starts with //
                if query.startswith('//'):
                    tag = query.split(':')[-1]
                    return root.findall(f'.//{{*}}{tag}')
            if namespaces:
                return root.findall(query, namespaces=namespaces)
            return root.findall(query)
        except Exception as e:
            vprint(f"safe_xpath error: {e}")
            return []

# Label priority constants (used in parse_labels_file)
# Lower priority values are preferred (1 is highest priority)
PRIORITY_VERBOSE_LABEL = 1
PRIORITY_ALT_LABEL = 2
PRIORITY_STANDARD_LABEL = 3
PRIORITY_INDUSTRY_LABEL = 4
PRIORITY_TERSE_LABEL = 5
PRIORITY_TOTAL_LABEL = 10
PRIORITY_DEFAULT = 99
PRIORITY_WORST = 100
PRIORITY_LEGACY_DEFAULT = 50
PRIORITY_GENERIC_PENALTY = 50  # Added to priority for generic labels like "Total"

# Common dimension axis and member mapping for Japanese fallback
COMMON_DIMENSION_MAPPING = {
    'ConsolidatedOrNonConsolidatedAxis': '連結・非連結',
    'ConsolidatedMember': '連結',
    'NonConsolidatedMember': '単体',
    'OperatingSegmentsAxis': '事業セグメント',
    'OperatingSegmentsMember': '事業セグメント',
    'BusinessSegmentsAxis': '事業セグメント',
    'BusinessSegmentsMember': '事業セグメント',
    'ReportableSegmentsAxis': '報告セグメント',
    'ReportableSegmentsMember': '報告セグメント',
    'OtherReportableSegmentsMember': 'その他事業',
    'EntityInformationAxis': '提出者情報',
    # Business Segment Standard Members (2024 Taxonomy)
    'OperatingSegmentsNotIncludedInReportableSegmentsAndOtherRevenueGeneratingBusinessActivitiesMember': '報告セグメント以外の全てのセグメント',
    'TotalOfReportableSegmentsAndOthersMember': '報告セグメント及びその他の合計',
    'UnallocatedAmountsAndEliminationMember': '全社・消去',
    'ReconcilingItemsMember': '調整項目',
}

# ============================================================================
# CONSTANTS & MAPPINGS
# ============================================================================
# 【将来の分割先】output/mappings.py
#
# IFRS/J-GAAPのラベルマッピング、シート名マッピングなど
# Excel生成で使用する定数を定義
#
# 注意: process_xbrl_zips 関数内にも SHEET_MAPPING, HEADING_DICT,
#       SEGMENT_DICT が定義されている（2580-2960行付近）
#       分割時はこれらも mappings.py に統合する
# ============================================================================

# IFRS account name mapping to match commercial tools
IFRS_LABEL_MAPPING = {
    'jpigp_cor_RevenueIFRS': '売上高',
    'jpigp_cor_Revenue': '売上高',
    'jpigp_cor_ProfitLossBeforeTaxIFRS': '税引前当期利益',
    'jpigp_cor_ProfitLossAttributableToOwnersOfParentIFRS': '親会社の所有者に帰属する当期利益',
    'jpigp_cor_ProfitLossAttributableToNonControllingInterestsIFRS': '非支配持分',
    'jpigp_cor_AssetsIFRS': '資産合計',
    'jpigp_cor_EquityIFRS': '資本合計',
    'jpigp_cor_EquityAttributableToOwnersOfParentIFRS': '親会社の所有者に帰属する持分合計',
    'jpigp_cor_OperatingProfitIFRS': '営業利益',
    'jpigp_cor_OperatingRevenueIFRS': '営業収益',
    'jpigp_cor_CostOfSalesIFRS': '売上原価',
    'jpigp_cor_GrossProfitIFRS': '売上総利益',
    'jpigp_cor_SellingGeneralAndAdministrativeExpensesIFRS': '販売費及び一般管理費',
    'jpigp_cor_InventoriesCAIFRS': '棚卸資産',
    'jpigp_cor_PropertyPlantAndEquipmentIFRS': '有形固定資産',
    'jpigp_cor_IntangibleAssetsIFRS': '無形資産',
    'jpigp_cor_CurrentAssetsIFRS': '流動資産合計',
    'jpigp_cor_NonCurrentAssetsIFRS': '非流動資産合計',
    'jpigp_cor_LiabilitiesIFRS': '負債合計',
}

# Helper to find specific linkbase/instance files in the unzipped folder
def find_xbrl_files(extract_dir):
    files = {'lab': []}
    
    # 1. Global Label Collection (Resilient to structure)
    # Collect ALL Japanese label linkbases from the entire package
    for root, _, filenames in os.walk(extract_dir):
        for f in filenames:
            fl = f.lower()
            if fl.endswith('_lab.xml') and not fl.endswith('_lab-en.xml'):
                files['lab'].append(os.path.join(root, f))
    
    # 2. Instance and Presentation Lookup
    # Prefer PublicDoc but fallback to any identified file
    base_path = None
    for root, dirs, _ in os.walk(extract_dir):
        if 'PublicDoc' in dirs:
            base_path = os.path.join(root, 'PublicDoc')
            break
    
    # Priority 1: Files in PublicDoc
    if base_path and os.path.exists(base_path):
        for f in os.listdir(base_path):
            fl = f.lower()
            full_path = os.path.join(base_path, f)
            if 'pre' not in files and fl.endswith('_pre.xml'):
                files['pre'] = full_path
            elif 'xbrl' not in files and fl.endswith('.xbrl'):
                files['xbrl'] = full_path
            elif fl.endswith('.htm') or fl.endswith('.html'):
                if 'ixbrl' not in files: files['ixbrl'] = []
                files['ixbrl'].append(full_path)
    
    # Priority 2: Fallback to global search if missing
    if 'pre' not in files or 'xbrl' not in files:
        for root, _, filenames in os.walk(extract_dir):
            # Skip AuditDoc for fallback instance search to avoid wrong facts
            if 'AuditDoc' in root: continue 
            for f in filenames:
                fl = f.lower()
                full_path = os.path.join(root, f)
                if 'pre' not in files and fl.endswith('_pre.xml'):
                    files['pre'] = full_path
                elif 'xbrl' not in files and fl.endswith('.xbrl'):
                    files['xbrl'] = full_path
    
    return files if 'pre' in files and 'xbrl' in files else None

def fetch_taxonomy_url(year):
    """Dynamically fetch EDINET taxonomy URL from FSA index page.

    This approach is more robust than hardcoded URLs because:
    1. FSA may update taxonomy URLs when they reorganize their site
    2. New years are automatically discovered
    3. Fallback to hardcoded URLs if fetching fails

    Args:
        year: Taxonomy year as string (e.g., '2025')

    Returns:
        str: Taxonomy ZIP URL, or None if not found
    """
    try:
        # Step 1: Fetch index page to find the year's detail page (with retry)
        index_url = 'https://www.fsa.go.jp/search/EDINET_Taxonomy_All.html'
        debug_log(f"Fetching taxonomy index from {index_url}")

        with http_request_with_retry(index_url, max_retries=3, timeout=10) as response:
            html = response.read().decode('utf-8', errors='ignore')

        # Step 2: Parse to find the link for requested year
        # Pattern: <a href="/search/YYYYMMDD.html">YYYY年版EDINETタクソノミ</a>
        # Note: We need to create a dynamic pattern with the year, so we use format string
        pattern = rf'<a href="(/search/\d+\.html)">{year}年版EDINETタクソノミ</a>'
        match = re.search(pattern, html)

        if not match:
            debug_log(f"Could not find {year} taxonomy link in index page")
            return None

        detail_path = match.group(1)
        detail_url = f'https://www.fsa.go.jp{detail_path}'
        debug_log(f"Found detail page: {detail_url}")

        # Step 3: Fetch detail page to find ZIP download link (with retry)
        with http_request_with_retry(detail_url, max_retries=3, timeout=10) as response:
            detail_html = response.read().decode('utf-8', errors='ignore')

        # Pattern: <a href="/search/YYYYMMDD/1c_Taxonomy.zip">
        # Use pre-compiled regex for performance
        zip_match = _RE_TAXONOMY_ZIP.search(detail_html)

        if zip_match:
            zip_path = zip_match.group(1)
            taxonomy_url = f'https://www.fsa.go.jp{zip_path}'
            debug_log(f"Found taxonomy URL: {taxonomy_url}")
            return taxonomy_url
        else:
            debug_log(f"Could not find Taxonomy.zip link in detail page")
            return None

    except Exception as e:
        debug_log(f"ERROR: Failed to fetch taxonomy URL for {year}: {e}")
        return None

def get_standard_labels(year, cache_dir=None):
    """Returns (all_labels, label_priorities) for the given taxonomy year.
    Uses cached standard_labels.json if it exists.
    """
    if cache_dir is None:
        cache_dir = os.path.join(SCRIPT_DIR, 'edinet_taxonomies')
    
    start_time = time.time()
    tax_dir = os.path.join(cache_dir, str(year))
    labels_cache_file = os.path.join(tax_dir, 'standard_labels.json')

    debug_log(f"Checking taxonomy cache: {labels_cache_file}")

    # Fallback URLs (hardcoded) - used if dynamic fetching fails
    # These are maintained for reliability and offline operation
    fallback_urls = {
        '2025': 'https://www.fsa.go.jp/search/20241112/1c_Taxonomy.zip',
        '2024': 'https://www.fsa.go.jp/search/20231211/1c_Taxonomy.zip',
        '2023': 'https://www.fsa.go.jp/search/20221108/1c_Taxonomy.zip',
        '2022': 'https://www.fsa.go.jp/search/20211109/1c_Taxonomy.zip',
        '2021': 'https://www.fsa.go.jp/search/20201110/1c_Taxonomy.zip',
        '2020': 'https://www.fsa.go.jp/search/20191101/1c_Taxonomy.zip',
        '2019': 'https://www.fsa.go.jp/search/20190228/1c_Taxonomy.zip',
        '2018': 'https://www.fsa.go.jp/search/20180228/1c_Taxonomy.zip',
    }

    # Try to fetch URL dynamically from FSA index page (more robust for future updates)
    taxonomy_url = fetch_taxonomy_url(year)

    # Fallback to hardcoded URLs if dynamic fetch fails
    if not taxonomy_url:
        if year in fallback_urls:
            taxonomy_url = fallback_urls[year]
            debug_log(f"Using fallback URL for {year}")
        else:
            vprint(f"Taxonomy for year {year} not found (neither dynamic nor fallback).")
            return {}, {}
    
    # Try to load from cache (fast path, no lock needed)
    if os.path.exists(labels_cache_file):
        try:
            with open(labels_cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict) and 'labels' in data:
                    res = data['labels'], data.get('priorities', {})
                    debug_log(f"SUCCESS: Loaded taxonomy cache for {year} in {time.time() - start_time:.2f}s")
                    return res
                else:
                    # Legacy format compatibility
                    priorities = {k: PRIORITY_LEGACY_DEFAULT for k in data}
                    debug_log(f"SUCCESS: Loaded legacy taxonomy cache for {year} in {time.time() - start_time:.2f}s")
                    return data, priorities
        except Exception as e:
            debug_log(f"ERROR: Cache read error for {year}: {e}")

    # Cache doesn't exist - acquire locks to prevent race conditions
    # Use both thread lock (for multi-threaded processes) and file lock (for multi-process environments)
    lock_file_path = os.path.join(tax_dir, f'.taxonomy_{year}.lock')
    with file_lock(lock_file_path):
        with _TAXONOMY_LOCK:
            # Double-check: another thread/process may have created the cache while we were waiting
            if os.path.exists(labels_cache_file):
                try:
                    with open(labels_cache_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        if isinstance(data, dict) and 'labels' in data:
                            res = data['labels'], data.get('priorities', {})
                            debug_log(f"SUCCESS: Loaded taxonomy cache for {year} (created by another thread/process) in {time.time() - start_time:.2f}s")
                            return res
                        else:
                            priorities = {k: PRIORITY_LEGACY_DEFAULT for k in data}
                            debug_log(f"SUCCESS: Loaded legacy taxonomy cache for {year} (created by another thread/process) in {time.time() - start_time:.2f}s")
                            return data, priorities
                except Exception as e:
                    debug_log(f"ERROR: Cache read error for {year} after lock: {e}")

            if not os.path.exists(tax_dir):
                try:
                    os.makedirs(tax_dir, exist_ok=True)
                    debug_log(f"Created taxonomy directory: {tax_dir}")
                except Exception as e:
                    debug_log(f"WARNING: Could not create tax_dir {tax_dir}, falling back to /tmp: {e}")
                    tax_dir = os.path.join('/tmp', 'edinet_taxonomies', str(year))
                    try:
                        os.makedirs(tax_dir, exist_ok=True)
                    except Exception as e:
                        vprint(f"Fallback to /tmp failed for {year}: {e}")
                    labels_cache_file = os.path.join(tax_dir, 'standard_labels.json')

            if not os.path.exists(labels_cache_file):
                zip_path = os.path.join(tax_dir, 'taxonomy.zip')
                if not os.path.exists(os.path.join(tax_dir, 'taxonomy')): # rudimentary check for extracted data
                    vprint(f"Downloading EDINET taxonomy for {year} (takes a moment)...")
                    try:
                        http_retrieve_with_retry(taxonomy_url, zip_path, max_retries=3, timeout=30)
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            # Check for ZIP bomb before extraction
                            check_zip_bomb(zip_ref)

                            # Robust extraction: manually decode filenames using CP932 (shift_jis)
                            # to avoid Mojibake on Linux/Unix systems that default to UTF-8
                            for info in zip_ref.infolist():
                                try:
                                    # infolist().filename is often bytes or interpreted as CP437
                                    # We re-encode and decode correctly
                                    filename_raw = info.filename.encode('cp437')
                                    filename = filename_raw.decode('cp932')
                                except Exception:
                                    filename = info.filename

                                target_path = os.path.join(tax_dir, filename)
                                validate_zip_path(target_path, tax_dir)
                                if info.is_dir():
                                    os.makedirs(target_path, exist_ok=True)
                                else:
                                    os.makedirs(os.path.dirname(target_path), exist_ok=True)
                                    with zip_ref.open(info) as source, open(target_path, 'wb') as target:
                                        shutil.copyfileobj(source, target)
                        os.remove(zip_path)

                        # Canonical normalization: Rename any Mojibake "タクソノミ" folder to "taxonomy"
                        for entry in os.listdir(tax_dir):
                            full_p = os.path.join(tax_dir, entry)
                            if os.path.isdir(full_p) and ('â^' in entry or 'タクソノミ' in entry):
                                new_p = os.path.join(tax_dir, 'taxonomy')
                                if not os.path.exists(new_p):
                                    os.rename(full_p, new_p)
                                    vprint(f"Normalized taxonomy directory: {entry} -> taxonomy")
                    except Exception as e:
                        vprint(f"Failed to download/extract taxonomy for {year}: {e}")
                        return {}, {}

            vprint(f"Parsing EDINET taxonomy labels for {year}... (First run only)")
            # Use os.walk() instead of glob.glob() for better performance with thousands of files
            # Performance: os.walk() ~0.003s vs glob.glob() ~0.007s for 2709 files
            lab_files = []
            for root, dirs, files in os.walk(tax_dir):
                for file in files:
                    if file.endswith('_lab.xml'):
                        lab_files.append(os.path.join(root, file))
            lab_files.sort()
            all_labels = {}
            label_priorities = {} # {element_name: priority}

            # Track which taxonomy types we're loading
            taxonomy_types = set()

            for lf in lab_files:
                if 'deprecated' in lf or 'dep' in lf or '-en.xml' in lf:
                    continue

                # Extract taxonomy type (jpigp, jppfs, jpcrp, etc.)
                basename = os.path.basename(lf)
                if '_lab.xml' in basename:
                    tax_type = basename.split('_')[0]
                    taxonomy_types.add(tax_type)

                try:
                    # Determine taxonomy type from filename for domain-specific weighting
                    tax_type = os.path.basename(lf).split('_')[0]
                    parsed_labels, parsed_priorities = parse_labels_file(lf)

                    for el, text in parsed_labels.items():
                        prio = parsed_priorities.get(el, 99)

                        # Element-Prefix Awareness: Boost priority if the taxonomy file matches the element's prefix
                        # e.g. jppfs label for jppfs element is better than general label
                        el_prefix = el.split('_')[0] if '_' in el else ""
                        if el_prefix and el_prefix == tax_type:
                            prio -= 0.5 # Slight boost for domain-exact match

                        current_prio = label_priorities.get(el, 100)
                        if (el not in all_labels) or (prio < current_prio) or (prio == current_prio and text < all_labels.get(el, "")):
                            all_labels[el] = text
                            label_priorities[el] = prio
                except Exception as e:
                    vprint(f"Error parsing labels from {lf}: {e}")

            # Report which taxonomies were loaded
            if taxonomy_types:
                vprint(f"  Loaded taxonomies: {', '.join(sorted(taxonomy_types))}")

            ifrs_count = sum(1 for k in all_labels if 'IFRS' in k)
            vprint(f"  Total labels: {len(all_labels)}, IFRS labels: {ifrs_count}")

            if all_labels:
                try:
                    with open(labels_cache_file, 'w', encoding='utf-8') as f:
                        json.dump({'labels': all_labels, 'priorities': label_priorities}, f, ensure_ascii=False, indent=2)
                    debug_log(f"SUCCESS: Saved taxonomy cache to {labels_cache_file} in {time.time() - start_time:.2f}s")
                except Exception as e:
                    debug_log(f"WARNING: Could not cache labels to {labels_cache_file}: {e}")

            return all_labels, label_priorities


def parse_labels_file(lab_file):
    """Parse an XBRL label linkbase using lxml for robust namespace handling.
    Returns (labels, priorities) where labels is a dict mapping element names to text,
    and priorities maps them to their best priority score.
    """
    labels = {}
    priorities = {}
    try:
        # Parse XML with lxml if available (ElementTree.XMLParser doesn't support 'recover')
        if HAS_LXML:
            # Secure parser against XXE attacks
            parser = etree.XMLParser(recover=True, resolve_entities=False, no_network=True)
            tree = etree.parse(lab_file, parser)
        else:
            tree = etree.parse(lab_file)
    except Exception as e:
        # If parsing fails, return empty mappings
        vprint(f"Error parsing {lab_file}: {e}")
        return labels, priorities

    # Namespace map for XBRL linkbase
    ns = {
        "link": "http://www.xbrl.org/2003/linkbase",
        "xlink": "http://www.w3.org/1999/xlink",
        "xml": "http://www.w3.org/XML/1998/namespace"
    }

    # 1. Locate all <link:loc> elements to map label IDs to element QNames
    href_to_label_id = {}
    for loc in safe_xpath(tree, "//link:loc", namespaces=ns):
        href = loc.get("{http://www.w3.org/1999/xlink}href")
        label_id = loc.get("{http://www.w3.org/1999/xlink}label")
        if href and label_id:
            # Element name may be a QName like jppfs_cor:CashAndDeposits
            element_name = href.split('#')[-1].replace(':', '_')
            href_to_label_id[label_id] = element_name

    # 2. Filter arcs to only concept‑label relationships (collect ALL associated resource IDs)
    label_id_to_res_ids = {}
    arc_xpath = "//link:labelArc[@xlink:arcrole='http://www.xbrl.org/2003/arcrole/concept-label']"
    for arc in safe_xpath(tree, arc_xpath, namespaces=ns):
        from_id = arc.get("{http://www.w3.org/1999/xlink}from")
        to_id = arc.get("{http://www.w3.org/1999/xlink}to")
        if from_id and to_id:
            if from_id not in label_id_to_res_ids:
                label_id_to_res_ids[from_id] = []
            label_id_to_res_ids[from_id].append(to_id)

    # 3. Gather label resources (<link:label>) with Japanese language
    res_id_to_text = {}
    res_id_to_priority = {}
    
    # Role priority: verboseLabel is standard for EDINET CSV output
    # XBRL Label Roles and their associated priority (lower is better)
    role_priority = {
        "http://www.xbrl.org/2003/role/verboseLabel": PRIORITY_VERBOSE_LABEL,
        "http://disclosure.edinet-fsa.go.jp/jpcrp/alt/role/label": PRIORITY_ALT_LABEL, # EDINET industry-specific alternate
        "http://www.xbrl.org/2003/role/label": PRIORITY_STANDARD_LABEL,
        "http://disclosure.edinet-fsa.go.jp/jppfs/ele/role/label": PRIORITY_INDUSTRY_LABEL, # Electric Power
        "http://disclosure.edinet-fsa.go.jp/jppfs/gas/role/label": PRIORITY_INDUSTRY_LABEL, # Gas
        "http://disclosure.edinet-fsa.go.jp/jppfs/sec/role/label": PRIORITY_INDUSTRY_LABEL, # Securities
        "http://disclosure.edinet-fsa.go.jp/jppfs/ins/role/label": PRIORITY_INDUSTRY_LABEL, # Insurance
        "http://disclosure.edinet-fsa.go.jp/jppfs/bnk/role/label": PRIORITY_INDUSTRY_LABEL, # Banking
        "http://www.xbrl.org/2003/role/terseLabel": PRIORITY_TERSE_LABEL,
        "http://www.xbrl.org/2003/role/totalLabel": PRIORITY_TOTAL_LABEL,
        "http://disclosure.edinet-fsa.go.jp/jpcrp/alt/role/totalLabel": 11,
    }

    GENERIC_LABELS = ('合計', '小計', '計', 'total', 'sum', 'subtotal', '金額')

    for res in safe_xpath(tree, "//link:label", namespaces=ns):
        lang = res.get("{http://www.w3.org/XML/1998/namespace}lang")
        if not lang or not lang.startswith('ja'):
            continue
        res_id = res.get("{http://www.w3.org/1999/xlink}label")
        role = res.get("{http://www.w3.org/1999/xlink}role")
        if not res_id:
            continue
            
        text = ''.join(res.itertext()).strip()
        if not text:
            continue
            
        priority = role_priority.get(role, PRIORITY_DEFAULT)
        # Demote verboseLabel if it contains structural markers like "、報告セグメント"
        # to prefer cleaner standard labels for segment names.
        if priority == PRIORITY_VERBOSE_LABEL:
            structural_markers = ['、報告セグメント', '、セグメント情報', '、事業セグメント', '、セグメント別情報']
            if any(s in text for s in structural_markers):
                priority = PRIORITY_INDUSTRY_LABEL  # Standard label will take precedence

        # Penalize generic labels to avoid "Total" appearing everywhere if a better name exists
        # Skip penalty if it's the high-priority verboseLabel
        if priority > PRIORITY_VERBOSE_LABEL and any(g in text.lower() for g in GENERIC_LABELS):
            priority += PRIORITY_GENERIC_PENALTY
            
        if (res_id not in res_id_to_text) or (priority < res_id_to_priority.get(res_id, PRIORITY_WORST)) or (priority == res_id_to_priority.get(res_id, PRIORITY_WORST) and text < res_id_to_text[res_id]):
            res_id_to_text[res_id] = text
            res_id_to_priority[res_id] = priority

    # 4. Build final mapping (pick the best label text among all resource IDs)
    for label_id, element_name in href_to_label_id.items():
        res_ids = label_id_to_res_ids.get(label_id, [])
        best_text = None
        best_priority = PRIORITY_WORST
        
        for res_id in res_ids:
            text = res_id_to_text.get(res_id)
            priority = res_id_to_priority.get(res_id, PRIORITY_DEFAULT)
            if text and priority < best_priority:
                best_text = text
                best_priority = priority
                
        if best_text:
            if element_name not in labels or best_priority < priorities.get(element_name, 100):
                labels[element_name] = best_text
                priorities[element_name] = best_priority
            # [REMOVED] mapping base name to labels[base] as it causes collisions
    return labels, priorities

def clean_label(text):
    """Clean structural markers and suffixes from labels to ensure consistency."""
    if not text:
        return ""
    import unicodedata
    # Normalize full-width/half-width characters (e.g. 0-9, A-Z)
    text = unicodedata.normalize('NFKC', text)
    # Remove standard structural markers
    text = text.replace(' [メンバー]', '').replace(' [軸]', '').replace(' [項目]', '').replace(' [区分]', '').replace(' [要素]', '').strip()
    
    # Remove common verbose suffixes that differentiate presentation labels from factual dimensions
    # especially for segments
    suffixes_to_remove = ['、報告セグメント', '、セグメント情報', '、事業セグメント', '、セグメント別情報', '、セグメント情報別']
    for s in suffixes_to_remove:
        if text.endswith(s):
            text = text[:-len(s)]
    
    # Also handle the variant without '、' just in case
    for s in [s.replace('、', '') for s in suffixes_to_remove if '、' in s]:
        if text.endswith(s):
            text = text[:-len(s)]
            
    return text.strip()

def convert_camel_case_to_title(name):
    # e.g. CashAndDeposits -> Cash And Deposits
    # Uses pre-compiled regexes for performance (called frequently in loops)
    s1 = _RE_CAMEL_CASE_1.sub(r'\1 \2', name)
    return _RE_CAMEL_CASE_2.sub(r'\1 \2', s1).title()

def parse_presentation_linkbase(pre_file):
    vprint(f"Parsing presentation linkbase... {os.path.basename(pre_file)}")
    try:
        # Use lxml for robust namespace handling if available
        if HAS_LXML:
            # Secure parser against XXE attacks
            parser = etree.XMLParser(recover=True, resolve_entities=False, no_network=True)
            tree = etree.parse(pre_file, parser)
        else:
            tree = etree.parse(pre_file)
    except Exception as e:
        vprint(f"Error parsing presentation linkbase: {e}")
        return {}

    ns = {
        "link": "http://www.xbrl.org/2003/linkbase",
        "xlink": "http://www.w3.org/1999/xlink"
    }

    # 1. Group by role URI first
    role_to_content = {} # {role_uri: {'locs': {label: element}, 'arcs': [arc_dicts]}}
    
    links = safe_xpath(tree, "//link:presentationLink", namespaces=ns)
    for link in links:
        role_uri = link.get("{http://www.w3.org/1999/xlink}role")
        if not role_uri:
            continue
        
        if role_uri not in role_to_content:
            role_to_content[role_uri] = {'locs': {}, 'arcs': []}
            
        # Map locators in this link
        locs = safe_xpath(link, "link:loc", namespaces=ns)
        for loc in locs:
            href = loc.get("{http://www.w3.org/1999/xlink}href")
            label = loc.get("{http://www.w3.org/1999/xlink}label")
            if href and label:
                # Normalize element name: replace ':' with '_' to match facts and labels
                element_name = href.split('#')[-1].replace(':', '_')
                role_to_content[role_uri]['locs'][label] = element_name
                
        # Map arcs in this link
        arcs = safe_xpath(link, "link:presentationArc", namespaces=ns)
        for arc in arcs:
            from_id = arc.get("{http://www.w3.org/1999/xlink}from")
            to_id = arc.get("{http://www.w3.org/1999/xlink}to")
            order = arc.get("order")
            pref_label = arc.get("preferredLabel")
            if from_id and to_id:
                role_to_content[role_uri]['arcs'].append({
                    'from': from_id,
                    'to': to_id,
                    'order': float(order) if order else 0.0,
                    'preferredLabel': pref_label
                })

    statement_trees = {} # {role_name: [arc_dicts]}
    
    for role_uri, content in role_to_content.items():
        role_name = role_uri.split('/')[-1]
        label_to_element = content['locs']
        
        parent_child = []
        for i, arc in enumerate(content['arcs']):
            p = label_to_element.get(arc['from'])
            c = label_to_element.get(arc['to'])
            if p and c:
                parent_child.append({
                    'parent': p,
                    'child': c,
                    'order': arc['order'],
                    'index': i,
                    'preferredLabel': arc.get('preferredLabel')
                })
                
        if not parent_child:
            continue
        
        # Original role
        statement_trees[role_name] = parent_child
        
        # Special logic for "jumbo" roles (e.g. Cabinet Office Ordinance Form 3)
        # These roles often contain many independent financial statements under specific Heading elements.
        jumbo_indicators = ['formno3', 'cabinetofficeordinance', 'annualsecuritiesreport']
        if any(ji in role_name.lower() for ji in jumbo_indicators):
            major_headings = [
                'ConsolidatedBalanceSheetHeading', 'ConsolidatedStatementOfIncomeHeading', 
                'ConsolidatedStatementOfCashFlowsHeading', 'ConsolidatedStatementOfChangesInEquityHeading',
                'BalanceSheetHeading', 'StatementOfIncomeHeading', 'StatementOfChangesInEquityHeading',
                'ConsolidatedStatementOfFinancialPositionIFRSHeading', 'ConsolidatedStatementOfProfitOrLossIFRSHeading',
                'ConsolidatedStatementOfCashFlowsIFRSHeading', 'ConsolidatedStatementOfChangesInEquityIFRSHeading',
                'ConsolidatedStatementOfFinancialPositionHeading', 
                'ConsolidatedStatementOfProfitOrLossHeading',
                'ConsolidatedStatementOfComprehensiveIncomeIFRSHeading',
                'SummaryOfBusinessResultsHeading', 'BusinessResultsOfGroupHeading', 'BusinessResultsOfReportingCompanyHeading'
            ]
            
            all_elements = label_to_element.values()
            for h in major_headings:
                # Look for ALL elements that end with the heading name (handles prefixes and underscores)
                h_elements = [el for el in all_elements if el.endswith(h)]
                
                for h_element in h_elements:
                    # Extract subtree starting from this heading
                    subtree_arcs = []
                    queue = [h_element]
                    seen = {h_element}
                    while queue:
                        curr_parent = queue.pop(0)
                        for arc in parent_child:
                            if arc['parent'] == curr_parent:
                                subtree_arcs.append(arc)
                                if arc['child'] not in seen:
                                    seen.add(arc['child'])
                                    queue.append(arc['child'])
                    
                    if subtree_arcs:
                        virtual_role = h_element
                        if virtual_role.endswith('Heading'):
                            virtual_role = virtual_role[:-7]
                        statement_trees[virtual_role] = subtree_arcs
                        
    return statement_trees

def parse_instance_contexts_and_units(xbrl_file, labels_map):
    vprint(f"Parsing XBRL contexts and units... {os.path.basename(xbrl_file)}")
    try:
        # Use lxml for robust namespace handling if available
        if HAS_LXML:
            # Secure parser against XXE attacks
            parser = etree.XMLParser(recover=True, resolve_entities=False, no_network=True)
            tree = etree.parse(xbrl_file, parser)
        else:
            tree = etree.parse(xbrl_file)
    except Exception as e:
        vprint(f"Error parsing XBRL instance: {e}")
        return {}, {}

    # Build suffix index for O(1) label lookups (performance optimization)
    # This converts O(N) suffix searches to O(1) hash lookups
    suffix_index = build_suffix_index(labels_map)

    # Standard namespaces for XBRL instance and dimensions
    ns = {
        "xbrli": "http://www.xbrl.org/2003/instance",
        "xbrldi": "http://xbrl.org/2006/xbrldi"
    }

    contexts = {}
    
    # 1. Parse contexts
    for ctx in safe_xpath(tree, "//xbrli:context", namespaces=ns):
        ctx_id = ctx.get('id')
        if not ctx_id:
            continue
            
        members = safe_xpath(ctx, ".//xbrldi:explicitMember", namespaces=ns)
        dim_vals = []
        for m in members:
            # Handle QNames in member text (e.g., jppfs_cor:EnergySegmentMember)
            m_text = m.text or ""
            
            # --- Axis Name Resolution ---
            dim_qname = m.get("dimension")
            dim_val = dim_qname.split(':')[-1] if dim_qname else ''
            
            # Resolve Axis Label
            prefixes = ['jpcrp_cor_', 'jppfs_cor_', 'jpigp_cor_', 'jpdei_cor_', '']
            axis_label = None
            if dim_val in COMMON_DIMENSION_MAPPING:
                axis_label = COMMON_DIMENSION_MAPPING[dim_val]
            else:
                for p in prefixes:
                    if p + dim_val in labels_map:
                        axis_label = clean_label(labels_map[p + dim_val])
                        break
                if not axis_label:
                    axis_label = convert_camel_case_to_title(dim_val.replace('Axis', '')) if dim_val else ''

            # --- Member Name Resolution ---
            member_val = m_text.split(':')[-1]
            label = None
            if member_val in COMMON_DIMENSION_MAPPING:
                label = COMMON_DIMENSION_MAPPING[member_val]
            else:
                # Try with standard prefixes first
                for p in prefixes:
                    if p + member_val in labels_map:
                        label = clean_label(labels_map[p + member_val])
                        break
                
                # If not found, use suffix index for O(1) lookup
                # (to catch standard elements from any taxonomy namespace)
                if not label and member_val in suffix_index:
                    _, label_text = suffix_index[member_val]
                    label = clean_label(label_text)

            # Fallback for company specific segment names found in _lab.xml
            if label: label = label.replace(' [メンバー]', '').replace(' [要素]', '').replace(' [区分]', '').strip()
            if not label and member_val in suffix_index:
                _, label = suffix_index[member_val]
            
            if label: label = label.replace(' [メンバー]', '').replace(' [要素]', '').replace(' [区分]', '').strip()
            if not label:
                if member_val.endswith('Member'):
                    label = convert_camel_case_to_title(member_val.replace('Member', ''))
                else:
                    label = member_val
            
            # Combine Axis and Member if useful
            skip_axes = ('報告セグメント', 'セグメント情報', '事業セグメント', '会計基準', '連結・単体', '連結・非連結', 
                         'ConsolidatedOrNonConsolidated', 'OperatingSegments', 'BusinessSegments', 'ReportableSegments')
            if axis_label and not any(sa in axis_label.replace(' ', '') for sa in skip_axes):
                dim_vals.append(f"{axis_label}：{label}")
            else:
                dim_vals.append(label)
                
        dim_str = "、".join(dim_vals) if dim_vals else "全体"
        # Clean up verbose XBRL labels
        dim_str = dim_str.replace('、報告セグメント', '').replace('非連結又は個別', '単体').replace('非連結', '単体')
        if dim_str == 'NonConsolidated' or dim_str == 'Non Consolidated':
            dim_str = '単体'
        if dim_str == 'Consolidated':
            dim_str = '連結'
            
        period_elem = safe_xpath(ctx, "xbrli:period", namespaces=ns)
        if period_elem:
            period_elem = period_elem[0]
            instant = safe_xpath(period_elem, "xbrli:instant", namespaces=ns)
            end_date = safe_xpath(period_elem, "xbrli:endDate", namespaces=ns)
            
            p_val = None
            start_val = None
            if instant:
                p_val = instant[0].text
            elif end_date:
                p_val = end_date[0].text
                start_elem = safe_xpath(period_elem, "xbrli:startDate", namespaces=ns)
                if start_elem:
                    start_val = start_elem[0].text
                
            if p_val:
                contexts[ctx_id] = (p_val, dim_str, start_val)
                
    units = {}
    # 2. Parse units
    for unit in safe_xpath(tree, "//xbrli:unit", namespaces=ns):
        unit_id = unit.get('id')
        if not unit_id:
            continue
        
        is_jpy = False
        # Only consider simple units (non‑divide) for JPY amount identification
        if not safe_xpath(unit, "xbrli:divide", namespaces=ns):
            measure = safe_xpath(unit, ".//xbrli:measure", namespaces=ns)
            if measure and 'JPY' in (measure[0].text or ""):
                is_jpy = True
                
        units[unit_id] = is_jpy
                
    return contexts, units

def parse_ixbrl_facts(ixbrl_files, contexts, units):
    t_start = time.time()
    parser_info = 'lxml' if HAS_LXML else 'html.parser'
    debug_log(f"Starting Inline XBRL parsing using {parser_info} for {len(ixbrl_files)} files")
    facts = []
    
    for f in ixbrl_files:
        size_mb = os.path.getsize(f) / (1024 * 1024)
        debug_log(f"  Parsing {os.path.basename(f)} ({size_mb:.2f} MB)...")
        try:
            with open(f, 'r', encoding='utf-8', errors='replace') as file:
                content = file.read()
            
            if HAS_LXML:
                try:
                    from lxml import html
                    # Secure parser against XXE attacks (Note: HTMLParser doesn't support resolve_entities)
                    parser = html.HTMLParser(no_network=True)
                    tree = html.fromstring(content, parser=parser)
                    # Use a more robust way to find tags that works with or without namespace awareness
                    tags = [t for t in tree.iter() if any(x in (t.tag if isinstance(t.tag, str) else "").lower() for x in ('nonfraction', 'nonnumeric'))]
                except Exception as e:
                    debug_log(f"  LXML fast-path failed: {e}. Falling back to BS4.")
                    HAS_LXML_LOCAL = False
                else:
                    HAS_LXML_LOCAL = True
            else:
                HAS_LXML_LOCAL = False

            if not HAS_LXML_LOCAL:
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(content, 'html.parser')
                def is_ix_tag(tag):
                    if not tag.name: return False
                    local = tag.name.split(':')[-1].lower()
                    return local in ('nonfraction', 'nonnumeric')
                tags = soup.find_all(is_ix_tag)

            elem_order_in_file = 0
            for tag in tags:
                if HAS_LXML_LOCAL:
                    # Access attributes robustly (handle namespaced keys like {uri}name or ix:name)
                    # LXML uses {uri}attribute_name format for namespaced attributes
                    attrs = {}
                    for k, v in tag.attrib.items():
                        # Extract local name: handle {uri}name and prefix:name
                        local_k = k
                        if '}' in local_k:
                            local_k = local_k.split('}')[-1]
                        if ':' in local_k:
                            local_k = local_k.split(':')[-1]
                        attrs[local_k.lower()] = v
                    ctx_ref = attrs.get('contextref')
                    if not ctx_ref or ctx_ref not in contexts: continue
                    
                    element_name = attrs.get('name')
                    if not element_name: continue
                    if ':' in element_name:
                        element_name = element_name.replace(':', '_')
                    
                    value = tag.text_content().strip() if hasattr(tag, 'text_content') else (tag.text or "").strip()
                    local_name = tag.tag.split('}')[-1].lower() if isinstance(tag.tag, str) and '}' in tag.tag else (tag.tag.split(':')[-1].lower() if isinstance(tag.tag, str) else "")
                    
                    unit_ref = attrs.get('unitref')
                    scale = attrs.get('scale', '0')
                    sign = attrs.get('sign', '')
                else:
                    # BS4 path
                    ctx_ref = None
                    for k, v in tag.attrs.items():
                        if k.lower() == 'contextref':
                            ctx_ref = v
                            break
                    if not ctx_ref or ctx_ref not in contexts: continue
                    
                    element_name = None
                    for k, v in tag.attrs.items():
                        if k.lower() == 'name':
                            element_name = v
                            break
                    if not element_name: continue
                    if ':' in element_name:
                        element_name = element_name.replace(':', '_')
                    
                    value = tag.get_text().strip()
                    local_name = tag.name.split(':')[-1].lower()

                if local_name == 'nonnumeric':
                    # Skip massive text blocks only if it's explicitly a TextBlock element
                    if 'TextBlock' in element_name:
                        continue
                
                valStr = ""
                if local_name == 'nonfraction':
                    if not HAS_LXML_LOCAL:
                        unit_ref = None
                        for k, v in tag.attrs.items():
                            if k.lower() == 'unitref':
                                unit_ref = v
                                break
                        scale = '0'
                        sign = ''
                        for k, v in tag.attrs.items():
                            if k.lower() == 'scale': scale = v
                            if k.lower() == 'sign': sign = v

                    is_jpy = units.get(unit_ref, False) if unit_ref else False
                    clean_val = value.replace(',', '').replace('△', '-').replace('▲', '-').replace('(', '-').replace(')', '').strip()
                    
                    try:
                        amt = float(clean_val)
                        if sign == '-': amt *= -1
                        amt *= (10 ** int(scale or 0))
                        if is_jpy: amt /= 1000000.0
                        valStr = str(int(amt)) if amt.is_integer() else str(amt)
                    except Exception:
                        valStr = value
                else:
                    valStr = value
                    
                f_data = {
                    'element': element_name,
                    'context': ctx_ref,
                    'period': contexts[ctx_ref][0],
                    'dimension': contexts[ctx_ref][1],
                    'value': valStr,
                    'source_file': f,
                    'elem_order': elem_order_in_file
                }
                if contexts[ctx_ref][2]: # start_date
                    f_data['start_date'] = contexts[ctx_ref][2]
                facts.append(f_data)
                elem_order_in_file += 1
            
            if elem_order_in_file > 0:
                vprint(f"  Extracted {elem_order_in_file} facts from {os.path.basename(f)}")
            
            if not HAS_LXML_LOCAL:
                soup.decompose()
                del soup
            else:
                del tree

        except Exception as e:
            debug_log(f"ERROR: Error parsing file {f}: {e}")
                
    debug_log(f"COMPLETED: Parsed all Inline XBRL facts in {time.time() - t_start:.2f}s")
    return facts


def parse_standard_xbrl_facts(f, contexts, units):
    """Parse facts from a standard XBRL instance (.xbrl file)."""
    facts = []
    if not f or not os.path.exists(f):
        return facts
    
    t_start = time.time()
    try:
        from lxml import etree
        tree = etree.parse(f)
        root = tree.getroot()
        
        # All non-metadata elements in the root are potential facts
        # Standard XBRL elements have contextRef and sometimes unitRef
        elem_order_in_file = 0
        for tag in root.iterchildren():
            # Skip metadata/infrastructure tags
            if '}' in tag.tag:
                local_name = tag.tag.split('}')[-1]
            else:
                local_name = tag.tag
            
            if local_name in ('context', 'unit', 'schemaRef', 'footnoteLink'):
                continue
            
            ctx_ref = tag.get('contextRef')
            if not ctx_ref or ctx_ref not in contexts:
                continue
            
            element_name = tag.tag
            if '}' in element_name:
                # Keep the prefix if possible or just use the local name with a synthetic prefix
                # In this script, we usually use prefix_localname
                # We need to find the prefix from the namespace map
                prefix = None
                ns_uri = tag.tag.split('}')[0][1:]
                for p, uri in tag.nsmap.items():
                    if uri == ns_uri:
                        prefix = p
                        break
                if prefix:
                    element_name = f"{prefix}_{local_name}"
                else:
                    element_name = local_name
            
            if ':' in element_name:
                element_name = element_name.replace(':', '_')
                
            value = (tag.text or "").strip()
            if not value and len(tag) > 0:
                # Handle cases where value might be in a child (rare in standard facts but possible)
                continue
            
            unit_ref = tag.get('unitRef')
            scale = tag.get('decimals', '0') # Standard XBRL often uses decimals instead of scale
            # But Hitachi's .xbrl might use scale if it's an extension?
            # Actually standard XBRL uses decimals or precision.
            # However, for simplicity and compatibility with the rest of the script:
            scale = tag.get('scale', '0') 
            sign = tag.get('sign', '')
            
            is_jpy = units.get(unit_ref, False) if unit_ref else False
            clean_val = value.replace(',', '').replace('△', '-').replace('▲', '-').replace('(', '-').replace(')', '').strip()
            
            valStr = ""
            try:
                if clean_val:
                    amt = float(clean_val)
                    if sign == '-': amt *= -1
                    # Note: Standard XBRL 'decimals' works differently than 'scale', 
                    # but many Japanese filings use 'scale' attribute even in .xbrl for consistency.
                    amt *= (10 ** int(scale or 0))
                    if is_jpy: amt /= 1000000.0
                    valStr = str(int(amt)) if amt.is_integer() else str(amt)
                else:
                    valStr = ""
            except Exception:
                valStr = value
                
            f_data = {
                'element': element_name,
                'context': ctx_ref,
                'period': contexts[ctx_ref][0],
                'dimension': contexts[ctx_ref][1],
                'value': valStr,
                'source_file': f,
                'elem_order': elem_order_in_file
            }
            if contexts[ctx_ref][2]: # start_date
                f_data['start_date'] = contexts[ctx_ref][2]
            facts.append(f_data)
            elem_order_in_file += 1
            
        debug_log(f"COMPLETED: Parsed {len(facts)} facts from standard XBRL {os.path.basename(f)} in {time.time() - t_start:.2f}s")
    except Exception as e:
        debug_log(f"ERROR: Error parsing standard XBRL {f}: {e}")
        
    return facts




def create_hierarchy(parent_child_arcs):
    """Create a flattened list representing the hierarchy traversal."""
    # Group by parent to easily find children
    adj = {}
    for arc in parent_child_arcs:
        p = arc['parent']
        if p not in adj: adj[p] = []
        adj[p].append(arc)
    
    # Sort children by presentation order (order attribute from linkbase)
    for p in adj:
        # Primary: order (explicit ordering from presentation linkbase, if defined)
        # Secondary: index (appearance order in XBRL)
        # Tertiary: child name (for stable sorting)
        adj[p].sort(key=lambda x: (x.get('order', 0), x.get('index', 0), x['child']))
        
    roots = set(arc['parent'] for arc in parent_child_arcs)
    children = set(arc['child'] for arc in parent_child_arcs)
    top_roots = sorted(list(roots - children))
    
    if not top_roots and parent_child_arcs:
        # If circular or no clear root, pick the parent of the first arc
        top_roots = [parent_child_arcs[0]['parent']]
        
    ordered_items = []
    seen = set()
    
    def traverse(node_name, path, depth, pref_label=None, ancestors=None):
        # Use a tuple of (node_name, pref_label) to allow the same element
        # to appear multiple times if it has different preferred labels
        # (common in Cash Flow for beginning/ending balance)
        node_id = (node_name, pref_label)
        if node_id in seen: return
        seen.add(node_id)

        full_path = path + "::" + node_name
        if pref_label:
            full_path += f"|{pref_label}"

        ordered_items.append((node_name, full_path, depth, pref_label))

        if node_name in adj:
            if ancestors is None:
                ancestors = set()
            if node_name in ancestors:
                return  # 循環参照を検出したらスキップ
            next_ancestors = ancestors | {node_name}
            for arc in adj[node_name]:
                traverse(arc['child'], full_path, depth + 1, arc.get('preferredLabel'), next_ancestors)
                
    for root in top_roots:
        traverse(root, "", 0)
        
    return ordered_items

def merge_sequences(master, new_seq):
    """Merge new_seq into master using 'append unknown items' logic.
    Since reports are processed newest to oldest, this ensures the latest order is at the front.
    """
    if not master: return new_seq
    res = list(master)
    for item in new_seq:
        if item and item not in res:
            res.append(item)
    return res

def merge_ordered_items(master, new_items):
    """Smart-merge create_hierarchy output tuples from new_items into master.

    Items are (node_name, full_path, depth, pref_label) tuples.
    Elements that exist in master are skipped.
    New elements are inserted at the position dictated by their neighbors in new_items
    that ARE already present in master — preserving relative order from the newer sequence.

    Called newest-year-first so the newest year's sequence is the authoritative order,
    and older years only contribute elements missing from newer years, placed correctly.
    """
    if not master:
        return list(new_items)

    # Index master by full_path for O(1) lookup
    path_to_pos = {item[1]: i for i, item in enumerate(master)}
    result = list(master)

    for ni, new_item in enumerate(new_items):
        new_path = new_item[1]
        if new_path in path_to_pos:
            continue  # Already present

        # Find the rightmost predecessor in new_items that is in result
        insert_after = -1
        for j in range(ni - 1, -1, -1):
            p = new_items[j][1]
            if p in path_to_pos:
                insert_after = path_to_pos[p]
                break

        # Find the leftmost successor in new_items that is in result
        insert_before = len(result)
        for j in range(ni + 1, len(new_items)):
            s = new_items[j][1]
            if s in path_to_pos:
                insert_before = path_to_pos[s]
                break

        insert_pos = min(insert_after + 1, insert_before)
        result.insert(insert_pos, new_item)

        # Rebuild index (insert shifts all positions at/after insert_pos)
        path_to_pos = {item[1]: i for i, item in enumerate(result)}

    return result

# ============================================================================
# CORE LAYER - Main Processing Pipeline
# ============================================================================
# 【将来の分割先】core/processor.py, core/pipeline.py
#
# この巨大関数（1357-3531行、約2200行）は以下のフェーズで構成されます:
#
# Phase 1: ファイル展開（1357-1600行付近）
#   - ZIP展開、ファイル検出
#   - 並列処理（ThreadPoolExecutor）
#   - 将来の分割先: load_phase()
#
# Phase 2: タクソノミ取得（1600-1700行付近）
#   - タクソノミラベルの取得
#   - 将来の分割先: taxonomy_phase()
#
# Phase 3: XBRL解析（1700-2300行付近）
#   - プレゼンテーション、コンテキスト、事実値の解析
#   - データマージ・重複排除
#   - 将来の分割先: parse_phase()
#
# Phase 4: Excel生成（2300-3500行付近）
#   - シート生成、データ書き込み、フォーマット
#   - 将来の分割先: output_phase() → excel/writer.py, sheets.py, formatter.py
#
# 分割時の最重要原則:
# 1. 各フェーズを完全独立関数にする
# 2. XBRLData（統一中間モデル）でデータを受け渡す
# 3. if文による分岐はStrategy Patternで解決（特にExcel層）
# ============================================================================

def process_xbrl_zips(zip_paths, output_dir=None):
    overall_start = time.time()
    if not zip_paths:
        return None
    zip_paths = sorted(zip_paths)
        
    global HAS_PANDAS, HAS_OPENPYXL
    # Delay loading heavy libraries inside the function
    try:
        import pandas as pd
        HAS_PANDAS = True
    except ImportError:
        pd = None
        HAS_PANDAS = False

    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        HAS_OPENPYXL = True
    except ImportError:
        Workbook = None
        HAS_OPENPYXL = False

    if not HAS_OPENPYXL:
        print("Error: openpyxl is not installed. Excel generation is impossible.", file=sys.stderr)
        return None
        
    global_element_period_values = {} # {element: {col_key: value}}
    merged_trees = {} # {role_name: {(parent, child): order}}
    per_year_role_arcs = {} # {role_name: [arcs_newest, arcs_year2, ...]} for smart ordering
    seen_children_in_role = {} # {role_name: set(children)}
    labels_map = {} # {element: label_text}
    labels_map_priorities = {} # {element: priority}
    master_member_seq = []
    
    periods_seen = set()
    all_facts = []  # Accumulate facts across all zips for fallback logic

    # Use provided output_dir for temp files if possible to avoid permission issues in /tmp
    parent_temp_dir = output_dir if output_dir and os.path.exists(output_dir) else None
    temp_base = tempfile.mkdtemp(dir=parent_temp_dir)

    try:
        # ========================================================================
        # Phase 1: ファイル展開・XBRL解析（並列処理）
        # ========================================================================
        # 【将来の分割先】load_phase() + parse_phase()
        #
        # 処理内容:
        # - 各ZIPファイルを並列で展開
        # - XBRLファイル（presentation, instance, iXBRL）を検出
        # - タクソノミラベルを取得
        # - プレゼンテーション階層、コンテキスト、事実値を解析
        # - スレッドごとに結果を集約
        #
        # 分割時の注意:
        # - ThreadPoolExecutor の管理は pipeline.py 内で隠蔽
        # - 各ワーカーは XBRLData を返すように変更
        # ========================================================================
        from concurrent.futures import ThreadPoolExecutor
        
        def process_single_zip(zip_idx, zip_path):
            thread_labels = {}
            thread_priorities = {}
            thread_facts = []
            thread_periods = set()
            thread_values = {} # {el: {col: val}}
            
            basename = os.path.basename(zip_path)
            debug_log(f"Starting worker for {basename}")
            if not os.path.exists(zip_path):
                return {'error': f"File not found: {basename}", 'zip_path': zip_path}
                
            extract_dir = os.path.join(temp_base, f"zip_{zip_idx}")
            
            # --- Retry logic for ZIP opening/extraction ---
            max_zip_retries = 2
            zip_ref = None
            for attempt in range(max_zip_retries + 1):
                try:
                    zip_ref = zipfile.ZipFile(zip_path, 'r')
                    check_zip_bomb(zip_ref)
                    
                    # Selective extraction
                    for info in zip_ref.infolist():
                        filename_lower = info.filename.lower()
                        if info.is_dir(): continue
                        
                        should_extract = (
                            (filename_lower.endswith('_lab.xml') and not filename_lower.endswith('_lab-en.xml')) or
                            filename_lower.endswith('_pre.xml') or
                            filename_lower.endswith('.xbrl') or
                            (filename_lower.endswith(('.htm', '.html')) and 'publicdoc' in filename_lower) or
                            filename_lower.endswith('manifest.xml')
                        )
                        if should_extract:
                            info.filename = info.filename.replace('\\', '/')
                            target_path = os.path.join(extract_dir, info.filename)
                            validate_zip_path(target_path, extract_dir)
                            zip_ref.extract(info, extract_dir)
                    break # Success
                except Exception as e:
                    if zip_ref: zip_ref.close()
                    if attempt < max_zip_retries:
                        debug_log(f"Retry ZIP extraction ({attempt+1}/{max_zip_retries}) for {basename}: {e}")
                        time.sleep(1)
                        if os.path.exists(extract_dir):
                            shutil.rmtree(extract_dir)
                        continue
                    else:
                        return {'error': f"Failed to extract {basename}: {str(e)}", 'zip_path': zip_path}
                finally:
                    if zip_ref: zip_ref.close()
                
            subdirs = [d for d in os.listdir(extract_dir) if os.path.isdir(os.path.join(extract_dir, d))]
            if len(subdirs) == 1:
                extract_dir = os.path.join(extract_dir, subdirs[0])
                
            xbrl_files = find_xbrl_files(extract_dir)
            if not xbrl_files:
                return None
                
            taxonomy_year = None
            if xbrl_files['pre']:
                with open(xbrl_files['pre'], 'r', encoding='utf-8') as f:
                    content = f.read(4000)
                    m = _RE_TAXONOMY_YEAR.search(content)
                    if m:
                        year_str = m.group(1)
                        taxonomy_year = '2021' if year_str == '2020' else year_str
            
            if taxonomy_year:
                # Auto-update edinet_taxonomy_dict.py if XBRL references a newer taxonomy year
                check_and_update_edinet_taxonomy(taxonomy_year)
                std_labels, std_priorities = get_standard_labels(taxonomy_year)
                thread_labels.update(std_labels)
                thread_priorities.update(std_priorities)

            # --- NEW: Detect Report-Level Accounting Standard (V13) ---
            report_std = None # Default: None (don't assume until detected)
            # Use list append + join for efficient string concatenation
            search_content_parts = []
            if xbrl_files.get('pre'):
                with open(xbrl_files['pre'], 'r', encoding='utf-8', errors='ignore') as f:
                    search_content_parts.append(f.read(40000).lower()) # Increased context
            if xbrl_files.get('ixbrl'):
                # Check first iXBRL file for standard indicators
                with open(xbrl_files['ixbrl'][0], 'r', encoding='utf-8', errors='ignore') as f:
                    search_content_parts.append(f.read(40000).lower())
            search_content = ''.join(search_content_parts)
            
            if 'jpigp' in search_content or 'ifrs.org' in search_content or 'ifrs-full' in search_content: 
                report_std = 'IFRS'
            elif 'jpusp' in search_content or 'us-gaap' in search_content: 
                report_std = 'US'
            elif 'jpmis' in search_content: 
                report_std = 'JMIS'
            elif 'jppfs' in search_content:
                report_std = 'JP'
            
            debug_log(f"  [DEBUG] Report standard detected as: {report_std} (from pre/ixbrl content)")


            for lf in xbrl_files.get('lab', []):
                local_labels, local_priorities = parse_labels_file(lf)
                for k, v in local_labels.items():
                    p = local_priorities.get(k, 99) - 1
                    if k not in thread_labels or p < thread_priorities.get(k, 100):
                        thread_labels[k] = v
                        thread_priorities[k] = p
            
            # Phase 2: Demote IFRS mapping priority
            for el_name, alias in IFRS_LABEL_MAPPING.items():
                if el_name not in thread_labels or 20 < thread_priorities.get(el_name, 100):
                    thread_labels[el_name] = alias
                    thread_priorities[el_name] = 20
            
            contexts, units = parse_instance_contexts_and_units(xbrl_files['xbrl'], thread_labels)
            
            # Phase 3: Selective Parsing (Case-insensitive extension and dual format support)
            public_doc_dir = os.path.dirname(xbrl_files['xbrl'])
            all_files = os.listdir(public_doc_dir)
            ix_files = []
            for f in all_files:
                fl = f.lower()
                if '_ixbrl' in fl and (fl.endswith('.htm') or fl.endswith('.html')):
                    ix_files.append(os.path.join(public_doc_dir, f))
            ix_files = sorted(ix_files)

            facts = parse_ixbrl_facts(ix_files, contexts, units) 
            # NEW: Also parse standard XBRL facts for non-consolidated data often missing from iXBRL
            std_facts = parse_standard_xbrl_facts(xbrl_files['xbrl'], contexts, units)
            
            # Merge facts, avoiding duplicates (favor Inline XBRL if both exist for same element/context)
            seen_facts = {}
            for f in facts:
                key = (f['element'], f['context'])
                seen_facts[key] = f
            
            for f in std_facts:
                key = (f['element'], f['context'])
                if key not in seen_facts:
                    facts.append(f)
                    seen_facts[key] = f

            thread_facts.extend(facts)
            debug_log(f"Worker for {os.path.basename(zip_path)} found {len(facts)} total facts ({len(std_facts)} from .xbrl)")
            
            for f in facts:
                el = f['element']
                period = f['period']
                dim = f.get('dimension', '')
                val = f['value']
                dim_label = dim if dim else "全体"
                
                # --- Granular Fact Tagging (V13) ---
                fact_std = None
                if el.startswith('jpigp_cor'): fact_std = 'IFRS'
                elif el.startswith('jppfs_cor'): fact_std = 'JP'
                elif el.startswith('jpusp_cor'): fact_std = 'US'
                elif el.startswith('jpmis_cor'): fact_std = 'JMIS'
                elif el.startswith('jpcrp_cor'):
                    if 'IFRS' in el: fact_std = 'IFRS'
                    elif 'USGAAP' in el: fact_std = 'US'
                    elif 'JMIS' in el: fact_std = 'JMIS'
                    else: fact_std = report_std # fallback to document standard for jpcrp elements (general metadata)
                else:
                    # Extension elements (e.g. E01766...)
                    fact_std = report_std
                
                # Use standard-aware column key to separate identical periods (e.g. 2020 JP vs 2020 IFRS)
                col_key = (fact_std, dim_label, period)
                if el not in thread_values: thread_values[el] = {}
                thread_values[el][col_key] = val
                thread_periods.add(col_key)
                # Store extra metadata (startDate) for periodStartLabel lookup
                if 'start_date' in f:
                    if '_metadata' not in thread_values: thread_values['_metadata'] = {}
                    thread_values['_metadata'][col_key] = f['start_date']
                
            trees = parse_presentation_linkbase(xbrl_files['pre'])

            from diversity_analysis import _extract_subsidiary_names_from_ixbrl
            subsidiary_row_names = _extract_subsidiary_names_from_ixbrl(ix_files)

            return {
                'labels': thread_labels,
                'priorities': thread_priorities,
                'facts': thread_facts,
                'periods': thread_periods,
                'values': thread_values,
                'trees': trees,
                'member_seq': [], # Will fill below
                'year': taxonomy_year,
                'report_std': report_std,
                'subsidiary_row_names': subsidiary_row_names,
            }

        # Multi-threading for performance (I/O and C-based lxml parsing)
        # Use a maximum of 4 workers to avoid memory exhaustion in CGI
        t_parallel_start = time.time()
        with ThreadPoolExecutor(max_workers=min(len(zip_paths), 4)) as executor:
            def process_single_zip_wrapper(p):
                try:
                    res = process_single_zip(p[0], p[1])
                    if res and 'error' not in res:
                        # Build suffix index for O(1) label lookups
                        res_suffix_index = build_suffix_index(res['labels'])

                        # Identify segment members in order from trees
                        local_seq = []
                        for role_name, arcs in res['trees'].items():
                            rn_lower = role_name.lower()
                            # Broaden detection to include Japanese terms and variants
                            if 'segment' in rn_lower or 'セグメント' in role_name or '事業' in role_name:
                                items = create_hierarchy(arcs)
                                for el, path, depth, pref in items:
                                    parts = el.split('_')
                                    base = parts[-1]
                                    label = None
                                    if base in COMMON_DIMENSION_MAPPING:
                                        label = COMMON_DIMENSION_MAPPING[base]
                                    else:
                                        for pr in ['', 'jpcrp_cor_', 'jppfs_cor_', 'jpigp_cor_', 'jpcrp030000-asr_']:
                                            if pr + base in res['labels']:
                                                label = res['labels'][pr + base]
                                                break
                                        
                                        if not label and base in res_suffix_index:
                                            # Use suffix index for O(1) lookup of company-specific members
                                            _, label = res_suffix_index[base]
                                    if label:
                                        label = clean_label(label)
                                        # Skip '全体' and headings that are likely just grouping nodes
                                        if label not in local_seq and label != '全体' and not el.endswith('Abstract') and not el.endswith('Heading'):
                                            local_seq.append(label)
                        res['member_seq'] = local_seq
                    return res
                except Exception as e:
                    debug_log(f"Worker failed for {p[1]}: {e}")
                    return {'error': str(e), 'zip_path': p[1]}
            results_raw = list(executor.map(process_single_zip_wrapper, enumerate(zip_paths)))

        errors = [r['error'] for r in results_raw if r and 'error' in r]
        results = [r for r in results_raw if r and 'error' not in r]

        debug_log(f"Parallel ZIP processing completed in {time.time() - t_parallel_start:.2f}s")

        # Sort results by taxonomy year DESCENDING to ensure latest structure is prioritized.
        # 同一taxonomy_yearのZIPが複数ある場合、実際の決算期（最大period）を2次キーにして
        # 新しい年度のZIPが先に処理されるようにする。これにより、新しいZIPの前期データが
        # segment_covered_dimsによって誤ってブロックされるのを防ぐ。
        t_merge_start = time.time()
        results = [r for r in results if r]
        for _res in results:
            _periods_in_res = [p for (_, _, p) in _res.get('periods', set()) if p]
            _res['_max_period'] = max(_periods_in_res, default='') if _periods_in_res else ''
        results.sort(key=lambda x: (str(x.get('year') or '0000'), x.get('_max_period', '')), reverse=True)

        for res in results:
            # Merge member sequences
            master_member_seq = merge_sequences(master_member_seq, res['member_seq'])

        report_stds = set()
        # セグメント事業区分ディメンションについて、新しいXBRLが既にカバーした
        # セグメントメンバーを古いXBRLで上書き/追加しないための辞書。
        # resultsは新しい年度順にソート済みなので、順番にカバー済み情報を蓄積する。
        #
        # セグメント軸の識別方法:
        #   parse_instance_contexts_and_units でセグメント軸（OperatingSegmentsAxis等）は
        #   skip_axes に含まれるため dim_str に「：」が入らない（例: "ゲーム事業"）。
        #   一方、株主資本変動・役員等の軸は「axis：member」形式（例: "Components Of Equity：..."）。
        #   よって dim_label に「：」が含まれるものはセグメント事業区分ではない。
        #
        # ブロック戦略: (fact_std, period) に対して新しいXBRLがカバーしたdim_labelの集合を記録。
        # 古いXBRLのdim_labelがその集合に含まれない場合のみブロックする。
        # これにより、セグメント区分が変わった年度の前期データ混入を防ぎつつ、
        # 新しいXBRLが持たない要素（例: 売上収益）を古いXBRLから補完できる。
        _NON_SEGMENT_DIMS = frozenset({'全体', '連結', '単体', '個別', ''})
        def _is_segment_dim(dim_label):
            """セグメント事業区分ディメンションか判定する。"""
            if dim_label in _NON_SEGMENT_DIMS:
                return False
            if '：' in dim_label:  # 「axis：member」形式 → セグメント事業区分ではない
                return False
            return True
        # --- 各ZIPの当期/前期ペアを構築（古→新の昇順）---
        # results は新しい年度順ソート済みなので reversed で古→新の順に走査する
        # current_dims / prior_dims: そのZIPの当期・前期に実在したセグメントdim名の集合
        # これをPPMシートでのフィルタリングに使い、他のZIPからのデータ混入を防ぐ
        filing_pairs = []
        for res in reversed(results):
            seg_periods = set()
            seg_dim_by_period = {}   # period_str -> set of dim_labels
            for el, el_vals in res.get('values', {}).items():
                if el == '_metadata':
                    continue
                for (std, dim, period) in el_vals.keys():
                    if _is_segment_dim(dim):
                        seg_periods.add(period)
                        if period not in seg_dim_by_period:
                            seg_dim_by_period[period] = set()
                        seg_dim_by_period[period].add(dim)
            sorted_p = sorted(seg_periods)
            if len(sorted_p) >= 2:
                cur_p = sorted_p[-1]
                pri_p = sorted_p[-2]
                filing_pairs.append({
                    'current': cur_p,
                    'prior': pri_p,
                    'current_dims': seg_dim_by_period.get(cur_p, set()),
                    'prior_dims':   seg_dim_by_period.get(pri_p, set()),
                })
            elif len(sorted_p) == 1:
                cur_p = sorted_p[0]
                filing_pairs.append({
                    'current': cur_p,
                    'prior': None,
                    'current_dims': seg_dim_by_period.get(cur_p, set()),
                    'prior_dims':   set(),
                })

        # taxonomy_year が同一のZIPが複数ある場合、reversed(results) では時系列順が保証されないため
        # current 日付の昇順にソートして確定的な年度順にする
        filing_pairs.sort(key=lambda fp: fp.get('current') or '')

        # (fact_std, period) -> set of dim_labels covered by newer XBRLs
        segment_covered_dims = {}  # type: dict

        # 各XBRLの「当期会計年度」を事前計算する。
        # 当期データはセグメント構造変更時でもブロックしない（前期データのみブロック対象）。
        # これにより、セグメント変更過渡年度において、新旧両方のセグメント構造データが
        # 分析シートに保持され、PPMシートがdimsフィルタで正しい列を選択できるようになる。
        for res in results:
            seg_periods_r = set()
            for el_r, el_vals_r in res.get('values', {}).items():
                if el_r == '_metadata':
                    continue
                for (std_r, dim_r, period_r) in el_vals_r.keys():
                    if _is_segment_dim(dim_r):
                        seg_periods_r.add(period_r)
            sorted_p_r = sorted(seg_periods_r)
            res['_ppm_cur_period'] = sorted_p_r[-1] if sorted_p_r else None

        # 一番古い財務諸表の当期末を特定する（resultsは新しい年度順ソート済み）。
        # 最古のZIPのみ前期・当期両方のデータを使用し、それ以外は当期データのみを使用する。
        # これにより、勘定科目名が変更された場合の前期データ二重計上を防ぐ。
        #
        # 注意: _max_period は提出日（例: 2025-06-19）が最大になる場合がある。
        # XBRLには FilingDateInstant 等の提出日コンテキストが含まれ、監査書類を含む場合は
        # 提出日の出現頻度が高くなるため、頻度ベースの判定は誤りやすい。
        #
        # より確実な判定方法:
        # 提出日は Instant コンテキストのみに使われ、Duration（期間）コンテキストには現れない。
        # 一方、財政年度末は損益計算書・CF計算書等の Duration ファクトの終了日として必ず使われる。
        # よって、Duration ファクト（start_date が存在するファクト）の終了日の最大値を当期末とする。
        def _calc_fs_cur_period(res):
            """当期財務期末日を算出する。
            Durationファクトの終了日（= 財政年度末）の最大値を返す。
            提出日は Instant のみに使われるため自動的に除外される。"""
            metadata = res.get('values', {}).get('_metadata', {})
            duration_end_dates = set()
            for col_key, start_date in metadata.items():
                if start_date:  # start_date が存在する = Duration ファクト
                    _, _, period = col_key
                    if period:
                        duration_end_dates.add(period)
            if duration_end_dates:
                return max(duration_end_dates)
            # フォールバック: Duration ファクトがない場合は _max_period を使う
            return res.get('_max_period', '')
        for _res in results:
            _res['_fs_cur_period'] = _calc_fs_cur_period(_res)
        debug_log(f"[PriorPeriodFilter] total results={len(results)}")
        for _res in results:
            debug_log(f"[PriorPeriodFilter]   zip _max_period={_res.get('_max_period')}, _fs_cur_period={_res.get('_fs_cur_period')}")

        def _get_precision(s):
            if not s or '.' not in s: return 0
            return len(s.split('.')[-1])

        def _merge_value(el, col_key, new_val):
            """global_element_period_valuesへ精度優先でマージする。"""
            if el not in global_element_period_values:
                global_element_period_values[el] = {}
            old_val = global_element_period_values[el].get(col_key)
            if old_val is None:
                global_element_period_values[el][col_key] = new_val
            elif _get_precision(new_val) > _get_precision(old_val):
                global_element_period_values[el][col_key] = new_val

        # ── Pass 1: ラベル・ファクト・当期データのみをマージ ───────────────────────
        # 各ZIPから _fs_cur_period（当期末日）に一致するperiodのデータのみを登録する。
        # 勘定科目名が変更された場合、前期データを新旧両方の要素名で重複登録しないための処置。
        for res in results:
            if res.get('report_std'):
                report_stds.add(res['report_std'])

            # Merge labels with priorities
            for k, v in res['labels'].items():
                p = res['priorities'].get(k, 100)
                if k not in labels_map or p < labels_map_priorities.get(k, 101):
                    labels_map[k] = v
                    labels_map_priorities[k] = p

            all_facts.extend(res['facts'])
            periods_seen.update(res['periods'])

            res_cur_period = res.get('_ppm_cur_period')
            res_max_period = res.get('_fs_cur_period', '')

            for el, vals in res['values'].items():
                for col_key, new_val in vals.items():
                    fact_std, dim_label, period = col_key
                    # セグメントフィルタ（既存ロジック維持）
                    if (el != '_metadata'
                            and _is_segment_dim(dim_label)
                            and period != res_cur_period
                            and (fact_std, period) in segment_covered_dims
                            and dim_label not in segment_covered_dims[(fact_std, period)]):
                        continue
                    # 当期データのみ登録（前期はPass 2で補完）
                    if el != '_metadata' and res_max_period and period != res_max_period:
                        continue
                    _merge_value(el, col_key, new_val)

            # セグメントカバー情報を更新（Pass 2のセグメントフィルタで使用）
            for col_key in res['periods']:
                _fs, _dim, _per = col_key
                if _is_segment_dim(_dim):
                    key = (_fs, _per)
                    if key not in segment_covered_dims:
                        segment_covered_dims[key] = set()
                    segment_covered_dims[key].add(_dim)

        # ── Pass 2: 前期データをギャップ補完 ──────────────────────────────────────
        # 例: IFRS開示がFY2019から始まった場合、FY2018のデータはFY2019申告の前期として取得する。
        # これにより、財務諸表の種類ごとに「一番古い当期を持つZIP」が自動的に異なっても正しく動作する。
        #
        # ギャップ判定の単位: (fact_std, dim_label, period) の組み合わせ。
        # いずれかのZIPの当期としてこの組み合わせがカバーされている場合は前期データをスキップする。
        # こうすることで、勘定科目名が変わった年度でも二重計上を防げる:
        #   例: 2024/3期のデータが OldElement (202403当期) に存在するとき、
        #       NewElement (202503前期) の 2024/3期データは (JP,全体,2024-03-31) がカバー済みなのでスキップ。
        # 財務諸表コア要素のプレフィックス（勘定科目として扱う要素）
        # jpcrp_cor は主要経営指標等（サマリー）要素を含むため除外する。
        # 例: jpcrp_cor:RevenueIFRSSummaryOfBusinessResults は IFRS 開示前の ZIP にも含まれており、
        #     これを covered_by_current に含めると、後続 ZIP の IFRS 財務諸表ギャップ補完を誤ってブロックする。
        _FS_CORE_PREFIXES = ('jpigp_cor_', 'jppfs_cor_', 'jpusp_cor_', 'jpmis_cor_')

        covered_by_current = set()  # Pass 1 でカバーされた (fact_std, dim_label, period)
        for res in results:
            res_max_period = res.get('_fs_cur_period', '')
            if not res_max_period:
                continue
            for el, vals in res['values'].items():
                if el == '_metadata':
                    continue
                # 財務諸表コア要素のみ対象（jpcrp_cor 等のサマリー要素は除外）
                if not any(el.startswith(p) for p in _FS_CORE_PREFIXES):
                    continue
                for col_key in vals:
                    fact_std, dim_label, period = col_key
                    if period == res_max_period:
                        covered_by_current.add(col_key)

        for res in results:
            res_cur_period = res.get('_ppm_cur_period')
            res_max_period = res.get('_fs_cur_period', '')

            for el, vals in res['values'].items():
                for col_key, new_val in vals.items():
                    fact_std, dim_label, period = col_key
                    # セグメントフィルタ（既存ロジック維持）
                    if (el != '_metadata'
                            and _is_segment_dim(dim_label)
                            and period != res_cur_period
                            and (fact_std, period) in segment_covered_dims
                            and dim_label not in segment_covered_dims[(fact_std, period)]):
                        continue
                    # 前期データのみ対象（当期はPass 1で処理済み）
                    if el == '_metadata' or not res_max_period or period == res_max_period:
                        continue
                    # (fact_std, dim_label, period) が当期カバー済みならスキップ
                    # ※ 勘定科目名変更によるOldElement/NewElement二重計上を防ぐ
                    # カバー済み判定は jpigp_cor_ 等のコア要素で行うが（covered_by_current の構築時）、
                    # ブロックはすべての要素に適用する。
                    # こうすることで、jpcrp030000-asr_ 等の会社固有拡張要素も二重計上を防げる。
                    if col_key in covered_by_current:
                        continue
                    _merge_value(el, col_key, new_val)
            
            # Merge presentation trees
            for role, tree_arcs in res['trees'].items():
                base_name = role.split('_')[-1]

                # Debug: Log BusinessResults roles
                if 'BusinessResults' in base_name:
                    debug_log(f"[Presentation-Tree] Found role {role} with {len(tree_arcs)} arcs from {res.get('year', 'unknown')}")
                    sample_arcs = list(tree_arcs)[:3]
                    for arc in sample_arcs:
                        debug_log(f"  arc: child={arc['child']}, order={arc['order']}")
                # Merge SegmentInformation variants into a single role
                sub_role_idx = 0
                if 'SegmentInformation' in base_name and '-' in base_name:
                    parts = base_name.rsplit('-', 1)
                    if parts[1].isdigit():
                        sub_role_idx = int(parts[1]) * 1000
                    role = parts[0]
                
                # Filter relevant roles
                is_accepted = (base_name.startswith('Consolidated') or base_name.startswith('Statement') or 
                               base_name.startswith('BalanceSheet') or base_name.startswith('Notes') or 
                               'BusinessResults' in base_name or 'SegmentInformation' in base_name or 
                               'AnalysisOfOperatingResults' in base_name)
                
                if not is_accepted:
                    continue
                
                # Normalize standalone roles: NonConsolidatedBalanceSheet -> BalanceSheet
                IFRS_ROLE_MAP = {
                    'StatementOfFinancialPosition': 'BalanceSheet',
                    'StatementOfProfitOrLoss': 'StatementOfIncome',
                    'StatementOfComprehensiveIncome': 'StatementOfComprehensiveIncome',
                    'StatementOfChangesInEquity': 'StatementOfChangesInEquity',
                    'StatementOfCashFlows': 'StatementOfCashFlows',
                }
                
                curr_base = base_name.replace('NonConsolidated', '')
                if curr_base in IFRS_ROLE_MAP:
                    new_base = IFRS_ROLE_MAP[curr_base]
                    role = role.replace(base_name, new_base)
                elif base_name.startswith('NonConsolidated'):
                    new_base = base_name[15:]
                    role = role.replace(base_name, new_base)
                
                if role not in merged_trees:
                    merged_trees[role] = {}
                    seen_children_in_role[role] = set()

                # Collect per-year arcs for smart ordering (skip SegmentInformation
                # which uses sub_role_idx offsets to separate sub-roles)
                if 'SegmentInformation' not in base_name:
                    if role not in per_year_role_arcs:
                        per_year_role_arcs[role] = []
                    per_year_role_arcs[role].append(list(tree_arcs))

                for arc in tree_arcs:
                    p, c, o, i, pl = arc['parent'], arc['child'], arc['order'], arc.get('index', 0), arc.get('preferredLabel')
                    # Unique key including preferredLabel to allow duplicates in CF statements
                    arc_key = (p, c, pl)
                    # Collect all arcs from all years into merged_trees for parent-child relationships.
                    # Ordering is handled separately via per_year_role_arcs + merge_ordered_items.
                    if arc_key not in merged_trees[role]:
                        merged_trees[role][arc_key] = (float(o) + sub_role_idx, i)

        # --- Build element-to-statement-type mapping (FIX V7 - IMPROVED) ---
        # Use a smarter approach: if an element appears in multiple statement types,
        # remove it from the mapping (it's a shared element like Abstract, Heading, etc.)
        element_to_statement_type = {}  # {element_name: statement_type or None}

        for role_name, arcs_dict in merged_trees.items():
            # Determine statement type from role name
            statement_type = None
            base_name = role_name.split('_')[-1]

            # Map role base name to statement type
            if 'ConsolidatedBalanceSheet' in base_name or 'BalanceSheet' in base_name:
                statement_type = 'BalanceSheet'
            elif 'ConsolidatedStatementOfIncome' in base_name or 'StatementOfIncome' in base_name:
                statement_type = 'StatementOfIncome'
            elif 'ConsolidatedStatementOfCashFlows' in base_name or 'StatementOfCashFlows' in base_name:
                statement_type = 'StatementOfCashFlows'
            elif 'ConsolidatedStatementOfChangesInEquity' in base_name or 'StatementOfChangesInEquity' in base_name or 'StatementOfChangesInNetAssets' in base_name:
                statement_type = 'StatementOfChangesInEquity'
            elif 'ConsolidatedStatementOfComprehensiveIncome' in base_name or 'StatementOfComprehensiveIncome' in base_name:
                statement_type = 'StatementOfComprehensiveIncome'
            elif 'Notes' in base_name or 'Segment' in base_name:
                statement_type = 'Notes'
            elif 'BusinessResults' in base_name:
                statement_type = 'BusinessResults'

            # For each element in this role, record or update its statement type
            if statement_type:
                for (parent, child, _), _ in arcs_dict.items():
                    # Process both parent and child
                    for element in [parent, child]:
                        if not element:
                            continue

                        # Skip structural elements that are legitimately shared across statements
                        # (Axis, Member, Abstract, Heading, TextBlock, LineItems, Table)
                        structural_suffixes = ('Axis', 'Member', 'Abstract', 'Heading', 'TextBlock', 'LineItems', 'Table')
                        if any(element.endswith(suffix) for suffix in structural_suffixes):
                            continue

                        if element in element_to_statement_type:
                            # Element already seen in another role
                            existing_type = element_to_statement_type[element]
                            if existing_type is not None:  # Not yet marked as shared
                                # Only mark as shared if both types are main financial statements (not Notes)
                                # Notes often reference main statement elements, but that shouldn't disqualify them
                                main_statement_types = {'BalanceSheet', 'StatementOfIncome', 'StatementOfCashFlows',
                                                       'StatementOfChangesInEquity', 'StatementOfComprehensiveIncome'}
                                if (existing_type in main_statement_types and
                                    statement_type in main_statement_types and
                                    existing_type != statement_type):
                                    # Different main statement types - mark as shared (None)
                                    element_to_statement_type[element] = None
                                    debug_log(f"  [Mapping] Element '{element}' appears in multiple statement types ({existing_type}, {statement_type}) - marked as shared")
                                # else: same type or one is Notes, keep existing mapping
                        else:
                            # First time seeing this element
                            element_to_statement_type[element] = statement_type

        # Count unique (non-shared) elements
        unique_elements = sum(1 for v in element_to_statement_type.values() if v is not None)
        shared_elements = sum(1 for v in element_to_statement_type.values() if v is None)
        debug_log(f"Merged total: {len(all_facts)} facts, {len(periods_seen)} periods, {len(merged_trees)} tree roles")
        debug_log(f"Element mapping: {unique_elements} unique elements, {shared_elements} shared elements")
    except Exception as e:
        debug_log(f"ERROR: Overall processing failure: {e}")
        import traceback
        debug_log(traceback.format_exc())
    finally:
        shutil.rmtree(temp_base)

    # Pre-collect elements with facts to identify non-abstract items (abstract=false)
    # This is used to allow relevant extension tags in major statements.
    elements_with_facts = {f['element'] for f in all_facts}

    # --- Fallback for old EDINET format (e.g. 2016-2018) ---
    # build synthetic roles from the element appearance order in the known ixbrl files.
    EDINET_DOC_ROLE_MAP = {
        '0105010': 'rol_ConsolidatedBalanceSheet',
        '0105020': 'rol_ConsolidatedStatementOfIncome',
        '0105025': 'rol_ConsolidatedStatementOfComprehensiveIncome',
        '0105040': 'rol_ConsolidatedStatementOfChangesInNetAssets',
        '0105050': 'rol_ConsolidatedStatementOfCashFlows',
        # Notes and Accounting Policies
        '0106010': 'rol_NotesAccountingPolicies',
        '0107010': 'rol_Notes',
        # Segment Information
        '0114010': 'rol_SegmentInformation',
    }

    # Always try to capture facts from these critical documents as a fallback for structure
    roles_to_fill = EDINET_DOC_ROLE_MAP

    if roles_to_fill:
        # First, build a map of element -> presentation linkbase order from merged_trees
        # This ensures we preserve the correct order from presentation linkbase
        elem_presentation_order = {}  # {element: min_presentation_order}
        for role, arcs_dict in merged_trees.items():
            for (parent, child, _), (order_val, _) in arcs_dict.items():
                if child and child not in elem_presentation_order:
                    elem_presentation_order[child] = order_val
                elif child and order_val < elem_presentation_order[child]:
                    elem_presentation_order[child] = order_val

        # Debug: Show some of the presentation orders
        debug_log(f"[Presentation-Order-Map] Built map with {len(elem_presentation_order)} elements")
        elements_with_facts = set(f['element'] for f in all_facts if f['value'].strip() != "")
            
        sample_elements = list(elem_presentation_order.items())[:10]
        for el, order in sample_elements:
            debug_log(f"  {el}: {order}")

        facts_by_doc = {}  # {doc_code: {element: min_order}}
        for f in all_facts:
            src = f.get('source_file', '')
            fname = os.path.basename(src)
            for doc_code in roles_to_fill:
                if re.match(r'^' + doc_code, fname):
                    if doc_code not in facts_by_doc:
                        facts_by_doc[doc_code] = {}
                    el = f['element']
                    # Prefer presentation linkbase order, fallback to elem_order (file appearance order)
                    if el in elem_presentation_order:
                        order = elem_presentation_order[el]
                    else:
                        order = f.get('elem_order', 0) + 10000  # Add offset to ensure presentation orders come first
                    if el not in facts_by_doc[doc_code] or order < facts_by_doc[doc_code][el]:
                        facts_by_doc[doc_code][el] = order

        for doc_code, role_name in roles_to_fill.items():
            if doc_code not in facts_by_doc:
                continue
            elem_order_map = facts_by_doc[doc_code]
            if not elem_order_map:
                continue

            # Debug: Show first few elements from this document
            debug_log(f"[Fallback-Build] Processing doc_code={doc_code}, role={role_name}, {len(elem_order_map)} elements")
            sample_elems = list(elem_order_map.items())[:5]
            for el, order in sample_elems:
                debug_log(f"  {el}: order={order}")

            sorted_elems = sorted(elem_order_map.items(), key=lambda x: x[1])
            
            # For combined filings (typically 0105010), split into separate statements
            # Auto-detect IFRS vs J-GAAP based on element name prefixes/keywords
            if doc_code == '0105010':
                is_ifrs_filing = any(
                    elem.startswith('jpigp') or 'IFRS' in elem
                    for elem in elem_order_map.keys()
                )
                
                if is_ifrs_filing:
                    headings_to_roles = {
                        'ConsolidatedStatementOfFinancialPositionIFRSHeading': 'rol_ConsolidatedStatementOfFinancialPositionIFRS',
                        'ConsolidatedStatementOfProfitOrLossIFRSHeading': 'rol_ConsolidatedStatementOfProfitOrLossIFRS',
                        'ConsolidatedStatementOfProfitOrLossAndOtherComprehensiveIncomeIFRSHeading': 'rol_ConsolidatedStatementOfProfitOrLossIFRS',
                        'ConsolidatedStatementOfCashFlowsIFRSHeading': 'rol_ConsolidatedStatementOfCashFlowsIFRS',
                        'ConsolidatedStatementOfChangesInEquityIFRSHeading': 'rol_ConsolidatedStatementOfChangesInEquityIFRS',
                        'StatementOfFinancialPositionIFRSHeading': 'rol_ConsolidatedStatementOfFinancialPositionIFRS',
                        'StatementOfProfitOrLossIFRSHeading': 'rol_ConsolidatedStatementOfProfitOrLossIFRS',
                        'NotesIFRSHeading': 'rol_NotesIFRS',
                        'SegmentInformationIFRSHeading': 'rol_SegmentInformationIFRS',
                        'RelatedInformationIFRSHeading': 'rol_RelatedInformationIFRS',
                        'InformationAboutReportableSegmentsIFRSHeading': 'rol_SegmentInformationIFRS',
                        'OperatingSegmentsIFRSHeading': 'rol_SegmentInformationIFRS',
                        'BusinessSegmentInformationIFRSHeading': 'rol_SegmentInformationIFRS',
                        'PropertyPlantAndEquipmentIFRSHeading': 'rol_PPEIFRS',
                        'IntangibleAssetsIFRSHeading': 'rol_IntangibleAssetsIFRS',
                        'InventoriesIFRSHeading': 'rol_InventoriesIFRS',
                        'FinancialInstrumentsIFRSHeading': 'rol_FinancialInstrumentsIFRS',
                        # Backup markers
                        'RevenueIFRS': 'rol_ConsolidatedStatementOfProfitOrLossIFRS',
                        'NetSalesIFRS': 'rol_ConsolidatedStatementOfProfitOrLossIFRS',
                        'NetCashProvidedByUsedInOperatingActivitiesIFRS': 'rol_ConsolidatedStatementOfCashFlowsIFRS',
                        'RetainedEarningsIFRS': 'rol_ConsolidatedStatementOfChangesInEquityIFRS',
                        'SegmentInformationIFRS': 'rol_SegmentInformationIFRS',
                    }
                    default_role = 'rol_ConsolidatedStatementOfFinancialPositionIFRS'
                else:
                    # J-GAAP combined filing — use standard J-GAAP role names
                    headings_to_roles = {
                        'ConsolidatedBalanceSheetTextBlock': 'rol_ConsolidatedBalanceSheet',
                        'ConsolidatedStatementOfIncomeTextBlock': 'rol_ConsolidatedStatementOfIncome',
                        'ConsolidatedStatementOfComprehensiveIncomeTextBlock': 'rol_ConsolidatedStatementOfComprehensiveIncome',
                        'ConsolidatedStatementOfChangesInNetAssetsTextBlock': 'rol_ConsolidatedStatementOfChangesInNetAssets',
                        'ConsolidatedStatementOfCashFlowsTextBlock': 'rol_ConsolidatedStatementOfCashFlows',
                        # Heading-style markers
                        'ConsolidatedBalanceSheetHeading': 'rol_ConsolidatedBalanceSheet',
                        'ConsolidatedStatementOfIncomeHeading': 'rol_ConsolidatedStatementOfIncome',
                        'ConsolidatedStatementOfComprehensiveIncomeHeading': 'rol_ConsolidatedStatementOfComprehensiveIncome',
                        'ConsolidatedStatementOfChangesInEquityHeading': 'rol_ConsolidatedStatementOfChangesInNetAssets',
                        'ConsolidatedStatementOfCashFlowsHeading': 'rol_ConsolidatedStatementOfCashFlows',
                        'NotesHeading': 'rol_Notes',
                    }
                    default_role = 'rol_ConsolidatedBalanceSheet'
                    print(f"[Fallback-Split] Detected J-GAAP filing for 0105010", file=sys.stderr)
                
                curr_role = default_role
                curr_arcs = []
                roles_created = set()
                
                for i, (elem, _order) in enumerate(sorted_elems):
                    base = elem.split('_')[-1]
                    if base in headings_to_roles:
                        new_role = headings_to_roles[base]
                        if new_role != curr_role:
                            if curr_arcs:
                                # Merge instead of Overwrite
                                # IMPORTANT: Preserve presentation linkbase order values (don't overwrite)
                                if curr_role not in merged_trees: merged_trees[curr_role] = {}
                                for arc in curr_arcs:
                                    key = (arc['parent'], arc['child'], arc['preferredLabel'])
                                    # Only add if not already exists (prioritize presentation linkbase)
                                    if key not in merged_trees[curr_role]:
                                        merged_trees[curr_role][key] = (arc['order'], arc['index'])
                                print(f"[Fallback-Split] Merged synthetic role {curr_role} (Phase 1)", file=sys.stderr)
                                roles_created.add(curr_role)
                            curr_role = new_role
                            curr_arcs = []
                    
                    # Namespace & Element Filter (Revision 3):
                    # For major financial statements (BS, PL, CF), only allow standard namespace elements.
                    # AND skip items that look like they belong in segments or detailed notes.
                    if any(x in curr_role for x in ('BalanceSheet', 'StatementOfIncome', 'StatementOfCashFlows', 'FinancialPosition', 'ProfitOrLoss')):
                        # 1. Namespace Check: Handle Standard Namespaces (jpcrp, jppfs, jpigp, jpusp, jpmis)
                        # Elements usually look like common_prefix_elementName
                        prefix = elem.split('_')[0] if '_' in elem else ""
                        standard_prefixes = ('jpcrp_cor', 'jppfs_cor', 'jpigp_cor', 'jpusp_cor', 'jpmis_cor')
                        is_standard_ns = any(p == prefix for p in standard_prefixes)

                        # Special case for jpigp (sometimes it might not have _cor if it's a heading?)
                        # Actually standard elements almost always have _cor.
                        if not is_standard_ns and prefix in ('jpcrp', 'jppfs', 'jpigp', 'jpusp', 'jpmis'):
                             # Backup check for prefix without _cor if it's very standard
                             is_standard_ns = True

                        # Extension namespaces (e.g. jpcrp030000-asr_E01737-000)
                        # [主表判定] Allow extension elements in main statements IF they have actual 
                        # fact values (non-abstract / abstract=false).
                        # Also allow specifically for Hitachi's TradeReceivablesAndContractAssets
                        if ('E' in prefix and '-' in prefix) or 'TradeReceivablesAndContractAssets' in elem:
                            if elem in elements_with_facts:
                                is_standard_ns = True # Explicitly allow if it has values
                            else:
                                is_standard_ns = False

                        # 2. Detail/Note Blacklist
                        el_lower = elem.lower()
                        is_detail_item = any(x in el_lower for x in (
                            'segment', 'externalcustomer', 'revenuefromexternal',
                            'acquisitioncost', 'accumulateddepreciation', 'accumulatedamortization',
                            'rawmaterials', 'workinprocess', 'merchandise', 'finishedgoods'
                        ))

                        # 3. CashFlow-specific filter (V13)
                        # For CashFlow statements, only allow CF-related elements
                        if 'CashFlow' in curr_role:
                            is_cf_element = any(suffix in elem for suffix in [
                                'OpeCFIFRS', 'InvCFIFRS', 'FinCFIFRS',  # IFRS CF suffixes
                                'OpeCF', 'InvCF', 'FinCF', 'CCE',       # JP-GAAP CF suffixes
                                'CashFlow', 'CashAndCashEquivalents'    # Generic CF terms
                            ])
                            is_structural_elem = any(keyword in elem for keyword in [
                                'Abstract', 'Heading', 'Table', 'LineItems', 'Axis', 'Member'
                            ])
                            # Skip non-CF elements unless structural
                            if not is_cf_element and not is_structural_elem:
                                debug_log(f"  [CF-Fallback-Filter] Skipped non-CF element: {elem}")
                                continue

                        if not is_standard_ns or is_detail_item:
                            # Skip this element - it's either an extension namespace (likely detail) or blacklisted detail
                            continue

                    curr_arcs.append({'parent': curr_role, 'child': elem, 'order': float(_order), 'index': i, 'preferredLabel': None})
                if curr_arcs:
                    # IMPORTANT: Preserve presentation linkbase order values (don't overwrite)
                    if curr_role not in merged_trees: merged_trees[curr_role] = {}
                    for arc in curr_arcs:
                        key = (arc['parent'], arc['child'], arc['preferredLabel'])
                        # Only add if not already exists (prioritize presentation linkbase)
                        if key not in merged_trees[curr_role]:
                            merged_trees[curr_role][key] = (arc['order'], arc['index'])
                    print(f"[Fallback-Split] Merged synthetic role {curr_role} (Phase 1)", file=sys.stderr)
            else:
                virtual_root = role_name
                arcs = []

                # Debug: Log first few elements for this role
                if 'BusinessResults' in role_name and len(sorted_elems) > 0:
                    debug_log(f"[Fallback-Simple] Creating {role_name} from {doc_code}, {len(sorted_elems)} elements")
                    sample_count = min(5, len(sorted_elems))
                    for i in range(sample_count):
                        elem, _order = sorted_elems[i]
                        debug_log(f"  {elem}: order={_order}")

                for i, (elem, _order) in enumerate(sorted_elems):
                    arcs.append({'parent': virtual_root, 'child': elem, 'order': float(_order), 'index': i, 'preferredLabel': None})

                if arcs:
                    # IMPORTANT: Preserve presentation linkbase order values (don't overwrite)
                    if role_name not in merged_trees: merged_trees[role_name] = {}
                    for arc in arcs:
                        key = (arc['parent'], arc['child'], arc['preferredLabel'])
                        # Only add if not already exists (prioritize presentation linkbase)
                        if key not in merged_trees[role_name]:
                            merged_trees[role_name][key] = (arc['order'], arc['index'])
                    print(f"[Fallback] Merged synthetic role {role_name} from {doc_code} (Phase 1)", file=sys.stderr)

    # --- Clean up stub taxonomy roles from jumbo roles (FIX V5) ---
    # Jumbo roles create virtual taxonomy roles (e.g., jppfs_cor_ConsolidatedBalanceSheet)
    # but these often only contain Heading and TextBlock elements.
    # If a fallback role exists with substantially more elements, remove the stub taxonomy role.
    # BUT FIRST: Merge order information from taxonomy role to fallback role to preserve correct ordering.
    DEDUP_PREFIXES = ('jppfs_cor_', 'jpigp_cor_', 'jpcrp_cor_', 'rol_')
    roles_to_clean = set()
    for role_name in list(merged_trees.keys()):
        # Only check taxonomy roles (not fallback roles)
        if role_name.startswith('rol_'):
            continue

        # Extract base name
        for pfx in ('jppfs_cor_', 'jpigp_cor_', 'jpcrp_cor_'):
            if role_name.startswith(pfx):
                base = role_name[len(pfx):]
                break
        else:
            continue  # Not a standard taxonomy role

        # Count elements in this taxonomy role
        tax_elem_count = len(merged_trees[role_name])

        # Check if a fallback role with the same base exists and has more elements
        fallback_role = 'rol_' + base.replace('-indirect', '').replace('-direct', '')
        if fallback_role in merged_trees:
            fallback_elem_count = len(merged_trees[fallback_role])

            # If taxonomy role is a stub (<=5 elements) and fallback has substantially more (>20),
            # remove the taxonomy role and keep the fallback
            if tax_elem_count <= 5 and fallback_elem_count > 20:
                # BEFORE deleting: Copy order information from taxonomy role to fallback role
                # This preserves the correct presentation order from presentation linkbase
                debug_log(f"[Stub-Cleanup] Analyzing {role_name} ({tax_elem_count} elems) vs {fallback_role} ({fallback_elem_count} elems)")

                # Build a map of child elements and their orders from the taxonomy role
                tax_child_orders = {}
                for (p, c, pl), (order_val, idx) in merged_trees[role_name].items():
                    if c:
                        tax_child_orders[c] = order_val
                        debug_log(f"  [Tax-Order] {c}: order={order_val}")

                # Update orders in fallback role for matching elements
                updates_count = 0
                for fb_key in list(merged_trees[fallback_role].keys()):
                    fb_p, fb_c, fb_pl = fb_key
                    if fb_c in tax_child_orders:
                        fb_order, fb_idx = merged_trees[fallback_role][fb_key]
                        new_order = tax_child_orders[fb_c]
                        merged_trees[fallback_role][fb_key] = (new_order, fb_idx)
                        debug_log(f"  [Order-Merge] Updated order for {fb_c}: {fb_order} -> {new_order}")
                        updates_count += 1

                debug_log(f"  [Order-Merge] Updated {updates_count} elements in {fallback_role}")
                roles_to_clean.add(role_name)
                debug_log(f"[Stub-Cleanup] Removing stub taxonomy role {role_name} ({tax_elem_count} elems) in favor of fallback role {fallback_role} ({fallback_elem_count} elems)")

    for role_name in roles_to_clean:
        del merged_trees[role_name]

    debug_log(f"Data merging and tree processing completed in {time.time() - t_merge_start:.2f}s")

    # Build hierarchical data structure for Excel sheets
    t_hierarchy_start = time.time()
    all_years_data = {} # {role_name: {hierarchical_key: {period: value}}}
    role_to_order = {} # {role_name: [hierarchical_key1, ...]}
    
    for role, pd_dict in merged_trees.items():
        tree_arcs = [{'parent': p, 'child': c, 'order': o_i[0], 'index': o_i[1], 'preferredLabel': pl}
                     for (p, c, pl), o_i in pd_dict.items()]

        # Debug: Log arcs for BusinessResults roles
        base_name = role.split('_')[-1]
        if 'BusinessResults' in base_name:
            debug_log(f"[Merged-Trees-Arcs] Role {role} has {len(tree_arcs)} arcs")
            for arc in tree_arcs:
                if 'Revenue' in arc['child'] or 'ProfitLoss' in arc['child'] or 'Tax' in arc['child']:
                    debug_log(f"  parent={arc['parent']}, child={arc['child']}, order={arc['order']}, pref={arc.get('preferredLabel')}")

        # Build ordered_items by smart-merging per-year sequences (newest year first).
        # This ensures the latest year's order is authoritative, while elements that
        # only appear in older years are inserted at their correct relative positions
        # (between the same neighboring elements as in the older year's linkbase).
        # For SegmentInformation (uses sub_role_idx), fall back to merged_trees ordering.
        if role in per_year_role_arcs:
            ordered_items = []
            for year_arcs in per_year_role_arcs[role]:
                year_ordered = create_hierarchy(year_arcs)
                ordered_items = merge_ordered_items(ordered_items, year_ordered)
        else:
            ordered_items = create_hierarchy(tree_arcs)

        # Debug: Log ordered items for BusinessResults roles
        if 'BusinessResults' in base_name and len(ordered_items) > 0:
            debug_log(f"[Hierarchy-Result] Role {role} has {len(ordered_items)} ordered items")
            sample_count = min(10, len(ordered_items))
            for i in range(sample_count):
                el, full_path, depth, pref_label = ordered_items[i]
                if 'Revenue' in el or 'ProfitLoss' in el or 'Tax' in el:
                    debug_log(f"  {i}: {el}")

        # FIX: For Cash Flow statements, ensure section totals appear AFTER their detail items
        # When merging multiple years, detail items from older years may have higher order values
        # than the total from the newest year, causing the total to appear before details
        base_name = role.split('_')[-1]
        if 'CashFlow' in base_name:
            # Define Cash Flow section total elements that must appear last within their sections
            # We identify sections by the total element itself, and find its siblings
            cf_section_totals = [
                'NetCashProvidedByUsedInOperatingActivities',
                'NetCashProvidedByUsedInInvestingActivities',
                'NetCashProvidedByUsedInFinancingActivities',
            ]

            # For each section total element, ensure it appears after all its sibling detail items
            for total_element_suffix in cf_section_totals:
                # Find the total element
                total_item = None
                total_item_index = None
                total_parent_path = None

                for i, (el, full_path, depth, pref_label) in enumerate(ordered_items):
                    # Check if this is the total element (ends with the total element name)
                    if el.endswith(total_element_suffix):
                        total_item = (el, full_path, depth, pref_label)
                        total_item_index = i
                        # Extract parent path (everything before the last "::")
                        if '::' in full_path:
                            total_parent_path = '::'.join(full_path.split('::')[:-1])
                        break

                # If we found a total element, find all its siblings (same parent, same or higher depth)
                if total_item and total_item_index is not None and total_parent_path is not None:
                    sibling_items = []

                    for i, (el, full_path, depth, pref_label) in enumerate(ordered_items):
                        # Skip the total element itself
                        if i == total_item_index:
                            continue

                        # Check if this item has the same parent path (i.e., it's a sibling)
                        # We identify siblings as items that have the total's parent in their path
                        # and are at the same or deeper level
                        if total_parent_path in full_path:
                            # Make sure it's a detail item, not a sub-section total
                            # (e.g., it shouldn't be another section's total or abstract)
                            is_detail = True
                            for other_total in cf_section_totals:
                                if other_total != total_element_suffix and el.endswith(other_total):
                                    is_detail = False
                                    break

                            if is_detail:
                                sibling_items.append(i)

                    # If there are siblings after the total, we need to move the total to the end
                    if sibling_items:
                        last_sibling_index = max(sibling_items)

                        # If the total appears before the last sibling, move it
                        if total_item_index < last_sibling_index:
                            # Remove the total from its current position
                            ordered_items.pop(total_item_index)

                            # Recalculate sibling positions after removal
                            new_sibling_items = []
                            for i, (el, full_path, depth, pref_label) in enumerate(ordered_items):
                                if i == total_item_index:
                                    continue

                                if total_parent_path in full_path:
                                    is_detail = True
                                    for other_total in cf_section_totals:
                                        if other_total != total_element_suffix and el.endswith(other_total):
                                            is_detail = False
                                            break

                                    if is_detail:
                                        new_sibling_items.append(i)

                            # Insert the total after the last sibling
                            if new_sibling_items:
                                insert_pos = max(new_sibling_items) + 1
                                ordered_items.insert(insert_pos, total_item)
                            else:
                                # If no siblings found after removal, just append at the end
                                ordered_items.append(total_item)

        # Determine this role's statement type for filtering
        current_role_type = None
        base_name = role.split('_')[-1]
        if 'ConsolidatedBalanceSheet' in base_name or 'BalanceSheet' in base_name:
            current_role_type = 'BalanceSheet'
        elif 'ConsolidatedStatementOfIncome' in base_name or 'StatementOfIncome' in base_name:
            current_role_type = 'StatementOfIncome'
        elif 'ConsolidatedStatementOfCashFlows' in base_name or 'StatementOfCashFlows' in base_name:
            current_role_type = 'StatementOfCashFlows'
        elif 'ConsolidatedStatementOfChangesInEquity' in base_name or 'StatementOfChangesInEquity' in base_name:
            current_role_type = 'StatementOfChangesInEquity'
        elif 'ConsolidatedStatementOfComprehensiveIncome' in base_name or 'StatementOfComprehensiveIncome' in base_name:
            current_role_type = 'StatementOfComprehensiveIncome'

        all_years_data[role] = {}
        role_to_order[role] = []

        for el, full_path, depth, pref_label in ordered_items:
            # Filter elements based on statement type mapping (FIX V9 - SKIP UNMAPPED, BREAK ON MISMATCH)
            # Skip elements not in mapping (shared/structural elements)
            # Stop only when a mapped element has a different statement type
            should_stop = False

            # --- STRICT FILTER: Explicit cross-statement element blacklist ---
            # Some XBRL files incorrectly include elements from other statements
            # Use explicit element name matching to filter these out (avoids false positives)

            # P/L elements that should never appear in Balance Sheet
            if current_role_type == 'BalanceSheet':
                pl_element_patterns = (
                    'OperatingRevenue1', 'OperatingRevenue2',  # 営業収益
                    'NetSales', 'OrdinaryIncome', 'OrdinaryLoss',  # 売上高、経常利益/損失
                    'OperatingProfit', 'OperatingLoss',  # 営業利益/損失
                    'GrossProfit', 'GrossLoss',  # 売上総利益/損失
                    'ProfitBeforeTax', 'LossBeforeTax',  # 税引前利益/損失
                    'BasicEarningsPerShare', 'DilutedEarningsPerShare',  # 1株当たり利益
                )
                if any(el.endswith(pattern) for pattern in pl_element_patterns):
                    debug_log(f"  [BS-Filter] Skipping P/L element '{el}' in BalanceSheet role '{role}'")
                    should_stop = True

            # Balance Sheet elements that should never appear in P/L
            elif current_role_type == 'StatementOfIncome':
                bs_element_patterns = (
                    'CashAndDeposits', 'CashAndCashEquivalents',  # 現金預金
                    'NotesAndAccountsReceivable', 'AccountsReceivable',  # 受取手形・売掛金
                    'Inventories', 'MerchandiseAndFinishedGoods',  # 棚卸資産
                    'PropertyPlantAndEquipment', 'IntangibleAssets',  # 有形固定資産、無形固定資産
                    'TotalAssets', 'TotalLiabilities',  # 資産合計、負債合計
                    'NotesAndAccountsPayable', 'AccountsPayable',  # 支払手形・買掛金
                    'TotalEquity', 'ShareCapital', 'RetainedEarnings',  # 純資産、資本金、利益剰余金
                )
                if any(el.endswith(pattern) for pattern in bs_element_patterns):
                    debug_log(f"  [PL-Filter] Skipping BS element '{el}' in StatementOfIncome role '{role}'")
                    should_stop = True

            if current_role_type and el in element_to_statement_type:
                element_type = element_to_statement_type[el]

                if element_type is None:
                    # Element is not in mapping (shared/structural) - skip judgment, continue output
                    debug_log(f"  [Skip-Judgment] Element '{el}' not in mapping (shared) - skipping judgment, continuing output")
                    # Do NOT stop, just continue to next element
                elif element_type != current_role_type and element_type != 'Notes':
                    # Element belongs to a different specific statement type - stop here
                    # EXCEPTION: Do NOT stop for P/L elements (GrossProfit, OperatingProfit, etc.)
                    # These may appear in multiple roles due to XBRL structure, but should not cause early termination
                    pl_element_suffixes = (
                        'GrossProfit', 'GrossLoss', 'OperatingProfit', 'OperatingLoss',
                        'OrdinaryIncome', 'OrdinaryLoss', 'ProfitBeforeTax', 'LossBeforeTax',
                        'NetSales', 'OperatingRevenue', 'Revenue',
                        'SellingGeneralAndAdministrativeExpenses',  # 販売費及び一般管理費
                        'NonOperatingIncome', 'NonOperatingExpenses',  # 営業外損益
                        'ExtraordinaryIncome', 'ExtraordinaryLosses'  # 特別損益
                    )
                    if any(el.endswith(suffix) for suffix in pl_element_suffixes):
                        debug_log(f"  [Type-Filter-Skip] P/L element '{el}' type mismatch ignored (expected: {current_role_type}, mapped: {element_type})")
                    else:
                        debug_log(f"  [Type-Filter] Found {element_type} element '{el}' in {current_role_type} role '{role}' - stopping output")
                        should_stop = True

            if should_stop:
                break

            role_to_order[role].append((full_path, pref_label))
            all_years_data[role][full_path] = {}
            if el in global_element_period_values:
                for period, val in global_element_period_values[el].items():
                    all_years_data[role][full_path][period] = val

    # --- Deduplicate overlapping roles (Fix B - Refined) ---
    # Group roles by their fundamental base name (ignoring prefixes like jppfs_cor_ and suffixes like -indirect)
    DEDUP_PREFIXES = ('jppfs_cor_', 'jpigp_cor_', 'jpcrp_cor_', 'rol_')
    roles_by_base = {}
    for r in role_to_order.keys():
        base = r
        for pfx in DEDUP_PREFIXES:
            if r.startswith(pfx):
                base = r[len(pfx):]
                break
        # Normalize base for matching: strip common suffix variants
        # Note: Do NOT strip IFRS/JMIS/US here if we want them as separate sheets
        match_base = base
        for sfx in ('-indirect', '-direct'):
            if match_base.endswith(sfx): match_base = match_base[:-len(sfx)]
        
        if match_base not in roles_by_base: roles_by_base[match_base] = []
        roles_by_base[match_base].append(r)
        
    roles_to_remove = set()
    for match_base, roles in roles_by_base.items():
        if len(roles) <= 1: continue

        # Pick a primary role to merge into
        # PRIORITY (REVISED V4):
        # After stub cleanup, stubs are gone. Prefer whichever role has more structure.
        # Generally fallback roles have fuller document-order structure.
        # 1. Fallback synthetic roles (rol_) - often have complete structure from document order
        # 2. Roles with -indirect suffix (for CF statements)
        # 3. Roles from standard taxonomy (jppfs_cor_, etc.) - use if fallback doesn't exist
        primary = None
        taxonomy_roles = [r for r in roles if not r.startswith('rol_')]
        fallback_roles = [r for r in roles if r.startswith('rol_')]

        # Prefer fallback roles (complete structure)
        if fallback_roles:
            for r in sorted(fallback_roles, key=lambda x: (0 if '-indirect' in x else 1, len(x))):
                if primary is None: primary = r
                if '-indirect' in r:
                    primary = r
                    break

        # Use taxonomy roles if no fallback
        if not primary and taxonomy_roles:
            for r in sorted(taxonomy_roles, key=lambda x: (0 if '-indirect' in x else 1, len(x))):
                if primary is None: primary = r
                if '-indirect' in r:
                    primary = r
                    break

        # Final fallback (should not happen)
        if not primary:
            primary = sorted(roles)[0]
        
        for r in roles:
            if r == primary: continue
            # Merge structure paths
            existing_paths = {p for p, _ in role_to_order[primary]}
            major_roles = ('ConsolidatedBalanceSheet', 'ConsolidatedStatementOfIncome', 'ConsolidatedStatementOfCashFlows', 'ConsolidatedStatementOfChangesInEquity', 
                           'BalanceSheet', 'StatementOfIncome', 'StatementOfCashFlows', 'StatementOfChangesInEquity')
            is_major = any(mr in primary for mr in major_roles)
            
            for full_path_data in role_to_order[r]:
                fp, pl = full_path_data
                if fp not in existing_paths:
                    # Conservative Merge (V4):
                    # With fallback cleanup, fallback roles should not contaminate taxonomy roles.
                    # Standard dedup logic applies.
                    is_synthetic_primary = primary.startswith('rol_')
                    if is_major and r.startswith('rol_') and (not is_synthetic_primary or len(existing_paths) > 20) and len(existing_paths) > 5:
                        continue

                    # Additional filter for CashFlow statements to exclude non-CF elements (V13)
                    # This prevents notes elements (e.g., sales expenses, finance income breakdown)
                    # from being merged into the CF statement
                    if 'CashFlow' in primary:
                        # Check if this is a CF-related element
                        is_cf_element = any(suffix in fp for suffix in [
                            'OpeCFIFRS', 'InvCFIFRS', 'FinCFIFRS',  # IFRS CF suffixes
                            'OpeCF', 'InvCF', 'FinCF', 'CCE',       # JP-GAAP CF suffixes
                            'CashFlow', 'CashAndCashEquivalents'    # Generic CF terms
                        ])
                        # Allow structural elements (Abstract, Heading, Table, Axis, Member)
                        is_structural = any(keyword in fp for keyword in [
                            'Abstract', 'Heading', 'Table', 'LineItems', 'Axis', 'Member'
                        ])

                        # Skip non-CF elements unless they are structural
                        if not is_cf_element and not is_structural:
                            debug_log(f"  [CF-Filter] Skipped non-CF element in CashFlow role: {fp}")
                            continue

                    role_to_order[primary].append(full_path_data)
                    existing_paths.add(fp)
            # Merge data values
            for fp, period_vals in all_years_data.get(r, {}).items():
                # No filtering needed here for now - data should follow structure
                # The structure merge above handles filtering

                # Apply same CF filter to data values (V13)
                if 'CashFlow' in primary:
                    is_cf_element = any(suffix in fp for suffix in [
                        'OpeCFIFRS', 'InvCFIFRS', 'FinCFIFRS', 'OpeCF', 'InvCF', 'FinCF', 'CCE',
                        'CashFlow', 'CashAndCashEquivalents'
                    ])
                    is_structural = any(keyword in fp for keyword in [
                        'Abstract', 'Heading', 'Table', 'LineItems', 'Axis', 'Member'
                    ])
                    if not is_cf_element and not is_structural:
                        continue  # Skip non-CF data values

                if fp not in all_years_data[primary]:
                    all_years_data[primary][fp] = {}
                for period, val in period_vals.items():
                    if period not in all_years_data[primary][fp]:
                        all_years_data[primary][fp][period] = val
            roles_to_remove.add(r)
            debug_log(f"[Dedup] Merged {r} into primary {primary} (Base: {match_base})")
            
    for r in roles_to_remove:
        del role_to_order[r]
        if r in all_years_data:
            del all_years_data[r]

    # Try to find company name for filename
    company_name = "企業名不明"
    name_suffixes = ['CompanyNameCoverPage', 'EntityNameCompanyName', 'EntityNameEntityReportingName']
    for suffix in name_suffixes:
        found = False
        for el_name, vals in global_element_period_values.items():
            if el_name.endswith(suffix):
                if vals:
                    # Pick a value (latest period if possible)
                    sorted_keys = sorted(vals.keys(), key=lambda x: x[1] if isinstance(x, tuple) else x, reverse=True)
                    company_name = vals[sorted_keys[0]]
                    found = True
                    break
        if found: break
    
    # Clean company name for filename (remove "株式会社", "（株）", "(株)")
    company_name = company_name.replace("株式会社", "").replace("（株）", "").replace("(株)", "").strip()
    
    # Debug: log top elements to see what was captured
    if VERBOSE_LOGGING or company_name == "企業名不明":
        top_el = sorted(list(global_element_period_values.keys()))[:30]
        debug_log(f"DEBUG: Company discovery failed. Top 30 elements: {top_el}")

    debug_log(f"Hierarchical data structure built in {time.time() - t_hierarchy_start:.2f}s")

    # ========================================================================
    # Phase 2: Excel生成
    # ========================================================================
    # 【将来の分割先】output_phase() → excel/writer.py, sheets.py, formatter.py
    #
    # 処理内容:
    # - Workbook作成
    # - シート生成（ロールごと）
    # - データ書き込み
    # - フォーマット適用（列幅、罫線、数値書式）
    # - ファイル保存
    #
    # 分割時の最重要原則:
    # 1. if文による分岐は禁止 → Strategy Pattern で解決
    # 2. writer.py: Workbook生成のみ
    # 3. sheets.py: シート構築（SHEET_BUILDERS辞書で振り分け）
    # 4. formatter.py: 見た目のみ（データ判断禁止）
    # ========================================================================
    t_excel_start = time.time()
    print(f"Generating Excel for {company_name}...", file=sys.stderr)
    # Note: write_only=True is faster but incompatible with sheet merging and formatting
    # Current implementation requires normal mode for merge operations
    wb = Workbook()
    default_sheet_removed = False

    # List to collect segment sheet information for later processing
    segment_sheets_info = []

    # Identify periods and standards for sheet planning
    t_sheet_planning_start = time.time()
    # Identify periods that are standalone (not consolidated)
    periods_with_standalone = set()
    for role, ordered_keys_dict in all_years_data.items():
        for full_path, p_dict in ordered_keys_dict.items():
            for c in p_dict.keys():
                # c is now (standard, dim, period)
                std, dim, period = c if len(c) == 3 else ("JP", c[0], c[1])
                if dim == '単体':
                    periods_with_standalone.add(period)
                    
    # Identify accounting standards for each period separately for Consolidated and Non-consolidated
    consolidated_standards = {} 
    non_consolidated_standards = {}

    # --- Refined Standard Detection (V13) ---
    # Use the fact_std already tagged in the keys (c)
    for el_name, vals in global_element_period_values.items():
        for c in vals:
            # c is (fact_std, dim, p)
            if len(c) == 3:
                f_std, dim, p = c
                if f_std:
                    if dim in ('全体', '連結', '全社', '連結財務諸表計上額'):
                        if p not in consolidated_standards: consolidated_standards[p] = set()
                        consolidated_standards[p].add(f_std)
                    elif dim == '単体':
                        if p not in non_consolidated_standards: non_consolidated_standards[p] = set()
                        non_consolidated_standards[p].add(f_std)
    
    # Global Fallback: Only if NO standards were detected for a period that has data
    all_explicit_stds = set()
    for s_set in consolidated_standards.values(): all_explicit_stds.update(s_set)
    for s_set in non_consolidated_standards.values(): all_explicit_stds.update(s_set)
    
    if not all_explicit_stds:
        # Fallback to JP only if document-wide report_std is JP or generic
        # or if absolutely nothing was found.
        # This is for pure JP-GAAP reports where no prefixes are found.
        for el_name, vals in global_element_period_values.items():
            for c in vals:
                # c is (fact_std, dim, p)
                f_std, dim, p = c if len(c) == 3 else (None, c[0], c[1])
                if dim in ('全体', '連結', '全社', '連結財務諸表計上額'):
                    if p not in consolidated_standards: consolidated_standards[p] = set(['JP'])
                elif dim == '単体':
                    if p not in non_consolidated_standards: non_consolidated_standards[p] = set(['JP'])

    debug_log(f"Consolidated Standards: {consolidated_standards}")
    debug_log(f"Non-consolidated Standards: {non_consolidated_standards}")

    sorted_periods = sorted(list(periods_seen))
    
    used_sheet_names = set()
    
    # Determine which standards to create for each role
    all_role_work = []
    
    # Pre-calculate all detected standards for the document (V13)
    doc_cons_stds = set()
    for s_set in consolidated_standards.values(): doc_cons_stds.update(s_set)
    doc_cons_stds = sorted([s for s in doc_cons_stds if s and s != 'JP_ALL'])
    if not doc_cons_stds: doc_cons_stds = ['JP']
    
    doc_noncons_stds = set()
    for s_set in non_consolidated_standards.values(): doc_noncons_stds.update(s_set)
    doc_noncons_stds = sorted([s for s in doc_noncons_stds if s and s != 'JP_ALL'])
    if not doc_noncons_stds: doc_noncons_stds = ['JP']

    # Select representive report_std from zips
    report_std = None
    if 'IFRS' in doc_cons_stds: report_std = 'IFRS'
    elif 'US' in doc_cons_stds: report_std = 'US'
    elif 'JMIS' in doc_cons_stds: report_std = 'JMIS'
    elif 'JP' in doc_cons_stds: report_std = 'JP'

    for role, ordered_keys in role_to_order.items():
        base_name = role.split('_')[-1]
        
        # --- Role-Based Standard Detection (V13) ---
        # 1. Scan elements in this role to see which standards are explicitly used
        role_detected_stds = set()
        for full_path, _ in ordered_keys:
            el_name = full_path.split('/')[-1]
            if el_name.startswith('jpigp_cor'): role_detected_stds.add('IFRS')
            elif el_name.startswith('jppfs_cor'): role_detected_stds.add('JP')
            elif el_name.startswith('jpusp_cor'): role_detected_stds.add('US')
            elif el_name.startswith('jpmis_cor'): role_detected_stds.add('JMIS')
            elif el_name.startswith('jpcrp_cor'):
                if 'IFRS' in el_name: role_detected_stds.add('IFRS')
                elif 'USGAAP' in el_name: role_detected_stds.add('US')
                elif 'JMIS' in el_name: role_detected_stds.add('JMIS')
        
        # 2. Decide standards to try for this role
        if base_name in ('SummaryOfBusinessResults', 'BusinessResultsOfGroup'):
            # Consolidated summaries need all detected consolidated standards
            standards_to_try = doc_cons_stds
        elif base_name == 'BusinessResultsOfReportingCompany':
            standards_to_try = ['JP_ALL']
        elif 'SegmentInformation' in base_name or 'AnalysisOfOperatingResults' in base_name:
            if 'IFRS' in base_name or 'IFRS' in role:
                standards_to_try = ['IFRS']
            else:
                # Use role-specific detection if possible, fallback to doc consolidated
                standards_to_try = sorted([s for s in role_detected_stds if s])
                if not standards_to_try: standards_to_try = doc_cons_stds
        else:
            is_standalone = 'Consolidated' not in base_name
            if is_standalone:
                # Standalone roles use non-consolidated standards (always 'JP_ALL' for now but kept dynamic)
                standards_to_try = ['JP_ALL']
            else:
                # Consolidated Financial Statements (BS, PL, CF)
                if 'IFRS' in base_name: standards_to_try = ['IFRS']
                elif 'JMIS' in base_name: standards_to_try = ['JMIS']
                elif 'US' in base_name: standards_to_try = ['US']
                else:
                    # Generic roles: use only detections within THIS role's tree
                    standards_to_try = sorted([s for s in role_detected_stds if s])
                    if not standards_to_try:
                        # Fallback for generic roles: prefer report standard, then doc consolidated standards
                        if report_std and report_std in doc_cons_stds:
                            standards_to_try = [report_std]
                        else:
                            standards_to_try = doc_cons_stds
        
        for std in standards_to_try:
            all_role_work.append((role, ordered_keys, std))

    debug_log(f"Sheet planning completed in {time.time() - t_sheet_planning_start:.2f}s ({len(all_role_work)} sheets to process)")

    # Pre-calculate PL minimum period per standard to filter CF and CE
    pl_start_periods = {}
    for r, o_keys, c_std in all_role_work:
        b_name = r.split('_')[-1]
        if 'ConsolidatedStatementOfIncome' in b_name or 'ConsolidatedStatementOfProfitOrLoss' in b_name or 'StatementOfIncome' in b_name or 'StatementOfProfitOrLoss' in b_name:
            for fp_data in o_keys:
                fp, _ = fp_data
                if fp in all_years_data[r]:
                    for c in all_years_data[r][fp].keys():
                        f_std, f_dim, f_p = c if len(c) == 3 else ("JP", c[0], c[1])
                        if f_dim not in ('全体', '連結', '全社', '単体', '個別'):
                            continue
                        if c_std != 'JP_ALL' and f_std is not None and f_std != c_std:
                            continue
                        
                        period = c[2] if len(c) == 3 else c[1]
                        if c_std not in pl_start_periods or period < pl_start_periods[c_std]:
                            pl_start_periods[c_std] = period

    # Generate Excel sheets
    t_sheet_generation_start = time.time()
    for role, ordered_keys, current_standard in all_role_work:
        # Clean role name for sheet
        base_name = role.split('_')[-1]

        # REORDERING: For Summary of Business Results (Reporting Company), 
        # ensure RevenueKeyFinancialData follows NetSalesSummaryOfBusinessResults
        if base_name == 'BusinessResultsOfReportingCompany':
            temp_keys = list(ordered_keys)
            target_el = 'jpcrp_cor_RevenueKeyFinancialData'
            base_el = 'jpcrp_cor_NetSalesSummaryOfBusinessResults'
            
            target_idx = -1
            base_idx = -1
            for i, (fp, _) in enumerate(temp_keys):
                if base_el in fp: base_idx = i
                if target_el in fp: target_idx = i
                
            if base_idx != -1 and target_idx != -1 and target_idx != base_idx + 1:
                item = temp_keys.pop(target_idx)
                if target_idx < base_idx: base_idx -= 1
                temp_keys.insert(base_idx + 1, item)
                ordered_keys = temp_keys

        sheet_mapping = {
            'ConsolidatedBalanceSheet': '連結貸借対照表',
            'ConsolidatedStatementOfIncome': '連結損益計算書',
            'ConsolidatedStatementOfComprehensiveIncome': '連結包括利益計算書',
            'ConsolidatedStatementOfChangesInEquity': '連結株主資本等変動計算書',
            'ConsolidatedStatementOfChangesInNetAssets': '連結株主資本等変動計算書',
            'ConsolidatedStatementOfCashFlows': '連結キャッシュ・フロー計算書',
            'ConsolidatedStatementOfCashFlows-indirect': '連結キャッシュ・フロー計算書',
            'ConsolidatedStatementOfCashFlows-direct': '連結キャッシュ・フロー計算書',
            'ConsolidatedStatementOfFinancialPositionIFRS': '連結財政状態計算書',
            'ConsolidatedStatementOfProfitOrLossIFRS': '連結損益計算書',
            'ConsolidatedStatementOfComprehensiveIncomeIFRS': '連結包括利益計算書',
            'ConsolidatedStatementOfChangesInEquityIFRS': '連結株主資本等変動計算書',
            'ConsolidatedStatementOfCashFlowsIFRS': '連結キャッシュ・フロー計算書',
            'BalanceSheet': '貸借対照表',
            'StatementOfIncome': '損益計算書',
            'StatementOfComprehensiveIncome': '包括利益計算書',
            'StatementOfChangesInEquity': '株主資本等変動計算書',
            'StatementOfChangesInNetAssets': '株主資本等変動計算書',
            'StatementOfCashFlows': 'キャッシュ・フロー計算書',
            'StatementOfCashFlows-indirect': 'キャッシュ・フロー計算書',
            'StatementOfCashFlows-direct': 'キャッシュ・フロー計算書',
            'SummaryOfBusinessResults': '主要な経営指標等の推移',
            'BusinessResultsOfGroup': '主要な経営指標等の推移（連結）',
            'BusinessResultsOfReportingCompany': '主要な経営指標等の推移（単体）',
            # Note / Segment keywords (without '注記_' prefix, as it's added by logic)
            'SegmentInformationConsolidatedFinancialStatementsIFRS': 'セグメント情報等',
            'AnalysisOfOperatingResultsConsolidatedFinancialStatementsIFRS': 'セグメント情報',
            'NotesSegmentInformationEtcConsolidatedFinancialStatements': 'セグメント情報等',
            'NotesAnalysisOfOperatingResultsConsolidatedFinancialStatements': 'セグメント情報',
            'StatementOfFinancialPositionIFRS': '連結財政状態計算書',
            'StatementOfProfitOrLossIFRS': '連結損益計算書',
            'InventoriesConsolidatedFinancialStatementsIFRS': '棚卸資産',
            'PropertyPlantAndEquipmentConsolidatedFinancialStatementsIFRS': '有形固定資産',
            'GoodwillAndIntangibleAssetsConsolidatedFinancialStatementsIFRS': 'のれん及び無形資産',
            'SellingGeneralAndAdministrativeExpensesConsolidatedFinancialStatementsIFRS': '販売費及び一般管理費',
            'FinanceIncomeAndFinanceCostsConsolidatedFinancialStatementsIFRS': '金融収益及び金融費用',
            'TradeAndOtherReceivablesConsolidatedFinancialStatementsIFRS': '営業債権及びその他の債権',
            'TradeAndOtherPayablesConsolidatedFinancialStatementsIFRS': '営業債務及びその他の債務',
            'OtherInvestmentsConsolidatedFinancialStatementsIFRS': 'その他の投資',
            'ExpensesByNatureConsolidatedFinancialStatementsIFRS': '費用の性質別内訳'
        }
        
        japanese_name = sheet_mapping.get(base_name)
        if not japanese_name:
            if base_name.startswith('Notes'):
                sub_name = base_name[5:] # remove 'Notes'
                # Dynamic lookup in labels_map for element names based on base_name
                # Try multiple prefixes for IFRS and J-GAAP
                prefixes = ["jpigp_cor_", "jpcrp_cor_", "jppfs_cor_"]
                # Possible variations: Prefix + role_name + suffix, or Prefix + sub_name + suffix
                search_terms = []
                for p in prefixes:
                    for suffix in ["Heading", "TextBlock", ""]:
                        search_terms.append(f"{p}{base_name}{suffix}")
                        if base_name.startswith('Notes'):
                            search_terms.append(f"{p}{base_name[5:]}{suffix}")
                        else:
                            search_terms.append(f"{p}Notes{base_name}{suffix}")
                
                for el in search_terms:
                    if el in labels_map:
                        raw_label = labels_map[el]
                        # Clean up: remove prefixes and standardize
                        # Example: "注記事項－..." or suffix phrases
                        cl_label = raw_label.split('、')[0].split(' [')[0].replace('注記事項－', '').strip()
                        if cl_label:
                            japanese_name = '注記_' + cl_label
                            break

                if not japanese_name:
                    lookup_name = base_name[5:] if base_name.startswith('Notes') else base_name
                    if 'SegmentInformation' in base_name:
                        # Normalize for lookup
                        m = _RE_SEGMENT_SUFFIX.search(lookup_name)
                        segment_dict = {
                            '01': '報告セグメントの概要等',
                            '02': 'セグメント情報',
                            '03': '差異調整事項等',
                            '04': '関連情報',
                            '05': '減損損失',
                            '06': 'のれん',
                            '07': '負ののれん'
                        }
                        if m and m.group(1) in segment_dict:
                            inner_v = segment_dict[m.group(1)]
                        elif m:
                            inner_v = f'セグメント情報{int(m.group(1))}'
                        else:
                            inner_v = 'セグメント情報'
                        
                        japanese_name = sheet_mapping.get(lookup_name, inner_v)
                    else:
                        japanese_name = sheet_mapping.get(lookup_name, lookup_name)
            else:
                japanese_name = base_name
                
        # --- Robust Naming Logic (V6) ---
        # 1. Standard Suffix Suffix placement
        suffix = ""
        is_ifrs = (current_standard == 'IFRS')
        is_jmis = (current_standard == 'JMIS')
        is_us = (current_standard == 'US')
        is_all = (current_standard == 'JP_ALL')
        
        if is_ifrs: suffix = '(IFRS)'
        elif is_jmis: suffix = '(JMIS)'
        elif is_us: suffix = '(US GAAP)'
        elif not is_all: suffix = '(日本基準)'
        
        # 2. Handle Analytical Suffix (_分析)
        analytical_suffix = ""
        if 'AnalysisOfOperatingResults' in base_name:
            analytical_suffix = "_分析"
            
        # 3. Assemble components
        # Avoid doubling suffix if already present
        if suffix and suffix in japanese_name:
            suffix = ""
            
        final_sheet_name = f"{japanese_name}{suffix}{analytical_suffix}"
        
        # 4. Final '注記_' Prefixing for notes/segments
        if base_name.startswith('Notes') or 'SegmentInformation' in base_name or 'AnalysisOfOperatingResults' in base_name:
            if not final_sheet_name.startswith('注記_'):
                final_sheet_name = '注記_' + final_sheet_name
                
        sheet_name = final_sheet_name
        
        # In Japanese, 31 characters maximum for sheet name
        if len(sheet_name) > 31:
            # If the name is too long, truncate it before the suffix and re-add suffix
            # This logic needs to be careful with analytical_suffix and standard suffix
            
            # Calculate length available for the base name
            total_suffix_len = len(suffix) + len(analytical_suffix)
            allowed_base_len = 31 - total_suffix_len
            
            # Truncate the base part of the name
            truncated_base_name = japanese_name[:allowed_base_len]
            
            # Reconstruct the sheet name
            sheet_name = f"{truncated_base_name}{suffix}{analytical_suffix}"
        
        # Collect columns relevant to THIS role based on sheet type
        is_segment = 'SegmentInformation' in base_name or 'AnalysisOfOperatingResults' in base_name
        is_consolidated = 'Consolidated' in base_name or 'Group' in base_name or 'SummaryOfBusinessResults' in base_name
        is_non_consolidated = not is_consolidated and not is_segment
        
        debug_log(f"[DEBUG] Processing sheet: {sheet_name} (role: {role}, std: {current_standard}, is_segment: {is_segment})")
        debug_log(f"  [DEBUG] Role has {len(ordered_keys)} elements in presentation tree")

        # --- Skip roles that only contain structural/non-data elements ---
        # Check if role contains only TextBlock, Abstract, Heading, Table, Axis, Member elements
        structural_suffixes = ('TextBlock', 'Abstract', 'Heading', 'Table', 'Axis', 'Member', 'LineItems')
        has_data_element = False
        for full_path, _ in ordered_keys:
            # Extract element name from full path (last component)
            element_name = full_path.split('::')[-1]
            # Remove dimension suffix if present (e.g., "ElementName|DimensionName")
            if '|' in element_name:
                element_name = element_name.split('|')[0]

            # Check if this is a structural element
            if not any(element_name.endswith(suffix) for suffix in structural_suffixes):
                has_data_element = True
                break

        if not has_data_element:
            debug_log(f"[Skip-Role] Role '{role}' contains only structural elements (TextBlock/Abstract/etc), skipping sheet creation")
            continue

        role_columns = set()
        if is_non_consolidated:
            # Merged dimensions for standalone: prioritize '単体' over '全体'
            period_to_best_dim = {}
            for full_path_data in ordered_keys:
                full_path, _ = full_path_data
                if full_path in all_years_data[role]:
                    for c in all_years_data[role][full_path].keys():
                        # c is (std, dim, period)
                        std, dim, period = c if len(c) == 3 else ("JP", c[0], c[1])
                        
                        if dim == '連結': continue
                        if dim not in ('全体', '単体'): continue
                        
                        if period not in period_to_best_dim:
                            period_to_best_dim[period] = (std, dim)
                        elif dim == '単体':
                            period_to_best_dim[period] = (std, dim)
            role_columns = set((s, d, p) for p, (s, d) in period_to_best_dim.items())
        else:
            for full_path_data in ordered_keys:
                full_path, pref_label = full_path_data
                if full_path in all_years_data[role]:
                    for c in all_years_data[role][full_path].keys():
                        # c is (std, dim, period)
                        std, dim, period = c if len(c) == 3 else ("JP", c[0], c[1])
                        
                        if is_segment:
                            if dim == '単体': continue
                        elif is_consolidated:
                            if dim not in ('全体', '連結', '全社'): continue
                        else: # other consolidated notes
                            if period in periods_with_standalone:
                                if dim == '単体': continue
                                
                        # Filtering by accounting standard (V13 Refined)
                        # c is (fact_std, dim, p)
                        fact_std, f_dim, f_p = c if len(c) == 3 else ("JP", c[0], c[1])

                        if current_standard == 'JP_ALL':
                            # Summary sheets merge all detected standards for the period
                            pass
                        else:
                            # Rule 1: If fact has an explicit standard, it MUST match the sheet's standard
                            if fact_std is not None and fact_std != current_standard:
                                continue
                            
                            # Rule 2: If fact has NO explicit standard (extension), 
                            # only include it if the sheet's standard is one of the primary standards for this period.
                            if fact_std is None:
                                p_stds = consolidated_standards.get(f_p, set())
                                if current_standard not in p_stds:
                                    continue
                            
                        role_columns.add(c)
        
        # Filter Sheets (BS, CE, CF) to ensures they don't precede the PL start period
        target_sheet_keywords = ['株主資本等変動計算書', 'キャッシュ・フロー計算書', '持分変動計算書', '貸借対照表', '財政状態計算書']
        if any(kw in sheet_name for kw in target_sheet_keywords):
            if current_standard in pl_start_periods:
                pl_min = pl_start_periods[current_standard]
                filtered_cols = set()
                for c in role_columns:
                    period = c[2] if len(c) == 3 else c[1]
                    # Only include periods that are on or after the PL start period
                    if period >= pl_min:
                        filtered_cols.add(c)
                role_columns = filtered_cols

        if not role_columns:
            continue
            
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Use existing seen_rows to avoid duplication across merged roles
            if not hasattr(ws, '_seen_rows'):
                ws._seen_rows = set()
            seen_rows = ws._seen_rows
            is_new_sheet = False
            debug_log(f"[Merge-Sheet] Merging data into existing sheet: {sheet_name}")
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws._seen_rows = set()
            seen_rows = ws._seen_rows
            is_new_sheet = True
            used_sheet_names.add(sheet_name)
            
        if not default_sheet_removed:
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            default_sheet_removed = True
        
        # Track separators so we only print them once per sheet
        seen_related = False
        seen_goodwill = False
        seen_negative_goodwill = False
        seen_impairment = False
        # seen_rows is now persistent per sheet
        
        # Sort columns logically
        def sort_col(c):
            # c is (std, dim, period)
            std, dim, period = c if len(c) == 3 else ("JP", c[0], c[1])
            
            # Specific fixed orders for standard structural members
            order = 500
            if dim == '単体':
                return (1, dim, period)
            
            # 1. Broad totals like 'Overall' or 'Consolidated'
            if any(s == dim for s in ('全体', '連結', '合計', '連結財務諸表計上額')):
                return (1000, dim, period)
            
            # 2. Adjustments and Eliminations
            if any(s in dim for s in ('調整', '調整額', '全社・消去', '消去又は全社', '調整項目')):
                return (980, dim, period)
            
            # 3. Intermediate totals
            if any(s in dim for s in ('報告セグメント及びその他の合計', '報告セグメント合計', '内部売上高又は振替高')):
                return (950, dim, period)

            if dim == '報告セグメント':
                return (900, dim, period)

            # 4. 'Others' category
            if any(s in dim for s in ('報告セグメント以外の全てのセグメント', '報告セグメント以外', 'その他')):
                return (940, dim, period)
            
            # 5. Members found in hierarchy (actual segments)
            if dim in master_member_seq:
                return (10 + master_member_seq.index(dim), dim, period)
                
            # 6. Fallback for everything else
            return (order, dim, period)
            
        sorted_role_cols = sorted(list(role_columns), key=sort_col)
        
        debug_log(f"  [DEBUG] Role columns: {sorted_role_cols}")
        if sheet_name == '貸借対照表':
             # Detailed trace for single BS
             debug_log(f"  [DEBUG-Trace-BS] is_non_consolidated={is_non_consolidated}")
        
        # '全体', '連結', '全社', '単体', '個別' などの基礎区分のみの場合はセグメント無し（単一階層ヘッダー）とする
        has_segments = any(c[1] not in ('全体', '連結', '全社', '単体', '個別') for c in sorted_role_cols)
        
        if is_new_sheet:
            if has_segments:
                # Two-tier header: Row 1 = Segments (dim), Row 2 = Dates (period)
                headers_row1 = ["", ""] + [c[1] for c in sorted_role_cols]
                headers_row2 = ["勘定科目", "項目（英名）"] + [c[2] for c in sorted_role_cols]
                ws.append(headers_row1)
                ws.append(headers_row2)
            else:
                # Single-tier header: Dates (period)
                headers = ["勘定科目", "項目（英名）"] + [c[2] for c in sorted_role_cols]
                ws.append(headers)
        elif ws.max_row == 0:
            # Handle edge case where sheet existed but was empty
            if has_segments:
                ws.append(["", ""] + [c[1] for c in sorted_role_cols])
                ws.append(["勘定科目", "項目（英名）"] + [c[2] for c in sorted_role_cols])
            else:
                ws.append(["勘定科目", "項目（英名）"] + [c[2] for c in sorted_role_cols])
        
        for full_path_data in ordered_keys:
            full_path, pref_label = full_path_data
            # Extract element name to get label
            el = full_path.split('::')[-1]
            if '|' in el: el = el.split('|')[0]

            # --- USER SUGGESTION: Skip irrelevant element types ---
            # Note: Keep Abstract and Heading elements for hierarchy display
            # Skip only TextBlock, Axis, Member, Table, LineItems
            if el.endswith(("TextBlock","Axis","Member","Table","LineItems")):
                continue

            # Determine if this element is a heading (Abstract or Heading suffix)
            is_heading = el.endswith(("Abstract", "Heading"))
            
            # Heading-specific terminology for Abstract elements
            heading_dict = {
                'AssetsIFRSAbstract': '資産',
                'AssetsAbstract': '資産',
                'CurrentAssetsIFRSAbstract': '流動資産',
                'CurrentAssetsAbstract': '流動資産',
                'NonCurrentAssetsIFRSAbstract': '非流動資産',
                'NonCurrentAssetsAbstract': '非流動資産',
                'NoncurrentAssetsAbstract': '非流動資産',
                'LiabilitiesAndEquityIFRSAbstract': '負債及び資本',
                'LiabilitiesAndNetAssetsAbstract': '負債及び純資産',
                'LiabilitiesIFRSAbstract': '負債',
                'LiabilitiesAbstract': '負債',
                'CurrentLiabilitiesIFRSAbstract': '流動負債',
                'CurrentLiabilitiesAbstract': '流動負債',
                'NonCurrentLiabilitiesIFRSAbstract': '非流動負債',
                'NonCurrentLiabilitiesAbstract': '非流動負債',
                'NoncurrentLiabilitiesAbstract': '非流動負債',
                'EquityIFRSAbstract': '資本',
                'NetAssetsAbstract': '純資産',
                'ShareholdersEquityAbstract': '株主資本',
            }

            # Common terminology translations as a fallback
            # Imported from edinet_taxonomy_dict.py (1,959 items)
            # Source: EDINET Taxonomy (https://disclosure2dl.edinet-fsa.go.jp/guide/static/disclosure/download/ESE140115.xlsx)
            common_dict = EDINET_COMMON_DICT
            
            parts = el.split('_')
            base_name = parts[-1] if len(parts) > 1 else el
            
            segment_dict = {
                # IFRS Specific Overrides for Segments
                'NetSalesIFRS': '売上高', 
                'IntersegmentSalesIFRS': 'セグメント間売上高',
                'ProfitLossBeforeTaxIFRS': '（税引前当期損益）',
                'AssetsIFRS': 'セグメント資産',
                'DepreciationAndAmortizationOperatingExpensesIFRS': '減価償却費及び償却費',
                'ImpairmentLossesOnNonFinancialAssetsPLIFRS': '非金融資産の減損損失',
                'OtherIncomeAndExpensesNetIFRS': 'その他の損益',
                'ShareOfProfitLossOfInvestmentsAccountedForUsingEquityMethodIFRS': '持分法による投資損益',
                'CapitalExpendituresIFRS': '資本的支出',
                'InvestmentsAccountedForUsingEquityMethodIFRS': '持分法で会計処理されている投資',
                'FinanceIncomeIFRS': '金融収益',
                'FinanceCostsIFRS': '金融費用',
                'ExternalRevenueIFRS': '外部売上高',
                'IntersegmentRevenueIFRS': 'セグメント間売上高',
                'SegmentProfitLossIFRS': 'セグメント利益又は損失（△）',
                'SegmentAssetsIFRS': 'セグメント資産',
                'OtherInformationIFRS': 'その他の情報',
                'DepreciationAndAmortisationIFRS': '減価償却費及び償却費',
                'OtherProfitLossIFRS': 'その他の損益',
                
                # J-GAAP Specific Overrides for Segments
                'NetSales': '計',
                'OrdinaryIncome': 'セグメント利益又は損失（△）',
                'OperatingIncome': 'セグメント利益又は損失（△）',
                'Assets': 'セグメント資産',
                'Liabilities': 'セグメント負債',
                'SalesToExternalCustomers': '外部顧客への売上高',
                'IntersegmentSalesOrTransfers': 'セグメント間の内部売上高又は振替高',
                'SegmentProfitLoss': 'セグメント利益又は損失（△）',
                'SegmentAssets': 'セグメント資産',
                'SegmentLiabilities': 'セグメント負債',
                'OtherItems': 'その他の項目',
                'DepreciationAndAmortization': '減価償却費',
                'AmortizationOfGoodwill': 'のれんの償却額',
                'InterestIncome': '受取利息',
                'InterestExpenses': '支払利息',
                'ShareOfProfitLossOfEntitiesAccountedForUsingEquityMethod': '持分法投資利益又は損失（△）',
                'InvestmentsInEntitiesAccountedForUsingEquityMethod': '持分法適用会社への投資額',
                'IncreaseInPropertyPlantAndEquipmentAndIntangibleAssets': '有形固定資産及び無形固定資産の増加額'
            }
            
            # Label resolution priority:
            # 1. Heading-specific dictionary (for Abstract/Heading elements)
            # 2. Segment-specific dictionary
            # 3. Common dictionary
            # 4. labels_map from XBRL
            # 5. CamelCase conversion
            if is_heading and el in heading_dict:
                label = heading_dict[el]
            elif is_segment and base_name in segment_dict:
                label = segment_dict[base_name]
            elif base_name in common_dict:
                label = common_dict[base_name]
            else:
                # 4. labels_map from XBRL (Local or Standard)
                label = labels_map.get(el)
                if label: 
                    label = label.replace(' [メンバー]', '').replace(' [要素]', '').replace(' [区分]', '').strip()
                else:
                    # 5. Fallback: IFRS Mapping
                    label = IFRS_LABEL_MAPPING.get(el)
            
            # 6. Final Fallback: CamelCase conversion
            if not label:
                label = convert_camel_case_to_title(base_name)
            
            # Append suffix for Cash Flow balances
            is_cf_sheet = 'キャッシュ・フロー' in sheet_name
            if is_cf_sheet:
                if pref_label == 'http://www.xbrl.org/2003/role/periodStartLabel':
                    label += "（期首残高）"
                elif pref_label in ('http://www.xbrl.org/2003/role/periodEndLabel', 'http://www.xbrl.org/2003/role/totalLabel'):
                    # Only append (期末残高) if the element is likely a balance element
                    if 'CashAndCashEquivalents' in el:
                        label += "（期末残高）"
                    
            # Insert Separator rows for Segment Notes subsets
            if 'SegmentInformation' in role:
                if not seen_related and ('NotesInformationAssociated' in el or 'DisclosureOfRelatedInformation' in el):
                    ws.append([])
                    ws.append(["【 注記：関連情報 】"])
                    seen_related = True
                elif 'Goodwill' in el and ('Amortization' in el or 'Negative' in el or 'Disclosure' in el):
                    if 'Negative' in el and not seen_negative_goodwill:
                        ws.append([])
                        ws.append(["【 注記：負ののれん 】"])
                        seen_negative_goodwill = True
                    elif 'Negative' not in el and not seen_goodwill:
                        ws.append([])
                        ws.append(["【 注記：のれん 】"])
                        seen_goodwill = True
                elif not seen_impairment and ('ImpairmentLoss' in el and 'Segment' in el):
                    ws.append([])
                    ws.append(["【 注記：減損損失 】"])
                    seen_impairment = True
            
            # --- 重複排除用のキー作成 (勘定科目名) ---
            # ここでは一旦スキップせず、データ収集後に判定する。
                    
            # Indent based on depth
            depth = len(full_path.split('::')) - 1
            indent_prefix = "　" * depth

            # Remove unwanted suffixes from label for Excel output
            display_label = label
            display_label = display_label.replace(' [目次項目]', '').replace(' [タイトル項目]', '')
            display_label = display_label.replace('（IFRS）', '').replace('(IFRS)', '')
            display_label = display_label.replace('、経営指標等', '')
            # Remove IFRS classification suffixes
            display_label = display_label.replace('、流動資産', '').replace('、非流動資産', '')
            display_label = display_label.replace('、流動負債', '').replace('、非流動負債', '')
            # Remove Cash Flow activity suffixes for IFRS
            display_label = display_label.replace('、営業活動によるキャッシュ・フロー', '')
            display_label = display_label.replace('、投資活動によるキャッシュ・フロー', '')
            display_label = display_label.replace('、財務活動によるキャッシュ・フロー', '')
            display_label = display_label.strip()

            row_data = [indent_prefix + display_label, el]
            
            has_numeric_data = False
            has_data = False
            for col_key in sorted_role_cols:
                # SPECIAL HANDLING for Cash Flow Beginning Balance
                # If this row is a periodStartLabel in a Cash Flow sheet, we pull data from the prior period's instant.
                val = ""
                is_cf_sheet = 'キャッシュ・フロー' in sheet_name
                is_start_row = pref_label == 'http://www.xbrl.org/2003/role/periodStartLabel'
                
                if is_cf_sheet and is_start_row:
                    # Find the startDate of the current column
                    current_start_date = global_element_period_values.get('_metadata', {}).get(col_key)
                    if current_start_date:
                        # Find the corresponding instant
                        # Cases:
                        # 1. Instant is exactly at current_start_date (e.g. 2024-04-01)
                        # 2. Instant is at the end of the previous day (e.g. 2024-03-31)
                        dates_to_try = [current_start_date]
                        try:
                            from datetime import datetime, timedelta
                            dt = datetime.strptime(current_start_date, '%Y-%m-%d')
                            prev_day = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
                            dates_to_try.append(prev_day)
                        except Exception:
                            pass
                        
                        found_val = False
                        for t_date in dates_to_try:
                            target_col_key = (col_key[0], t_date)
                            if el in global_element_period_values and target_col_key in global_element_period_values[el]:
                                val = global_element_period_values[el][target_col_key]
                                found_val = True
                                break
                        
                        if not found_val:
                            # Search for any dimension matching if exact dim fails
                            for t_date in dates_to_try:
                                for k, v in global_element_period_values.get(el, {}).items():
                                    if k[1] == t_date: # Match date, ignore dimension if needed? 
                                        # (Actually dimension should match, but sometimes it's "全体" vs "連結")
                                        val = v
                                        found_val = True
                                        break
                                if found_val: break
                
                # Only use fallback if this isn't a CF start row (to avoid pulling ending balance)
                if val == "" and not (is_cf_sheet and is_start_row):
                    # Try exact match first (standard, dimension, period)
                    val = all_years_data[role][full_path].get(col_key, "")
                    
                    if "TradeReceivablesAndContractAssets" in el:
                        debug_log(f"  [DEBUG-Trace-BS-Deep] el={el}, col_key={col_key}, val='{val}'")
                        debug_log(f"  [DEBUG-Trace-BS-Deep] all_keys={list(all_years_data[role][full_path].keys())}")
                        
                    # FALLBACK: If current standard is JP_ALL and exact match failed, 
                    # try matching by (None, dimension, period) or any other standard for extension tags
                    # Hitachi case: TradeReceivablesAndContractAssets might be tagged as IFRS but used in JP sheet
                    if val == "" and current_standard == 'JP_ALL':
                        for s in [None, 'JP', 'IFRS', 'US', 'JMIS', 'Generic']:
                            if s == col_key[0]: continue
                            test_key = (s, col_key[1], col_key[2])
                            v = all_years_data[role][full_path].get(test_key, "")
                            if v != "":
                                val = v
                                break
                    
                    if val == "" and (is_segment and 'AnalysisOfOperatingResults' in base_name):
                        # Fallback for analysis roles: if exact standard fails, try others for this dim/period
                        # because these roles are often sparsely populated across standards in some XBRL sets
                        for s in ['JP', 'IFRS', 'US', 'JMIS']:
                            if s == col_key[0]: continue
                            test_key = (s, col_key[1], col_key[2])
                            v = all_years_data[role][full_path].get(test_key, "")
                            if v != "":
                                val = v
                                break
                
                # Clean numeric values
                if val:
                    # Handle full-width characters and commas
                    import unicodedata
                    val_clean = unicodedata.normalize('NFKC', str(val)).replace(',', '').strip()

                    # Handle "円銭" format (e.g., "3,422円93銭" → 3422.93)
                    # This appears in per-share items like NetAssetsPerShare, EarningsPerShare
                    yen_sen_match = re.match(r'^([0-9]+)円([0-9]+)銭$', val_clean)
                    if yen_sen_match:
                        yen_part = yen_sen_match.group(1)
                        sen_part = yen_sen_match.group(2)
                        val = float(yen_part) + float(sen_part) / 100
                        has_numeric_data = True
                    else:
                        try:
                            if val_clean and not any(c.isalpha() for c in val_clean):
                                val = float(val_clean)
                                has_numeric_data = True
                        except Exception:
                            pass
                row_data.append(val)
                if val != "" :
                    has_data = True
                
            # Display heading elements even if they have no data (for hierarchy structure)
            # Display data elements only if they have at least one value
            if has_data or is_heading:
                # --- IFRSシート専用フィルタ: 英語勘定名に"IFRS"が含まれない行を除外 ---
                # IFRS sheets should only show elements with "IFRS" in their English name
                # Exception: Keep heading elements (Abstract/Heading) for hierarchy structure
                if is_ifrs and not is_heading:
                    if 'IFRS' not in el:
                        continue

                # --- セグメント情報や財務諸表の文字情報の除外 (Current Refinement) ---
                # Remove unwanted text blocks like *FinancialInformation or long descriptions
                # But keep heading elements for hierarchy structure
                is_financial_statement = any(kw in sheet_name for kw in ('貸借対照表', '損益計算書', '包括利益', 'キャッシュ・フロー', '株主資本'))
                if is_financial_statement or is_segment:
                    is_text_info = (
                        el.endswith('FinancialInformation') or
                        el.startswith(('jpcrp_cor_Description', 'jpcrp_cor_Note', 'jpcrp_cor_Regulations', 'jpcrp_cor_RemarkableEfforts'))
                    )
                    if is_text_info and not has_numeric_data:
                        continue
                    # Skip non-numeric data elements in segments, but keep headings for structure
                    if is_segment and not has_numeric_data and not is_heading:
                        continue
                # --- 重複排除 (勘定科目名と数値が完全に一致する行をスキップ) ---
                row_values_tuple = tuple(row_data[2:])
                row_key = (display_label, row_values_tuple)
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)

                # --- SPECIAL: Add beginning balance row for IFRS Cash Flow ---
                # If this is CashAndCashEquivalents with periodEndLabel in CF sheet,
                # first add a beginning balance row using prior period's ending values
                if (is_cf_sheet and 'CashAndCashEquivalents' in el and
                    pref_label and pref_label.endswith(('periodEndLabel', 'totalLabel'))):
                    # Create beginning balance row
                    beginning_label = label.replace('（期末残高）', '（期首残高）')
                    # Clean up IFRS suffixes from beginning label
                    beginning_label = beginning_label.replace('（IFRS）', '').replace('(IFRS)', '')
                    beginning_label = beginning_label.replace('、流動資産', '').replace('、非流動資産', '')
                    beginning_label = beginning_label.replace('、流動負債', '').replace('、非流動負債', '')
                    beginning_label = beginning_label.strip()

                    beginning_row = [indent_prefix + beginning_label, el]

                    # For each column, use the prior period's value
                    for i, col_key in enumerate(sorted_role_cols):
                        beginning_val = ""
                        # Get the period date for this column
                        if len(col_key) == 2:  # (dim, period)
                            period_date = col_key[1]
                        elif len(col_key) == 3:  # (std, dim, period)
                            period_date = col_key[2]
                        else:
                            period_date = None

                        if period_date:
                            # Find the startDate for this period
                            from datetime import datetime, timedelta
                            try:
                                # Look up the start date from metadata
                                start_date = global_element_period_values.get('_metadata', {}).get(col_key)
                                if start_date:
                                    # Try the start date and the day before
                                    dates_to_try = [start_date]
                                    try:
                                        dt = datetime.strptime(start_date, '%Y-%m-%d')
                                        prev_day = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
                                        dates_to_try.append(prev_day)
                                    except Exception:
                                        pass

                                    # Try to find instant value at beginning of period
                                    for t_date in dates_to_try:
                                        if len(col_key) == 2:
                                            target_key = (col_key[0], t_date)
                                        else:  # len == 3
                                            target_key = (col_key[0], col_key[1], t_date)

                                        if el in global_element_period_values:
                                            if target_key in global_element_period_values[el]:
                                                beginning_val = global_element_period_values[el][target_key]
                                                break
                                            # Try matching just the date
                                            for k, v in global_element_period_values[el].items():
                                                if isinstance(k, tuple) and k[-1] == t_date:
                                                    beginning_val = v
                                                    break
                                        if beginning_val:
                                            break
                            except Exception:
                                pass

                        # Clean and convert to numeric if possible
                        if beginning_val:
                            import unicodedata
                            val_clean = unicodedata.normalize('NFKC', str(beginning_val)).replace(',', '').strip()
                            try:
                                if val_clean and not any(c.isalpha() for c in val_clean):
                                    beginning_val = float(val_clean)
                            except Exception:
                                pass

                        beginning_row.append(beginning_val)

                    # Only add if it has some data
                    if any(v != "" for v in beginning_row[2:]):
                        beginning_row_key = (beginning_label, tuple(beginning_row[2:]))
                        if beginning_row_key not in seen_rows:
                            ws.append(beginning_row)
                            seen_rows.add(beginning_row_key)

                ws.append(row_data)

                # --- TAXONOMY STRUCTURE-BASED: Stop at appropriate end items ---
                # Use hierarchy depth and preferredLabel to determine natural end points

                is_cf_sheet = 'キャッシュ・フロー' in sheet_name
                is_pl_sheet = '損益計算書' in sheet_name

                # Helper: Check if this is the last significant item at current depth
                def is_end_of_statement(current_idx, ordered_keys_list):
                    """Check if next items are shallower (returning to parent level) or end of list"""
                    if current_idx >= len(ordered_keys_list) - 1:
                        return True

                    current_depth = len(full_path.split('::')) - 1

                    # Look ahead to see if we're returning to parent level
                    for next_idx in range(current_idx + 1, len(ordered_keys_list)):
                        next_fp, _ = ordered_keys_list[next_idx]
                        next_el_name = next_fp.split('::')[-1]
                        if '|' in next_el_name:
                            next_el_name = next_el_name.split('|')[0]

                        # Skip Abstract, TextBlock, Table, Axis, Member
                        if next_el_name.endswith(("Abstract", "TextBlock", "Table", "Axis", "Member")):
                            continue

                        next_depth = len(next_fp.split('::')) - 1

                        # If next real item is at same or shallower depth, we've reached end
                        if next_depth <= current_depth:
                            return True
                        # If deeper, there are more substantive items to come
                        return False

                    return True

                current_idx = ordered_keys.index(full_path_data)

                # Cash Flow: Stop AFTER CashAndCashEquivalents with periodEndLabel/totalLabel
                # (allow both periodStartLabel and periodEndLabel to be displayed)
                if is_cf_sheet and 'CashAndCashEquivalents' in el:
                    # Only break if this is periodEndLabel (not periodStartLabel)
                    # AND it's at the natural end of the statement
                    if pref_label and pref_label.endswith(('periodEndLabel', 'totalLabel')):
                        # Check if there are more CashAndCash items with periodStartLabel ahead
                        has_more_cash_items = False
                        for next_idx in range(current_idx + 1, len(ordered_keys)):
                            next_fp, next_pref = ordered_keys[next_idx]
                            next_el = next_fp.split('::')[-1]
                            if '|' in next_el:
                                next_el = next_el.split('|')[0]
                            # Skip non-substantive items
                            if next_el.endswith(("Abstract", "TextBlock", "Table", "Axis", "Member")):
                                continue
                            # If we find another CashAndCash item, don't break yet
                            if 'CashAndCashEquivalents' in next_el:
                                has_more_cash_items = True
                                break
                            # If we find a different substantive item, we can break
                            break

                        if not has_more_cash_items and is_end_of_statement(current_idx, ordered_keys):
                            break

                # Profit/Loss Statement: Stop at final profit or EPS items
                if is_pl_sheet:
                    # Look for final profit or EPS items with totalLabel or at natural end
                    # IMPORTANT: Only match FINAL profit items, not intermediate ones like GrossProfit, OperatingProfit
                    is_profit_item = any(keyword in el for keyword in ['ProfitLoss', 'NetIncome']) and not any(kw in el for kw in ['Gross', 'Operating', 'Ordinary'])
                    is_eps_item = any(keyword in el for keyword in ['EarningsPerShare', 'EarningsLossPerShare'])

                    if is_profit_item or is_eps_item:
                        # If this has totalLabel OR is at end of hierarchy, consider it as endpoint
                        if pref_label and 'totalLabel' in pref_label:
                            if is_end_of_statement(current_idx, ordered_keys):
                                break
                        # Even without totalLabel, if it's a key final item at natural end
                        elif is_eps_item and is_end_of_statement(current_idx, ordered_keys):
                            break

        # Apply formatting and column widths
        # Define exact match patterns
        ratio_elements_exact = {
            # Standard & Japanese GAAP
            'EquityToAssetRatioSummaryOfBusinessResults',
            'RateOfReturnOnEquitySummaryOfBusinessResults',
            'CapitalAdequacyRatioInternationalStandardSummaryOfBusinessResults',
            'CapitalAdequacyRatioDomesticStandardSummaryOfBusinessResults',
            'CapitalAdequacyRatioBISStandardSummaryOfBusinessResults',
            'CapitalAdequacyRatioDomesticStandard2SummaryOfBusinessResults',
            'PayoutRatioSummaryOfBusinessResults',
            'TotalShareholderReturnSummaryOfBusinessResults',
            'ComparativeTotalShareholderReturnSummaryOfBusinessResults',
            'jpcrp_cor_TotalShareholderReturn',
            'jpcrp_cor_TotalReturnOnSharePriceIndex',

            # IFRS Variations
            'RatioOfOwnersEquityToGrossAssetsIFRSSummaryOfBusinessResults',
            'RateOfReturnOnEquityIFRSSummaryOfBusinessResults',

            # JMIS Variations
            'RatioOfOwnersEquityToGrossAssetsJMISSummaryOfBusinessResults',
            'RateOfReturnOnEquityJMISSummaryOfBusinessResults',

            # US GAAP Variations
            'EquityToAssetRatioUSGAAPSummaryOfBusinessResults',
            'RateOfReturnOnEquityUSGAAPSummaryOfBusinessResults',

            # Industry Specific (Insurance, etc.)
            'NetLossRatioSummaryOfBusinessResultsINS',
            'NetOperatingExpenseRatioSummaryOfBusinessResultsINS'
        }

        # Define wildcard patterns (regex patterns for company-specific elements)
        # Pattern: jpcrp030000-asr_E*****-000_ElementName
        ratio_patterns_wildcard = [
            r'jpcrp030000-asr_E\d{5}-000_TotalShareholderReturn$',
            r'jpcrp030000-asr_E\d{5}-000_TotalReturnOnSharePriceIndex$',
        ]
        ratio_regex_compiled = [re.compile(pattern) for pattern in ratio_patterns_wildcard]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            # Element name is in column B (index 1)
            el_name = row[1].value if len(row) > 1 else None
            is_ratio = False
            if el_name:
                el_str = str(el_name)
                # Remove namespace prefix if present (e.g., "jpcrp030000-asr:E04509-000_..." → "E04509-000_...")
                el_no_ns = el_str.split(':')[-1] if ':' in el_str else el_str
                el_no_ns = el_no_ns.split('_', 1)[-1] if '_' in el_no_ns else el_no_ns

                # Check exact matches
                for r in ratio_elements_exact:
                    if el_str == r or el_str.endswith(':' + r) or el_str.endswith('_' + r):
                        is_ratio = True
                        break

                # Check wildcard patterns (only if not already matched)
                if not is_ratio:
                    for regex in ratio_regex_compiled:
                        if regex.search(el_str) or regex.search(el_no_ns):
                            is_ratio = True
                            break

            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if is_ratio:
                        cell.number_format = '0.0%'
                    else:
                        # Format: #,##0_;[Red]-#,##0
                        cell.number_format = r'#,##0_ ;[Red]\-#,##0 '
                    
        # --- Collect Segment Information for Later Analysis ---
        if is_segment:
            # Collect information needed for segment analysis
            segment_sheets_info.append({
                'sheet_name': sheet_name,
                'ordered_keys': ordered_keys,
                'all_years_data': all_years_data,
                'role': role,
                'sorted_role_cols': sorted_role_cols,
                'role_columns': role_columns,
                'current_standard': current_standard,
                'segment_dict': segment_dict,
                'common_dict': common_dict,
                'labels_map': labels_map,
                'used_sheet_names': used_sheet_names,
                'global_element_period_values': global_element_period_values,
                'filing_pairs': filing_pairs,
            })

    debug_log(f"Sheet generation completed in {time.time() - t_sheet_generation_start:.2f}s")

    # Auto-adjust or fix column widths
    t_colwidth_start = time.time()
    MAX_SAMPLE_ROWS = 100  # Only check first 100 rows for width calculation
    for out_ws in wb.worksheets:
        is_analysis = "_分析" in out_ws.title
        
        # 分析以外の財務諸表シート（標準シート）の設定
        if not is_analysis:
            # B列（項目ID）を非表示
            out_ws.column_dimensions['B'].hidden = True
            # ウィンドウ枠の固定 (A列と1行目を固定: セルB2を基準に左・上を固定)
            out_ws.freeze_panes = 'B2'
        else:
            # 分析シートのウィンドウ枠の固定 (A列・B列・1行目を固定: セルC2を基準に左・上を固定)
            out_ws.freeze_panes = 'C2'
            
        for col in out_ws.columns:
            col_idx = col[0].column  # 1-indexed: A=1, B=2, C=3...
            column_letter = col[0].column_letter
            
            # 分析以外の財務諸表シートではC列以降（時系列データ）の幅を12に固定
            if not is_analysis and col_idx >= 3:
                out_ws.column_dimensions[column_letter].width = 12
                continue
            
            # それ以外（分析シートの全列、または標準シートのA列など）は自動調整
            max_length = 0
            # Optimization: Only sample first 100 rows instead of entire column
            for cell in col[:MAX_SAMPLE_ROWS]:
                try:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        if len(val_str) > max_length:
                            max_length = len(val_str)
                except Exception:
                    pass
            # Add a little extra padding, especially for Japanese characters
            adjusted_width = (max_length + 2) * 1.2
            # Cap width to prevent massive columns from long text
            if adjusted_width > 50:
                adjusted_width = 50
            out_ws.column_dimensions[column_letter].width = adjusted_width
    debug_log(f"Column width adjustment completed in {time.time() - t_colwidth_start:.2f}s")

    # シートの並び替え
    def get_sheet_order(title):
        is_note = '注記' in title
        group = 0
        if not is_note:
            group = 1 if '連結' in title else 3
        else:
            # 単体の財務諸表に対応する注記名かどうかで判定
            non_consolidated_notes = ['注記_貸借対照表', '注記_損益計算書', '注記_株主資本等変動計算書', '注記_包括利益計算書', '注記_キャッシュ・フロー計算書', '注記_製造原価明細書']
            if '連結' not in title and any(n in title for n in non_consolidated_notes):
                group = 4
            else:
                group = 2
                
        # 財務諸表の種別による並び順
        stmt_order = 99
        if '経営指標' in title:
            stmt_order = 0
        elif 'セグメント' in title:
            stmt_order = 1
        elif '貸借対照表' in title or '財政状態' in title:
            stmt_order = 2
        elif '損益' in title or ('利益' in title and '包括' not in title):
            stmt_order = 3
        elif '包括' in title:
            stmt_order = 4
        elif '変動' in title:
            stmt_order = 5
        elif 'キャッシュ' in title:
            stmt_order = 6
            
        # 基準ごとの並び順 (日本基準が先)
        std_order = 2 if '(IFRS)' in title else 1

        # セグメント情報の場合の特殊な並び順
        # 日本基準セグメント -> IFRSセグメント -> 日本基準分析 -> IFRS分析
        if 'セグメント' in title:
            is_analysis = '_分析' in title
            # (group, stmt_order, is_analysis, std_order)
            # is_analysis: 0 (Original), 1 (Analysis)
            # std_order: 1 (JP), 2 (IFRS)
            # Result: (2, 0, 0, 1) -> (2, 0, 0, 2) -> (2, 0, 1, 1) -> (2, 0, 1, 2)
            return (group, stmt_order, 1 if is_analysis else 0, std_order)

        return (group, stmt_order, 0, std_order)
                
    wb._sheets.sort(key=lambda s: get_sheet_order(s.title))

    # Remove sheets with no numeric data (e.g., text-only note sheets)
    sheets_to_remove = []
    for out_ws in wb.worksheets:
        # Skip if sheet has very few rows (likely no meaningful data)
        if out_ws.max_row <= 2:
            has_numeric = False
            # Check if any cell contains numeric data
            for row in out_ws.iter_rows(min_row=2, max_row=out_ws.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        has_numeric = True
                        break
                if has_numeric:
                    break

            if not has_numeric:
                sheets_to_remove.append(out_ws.title)
                debug_log(f"[Remove-Sheet] Sheet '{out_ws.title}' has no numeric data (only {out_ws.max_row} rows)")

    # Remove sheets without numeric data
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])

    # Log summary of sheets for verification
    debug_log("Excel Sheet Summary:")
    for out_ws in wb.worksheets:
        debug_log(f"  - {out_ws.title}: {out_ws.max_row} rows")

    # ============================================================================
    # FINANCIAL ANALYSIS SHEETS (ROE分析など)
    # ============================================================================
    # 財務分析シートを追加（financial_analysis.pyモジュールを使用）
    add_financial_analysis_sheets(wb, debug_log)

    # ============================================================================
    # CCC ANALYSIS SHEET (CCC分析)
    # ============================================================================
    add_ccc_analysis_sheets(wb, debug_log)

    # ============================================================================
    # SEGMENT ANALYSIS SHEETS (セグメント分析)
    # ============================================================================
    # セグメント分析シートを追加（segment_analysis.pyモジュールを使用）
    bubble_label_info = add_segment_analysis_sheets(wb, segment_sheets_info, debug_log)

    # ============================================================================
    # DIVERSITY SHEET (ダイバーシティ)
    # ============================================================================
    # 連結子会社名マップを集約（新しい年度のデータを優先）
    merged_subsidiary_row_names = {}
    for res in reversed(results):  # results は新しい順なので reversed で古→新の順に上書き
        merged_subsidiary_row_names.update(res.get('subsidiary_row_names', {}))
    add_diversity_sheet(wb, global_element_period_values, debug_log, subsidiary_row_names=merged_subsidiary_row_names)
    add_human_capital_sheet(wb, global_element_period_values, debug_log)

    # 最新の期末を取得してファイル名に追加
    latest_period = ''
    try:
        from edinet_cache import EdinetCache
        cache = EdinetCache()
        results = cache.search_by_company_name(company_name)
        if results and results[0].get('latest_period_end'):
            latest_period_end = results[0]['latest_period_end']
            # YYYY-MM-DD -> YYYYMM 形式に変換
            if isinstance(latest_period_end, str) and '-' in latest_period_end:
                latest_period = f"_{latest_period_end.replace('-', '')[:6]}"
    except Exception as e:
        debug_log(f"Failed to fetch latest_period_end from edinet_cache: {e}")

    # 万が一edinet_cacheから取得できなかった場合のフォールバック（従来の方法）
    if not latest_period and periods_seen:
        # periods_seenは (fact_std, dim_label, period) のタプルのセット
        # 期間部分（3番目の要素）だけを取り出してソート
        period_dates = [p[2] if isinstance(p, tuple) else p for p in periods_seen]
        period_dates = [d for d in period_dates if d and isinstance(d, str)]  # 空文字列を除外、文字列のみ
        if period_dates:
            sorted_dates = sorted(period_dates)
            latest_date = sorted_dates[-1]  # 最新の日付
            # YYYY-MM-DD -> YYYYMM 形式に変換
            if isinstance(latest_date, str) and '-' in latest_date:
                latest_period = f"_{latest_date.replace('-', '')[:6]}"

    out_file = f'有報_{company_name}{latest_period}.xlsx'
    if output_dir:
        out_file = os.path.join(output_dir, out_file)

    debug_log(f"Excel generation (structure) completed in {time.time() - t_excel_start:.2f}s")
    t_save = time.time()
    wb.save(out_file)
    debug_log(f"Excel file write (wb.save) completed in {time.time() - t_save:.2f}s")

    # PPMバブルチャートにセグメント名ラベルを注入
    if bubble_label_info:
        try:
            import zipfile as _zipfile
            import re as _re
            from ppm_bubble_labels import inject_bubble_labels
            with _zipfile.ZipFile(out_file, 'r') as _z:
                all_chart_files = sorted(
                    [f for f in _z.namelist() if _re.match(r'xl/charts/chart\d+\.xml$', f)],
                    key=lambda x: int(_re.search(r'\d+', x).group())
                )
                bubble_chart_files = [f for f in all_chart_files if b'bubbleChart' in _z.read(f)]
            for _i, _seg_names in enumerate(bubble_label_info):
                if _i < len(bubble_chart_files):
                    inject_bubble_labels(out_file, [_seg_names], bubble_chart_files[_i])
                    debug_log(f"[PPM] Injected bubble labels into {bubble_chart_files[_i]}")
        except Exception as _e:
            debug_log(f"[PPM] bubble label injection failed: {_e}")

    debug_log(f"SUCCESS: Excel saved to {out_file} in {time.time() - t_excel_start:.2f}s")
    debug_log(f"TOTAL: process_xbrl_zips completed in {time.time() - overall_start:.2f}s")
    
    return {
        "excel_path": out_file,
        "errors": errors
    }

# ============================================================================
# CLI ENTRY POINT
# ============================================================================
# 【将来の分割先】cli.py
#
# コマンドライン引数の解析とprocess_xbrl_zipsの呼び出しのみ
# ビジネスロジックは含めない
# ============================================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python convert_xbrl_to_excel.py <path_to_zip_or_dir1> [<path_to_zip_or_dir2> ...]", file=sys.stderr)
        sys.exit(1)
        
    input_paths = sys.argv[1:]
    zip_files = []
    for p in input_paths:
        if os.path.isfile(p) and p.lower().endswith('.zip'):
            zip_files.append(p)
        elif os.path.isdir(p):
            # Recursively find all ZIP files in the directory
            for root, _, filenames in os.walk(p):
                for f in filenames:
                    if f.lower().endswith('.zip'):
                        zip_files.append(os.path.join(root, f))
    
    if not zip_files:
        print("Error: No ZIP files found in provided paths.", file=sys.stderr)
        sys.exit(1)
        
    process_xbrl_zips(zip_files)

if __name__ == "__main__":
    main()