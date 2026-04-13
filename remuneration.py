"""
Remuneration Analysis Module for XBRL to Excel Conversion - V4 (Swapped Layout)

役員区分ごとの報酬等（報酬総額、報酬内訳、員数）のシートを生成するモジュール
ユーザーの要望に応じ、役員区分を縦（行）、報酬区分を横（列）に配置した時系列レイアウト。
"""

import re
import unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _parse_value(val_str):
    """XBRL値文字列を float に変換。"""
    if val_str is None or val_str == '':
        return None
    s = unicodedata.normalize('NFKC', str(val_str)).replace(',', '').strip()
    s = s.replace('－', '-').replace('−', '-')
    if s == '-' or s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _format_date(period_str):
    """YYYY-MM-DD -> YYYY/M/D"""
    if not period_str or not isinstance(period_str, str):
        return period_str
    parts = period_str.split('-')
    if len(parts) == 3:
        try:
            return f"{int(parts[0])}/{int(parts[1])}/{int(parts[2])}"
        except ValueError:
            pass
    return period_str


def _get_best_value_for_element(global_element_period_values, el_name, member_keywords):
    """
    指定された要素について、各年度(period)で最も適切なメンバー(dim_label)の値を取得する。
    """
    vals = global_element_period_values.get(el_name, {})
    if not vals:
        return {}
        
    period_value_map = {} # {period: value}
    for (fact_std, dim_label, period), raw_val in vals.items():
        if not period:
            continue
        v = _parse_value(raw_val)
        if v is None:
            continue
            
        dim_l = str(dim_label).lower()
        match = False
        for kw in member_keywords:
            if kw.lower() in dim_l:
                match = True
                break
        
        if match:
            if period not in period_value_map:
                period_value_map[period] = v
    return period_value_map


def add_remuneration_sheet(workbook, global_element_period_values, debug_log=None):
    """
    「役員区分ごとの報酬等」シートを生成して追加する。
    縦：役員区分・年度、横：報酬区分
    """
    if debug_log is None:
        def debug_log(msg): pass

    sheet_name = '役員区分ごとの報酬等'
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # --- カテゴリー（行）定義 ---
    categories = [
        ('取締役（社外除く）', ['DirectorsExcludingOutsideDirectorsMember', '社外取締役を除く']),
        ('監査役（社外除く）', ['CorporateAuditorsExcludingOutsideCorporateAuditorsMember', '社外監査役を除く']),
        ('社外取締役', ['OutsideDirectorsMember', '社外取締役']),
        ('社外監査役', ['OutsideCorporateAuditorsMember', '社外監査役']),
    ]

    # --- 項目（列）定義 ---
    metrics = [
        ('報酬等の総額（円）', 'TotalAmountOfRemunerationEtcRemunerationEtcByCategoryOfDirectorsAndOtherOfficers', '#,##0'),
        ('固定報酬（円）', 'FixedRemunerationRemunerationByCategoryOfDirectorsAndOtherOfficers', '#,##0'),
        ('業績連動報酬（円）', 'PerformanceBasedRemunerationRemunerationByCategoryOfDirectorsAndOtherOfficers', '#,##0'),
        ('非金銭報酬等（円）', 'NonMonetaryRemunerationRemunerationByCategoryOfDirectorsAndOtherOfficers', '#,##0'),
        ('退職慰労金（円）', 'RetirementBenefitsRemunerationEtcByCategoryOfDirectorsAndOtherOfficers', '#,##0'),
        ('対象員数（人）', 'NumberOfDirectorsAndOtherOfficersRemunerationEtcByCategoryOfDirectorsAndOtherOfficers', '#,##0'),
    ]

    # --- スタイル ---
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    section_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    current_row = 1
    # --- タイトル: 実数の表 ---
    ws.cell(row=current_row, column=1, value='【報酬等の実数 (円、人)】').font = Font(bold=True, size=11)
    current_row += 1

    # ヘッダー作成 (A列:役員区分, B列:年度, C列以降:報酬区分)
    header_row = current_row
    headers = [('役員区分', 1), ('年度', 2)]
    for i, (m_label, _, _) in enumerate(metrics, start=3):
        headers.append((m_label, i))

    for text, col in headers:
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    current_row += 1

    # 全期間を収集 (2015-2025)
    all_years = range(2015, 2026)
    
    # データ抽出
    data_store = {} # {m_el: {cat_name: {period: val}}}
    periods_seen = set()

    for m_label, m_el, m_fmt in metrics:
        el_name = f"jpcrp_cor_{m_el}"
        data_store[el_name] = {}
        for cat_name, keywords in categories:
            best_vals = _get_best_value_for_element(global_element_period_values, el_name, keywords)
            data_store[el_name][cat_name] = best_vals
            periods_seen.update(best_vals.keys())

    # 各行の対応関係を記録するマップ (cat_name, display_date) -> row_index
    data_row_map = {}

    # --- 1. 実数の表の作成 ---
    for cat_name, keywords in categories:
        cat_start_row = current_row
        
        for y in all_years:
            period_to_use = None
            matches = [p for p in periods_seen if p.startswith(str(y))]
            if matches:
                period_to_use = sorted(matches)[-1]
            
            # A列: 役員区分
            cell_cat = ws.cell(row=current_row, column=1, value=cat_name)
            cell_cat.font = section_font
            cell_cat.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell_cat.border = thin_border
            
            # B列: 年度
            display_date = _format_date(period_to_use) if period_to_use else f"{y}/3/31"
            data_row_map[(cat_name, display_date)] = current_row # 行番号を記録

            cell_year = ws.cell(row=current_row, column=2, value=display_date)
            cell_year.font = normal_font
            cell_year.alignment = center_align
            cell_year.border = thin_border
            
            # C列以降: 各指標
            for col_idx, (m_label, m_el, m_fmt) in enumerate(metrics, start=3):
                el_name = f"jpcrp_cor_{m_el}"
                val = data_store[el_name][cat_name].get(period_to_use) if period_to_use else None
                
                cell_val = ws.cell(row=current_row, column=col_idx, value=val)
                cell_val.font = normal_font
                cell_val.alignment = right_align
                cell_val.number_format = m_fmt
                cell_val.border = thin_border
                
            current_row += 1
        
        ws.merge_cells(start_row=cat_start_row, end_row=current_row-1, start_column=1, end_column=1)
        current_row += 1 # 余白

    # --- 2. 比率の表の作成 ---
    current_row += 1
    ws.cell(row=current_row, column=1, value='【報酬区分ごとの比率 (対報酬等の総額)】').font = Font(bold=True, size=11)
    current_row += 1
    
    # ヘッダー (再利用)
    header_row_ratio = current_row
    for text, col in headers:
        cell = ws.cell(row=header_row_ratio, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 1

    for cat_name, keywords in categories:
        cat_start_row = current_row
        
        for y in all_years:
            # A列: 役員区分
            cell_cat = ws.cell(row=current_row, column=1, value=cat_name)
            cell_cat.font = section_font
            cell_cat.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell_cat.border = thin_border
            
            # B列: 年度
            period_to_use = None
            matches = [p for p in periods_seen if p.startswith(str(y))]
            if matches: period_to_use = sorted(matches)[-1]
            display_date = _format_date(period_to_use) if period_to_use else f"{y}/3/31"
            
            cell_year = ws.cell(row=current_row, column=2, value=display_date)
            cell_year.font = normal_font
            cell_year.alignment = center_align
            cell_year.border = thin_border
            
            # C列以降: 比率の計算式
            # 参照先の行番号を取得
            ref_row = data_row_map.get((cat_name, display_date))
            
            total_col = "C"
            for col_idx, (m_label, m_el, m_fmt) in enumerate(metrics, start=3):
                col_letter = chr(64 + col_idx)
                if ref_row:
                    # Formula references the row in the FIRST table
                    formula = f'=IF(OR(${total_col}{ref_row}="", {col_letter}{ref_row}="", ${total_col}{ref_row}=0), "", {col_letter}{ref_row}/${total_col}{ref_row})'
                else:
                    formula = ""
                
                cell_ratio = ws.cell(row=current_row, column=col_idx, value=formula)
                cell_ratio.font = normal_font
                cell_ratio.alignment = right_align
                cell_ratio.number_format = '0.0%'
                cell_ratio.border = thin_border
                
            current_row += 1

        ws.merge_cells(start_row=cat_start_row, end_row=current_row-1, start_column=1, end_column=1)
        current_row += 1 # 余白

    # 列幅の設定
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for i in range(3, 3 + len(metrics)):
        ws.column_dimensions[chr(64 + i)].width = 22
    
    ws.freeze_panes = 'C2'
    debug_log(f"[Remuneration] '{sheet_name}' 出力完了 (実数・比率併記版)")
