"""
Financial Analysis Module for XBRL to Excel Conversion

ROE分析やその他の財務分析機能を提供するモジュール
"""

import openpyxl.utils
import re

def create_roe_analysis_sheet(workbook, source_sheet_name, debug_log=None):
    """
    主要な経営指標等の推移シートからROE分析シートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "主要な経営指標等の推移（連結）(日本基準)"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]

    # 正規表現で「最初のカッコ内」と「それ以降」を抽出
    # \1: 最初のカッコの中身（連結）
    # \2: それ以降の文字列（(日本基準)）
    match = re.search(r'[（\(](.*?)[）\)](.*)', source_sheet_name)

    if match:
        # 「連結」 + 「(日本基準)」 の形にする
        short_name = f"{match.group(1)}{match.group(2)}"
    else:
        # カッコが含まれない場合のフォールバック
        short_name = source_sheet_name

    analysis_sheet_name = f"{short_name}_ROE分析"

    # シート名31文字制限対策（念のため）
    analysis_sheet_name = analysis_sheet_name[:31]

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 列数を取得
    # Note: num_cols will be adjusted later based on kikan
    # For now, just use source columns
    source_cols = source_ws.max_column
    num_cols = source_cols  # Will be adjusted later if kikan >= 10

    # 参照する行番号を特定（元シートから）
    row_mapping = {}  # 英語名 -> 行番号のマッピング
    for row in range(2, source_ws.max_row + 1):
        english_name = source_ws.cell(row, 2).value  # B列: 項目（英名）
        if english_name:
            row_mapping[english_name] = row

    def find_all_rows_by_keywords(keywords_list, item_name):
        """
        複数のキーワード候補で見つかったすべての行を検索（部分一致）
        """
        all_found = []
        for keywords in keywords_list:
            if isinstance(keywords, str):
                keywords = [keywords]

            for eng_name, row_num in row_mapping.items():
                if all(kw in eng_name for kw in keywords):
                    if row_num not in all_found:
                        all_found.append(row_num)
        
        if all_found:
            debug_log(f"{item_name}: found {len(all_found)} rows: {all_found}")
        return all_found

    def find_row_by_keywords(keywords_list, item_name):
        rows = find_all_rows_by_keywords(keywords_list, item_name)
        return rows[0] if rows else None

    # IFRS判定
    is_ifrs = 'IFRS' in source_sheet_name

    # 必要な勘定科目のキーワード候補（優先度順）
    # いずれの標準でも、推移表では「売上高」と「売上収益」の両方が現れうるため、混合して定義する
    sales_keywords_all = [
        ['Revenue', 'KeyFinancialData'],  # 売上収益（主要指標用）
        ['NetSales', 'KeyFinancialData'], # 売上高（主要指標用）
        ['Revenue', 'Summary'],  # 売上収益
        ['NetSales', 'Summary'],  # 売上高
        ['Revenue'],
        ['NetSales'],
        ['Sales', 'Summary'],
    ]
    
    profit_keywords_all = [
        ['ProfitLoss', 'OwnersOfParent', 'KeyFinancialData'],
        ['NetIncome', 'KeyFinancialData'],
        ['ProfitLoss', 'OwnersOfParent', 'Summary'],
        ['ProfitLoss', 'Attributable', 'OwnersOfParent'],
        ['ProfitLoss', 'Parent', 'Summary'],
        ['Profit', 'OwnersOfParent'],
        ['ProfitLoss', 'Summary'],
    ]

    if is_ifrs:
        # IFRS用キーワード (優先度順)
        sales_keywords = sales_keywords_all
        profit_keywords = profit_keywords_all

        # IFRS用: 親会社の所有者に帰属する持分（自己資本の代わり）
        equity_keywords = [
            ['EquityAttributableToOwnersOfParent', 'IFRS'],  # 親会社の所有者に帰属する持分
            ['EquityAttributable', 'OwnersOfParent'],  # 代替パターン
            ['Equity', 'OwnersOfParent'],  # 汎用
        ]

        total_assets_keywords = [
            ['TotalAssets', 'IFRS', 'Summary'],  # 総資産額（IFRS）
            ['TotalAssets', 'Summary'],  # 総資産額（汎用）
            ['TotalAssets'],  # 最低限
        ]

        equity_ratio_keywords = [
            ['RatioOfOwnersEquity', 'IFRS'],  # 親会社所有者帰属持分比率（IFRS）
            ['Ratio', 'OwnersEquity'],  # 代替パターン
            ['EquityRatio'],  # 汎用
        ]

        roe_keywords = [
            ['RateOfReturnOnEquity', 'IFRS'],  # 親会社所有者帰属持分利益率（IFRS ROE）
            ['ReturnOnEquity', 'Summary'],  # 代替パターン
            ['ROE'],  # 汎用
        ]
    else:
        # 日本基準用キーワード
        sales_keywords = sales_keywords_all
        profit_keywords = profit_keywords_all

        net_assets_keywords = [
            ['NetAssets', 'Summary'],  # 純資産額（標準）
            ['NetAssets'],  # 純資産額（汎用）
        ]

        total_assets_keywords = [
            ['TotalAssets', 'Summary'],  # 総資産額（標準）
            ['TotalAssets'],  # 総資産額（汎用）
        ]

        equity_ratio_keywords = [
            ['EquityToAssetRatio', 'Summary'],  # 自己資本比率（標準）
            ['EquityToAsset', 'Summary'],  # 代替パターン
            ['EquityRatio'],  # 汎用
        ]

        roe_keywords = [
            ['RateOfReturnOnEquity', 'Summary'],  # 自己資本利益率（標準）
            ['ReturnOnEquity', 'Summary'],  # 代替パターン
            ['ROE'],  # 汎用
        ]

    # 元シートの行番号を取得（部分一致検索）
    sales_rows = find_all_rows_by_keywords(sales_keywords, '売上高/売上収益')
    
    # 優先順位付け: IFRS(Revenue)を優先する (年度の新しいほうを採用)
    def sales_priority_roe(row_num):
        # source_wsのB列（英名）で判定
        b_val = str(source_ws.cell(row_num, 2).value or "")
        if 'Revenue' in b_val or 'IFRS' in b_val:
            return 0
        return 1
    
    if sales_rows:
        sales_rows.sort(key=sales_priority_roe)
        debug_log(f"ROE Sales rows (prioritized): {sales_rows}")
        
    sales_row = sales_rows[0] if sales_rows else None
    profit_row = find_row_by_keywords(profit_keywords, '当期純利益/当期利益')

    if is_ifrs:
        # IFRS: 親会社の所有者に帰属する持分を直接使用
        equity_row = find_row_by_keywords(equity_keywords, '親会社の所有者に帰属する持分')
        net_assets_row = None  # IFRSでは使用しない
    else:
        # 日本基準: 純資産額を使用
        net_assets_row = find_row_by_keywords(net_assets_keywords, '純資産額')
        equity_row = None  # 日本基準では計算で求める

    total_assets_row = find_row_by_keywords(total_assets_keywords, '総資産額')
    equity_ratio_row = find_row_by_keywords(equity_ratio_keywords, '自己資本比率/親会社所有者帰属持分比率')
    roe_row = find_row_by_keywords(roe_keywords, 'ROE')

    # 売上高が見つからない場合は、セクションヘッダーの次の行を使用
    if sales_row is None:
        # ヘッダーを部分一致で検索
        header_keywords = [
            ['BusinessResults', 'Heading'],
            ['BusinessResults'],
        ]
        header_row_num = find_row_by_keywords(header_keywords, 'セクションヘッダー')
        if header_row_num:
            # ヘッダーの次の行を売上高として使用
            sales_row = header_row_num + 1
            debug_log(f"Sales row not found, using first item after header (row {sales_row})")

    # ROE分析に必要な項目がすべて存在するかチェック
    if is_ifrs:
        required_items = {
            '売上収益': sales_row,
            '当期利益（親会社帰属）': profit_row,
            '総資産額': total_assets_row,
            '親会社所有者帰属持分比率': equity_ratio_row,
            'ROE': roe_row
        }
    else:
        required_items = {
            '売上高': sales_row,
            '当期純利益': profit_row,
            '総資産額': total_assets_row,
            '自己資本比率': equity_ratio_row,
            'ROE': roe_row
        }
    missing_items = [name for name, row in required_items.items() if row is None]
    if missing_items:
        debug_log(f"ROE analysis skipped for '{source_sheet_name}': missing items: {', '.join(missing_items)}")
        return

    # ヘッダー行と基本指標をすべて元シートから参照する形で追加
    def add_reference_row_full(source_row_num):
        """元シートの行全体を参照する行を追加"""
        if source_row_num is None:
            debug_log(f"Skipping row addition: source_row_num is None")
            return
        row_data = []
        for col in range(1, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            # 元シートのセルを参照する数式
            formula = f"='{source_sheet_name}'!{col_letter}{source_row_num}"
            row_data.append(formula)
        analysis_ws.append(row_data)

    # 1行目: ヘッダー（勘定科目、項目（英名）など）
    add_reference_row_full(1)

    # 2行目: セクションヘッダー（連結経営指標等）
    # A列とB列のみ参照、C列以降は空欄
    header_keywords = [
        ['BusinessResults', 'Heading'],
        ['BusinessResults'],
    ]
    header_row_num = find_row_by_keywords(header_keywords, 'セクションヘッダー')
    if header_row_num:
        # A列とB列のみ参照
        row_data = []
        for col in range(1, 3):  # A列とB列のみ
            col_letter = openpyxl.utils.get_column_letter(col)
            formula = f"='{source_sheet_name}'!{col_letter}{header_row_num}"
            row_data.append(formula)
        # C列以降は空欄
        for col in range(3, num_cols + 1):
            row_data.append('')
        analysis_ws.append(row_data)

    # 3-8行目: 基本指標（元シートから参照）
    if len(sales_rows) > 1:
        # 複数行ある場合は優先順位に従って「年度の新しいほう（IFRS等）」を採用する (Deduplication)
        # IF(収益<>0, 収益, 売上高) の形式で数式を作成
        row_data = [re.sub(r'\(.*?\)', '', source_ws.cell(sales_rows[0], 1).value).strip() if source_ws.cell(sales_rows[0], 1).value else '売上高', ' / '.join([f"'{source_sheet_name}'!B{r}" for r in sales_rows])]
        for col in range(3, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            # 優先順位に従った数式 (売上高1<>"", 売上高1, 売上高2...)
            formula = f"'{source_sheet_name}'!{col_letter}{sales_rows[-1]}"
            for r in reversed(sales_rows[:-1]):
                curr_cell = f"'{source_sheet_name}'!{col_letter}{r}"
                formula = f"IF({curr_cell}<>\"\",{curr_cell},{formula})"
            row_data.append(f"={formula}")
        analysis_ws.append(row_data)
    else:
        add_reference_row_full(sales_row)
    add_reference_row_full(profit_row)
    if is_ifrs:
        # IFRS: 親会社の所有者に帰属する持分を追加
        if equity_row is not None:
            add_reference_row_full(equity_row)
        else:
            # 持分が元データにない場合は、総資産額 × 親会社持分比率 で計算行を作成
            row_data = ['　親会社の所有者に帰属する持分（計算値）', '']
            for col in range(3, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                formula = f"='{source_sheet_name}'!{col_letter}{total_assets_row}*'{source_sheet_name}'!{col_letter}{equity_ratio_row}"
                row_data.append(formula)
            analysis_ws.append(row_data)
    else:
        # 日本基準: 純資産額を追加
        add_reference_row_full(net_assets_row)
    add_reference_row_full(total_assets_row)
    add_reference_row_full(equity_ratio_row)
    add_reference_row_full(roe_row)

    # 現在の行番号を記録（上記で追加した行の位置）
    current_row = analysis_ws.max_row
    sales_analysis_row = current_row - 5
    profit_analysis_row = current_row - 4
    if is_ifrs:
        equity_analysis_row = current_row - 3  # IFRS: 親会社の所有者に帰属する持分
        net_assets_analysis_row = None
    else:
        net_assets_analysis_row = current_row - 3  # 日本基準: 純資産額
        equity_analysis_row = None
    total_assets_analysis_row = current_row - 2
    equity_ratio_analysis_row = current_row - 1
    roe_analysis_row = current_row

    # 空行
    analysis_ws.append([''] * num_cols)

    # 計算指標
    # 自己資本の計算
    equity_calc_row_num = analysis_ws.max_row + 1
    equity_calc_row = ['　　　自己資本', '']
    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        if is_ifrs:
            # IFRS: 親会社の所有者に帰属する持分を直接参照
            formula = f"={col_letter}{equity_analysis_row}"
        else:
            # 日本基準: 総資産額 × 自己資本比率
            formula = f"={col_letter}{total_assets_analysis_row}*{col_letter}{equity_ratio_analysis_row}"
        equity_calc_row.append(formula)
    analysis_ws.append(equity_calc_row)

    # 自己資本（平均） = AVERAGE(前期:当期)
    equity_avg_row_num = analysis_ws.max_row + 1
    equity_avg_row = ['　　　自己資本（平均）', '']
    # 最初の期（C列）は計算不可
    equity_avg_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
        formula = f"=AVERAGE({prev_col_letter}{equity_calc_row_num}:{col_letter}{equity_calc_row_num})"
        equity_avg_row.append(formula)
    analysis_ws.append(equity_avg_row)

    # 総資産（平均） = AVERAGE(前期:当期)
    total_assets_avg_row_num = analysis_ws.max_row + 1
    total_assets_avg_row = ['　　　総資産（平均）', '']
    # 最初の期（C列）は計算不可
    total_assets_avg_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
        formula = f"=AVERAGE({prev_col_letter}{total_assets_analysis_row}:{col_letter}{total_assets_analysis_row})"
        total_assets_avg_row.append(formula)
    analysis_ws.append(total_assets_avg_row)

    # 空行
    analysis_ws.append([''] * num_cols)

    # ROE分析指標
    # ROE = 元シートのROE
    roe_calc_row_num = analysis_ws.max_row + 1
    roe_calc_row = ['　　　自己資本利益率(ROE)', '']
    # 最初の期（C列）は計算不可（平均が必要なため）
    roe_calc_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{roe_analysis_row}"
        roe_calc_row.append(formula)
    analysis_ws.append(roe_calc_row)

    # ROS = 当期純利益 / 売上高
    ros_row_num = analysis_ws.max_row + 1
    ros_row = ['　　　売上高利益率(ROS)', '']
    ros_row.append('')  # 最初の期
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{profit_analysis_row}/{col_letter}{sales_analysis_row}"
        ros_row.append(formula)
    analysis_ws.append(ros_row)

    # TOR = 売上高 / 総資産（平均）
    tor_row_num = analysis_ws.max_row + 1
    tor_row = ['　　　総資産回転率(TOR)', '']
    tor_row.append('')  # 最初の期
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{sales_analysis_row}/{col_letter}{total_assets_avg_row_num}"
        tor_row.append(formula)
    analysis_ws.append(tor_row)

    # LRV = 総資産（平均） / 自己資本（平均）
    lrv_row_num = analysis_ws.max_row + 1
    lrv_row = ['　　　レバレッジ(LEV)', '']
    lrv_row.append('')  # 最初の期
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{total_assets_avg_row_num}/{col_letter}{equity_avg_row_num}"
        lrv_row.append(formula)
    analysis_ws.append(lrv_row)

    # 検算1: ROS * TOR * LRV = ROE
    check1_row_num = analysis_ws.max_row + 1
    check1_row = ['　　　検算1(ROS*TOR*LEV=ROE)', '']
    check1_row.append('')  # 最初の期
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"=PRODUCT({col_letter}{ros_row_num}:{col_letter}{lrv_row_num})"
        check1_row.append(formula)
    analysis_ws.append(check1_row)

    # 検算2: 検算1 = ROE（TRUE/FALSE）
    check2_row_num = analysis_ws.max_row + 1
    check2_row = ['　　　検算2(検算1=ROE)', '']
    check2_row.append('')  # 最初の期
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
    #    formula = f"=ROUND({col_letter}{roe_calc_row_num},3)=ROUND({col_letter}{check1_row_num},3)"
        formula = f"=ABS(ROUND({col_letter}{roe_calc_row_num},3)-ROUND({col_letter}{check1_row_num},3))<=0.001"
        check2_row.append(formula)
    analysis_ws.append(check2_row)

    # ROA = 当期純利益 / 総資産（平均）
    roa_row_num = analysis_ws.max_row + 1
    roa_row = ['　　　ROA(総資産利益率)', '']
    roa_row.append('')  # 最初の期
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{profit_analysis_row}/{col_letter}{total_assets_avg_row_num}"
        roa_row.append(formula)
    analysis_ws.append(roa_row)

    # ============================================================================
    # セルの表示形式を設定
    # ============================================================================
    # 数値フォーマット定義
    number_format_integer = r'#,##0_ ;[Red]\-#,##0\ '  # 整数（カンマ区切り）
    number_format_decimal = r'#,##0_);[Red](#,##0)'  # 整数（カンマ区切り、負数は括弧）
    number_format_decimal2 = r'#,##0.00_);[Red](#,##0.00)'  # 小数2桁
    number_format_percent = r'0.0%'  # パーセント（小数1桁）

    # 基本指標の表示形式（元シートから参照している行）
    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)

        # 売上高、当期純利益、純資産額/親会社持分、総資産額: 整数カンマ区切り
        for row_num in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row, equity_analysis_row, total_assets_analysis_row]:
            if row_num is not None:
                analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_integer

        # 自己資本比率、自己資本利益率: パーセント
        for row_num in [equity_ratio_analysis_row, roe_analysis_row]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_percent

    # 計算指標の表示形式
    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)

        # 自己資本、自己資本（平均）、総資産（平均）: 整数カンマ区切り（括弧）
        for row_num in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_decimal

    # ROE分析指標の表示形式
    for col in range(4, num_cols + 1):  # D列から（C列は空）
        col_letter = openpyxl.utils.get_column_letter(col)

        # ROE、ROS、ROA: パーセント
        for row_num in [roe_calc_row_num, ros_row_num, roa_row_num]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_percent

        # TOR、LRV: 小数2桁
        for row_num in [tor_row_num, lrv_row_num]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_decimal2

        # 検算1: パーセント
        analysis_ws[f'{col_letter}{check1_row_num}'].number_format = number_format_percent

        # 検算2: パーセント（実際はTRUE/FALSEだが元シートに合わせる）
        analysis_ws[f'{col_letter}{check2_row_num}'].number_format = number_format_percent

    # ============================================================================
    # 列幅の設定とウィンドウ枠の固定
    # ============================================================================
    # A列: 幅28
    analysis_ws.column_dimensions['A'].width = 28

    # B列: 非表示
    analysis_ws.column_dimensions['B'].hidden = True

    # C列以降（年度の列）: 幅12
    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # B2でウィンドウ枠を固定
    analysis_ws.freeze_panes = 'B2'

    # ============================================================================
    # 対前年増加率セクション（A23以降）
    # ============================================================================
    # 空行（2行）
    analysis_ws.append([''] * num_cols)
    analysis_ws.append([''] * num_cols)

    # A23行: "　対前年増加率" ヘッダー
    yoy_header_row_num = analysis_ws.max_row + 1
    yoy_header_row = ['　対前年増加率', '']
    # C列は空欄（対前年増加率の初年度なので前年がない）
    yoy_header_row.append('')  # C列
    # D列以降は1行目を参照する数式
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        yoy_header_row.append(f'={col_letter}1')
    analysis_ws.append(yoy_header_row)

    # IFRS/日本基準で使用する行を決定
    if is_ifrs:
        # IFRS: 親会社の所有者に帰属する持分を使用
        equity_or_net_assets_row = equity_analysis_row
    else:
        # 日本基準: 純資産額を使用
        equity_or_net_assets_row = net_assets_analysis_row

    # 対前年増加率を計算する行の定義
    # 行24-29: 基本指標（売上高、当期純利益、純資産額/親会社持分、総資産額、自己資本比率、ROE）
    yoy_rows_basic = []
    for source_row in [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                       total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
        yoy_row_num = analysis_ws.max_row + 1
        yoy_rows_basic.append(yoy_row_num)
        yoy_row = [f'=A{source_row}', '']  # A列は元の行を参照
        yoy_row.append('')  # C列は空（初年度は前年がない）
        yoy_row.append('')  # D列は空（基準年なので前年がない）
        # E列以降: =E{source_row}/D{source_row}-1
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}/{prev_col_letter}{source_row}-1"
            yoy_row.append(formula)
        analysis_ws.append(yoy_row)

    # 空行
    analysis_ws.append([''] * num_cols)

    # 行31-33: 計算指標（自己資本、自己資本（平均）、総資産（平均））
    yoy_rows_calc = []
    for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
        yoy_row_num = analysis_ws.max_row + 1
        yoy_rows_calc.append(yoy_row_num)
        yoy_row = [f'=A{source_row}', '']
        yoy_row.append('')  # C列は空
        yoy_row.append('')  # D列は空（基準年なので前年がない）
        # E列以降: =E{source_row}/D{source_row}-1
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}/{prev_col_letter}{source_row}-1"
            yoy_row.append(formula)
        analysis_ws.append(yoy_row)

    # 空行
    analysis_ws.append([''] * num_cols)

    # 行35-41: ROE分析指標
    yoy_rows_roe = []
    for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                       check1_row_num, check2_row_num, roa_row_num]:
        yoy_row_num = analysis_ws.max_row + 1
        yoy_rows_roe.append(yoy_row_num)
        yoy_row = [f'=A{source_row}', '']
        yoy_row.append('')  # C列は空

        # check2_row_num（検算2）の場合は全列空欄
        if source_row == check2_row_num:
            for col in range(4, num_cols + 1):
                yoy_row.append('')
        else:
            yoy_row.append('')  # D列は空（基準年なので前年がない）
            # E列以降: =E{source_row}/D{source_row}-1
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
                formula = f"={col_letter}{source_row}/{prev_col_letter}{source_row}-1"
                yoy_row.append(formula)
        analysis_ws.append(yoy_row)

    # 対前年増加率セクションの表示形式を設定（パーセント）
    # E列以降（D列は空欄なので除外）
    for col in range(5, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row_num in yoy_rows_basic + yoy_rows_calc + yoy_rows_roe:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_percent


    # ============================================================================
    # 対前年差セクション
    # ============================================================================
    analysis_ws.append([''] * num_cols)
    analysis_ws.append([''] * num_cols)

    diff_yoy_header_row_num = analysis_ws.max_row + 1
    diff_yoy_header_row = ['　対前年差', '']
    diff_yoy_header_row.append('')  # C列
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        diff_yoy_header_row.append(f'={col_letter}1')
    analysis_ws.append(diff_yoy_header_row)

    # 行ごとの書式設定マッピング
    row_formats_map = {
        sales_analysis_row: number_format_integer,
        profit_analysis_row: number_format_integer,
        equity_or_net_assets_row: number_format_integer,
        total_assets_analysis_row: number_format_integer,
        equity_ratio_analysis_row: number_format_percent,
        roe_analysis_row: number_format_percent,
        equity_calc_row_num: number_format_decimal,
        equity_avg_row_num: number_format_decimal,
        total_assets_avg_row_num: number_format_decimal,
        roe_calc_row_num: number_format_percent,
        ros_row_num: number_format_percent,
        tor_row_num: number_format_decimal2,
        lrv_row_num: number_format_decimal2,
        check1_row_num: number_format_percent,
        check2_row_num: number_format_percent,
        roa_row_num: number_format_percent,
    }

    # 各項目を追加
    diff_rows_basic = []
    for source_row in [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                       total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
        diff_row_num = analysis_ws.max_row + 1
        diff_rows_basic.append(diff_row_num)
        diff_row = [f'=A{source_row}', '']
        diff_row.append('')  # C列
        diff_row.append('')  # D列は空
        # E列以降: =E{source_row}-D{source_row}
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}-{prev_col_letter}{source_row}"
            diff_row.append(formula)
        analysis_ws.append(diff_row)
        # 書式設定の適用
        fmt = row_formats_map.get(source_row)
        if fmt:
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                analysis_ws[f'{col_letter}{diff_row_num}'].number_format = fmt

    # 空行
    analysis_ws.append([''] * num_cols)

    diff_rows_calc = []
    for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
        diff_row_num = analysis_ws.max_row + 1
        diff_rows_calc.append(diff_row_num)
        diff_row = [f'=A{source_row}', '']
        diff_row.append('')  # C列
        diff_row.append('')  # D列は空
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}-{prev_col_letter}{source_row}"
            diff_row.append(formula)
        analysis_ws.append(diff_row)
        fmt = row_formats_map.get(source_row)
        if fmt:
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                analysis_ws[f'{col_letter}{diff_row_num}'].number_format = fmt

    # 空行
    analysis_ws.append([''] * num_cols)

    diff_rows_roe = []
    for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                       check1_row_num, check2_row_num, roa_row_num]:
        diff_row_num = analysis_ws.max_row + 1
        diff_rows_roe.append(diff_row_num)
        diff_row = [f'=A{source_row}', '']
        diff_row.append('')  # C列
        if source_row == check2_row_num:
            for col in range(4, num_cols + 1):
                diff_row.append('')
        else:
            diff_row.append('')  # D列
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
                formula = f"={col_letter}{source_row}-{prev_col_letter}{source_row}"
                diff_row.append(formula)
        analysis_ws.append(diff_row)
        fmt = row_formats_map.get(source_row)
        if fmt:
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                analysis_ws[f'{col_letter}{diff_row_num}'].number_format = fmt


    # ============================================================================
    # 10年前からの増加率計算（Q列）
    # ============================================================================
    # 最新の年の列を特定（最後のデータ列）
    # source_colsは元シートの実際の列数、num_colsはパディングを含む列数
    latest_col = source_cols
    latest_col_letter = openpyxl.utils.get_column_letter(latest_col)

    # 基準となる5つの指標がすべて揃っているかチェック
    target_rows = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                   total_assets_analysis_row, equity_ratio_analysis_row]

    # Helper function to check if a column has all required data
    def has_all_data(col_num):
        """Check if the specified column has all required data in the source sheet"""
        col_letter = openpyxl.utils.get_column_letter(col_num)
        for row_num in target_rows:
            cell_formula = analysis_ws[f'{col_letter}{row_num}'].value
            if not cell_formula:
                return False

            # Check the actual data in source sheet
            if isinstance(cell_formula, str) and cell_formula.startswith('='):
                try:
                    parts = cell_formula.split('!')
                    if len(parts) == 2:
                        source_cell_ref = parts[1].strip()
                        source_cell_value = source_ws[source_cell_ref].value
                        if source_cell_value is None or source_cell_value == '':
                            return False
                except Exception:
                    return False
        return True

    # Step 1: Find the oldest column (FYa) with all required data
    oldest_col = None
    for col in range(3, latest_col):
        if has_all_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        # No valid data found
        base_col = None
        kikan = 0  # No data available
        num_cols = source_cols  # Keep original columns
    else:
        # Step 2: Calculate FYb = FYa + 1 (need previous year for average calculation)
        fyb_col = oldest_col + 1

        # Step 3: Calculate period: kikan = FYc - FYb
        kikan = latest_col - fyb_col

        # Adjust num_cols based on kikan
        # For kikan >= 10: need padding to at least 16 columns for 10-year comparison
        # For kikan < 10: use actual source columns only (no padding)
        if kikan >= 10:
            num_cols = max(source_cols, 16)  # Ensure at least P column (16) for 10-year data
        else:
            num_cols = source_cols  # No padding for < 10 year data

        # Step 4: Determine base column based on period
        if kikan >= 10:
            # Use 10-year comparison: base = FYc - 10
            candidate_base_col = latest_col - 10

            # Handle fiscal year change: if multiple periods exist for the same fiscal year,
            # prefer the earlier one (which represents a full 12-month period)
            # Example: If 2014/12 and 2014/03 both exist, prefer 2014/03
            # Check if the previous column (candidate_base_col - 1) has the same year
            if candidate_base_col > 3:  # C列より右であれば前の列をチェック可能
                try:
                    # Get year from candidate and previous column
                    candidate_letter = openpyxl.utils.get_column_letter(candidate_base_col)
                    prev_letter = openpyxl.utils.get_column_letter(candidate_base_col - 1)

                    # Use YEAR function in formula to compare years
                    # Check if previous column has data and same year
                    if has_all_data(candidate_base_col - 1):
                        # Get the header cell references
                        candidate_ref = f"'{source_sheet_name}'!{candidate_letter}1"
                        prev_ref = f"'{source_sheet_name}'!{prev_letter}1"

                        # Check year values from source sheet
                        candidate_header = analysis_ws[f'{candidate_letter}1'].value
                        if isinstance(candidate_header, str) and candidate_header.startswith('='):
                            # Extract cell reference
                            parts = candidate_header.split('!')
                            if len(parts) == 2:
                                source_cell_ref = parts[1].strip().replace("'", "")
                                candidate_date = source_ws[source_cell_ref].value
                                prev_date = source_ws[f'{prev_letter[0] if len(prev_letter)==1 else prev_letter}1'].value

                                # If both dates exist and have same year, use the earlier month
                                if candidate_date and prev_date:
                                    if hasattr(candidate_date, 'year') and hasattr(prev_date, 'year'):
                                        if candidate_date.year == prev_date.year:
                                            # Same year - use previous column (earlier month = full year period)
                                            base_col = candidate_base_col - 1
                                            debug_log(f"Fiscal year change detected: using {prev_letter} instead of {candidate_letter} (same year {candidate_date.year})")
                                        else:
                                            base_col = candidate_base_col
                                    else:
                                        base_col = candidate_base_col
                                else:
                                    base_col = candidate_base_col
                        else:
                            base_col = candidate_base_col
                    else:
                        base_col = candidate_base_col
                except Exception:
                    # If any error occurs, fall back to original candidate
                    base_col = candidate_base_col
            else:
                base_col = candidate_base_col

            base_col_letter = openpyxl.utils.get_column_letter(base_col)
            debug_log(f"Using 10-year comparison: kikan={kikan}, base column {base_col_letter}")
        else:
            # Use longest available period: base = FYb
            base_col = fyb_col
            base_col_letter = openpyxl.utils.get_column_letter(base_col)
            oldest_col_letter = openpyxl.utils.get_column_letter(oldest_col)
            debug_log(f"Using longest available period: kikan={kikan}, oldest={oldest_col_letter}, base={base_col_letter}")

    # 増加率列を追加（Q列 = num_cols + 1）
    if base_col is not None:
        growth_col = num_cols + 1
        growth_col_letter = openpyxl.utils.get_column_letter(growth_col)
        base_col_letter = openpyxl.utils.get_column_letter(base_col)

        # Q1: 期間表示（例: "2024/2014"）
        period_formula = f"=YEAR({latest_col_letter}1) & \"/\" & YEAR({base_col_letter}1)"
        analysis_ws[f'{growth_col_letter}1'] = period_formula

        # Q2: 空（ヘッダー行）
        analysis_ws[f'{growth_col_letter}2'] = ''

        # Q3-Q20: 簡単な比率計算（最新値/基準値）
        # 基本指標: Q3-Q8 (売上高、当期純利益、純資産額、総資産額、自己資本比率、ROE)
        for source_row in [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                          total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
            ratio_formula = f"={latest_col_letter}{source_row}/{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{source_row}'] = ratio_formula
            # 書式設定: #,##0.00;[Red]-#,##0.00
            analysis_ws[f'{growth_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

        # Q9: 空行（対応する行9が空行）

        # 計算指標: Q10-Q12 (自己資本、自己資本（平均）、総資産（平均）)
        for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
            ratio_formula = f"={latest_col_letter}{source_row}/{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{source_row}'] = ratio_formula
            analysis_ws[f'{growth_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

        # Q13: 空行（対応する行13が空行）

        # ROE分析指標: Q14-Q20 (ROE、ROS、TOR、LRV、検算1、検算2、ROA)
        for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                          check1_row_num, check2_row_num, roa_row_num]:
            # check2_row_num（検算2）の場合は空欄
            if source_row == check2_row_num:
                continue
            ratio_formula = f"={latest_col_letter}{source_row}/{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{source_row}'] = ratio_formula
            analysis_ws[f'{growth_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

        # Q23: "　対前年増加率" ヘッダー（Q1を参照）
        analysis_ws[f'{growth_col_letter}{yoy_header_row_num}'] = f'={growth_col_letter}1'

        # Q24-Q41: 対前年増加率セクションの年平均増加率（Q3-Q20を移動）
        # CAGR計算: =(最新値/基準値)^(1/(YEAR(最新)-YEAR(基準)))-1
        # Q24-Q29: 基本指標
        for idx, row_num in enumerate(yoy_rows_basic):
            source_cagr_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                             total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
            cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                          f"^(1/(YEAR({latest_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
            analysis_ws[f'{growth_col_letter}{row_num}'] = cagr_formula

        # Q30: 空行（対応する行30が空行）
        # （analysis_wsの行30は空行なので何もしない）

        # Q31-Q33: 計算指標
        for idx, row_num in enumerate(yoy_rows_calc):
            source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
            cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                          f"^(1/(YEAR({latest_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
            analysis_ws[f'{growth_col_letter}{row_num}'] = cagr_formula

        # Q34: 空行（対応する行34が空行）

        # Q35-Q41: ROE分析指標（check2_row_numは除外）
        source_rows_roe = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                          check1_row_num, check2_row_num, roa_row_num]
        for idx, row_num in enumerate(yoy_rows_roe):
            source_cagr_row = source_rows_roe[idx]
            # check2_row_num（検算2）の場合は空欄
            if source_cagr_row == check2_row_num:
                continue
            cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                          f"^(1/(YEAR({latest_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
            analysis_ws[f'{growth_col_letter}{row_num}'] = cagr_formula

        # Q列の列幅を12に設定
        analysis_ws.column_dimensions[growth_col_letter].width = 12

        # Q24-Q41: 対前年増加率セクションの年平均増加率の表示形式を設定（パーセント）
        for row_num in yoy_rows_basic + yoy_rows_calc + yoy_rows_roe:
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = number_format_percent

        # Q列: 対前年差セクションの追加 (latest - base)
        analysis_ws[f'{growth_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({latest_col_letter}1) & \"-\" & YEAR({base_col_letter}1)"

        for idx, row_num in enumerate(diff_rows_basic):
            source_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                         total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
            diff_formula = f"={latest_col_letter}{source_row}-{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{row_num}'] = diff_formula
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

        for idx, row_num in enumerate(diff_rows_calc):
            source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
            diff_formula = f"={latest_col_letter}{source_row}-{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{row_num}'] = diff_formula
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

        for idx, row_num in enumerate(diff_rows_roe):
            source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                         check1_row_num, check2_row_num, roa_row_num][idx]
            if source_row == check2_row_num:
                continue
            diff_formula = f"={latest_col_letter}{source_row}-{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{row_num}'] = diff_formula
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

        # ============================================================================
        # R列とS列: 5年間の比較
        # ============================================================================
        # kikanに基づいて5年間の比較を追加
        # - kikan >= 10: R列=前半5年(base to mid), S列=後半5年(mid to latest)
        # - 5 <= kikan < 10: R列=最新5年(latest-5 to latest)のみ
        # - kikan < 5: R列S列なし

        if kikan >= 5:
            if kikan >= 10:
                # Case 1: kikan >= 10
                # R列: 前半5年間 (base to mid)
                # S列: 後半5年間 (mid to latest)
                mid_col = latest_col - 5  # 中間地点（5年前）
                mid_col_letter = openpyxl.utils.get_column_letter(mid_col)

                # R列 (num_cols + 2)
                r_col = num_cols + 2
                r_col_letter = openpyxl.utils.get_column_letter(r_col)

                # R1: 期間表示 (mid/base)
                r_period_formula = f"=YEAR({mid_col_letter}1) & \"/\" & YEAR({base_col_letter}1)"
                analysis_ws[f'{r_col_letter}1'] = r_period_formula

                # R2: 空
                analysis_ws[f'{r_col_letter}2'] = ''

                # R3-R20: 比率計算
                for source_row in [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                  total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
                    ratio_formula = f"={mid_col_letter}{source_row}/{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
                    ratio_formula = f"={mid_col_letter}{source_row}/{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                  check1_row_num, check2_row_num, roa_row_num]:
                    if source_row == check2_row_num:
                        continue
                    ratio_formula = f"={mid_col_letter}{source_row}/{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                # R23: ヘッダー（R1を参照）
                analysis_ws[f'{r_col_letter}{yoy_header_row_num}'] = f'={r_col_letter}1'

                # R列: 対前年差セクションの追加 (mid - base)
                analysis_ws[f'{r_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({mid_col_letter}1) & \"-\" & YEAR({base_col_letter}1)"

                for idx, row_num in enumerate(diff_rows_basic):
                    source_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                 total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    diff_formula = f"={mid_col_letter}{source_row}-{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_calc):
                    source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    diff_formula = f"={mid_col_letter}{source_row}-{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_roe):
                    source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                 check1_row_num, check2_row_num, roa_row_num][idx]
                    if source_row == check2_row_num:
                        continue
                    diff_formula = f"={mid_col_letter}{source_row}-{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                # R24-R41: CAGR計算 (base to mid)
                for idx, row_num in enumerate(yoy_rows_basic):
                    source_cagr_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                     total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    cagr_formula = (f"=({mid_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({mid_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_calc):
                    source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    cagr_formula = (f"=({mid_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({mid_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_roe):
                    source_cagr_row = source_rows_roe[idx]
                    if source_cagr_row == check2_row_num:
                        continue
                    cagr_formula = (f"=({mid_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({mid_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                # R列の列幅を12に設定
                analysis_ws.column_dimensions[r_col_letter].width = 12

                # S列 (num_cols + 3)
                s_col = num_cols + 3
                s_col_letter = openpyxl.utils.get_column_letter(s_col)

                # S1: 期間表示 (latest/mid)
                s_period_formula = f"=YEAR({latest_col_letter}1) & \"/\" & YEAR({mid_col_letter}1)"
                analysis_ws[f'{s_col_letter}1'] = s_period_formula

                # S2: 空
                analysis_ws[f'{s_col_letter}2'] = ''

                # S3-S20: 比率計算
                for source_row in [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                  total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{s_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{s_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                  check1_row_num, check2_row_num, roa_row_num]:
                    if source_row == check2_row_num:
                        continue
                    ratio_formula = f"={latest_col_letter}{source_row}/{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{s_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                # S23: ヘッダー（S1を参照）
                analysis_ws[f'{s_col_letter}{yoy_header_row_num}'] = f'={s_col_letter}1'

                # S列: 対前年差セクションの追加 (latest - mid)
                analysis_ws[f'{s_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({latest_col_letter}1) & \"-\" & YEAR({mid_col_letter}1)"

                for idx, row_num in enumerate(diff_rows_basic):
                    source_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                 total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_calc):
                    source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_roe):
                    source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                 check1_row_num, check2_row_num, roa_row_num][idx]
                    if source_row == check2_row_num:
                        continue
                    diff_formula = f"={latest_col_letter}{source_row}-{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)


                # S24-S41: CAGR計算 (mid to latest)
                # Note: S列のCAGRは (latest/mid)^(1/(latest_year - mid_year)) - 1
                for idx, row_num in enumerate(yoy_rows_basic):
                    source_cagr_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                     total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{mid_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({mid_col_letter}$1)))-1")
                    analysis_ws[f'{s_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_calc):
                    source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{mid_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({mid_col_letter}$1)))-1")
                    analysis_ws[f'{s_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_roe):
                    source_cagr_row = source_rows_roe[idx]
                    if source_cagr_row == check2_row_num:
                        continue
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{mid_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({mid_col_letter}$1)))-1")
                    analysis_ws[f'{s_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = number_format_percent

                # S列の列幅を12に設定
                analysis_ws.column_dimensions[s_col_letter].width = 12

                debug_log(f"Added R and S columns for 10+ year data: R={mid_col_letter}/{base_col_letter}, S={latest_col_letter}/{mid_col_letter}")

            else:
                # Case 2: 5 <= kikan < 10
                # R列のみ: 最新5年間 (latest-5 to latest)
                five_years_ago_col = latest_col - 5
                five_years_col_letter = openpyxl.utils.get_column_letter(five_years_ago_col)

                # R列 (num_cols + 2)
                r_col = num_cols + 2
                r_col_letter = openpyxl.utils.get_column_letter(r_col)

                # R1: 期間表示 (latest/5years_ago)
                r_period_formula = f"=YEAR({latest_col_letter}1) & \"/\" & YEAR({five_years_col_letter}1)"
                analysis_ws[f'{r_col_letter}1'] = r_period_formula

                # R2: 空
                analysis_ws[f'{r_col_letter}2'] = ''

                # R3-R20: 比率計算
                for source_row in [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                  total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                  check1_row_num, check2_row_num, roa_row_num]:
                    if source_row == check2_row_num:
                        continue
                    ratio_formula = f"={latest_col_letter}{source_row}/{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                # R23: ヘッダー（R1を参照）
                analysis_ws[f'{r_col_letter}{yoy_header_row_num}'] = f'={r_col_letter}1'

                # R列: 対前年差セクションの追加 (latest - 5years_ago)
                analysis_ws[f'{r_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({latest_col_letter}1) & \"-\" & YEAR({five_years_col_letter}1)"

                for idx, row_num in enumerate(diff_rows_basic):
                    source_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                 total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_calc):
                    source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_roe):
                    source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                 check1_row_num, check2_row_num, roa_row_num][idx]
                    if source_row == check2_row_num:
                        continue
                    diff_formula = f"={latest_col_letter}{source_row}-{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                # R24-R41: CAGR計算 (5years_ago to latest)
                for idx, row_num in enumerate(yoy_rows_basic):
                    source_cagr_row = [sales_analysis_row, profit_analysis_row, equity_or_net_assets_row,
                                     total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{five_years_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({five_years_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_calc):
                    source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{five_years_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({five_years_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_roe):
                    source_cagr_row = source_rows_roe[idx]
                    if source_cagr_row == check2_row_num:
                        continue
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{five_years_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({five_years_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                # R列の列幅を12に設定
                analysis_ws.column_dimensions[r_col_letter].width = 12

                debug_log(f"Added R column for 5-9 year data: R={latest_col_letter}/{five_years_col_letter}")

    debug_log(f"ROE analysis sheet created: {analysis_sheet_name}")



def create_roe_analysis_sheet_non_consolidated(workbook, source_sheet_name, debug_log=None):
    """
    主要な経営指標等の推移（単体）シートからROE分析シートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "主要な経営指標等の推移（単体）"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]

    # 正規表現で「最初のカッコ内」と「それ以降」を抽出
    # \1: 最初のカッコの中身（連結）
    # \2: それ以降の文字列（(日本基準)）
    match = re.search(r'[（\(](.*?)[）\)](.*)', source_sheet_name)

    if match:
        # 単体の形にする
        short_name = f"{match.group(1)}"
    else:
        # カッコが含まれない場合のフォールバック
        short_name = source_sheet_name

    analysis_sheet_name = f"{short_name}_ROE分析"

    # シート名31文字制限対策（念のため）
    analysis_sheet_name = analysis_sheet_name[:31]


    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 列数を取得
    source_cols = source_ws.max_column
    num_cols = source_cols

    # 参照する行番号を特定（元シートから）
    row_mapping = {}  # 英語名 -> 行番号のマッピング
    for row in range(2, source_ws.max_row + 1):
        english_name = source_ws.cell(row, 2).value  # B列: 項目（英名）
        if english_name:
            row_mapping[english_name] = row
            
    debug_log(f"ROE Standalone: Mapped {len(row_mapping)} element names from {source_ws.max_row} rows")
    debug_log(f"ROE Standalone: row_mapping keys: {list(row_mapping.keys())[:10]}")

    def find_all_rows_by_keywords(keywords_list, item_name):
        """
        複数のキーワード候補をすべて検索し、行番号のリストを返す（部分一致）
        """
        row_nums = []
        for keywords in keywords_list:
            if isinstance(keywords, str):
                keywords = [keywords]
            found_for_these_kws = False
            for eng_name, row_num in row_mapping.items():
                if all(kw in str(eng_name) for kw in keywords):
                    if row_num not in row_nums:
                        row_nums.append(row_num)
                        found_for_these_kws = True
            
            if found_for_these_kws:
                debug_log(f"  [Matching] Keyword set {keywords} found rows: {row_nums}")
        
        if row_nums:
            debug_log(f"{item_name}: overall found {len(row_nums)} rows: {row_nums}")
        else:
            debug_log(f"{item_name}: NOT FOUND among {len(row_mapping)} mapped elements")
        return row_nums

    def find_row_by_keywords(keywords_list, item_name):
        res = find_all_rows_by_keywords(keywords_list, item_name)
        return res[0] if res else None

    # 単体用の勘定科目のキーワード候補
    # 売上高と売上収益を混合して定義
    sales_keywords = [
        ['Revenue', 'KeyFinancialData'],
        ['NetSales', 'KeyFinancialData'],
        ['Revenue', 'Summary'],
        ['NetSales', 'Summary'],
        ['Revenue'],
        ['NetSales'],
        ['Sales', 'Summary'],
    ]

    profit_keywords = [
        ['NetIncome', 'KeyFinancialData'],
        ['NetIncome', 'Summary'],
        ['ProfitLoss', 'OwnersOfParent', 'Summary'],
        ['ProfitLoss', 'Summary'],
        ['NetIncomeLoss', 'Summary'],
        ['NetIncomeLoss'],
    ]

    net_assets_keywords = [
        ['NetAssets', 'KeyFinancialData'],
        ['NetAssets', 'Summary'],
        ['NetAssets'],
    ]

    total_assets_keywords = [
        ['TotalAssets', 'KeyFinancialData'],
        ['TotalAssets', 'Summary'],
        ['TotalAssets'],
    ]

    equity_ratio_keywords = [
        ['EquityToAssetRatio', 'KeyFinancialData'],
        ['EquityToAssetRatio', 'Summary'],
        ['EquityRatio'],
    ]

    roe_keywords = [
        ['RateOfReturnOnEquity', 'KeyFinancialData'],
        ['ReturnOnEquity', 'Summary'],
        ['ROE'],
    ]

    # 元シートの行番号を取得
    sales_rows = find_all_rows_by_keywords(sales_keywords, '売上高/売上収益')
    
    # 優先順位付け: IFRS(Revenue)を優先する (年度の新しいほうを採用)
    def sales_priority_stand(row_num):
        b_val = str(source_ws.cell(row_num, 2).value or "")
        if 'Revenue' in b_val or 'IFRS' in b_val:
            return 0
        return 1
    
    if sales_rows:
        sales_rows.sort(key=sales_priority_stand)
        debug_log(f"ROE Standalone Sales rows (prioritized): {sales_rows}")
        
    sales_row = sales_rows[0] if sales_rows else None
    profit_row = find_row_by_keywords(profit_keywords, '当期純利益')
    net_assets_row = find_row_by_keywords(net_assets_keywords, '純資産額')
    total_assets_row = find_row_by_keywords(total_assets_keywords, '総資産額')
    equity_ratio_row = find_row_by_keywords(equity_ratio_keywords, '自己資本比率')
    roe_row = find_row_by_keywords(roe_keywords, 'ROE')

    # 売上高が見つからない場合は、セクションヘッダーの次の行を使用
    if sales_row is None:
        # ヘッダーを部分一致で検索
        header_keywords = [
            ['BusinessResults', 'ReportingCompany', 'Heading'],  # 単体用ヘッダー
            ['BusinessResults', 'Heading'],
            ['BusinessResults'],
        ]
        header_row_num = find_row_by_keywords(header_keywords, 'セクションヘッダー')
        if header_row_num:
            # ヘッダーの次の行を売上高として使用
            sales_row = header_row_num + 1
            debug_log(f"Sales row not found, using first item after header (row {sales_row})")

    # ROE分析に必要な項目がすべて存在するかチェック
    required_items = {
        '売上高': sales_row,
        '当期純利益': profit_row,
        '総資産額': total_assets_row,
        '自己資本比率': equity_ratio_row,
        'ROE': roe_row
    }
    missing_items = [name for name, row in required_items.items() if row is None]
    if missing_items:
        debug_log(f"ROE analysis skipped for '{source_sheet_name}': missing items: {', '.join(missing_items)}")
        return

    # ここから先は連結と全く同じロジックなので、create_roe_analysis_sheetと共通化
    # 現在は冗長ですが、明確性のため一旦全コピー
    def add_reference_row_full(source_row_num):
        """元シートの行全体を参照する行を追加"""
        row_data = []
        for col in range(1, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            formula = f"='{source_sheet_name}'!{col_letter}{source_row_num}"
            row_data.append(formula)
        analysis_ws.append(row_data)

    # 1行目: ヘッダー
    add_reference_row_full(1)

    # 2行目: セクションヘッダー
    header_keywords = [
        ['BusinessResults', 'ReportingCompany', 'Heading'],
        ['BusinessResults', 'Heading'],
        ['BusinessResults'],
    ]
    header_row_num = find_row_by_keywords(header_keywords, 'セクションヘッダー')
    if header_row_num:
        row_data = []
        for col in range(1, 3):
            col_letter = openpyxl.utils.get_column_letter(col)
            formula = f"='{source_sheet_name}'!{col_letter}{header_row_num}"
            row_data.append(formula)
        for col in range(3, num_cols + 1):
            row_data.append('')
        analysis_ws.append(row_data)

    # 3行目: 売上高
    if len(sales_rows) > 1:
        # 複数行ある場合は優先順位に従って「年度の新しいほう（IFRS等）」を採用する (Deduplication)
        # IF(収益<>"", 収益, 売上高) の形式で数式を作成
        row_data = ['売上高', ' / '.join([f"'{source_sheet_name}'!B{r}" for r in sales_rows])]
        for col in range(3, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            # 優先順位に従った数式 (収益<>"", 収益, 売上高)
            formula = f"'{source_sheet_name}'!{col_letter}{sales_rows[-1]}"
            for r in reversed(sales_rows[:-1]):
                curr_cell = f"'{source_sheet_name}'!{col_letter}{r}"
                formula = f"IF({curr_cell}<>\"\",{curr_cell},{formula})"
            row_data.append(f"={formula}")
        analysis_ws.append(row_data)
    else:
        add_reference_row_full(sales_row)
    add_reference_row_full(profit_row)
    add_reference_row_full(net_assets_row)
    add_reference_row_full(total_assets_row)
    add_reference_row_full(equity_ratio_row)
    add_reference_row_full(roe_row)

    current_row = analysis_ws.max_row
    sales_analysis_row = current_row - 5
    profit_analysis_row = current_row - 4
    net_assets_analysis_row = current_row - 3
    total_assets_analysis_row = current_row - 2
    equity_ratio_analysis_row = current_row - 1
    roe_analysis_row = current_row

    # 空行
    analysis_ws.append([''] * num_cols)

    # 自己資本 = 総資産額 × 自己資本比率
    equity_calc_row_num = analysis_ws.max_row + 1
    equity_row = ['　　　自己資本', '']
    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{total_assets_analysis_row}*{col_letter}{equity_ratio_analysis_row}"
        equity_row.append(formula)
    analysis_ws.append(equity_row)

    # 自己資本（平均）
    equity_avg_row_num = analysis_ws.max_row + 1
    equity_avg_row = ['　　　自己資本（平均）', '']
    equity_avg_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
        formula = f"=AVERAGE({prev_col_letter}{equity_calc_row_num}:{col_letter}{equity_calc_row_num})"
        equity_avg_row.append(formula)
    analysis_ws.append(equity_avg_row)

    # 総資産（平均）
    total_assets_avg_row_num = analysis_ws.max_row + 1
    total_assets_avg_row = ['　　　総資産（平均）', '']
    total_assets_avg_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
        formula = f"=AVERAGE({prev_col_letter}{total_assets_analysis_row}:{col_letter}{total_assets_analysis_row})"
        total_assets_avg_row.append(formula)
    analysis_ws.append(total_assets_avg_row)

    # 空行
    analysis_ws.append([''] * num_cols)

    # ROE分析指標
    roe_calc_row_num = analysis_ws.max_row + 1
    roe_calc_row = ['　　　自己資本利益率(ROE)', '']
    roe_calc_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{roe_analysis_row}"
        roe_calc_row.append(formula)
    analysis_ws.append(roe_calc_row)

    # ROS = 当期純利益 / 売上高
    ros_row_num = analysis_ws.max_row + 1
    ros_row = ['　　　売上高利益率(ROS)', '']
    ros_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{profit_analysis_row}/{col_letter}{sales_analysis_row}"
        ros_row.append(formula)
    analysis_ws.append(ros_row)

    # TOR = 売上高 / 総資産（平均）
    tor_row_num = analysis_ws.max_row + 1
    tor_row = ['　　　総資産回転率(TOR)', '']
    tor_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{sales_analysis_row}/{col_letter}{total_assets_avg_row_num}"
        tor_row.append(formula)
    analysis_ws.append(tor_row)

    # LEV = 総資産（平均） / 自己資本（平均）
    lrv_row_num = analysis_ws.max_row + 1
    lrv_row = ['　　　レバレッジ(LEV)', '']
    lrv_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{total_assets_avg_row_num}/{col_letter}{equity_avg_row_num}"
        lrv_row.append(formula)
    analysis_ws.append(lrv_row)

    # 検算1: ROS * TOR * LEV = ROE
    check1_row_num = analysis_ws.max_row + 1
    check1_row = ['　　　検算1(ROS*TOR*LEV=ROE)', '']
    check1_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"=PRODUCT({col_letter}{ros_row_num}:{col_letter}{lrv_row_num})"
        check1_row.append(formula)
    analysis_ws.append(check1_row)

    # 検算2: 検算1 = ROE
    check2_row_num = analysis_ws.max_row + 1
    check2_row = ['　　　検算2(検算1=ROE)', '']
    check2_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        #formula = f"=ROUND({col_letter}{roe_calc_row_num},3)=ROUND({col_letter}{check1_row_num},3)"
        formula = f"=ABS(ROUND({col_letter}{roe_calc_row_num},3)-ROUND({col_letter}{check1_row_num},3))<=0.001"
        check2_row.append(formula)
    analysis_ws.append(check2_row)

    # ROA = 当期純利益 / 総資産（平均）
    roa_row_num = analysis_ws.max_row + 1
    roa_row = ['　　　ROA(総資産利益率)', '']
    roa_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"={col_letter}{profit_analysis_row}/{col_letter}{total_assets_avg_row_num}"
        roa_row.append(formula)
    analysis_ws.append(roa_row)

    # 表示形式設定
    number_format_integer = r'#,##0_ ;[Red]\-#,##0\ '
    number_format_decimal = r'#,##0_);[Red](#,##0)'
    number_format_decimal2 = r'#,##0.00_);[Red](#,##0.00)'
    number_format_percent = r'0.0%'

    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row_num in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row, total_assets_analysis_row]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_integer
        for row_num in [equity_ratio_analysis_row, roe_analysis_row]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_percent

    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row_num in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_decimal

    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row_num in [roe_calc_row_num, ros_row_num, roa_row_num]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_percent
        for row_num in [tor_row_num, lrv_row_num]:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_decimal2
        analysis_ws[f'{col_letter}{check1_row_num}'].number_format = number_format_percent
        analysis_ws[f'{col_letter}{check2_row_num}'].number_format = number_format_percent

    # 列幅設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True
    for col in range(3, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠固定
    analysis_ws.freeze_panes = 'B2'

    # 対前年増加率セクション
    analysis_ws.append([''] * num_cols)
    analysis_ws.append([''] * num_cols)

    yoy_header_row_num = analysis_ws.max_row + 1
    yoy_header_row = ['　対前年増加率', '']
    yoy_header_row.append('')
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        yoy_header_row.append(f'={col_letter}1')
    analysis_ws.append(yoy_header_row)

    yoy_rows_basic = []
    for source_row in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                       total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
        yoy_row_num = analysis_ws.max_row + 1
        yoy_rows_basic.append(yoy_row_num)
        yoy_row = [f'=A{source_row}', '']
        yoy_row.append('')
        yoy_row.append('')
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}/{prev_col_letter}{source_row}-1"
            yoy_row.append(formula)
        analysis_ws.append(yoy_row)

    analysis_ws.append([''] * num_cols)

    yoy_rows_calc = []
    for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
        yoy_row_num = analysis_ws.max_row + 1
        yoy_rows_calc.append(yoy_row_num)
        yoy_row = [f'=A{source_row}', '']
        yoy_row.append('')
        yoy_row.append('')
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}/{prev_col_letter}{source_row}-1"
            yoy_row.append(formula)
        analysis_ws.append(yoy_row)

    analysis_ws.append([''] * num_cols)

    yoy_rows_roe = []
    for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                       check1_row_num, check2_row_num, roa_row_num]:
        yoy_row_num = analysis_ws.max_row + 1
        yoy_rows_roe.append(yoy_row_num)
        yoy_row = [f'=A{source_row}', '']
        yoy_row.append('')
        if source_row == check2_row_num:
            for col in range(4, num_cols + 1):
                yoy_row.append('')
        else:
            yoy_row.append('')
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
                formula = f"={col_letter}{source_row}/{prev_col_letter}{source_row}-1"
                yoy_row.append(formula)
        analysis_ws.append(yoy_row)

    for col in range(5, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row_num in yoy_rows_basic + yoy_rows_calc + yoy_rows_roe:
            analysis_ws[f'{col_letter}{row_num}'].number_format = number_format_percent

    # ============================================================================
    # 対前年差セクション
    # ============================================================================
    analysis_ws.append([''] * num_cols)
    analysis_ws.append([''] * num_cols)

    diff_yoy_header_row_num = analysis_ws.max_row + 1
    diff_yoy_header_row = ['　対前年差', '']
    diff_yoy_header_row.append('')  # C列
    for col in range(4, num_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        diff_yoy_header_row.append(f'={col_letter}1')
    analysis_ws.append(diff_yoy_header_row)

    # 行ごとの書式設定マッピング
    row_formats_map = {
        sales_analysis_row: number_format_integer,
        profit_analysis_row: number_format_integer,
        net_assets_analysis_row: number_format_integer,
        total_assets_analysis_row: number_format_integer,
        equity_ratio_analysis_row: number_format_percent,
        roe_analysis_row: number_format_percent,
        equity_calc_row_num: number_format_decimal,
        equity_avg_row_num: number_format_decimal,
        total_assets_avg_row_num: number_format_decimal,
        roe_calc_row_num: number_format_percent,
        ros_row_num: number_format_percent,
        tor_row_num: number_format_decimal2,
        lrv_row_num: number_format_decimal2,
        check1_row_num: number_format_percent,
        check2_row_num: number_format_percent,
        roa_row_num: number_format_percent,
    }

    # 各項目を追加
    diff_rows_basic = []
    for source_row in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                       total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
        diff_row_num = analysis_ws.max_row + 1
        diff_rows_basic.append(diff_row_num)
        diff_row = [f'=A{source_row}', '']
        diff_row.append('')  # C列
        diff_row.append('')  # D列は空
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}-{prev_col_letter}{source_row}"
            diff_row.append(formula)
        analysis_ws.append(diff_row)
        fmt = row_formats_map.get(source_row)
        if fmt:
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                analysis_ws[f'{col_letter}{diff_row_num}'].number_format = fmt

    # 空行
    analysis_ws.append([''] * num_cols)

    diff_rows_calc = []
    for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
        diff_row_num = analysis_ws.max_row + 1
        diff_rows_calc.append(diff_row_num)
        diff_row = [f'=A{source_row}', '']
        diff_row.append('')  # C列
        diff_row.append('')  # D列は空
        for col in range(5, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
            formula = f"={col_letter}{source_row}-{prev_col_letter}{source_row}"
            diff_row.append(formula)
        analysis_ws.append(diff_row)
        fmt = row_formats_map.get(source_row)
        if fmt:
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                analysis_ws[f'{col_letter}{diff_row_num}'].number_format = fmt

    # 空行
    analysis_ws.append([''] * num_cols)

    diff_rows_roe = []
    for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                       check1_row_num, check2_row_num, roa_row_num]:
        diff_row_num = analysis_ws.max_row + 1
        diff_rows_roe.append(diff_row_num)
        diff_row = [f'=A{source_row}', '']
        diff_row.append('')  # C列
        if source_row == check2_row_num:
            for col in range(4, num_cols + 1):
                diff_row.append('')
        else:
            diff_row.append('')  # D列
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                prev_col_letter = openpyxl.utils.get_column_letter(col - 1)
                formula = f"={col_letter}{source_row}-{prev_col_letter}{source_row}"
                diff_row.append(formula)
        analysis_ws.append(diff_row)
        fmt = row_formats_map.get(source_row)
        if fmt:
            for col in range(5, num_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                analysis_ws[f'{col_letter}{diff_row_num}'].number_format = fmt


    # 10年前からの増加率計算（Q列以降）
    latest_col = source_cols
    latest_col_letter = openpyxl.utils.get_column_letter(latest_col)

    target_rows = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                   total_assets_analysis_row, equity_ratio_analysis_row]

    def has_all_data(col_num):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        for row_num in target_rows:
            cell_formula = analysis_ws[f'{col_letter}{row_num}'].value
            if not cell_formula:
                return False
            if isinstance(cell_formula, str) and cell_formula.startswith('='):
                try:
                    parts = cell_formula.split('!')
                    if len(parts) == 2:
                        source_cell_ref = parts[1].strip()
                        source_cell_value = source_ws[source_cell_ref].value
                        if source_cell_value is None or source_cell_value == '':
                            return False
                except Exception:
                    return False
        return True

    oldest_col = None
    for col in range(3, latest_col):
        if has_all_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        base_col = None
        kikan = 0
        num_cols = source_cols
    else:
        fyb_col = oldest_col + 1
        kikan = latest_col - fyb_col

        if kikan >= 10:
            num_cols = max(source_cols, 16)
        else:
            num_cols = source_cols

        if kikan >= 10:
            candidate_base_col = latest_col - 10
            if candidate_base_col > 3:
                try:
                    candidate_letter = openpyxl.utils.get_column_letter(candidate_base_col)
                    prev_letter = openpyxl.utils.get_column_letter(candidate_base_col - 1)
                    if has_all_data(candidate_base_col - 1):
                        candidate_header = analysis_ws[f'{candidate_letter}1'].value
                        if isinstance(candidate_header, str) and candidate_header.startswith('='):
                            parts = candidate_header.split('!')
                            if len(parts) == 2:
                                source_cell_ref = parts[1].strip().replace("'", "")
                                candidate_date = source_ws[source_cell_ref].value
                                prev_date = source_ws[f'{prev_letter[0] if len(prev_letter)==1 else prev_letter}1'].value
                                if candidate_date and prev_date:
                                    if hasattr(candidate_date, 'year') and hasattr(prev_date, 'year'):
                                        if candidate_date.year == prev_date.year:
                                            base_col = candidate_base_col - 1
                                            debug_log(f"Fiscal year change detected: using {prev_letter} instead of {candidate_letter} (same year {candidate_date.year})")
                                        else:
                                            base_col = candidate_base_col
                                    else:
                                        base_col = candidate_base_col
                                else:
                                    base_col = candidate_base_col
                        else:
                            base_col = candidate_base_col
                    else:
                        base_col = candidate_base_col
                except Exception:
                    base_col = candidate_base_col
            else:
                base_col = candidate_base_col
            base_col_letter = openpyxl.utils.get_column_letter(base_col)
            debug_log(f"Using 10-year comparison: kikan={kikan}, base column {base_col_letter}")
        else:
            base_col = fyb_col
            base_col_letter = openpyxl.utils.get_column_letter(base_col)
            oldest_col_letter = openpyxl.utils.get_column_letter(oldest_col)
            debug_log(f"Using longest available period: kikan={kikan}, oldest={oldest_col_letter}, base={base_col_letter}")

    if base_col is not None:
        growth_col = num_cols + 1
        growth_col_letter = openpyxl.utils.get_column_letter(growth_col)
        base_col_letter = openpyxl.utils.get_column_letter(base_col)

        period_formula = f"=YEAR({latest_col_letter}1) & \"/\" & YEAR({base_col_letter}1)"
        analysis_ws[f'{growth_col_letter}1'] = period_formula
        analysis_ws[f'{growth_col_letter}2'] = ''

        for source_row in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                          total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
            ratio_formula = f"={latest_col_letter}{source_row}/{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{source_row}'] = ratio_formula
            analysis_ws[f'{growth_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

        for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
            ratio_formula = f"={latest_col_letter}{source_row}/{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{source_row}'] = ratio_formula
            analysis_ws[f'{growth_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

        for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                          check1_row_num, check2_row_num, roa_row_num]:
            if source_row == check2_row_num:
                continue
            ratio_formula = f"={latest_col_letter}{source_row}/{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{source_row}'] = ratio_formula
            analysis_ws[f'{growth_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

        analysis_ws[f'{growth_col_letter}{yoy_header_row_num}'] = f'={growth_col_letter}1'

        for idx, row_num in enumerate(yoy_rows_basic):
            source_cagr_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                             total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
            cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                          f"^(1/(YEAR({latest_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
            analysis_ws[f'{growth_col_letter}{row_num}'] = cagr_formula

        for idx, row_num in enumerate(yoy_rows_calc):
            source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
            cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                          f"^(1/(YEAR({latest_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
            analysis_ws[f'{growth_col_letter}{row_num}'] = cagr_formula

        source_rows_roe = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                          check1_row_num, check2_row_num, roa_row_num]
        for idx, row_num in enumerate(yoy_rows_roe):
            source_cagr_row = source_rows_roe[idx]
            if source_cagr_row == check2_row_num:
                continue
            cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                          f"^(1/(YEAR({latest_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
            analysis_ws[f'{growth_col_letter}{row_num}'] = cagr_formula

        analysis_ws.column_dimensions[growth_col_letter].width = 12

        for row_num in yoy_rows_basic + yoy_rows_calc + yoy_rows_roe:
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = number_format_percent

        # Q列: 対前年差セクションの追加 (latest - base)
        analysis_ws[f'{growth_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({latest_col_letter}1) & \"-\" & YEAR({base_col_letter}1)"

        for idx, row_num in enumerate(diff_rows_basic):
            source_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                         total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
            diff_formula = f"={latest_col_letter}{source_row}-{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{row_num}'] = diff_formula
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

        for idx, row_num in enumerate(diff_rows_calc):
            source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
            diff_formula = f"={latest_col_letter}{source_row}-{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{row_num}'] = diff_formula
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

        for idx, row_num in enumerate(diff_rows_roe):
            source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                         check1_row_num, check2_row_num, roa_row_num][idx]
            if source_row == check2_row_num:
                continue
            diff_formula = f"={latest_col_letter}{source_row}-{base_col_letter}{source_row}"
            analysis_ws[f'{growth_col_letter}{row_num}'] = diff_formula
            analysis_ws[f'{growth_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

        # R列とS列: 5年間の比較
        if kikan >= 5:
            if kikan >= 10:
                mid_col = latest_col - 5
                mid_col_letter = openpyxl.utils.get_column_letter(mid_col)
                r_col = num_cols + 2
                r_col_letter = openpyxl.utils.get_column_letter(r_col)

                r_period_formula = f"=YEAR({mid_col_letter}1) & \"/\" & YEAR({base_col_letter}1)"
                analysis_ws[f'{r_col_letter}1'] = r_period_formula
                analysis_ws[f'{r_col_letter}2'] = ''

                for source_row in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                  total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
                    ratio_formula = f"={mid_col_letter}{source_row}/{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
                    ratio_formula = f"={mid_col_letter}{source_row}/{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                # R23: ヘッダー（R1を参照）
                analysis_ws[f'{r_col_letter}{yoy_header_row_num}'] = f'={r_col_letter}1'

                # R列: 対前年差セクションの追加 (mid - base)
                analysis_ws[f'{r_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({mid_col_letter}1) & \"-\" & YEAR({base_col_letter}1)"

                for idx, row_num in enumerate(diff_rows_basic):
                    source_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                 total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    diff_formula = f"={mid_col_letter}{source_row}-{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_calc):
                    source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    diff_formula = f"={mid_col_letter}{source_row}-{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_roe):
                    source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                 check1_row_num, check2_row_num, roa_row_num][idx]
                    if source_row == check2_row_num:
                        continue
                    diff_formula = f"={mid_col_letter}{source_row}-{base_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(yoy_rows_basic):
                    source_cagr_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                     total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    cagr_formula = (f"=({mid_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({mid_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_calc):
                    source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    cagr_formula = (f"=({mid_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({mid_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_roe):
                    source_cagr_row = source_rows_roe[idx]
                    if source_cagr_row == check2_row_num:
                        continue
                    cagr_formula = (f"=({mid_col_letter}{source_cagr_row}/{base_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({mid_col_letter}$1)-YEAR({base_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                analysis_ws.column_dimensions[r_col_letter].width = 12

                s_col = num_cols + 3
                s_col_letter = openpyxl.utils.get_column_letter(s_col)

                s_period_formula = f"=YEAR({latest_col_letter}1) & \"/\" & YEAR({mid_col_letter}1)"
                analysis_ws[f'{s_col_letter}1'] = s_period_formula
                analysis_ws[f'{s_col_letter}2'] = ''

                for source_row in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                  total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{s_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{s_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                  check1_row_num, check2_row_num, roa_row_num]:
                    if source_row == check2_row_num:
                        continue
                    ratio_formula = f"={latest_col_letter}{source_row}/{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{s_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'
                # S23: ヘッダー（S1を参照）
                analysis_ws[f'{s_col_letter}{yoy_header_row_num}'] = f'={s_col_letter}1'

                # S列: 対前年差セクションの追加 (latest - mid)
                analysis_ws[f'{s_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({latest_col_letter}1) & \"-\" & YEAR({mid_col_letter}1)"

                for idx, row_num in enumerate(diff_rows_basic):
                    source_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                 total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_calc):
                    source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_roe):
                    source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                 check1_row_num, check2_row_num, roa_row_num][idx]
                    if source_row == check2_row_num:
                        continue
                    diff_formula = f"={latest_col_letter}{source_row}-{mid_col_letter}{source_row}"
                    analysis_ws[f'{s_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(yoy_rows_basic):
                    source_cagr_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                     total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{mid_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({mid_col_letter}$1)))-1")
                    analysis_ws[f'{s_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_calc):
                    source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{mid_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({mid_col_letter}$1)))-1")
                    analysis_ws[f'{s_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_roe):
                    source_cagr_row = source_rows_roe[idx]
                    if source_cagr_row == check2_row_num:
                        continue
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{mid_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({mid_col_letter}$1)))-1")
                    analysis_ws[f'{s_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{s_col_letter}{row_num}'].number_format = number_format_percent

                analysis_ws.column_dimensions[s_col_letter].width = 12

                debug_log(f"Added R and S columns for 10+ year data: R={mid_col_letter}/{base_col_letter}, S={latest_col_letter}/{mid_col_letter}")

            else:
                five_years_ago_col = latest_col - 5
                five_years_col_letter = openpyxl.utils.get_column_letter(five_years_ago_col)
                r_col = num_cols + 2
                r_col_letter = openpyxl.utils.get_column_letter(r_col)

                r_period_formula = f"=YEAR({latest_col_letter}1) & \"/\" & YEAR({five_years_col_letter}1)"
                analysis_ws[f'{r_col_letter}1'] = r_period_formula
                analysis_ws[f'{r_col_letter}2'] = ''

                for source_row in [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                  total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num]:
                    ratio_formula = f"={latest_col_letter}{source_row}/{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                for source_row in [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                  check1_row_num, check2_row_num, roa_row_num]:
                    if source_row == check2_row_num:
                        continue
                    ratio_formula = f"={latest_col_letter}{source_row}/{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{source_row}'] = ratio_formula
                    analysis_ws[f'{r_col_letter}{source_row}'].number_format = '#,##0.00_);[Red](#,##0.00)'

                # R23: ヘッダー（R1を参照）
                analysis_ws[f'{r_col_letter}{yoy_header_row_num}'] = f'={r_col_letter}1'

                # R列: 対前年差セクションの追加 (latest - 5years_ago)
                analysis_ws[f'{r_col_letter}{diff_yoy_header_row_num}'] = f"=YEAR({latest_col_letter}1) & \"-\" & YEAR({five_years_col_letter}1)"

                for idx, row_num in enumerate(diff_rows_basic):
                    source_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                 total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_calc):
                    source_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    diff_formula = f"={latest_col_letter}{source_row}-{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(diff_rows_roe):
                    source_row = [roe_calc_row_num, ros_row_num, tor_row_num, lrv_row_num,
                                 check1_row_num, check2_row_num, roa_row_num][idx]
                    if source_row == check2_row_num:
                        continue
                    diff_formula = f"={latest_col_letter}{source_row}-{five_years_col_letter}{source_row}"
                    analysis_ws[f'{r_col_letter}{row_num}'] = diff_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = row_formats_map.get(source_row)

                for idx, row_num in enumerate(yoy_rows_basic):
                    source_cagr_row = [sales_analysis_row, profit_analysis_row, net_assets_analysis_row,
                                     total_assets_analysis_row, equity_ratio_analysis_row, roe_analysis_row][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{five_years_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({five_years_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_calc):
                    source_cagr_row = [equity_calc_row_num, equity_avg_row_num, total_assets_avg_row_num][idx]
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{five_years_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({five_years_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                for idx, row_num in enumerate(yoy_rows_roe):
                    source_cagr_row = source_rows_roe[idx]
                    if source_cagr_row == check2_row_num:
                        continue
                    cagr_formula = (f"=({latest_col_letter}{source_cagr_row}/{five_years_col_letter}{source_cagr_row})"
                                  f"^(1/(YEAR({latest_col_letter}$1)-YEAR({five_years_col_letter}$1)))-1")
                    analysis_ws[f'{r_col_letter}{row_num}'] = cagr_formula
                    analysis_ws[f'{r_col_letter}{row_num}'].number_format = number_format_percent

                analysis_ws.column_dimensions[r_col_letter].width = 12

                debug_log(f"Added R column for 5-9 year data: R={latest_col_letter}/{five_years_col_letter}")

    debug_log(f"ROE analysis sheet created: {analysis_sheet_name}")



def create_percentage_bs_sheet(workbook, source_sheet_name, debug_log=None):
    """
    連結貸借対照表から百分率BSシートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "連結貸借対照表(日本基準)"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    analysis_sheet_name = f"{source_sheet_name}_分析_百分率BS"

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 元シートの行数と列数を取得
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    # Find Assets row (資産合計) - the total assets line
    # Look for jppfs_cor_Assets or jppfs_cor_TotalAssets (exact match, not ending with)
    total_assets_row = None
    for row in range(2, min(100, max_row + 1)):
        b_val = source_ws.cell(row, 2).value
        if b_val:
            b_str = str(b_val)
            # Match patterns like "jppfs_cor_Assets" or "jppfs_cor_TotalAssets"
            # But NOT "jppfs_cor_CurrentAssets" or "jppfs_cor_NoncurrentAssets" etc.
            if b_str in ('jppfs_cor_Assets', 'jppfs_cor_TotalAssets') or \
               (b_str.endswith('_Assets') and 'Current' not in b_str and 'Noncurrent' not in b_str and
                'Abstract' not in b_str and 'Lease' not in b_str and 'Property' not in b_str and
                'Intangible' not in b_str and 'Investments' not in b_str and 'Deferred' not in b_str):
                total_assets_row = row
                debug_log(f"Found Assets row: {row} ({b_val})")
                break

    if total_assets_row is None:
        debug_log(f"Percentage BS sheet skipped: Assets row not found in '{source_sheet_name}'")
        return

    # 1行目: ヘッダー行（元シートを参照）
    header_row = []
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"='{source_sheet_name}'!{col_letter}1"
        header_row.append(formula)
    analysis_ws.append(header_row)

    # 2行目以降: データ行
    for row in range(2, max_row + 1):
        data_row = []
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)

            if col <= 2:
                # A列とB列は元シートを参照
                formula = f"='{source_sheet_name}'!{col_letter}{row}"
                data_row.append(formula)
            else:
                # C列以降: 百分率計算
                # =IF(OR('連結貸借対照表(日本基準)'!C$35="", '連結貸借対照表(日本基準)'!C2="", '連結貸借対照表(日本基準)'!C$35=0), "", '連結貸借対照表(日本基準)'!C2/'連結貸借対照表(日本基準)'!C$35)
                formula = (f"=IF(OR('{source_sheet_name}'!{col_letter}${total_assets_row}=\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}=\"\","
                          f"'{source_sheet_name}'!{col_letter}${total_assets_row}=0),"
                          f"\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}/'{source_sheet_name}'!{col_letter}${total_assets_row})")
                data_row.append(formula)

        analysis_ws.append(data_row)

    # データ列のどれが有効かを判定（資産合計があるか）
    def has_assets_data(col_num):
        """指定した列に資産合計データがあるかチェック"""
        if col_num < 3:
            return False
        try:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            value = source_ws[f'{col_letter}{total_assets_row}'].value
            return value is not None and value != ''
        except Exception:
            return False

    # 最古と最新のデータ列を特定
    oldest_col = None
    latest_col = max_col

    for col in range(3, max_col + 1):
        if has_assets_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        debug_log(f"Percentage BS sheet created but no valid data columns found")
        return

    # 期間を計算 (kikan = 最新列 - 最古列)
    kikan = latest_col - oldest_col

    # 比較列を追加
    # 10年データがある場合: N-D(10年), I-D(5年前半), N-I(5年後半)
    # 5-9年データがある場合: N-D(最大期間), I-D(5年前), N-I(残り期間) ※Iは存在する場合
    # 5年未満: なし

    comparison_cols = []

    if kikan >= 10:
        # 10年以上データがある場合
        # Column O: N-D (直近 - 10年前)
        ten_years_ago_col = latest_col - 10
        if ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 1,
                'latest': latest_col,
                'base': ten_years_ago_col
            })

        # Column P: I-D (5年前 - 10年前)  ※Iは直近から5年前
        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col and ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': ten_years_ago_col
            })

        # Column Q: N-I (直近 - 5年前)
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })
    elif kikan >= 5:
        # 5-9年データがある場合
        # Column N: M-C (直近 - 最古)
        comparison_cols.append({
            'col': max_col + 1,
            'latest': latest_col,
            'base': oldest_col
        })

        # Column O: H-C (5年前 - 最古) ※5年前がある場合
        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': oldest_col
            })

        # Column P: M-H (直近 - 5年前) ※5年前がある場合
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })

    # 比較列を追加
    for comp in comparison_cols:
        comp_col = comp['col']
        comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
        latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
        base_letter = openpyxl.utils.get_column_letter(comp['base'])

        # 1行目: 年度表示 (例: "2025-2015")
        year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
        analysis_ws[f'{comp_col_letter}1'] = year_formula

        # 2行目以降: 差分計算
        for row in range(2, max_row + 1):
            diff_formula = f"=IF(OR({latest_letter}{row}=\"\",{base_letter}{row}=\"\"),\"\",{latest_letter}{row}-{base_letter}{row})"
            analysis_ws[f'{comp_col_letter}{row}'] = diff_formula

    # 表示形式の設定
    number_format_percent = r'0.0%;[Red]-0.0%'

    # C列以降のすべてのデータセル（比較列含む）にパーセント表示を設定
    total_cols = max_col + len(comparison_cols)
    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(2, max_row + 1):
            analysis_ws[f'{col_letter}{row}'].number_format = number_format_percent

    # 列幅の設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True

    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠の固定 (B2)
    analysis_ws.freeze_panes = 'B2'

    debug_log(f"Percentage BS sheet created: {analysis_sheet_name}")


def create_percentage_pl_sheet(workbook, source_sheet_name, debug_log=None):
    """
    連結損益計算書から百分率PLシートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "連結損益計算書(日本基準)"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    analysis_sheet_name = f"{source_sheet_name} 分析_百分率PL"

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 元シートの行数と列数を取得
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    # Find NetSales row (売上高)
    net_sales_row = None
    for row in range(2, min(100, max_row + 1)):
        b_val = source_ws.cell(row, 2).value
        if b_val and 'NetSales' in str(b_val):
            net_sales_row = row
            debug_log(f"Found NetSales row: {row}")
            break

    if net_sales_row is None:
        debug_log(f"Percentage PL sheet skipped: NetSales row not found in '{source_sheet_name}'")
        return

    # 1行目: ヘッダー行（元シートを参照）
    header_row = []
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"='{source_sheet_name}'!{col_letter}1"
        header_row.append(formula)
    analysis_ws.append(header_row)

    # 2行目以降: データ行
    for row in range(2, max_row + 1):
        data_row = []
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)

            if col <= 2:
                # A列とB列は元シートを参照
                formula = f"='{source_sheet_name}'!{col_letter}{row}"
                data_row.append(formula)
            else:
                # C列以降: 百分率計算
                # =IF(OR(ISBLANK('連結損益計算書(日本基準)'!C2),ISBLANK('連結損益計算書(日本基準)'!C$3)),"",'連結損益計算書(日本基準)'!C2/'連結損益計算書(日本基準)'!C$3)
                formula = (f"=IF(OR(ISBLANK('{source_sheet_name}'!{col_letter}{row}),"
                          f"ISBLANK('{source_sheet_name}'!{col_letter}${net_sales_row})),"
                          f"\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}/'{source_sheet_name}'!{col_letter}${net_sales_row})")
                data_row.append(formula)

        analysis_ws.append(data_row)

    # データ列のどれが有効かを判定（売上高があるか）
    def has_sales_data(col_num):
        """指定した列に売上高データがあるかチェック"""
        if col_num < 3:
            return False
        try:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            value = source_ws[f'{col_letter}{net_sales_row}'].value
            return value is not None and value != ''
        except Exception:
            return False

    # 最古と最新のデータ列を特定
    oldest_col = None
    latest_col = max_col

    for col in range(3, max_col + 1):
        if has_sales_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        debug_log(f"Percentage PL sheet created but no valid data columns found")
        return

    # 期間を計算 (kikan = 最新列 - 最古列)
    kikan = latest_col - oldest_col

    # 比較列を追加
    # 10年データがある場合: M-C(10年), H-C(5年前半), M-H(5年後半)
    # 5-9年データがある場合: M-C(最大期間), H-C(5年前), M-H(残り期間) ※Hは存在する場合
    # 5年未満: なし

    comparison_cols = []

    if kikan >= 10:
        # 10年以上データがある場合
        # Column N: M-C (直近 - 10年前)
        ten_years_ago_col = latest_col - 10
        if ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 1,
                'latest': latest_col,
                'base': ten_years_ago_col
            })

        # Column O: H-C (5年前 - 10年前)  ※Hは直近から5年前
        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col and ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': ten_years_ago_col
            })

        # Column P: M-H (直近 - 5年前)
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })
    elif kikan >= 5:
        # 5-9年データがある場合
        # Column N: M-C (直近 - 最古)
        comparison_cols.append({
            'col': max_col + 1,
            'latest': latest_col,
            'base': oldest_col
        })

        # Column O: H-C (5年前 - 最古) ※5年前がある場合
        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': oldest_col
            })

        # Column P: M-H (直近 - 5年前) ※5年前がある場合
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })

    # 比較列を追加
    for comp in comparison_cols:
        comp_col = comp['col']
        comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
        latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
        base_letter = openpyxl.utils.get_column_letter(comp['base'])

        # 1行目: 年度表示 (例: "2025-2015")
        year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
        analysis_ws[f'{comp_col_letter}1'] = year_formula

        # 2行目以降: 差分計算
        for row in range(2, max_row + 1):
            diff_formula = f"=IF(OR({latest_letter}{row}=\"\",{base_letter}{row}=\"\"),\"\",{latest_letter}{row}-{base_letter}{row})"
            analysis_ws[f'{comp_col_letter}{row}'] = diff_formula

    # 表示形式の設定
    number_format_percent = r'0.0%;[Red]-0.0%'

    # C列以降のすべてのデータセル（比較列含む）にパーセント表示を設定
    total_cols = max_col + len(comparison_cols)
    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(2, max_row + 1):
            analysis_ws[f'{col_letter}{row}'].number_format = number_format_percent

    # 列幅の設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True

    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠の固定 (B2)
    analysis_ws.freeze_panes = 'B2'

    debug_log(f"Percentage PL sheet created: {analysis_sheet_name}")


def create_percentage_bs_sheet_non_consolidated(workbook, source_sheet_name, debug_log=None):
    """
    貸借対照表（単体）から百分率BSシートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "貸借対照表"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    analysis_sheet_name = f"{source_sheet_name}_分析_百分率BS"

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 元シートの行数と列数を取得
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    # Find Assets row (資産合計)
    total_assets_row = None
    for row in range(2, min(100, max_row + 1)):
        b_val = source_ws.cell(row, 2).value
        if b_val:
            b_str = str(b_val)
            if b_str in ('jppfs_cor_Assets', 'jppfs_cor_TotalAssets') or \
               (b_str.endswith('_Assets') and 'Current' not in b_str and 'Noncurrent' not in b_str and
                'Abstract' not in b_str and 'Lease' not in b_str and 'Property' not in b_str and
                'Intangible' not in b_str and 'Investments' not in b_str and 'Deferred' not in b_str):
                total_assets_row = row
                debug_log(f"Found Assets row: {row} ({b_val})")
                break

    if total_assets_row is None:
        debug_log(f"Percentage BS sheet (non-consolidated) skipped: Assets row not found in '{source_sheet_name}'")
        return

    # 1行目: ヘッダー行（元シートを参照）
    header_row = []
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"='{source_sheet_name}'!{col_letter}1"
        header_row.append(formula)
    analysis_ws.append(header_row)

    # 2行目以降: データ行
    for row in range(2, max_row + 1):
        data_row = []
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)

            if col <= 2:
                # A列とB列は元シートを参照
                formula = f"='{source_sheet_name}'!{col_letter}{row}"
                data_row.append(formula)
            else:
                # C列以降: 百分率計算
                formula = (f"=IF(OR('{source_sheet_name}'!{col_letter}${total_assets_row}=\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}=\"\","
                          f"'{source_sheet_name}'!{col_letter}${total_assets_row}=0),"
                          f"\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}/'{source_sheet_name}'!{col_letter}${total_assets_row})")
                data_row.append(formula)

        analysis_ws.append(data_row)

    # データ列のどれが有効かを判定（資産合計があるか）
    def has_assets_data(col_num):
        """指定した列に資産合計データがあるかチェック"""
        if col_num < 3:
            return False
        try:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            value = source_ws[f'{col_letter}{total_assets_row}'].value
            return value is not None and value != ''
        except Exception:
            return False

    # 最古と最新のデータ列を特定
    oldest_col = None
    latest_col = max_col

    for col in range(3, max_col + 1):
        if has_assets_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        debug_log(f"Percentage BS sheet (non-consolidated) created but no valid data columns found")
        return

    # 期間を計算 (kikan = 最新列 - 最古列)
    kikan = latest_col - oldest_col

    # 比較列を追加
    comparison_cols = []

    if kikan >= 10:
        # 10年以上データがある場合
        ten_years_ago_col = latest_col - 10
        if ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 1,
                'latest': latest_col,
                'base': ten_years_ago_col
            })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col and ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': ten_years_ago_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })
    elif kikan >= 5:
        # 5-9年データがある場合
        comparison_cols.append({
            'col': max_col + 1,
            'latest': latest_col,
            'base': oldest_col
        })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': oldest_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })

    # 比較列を追加
    for comp in comparison_cols:
        comp_col = comp['col']
        comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
        latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
        base_letter = openpyxl.utils.get_column_letter(comp['base'])

        # 1行目: 年度表示 (例: "2025-2015")
        year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
        analysis_ws[f'{comp_col_letter}1'] = year_formula

        # 2行目以降: 差分計算
        for row in range(2, max_row + 1):
            diff_formula = f"=IF(OR({latest_letter}{row}=\"\",{base_letter}{row}=\"\"),\"\",{latest_letter}{row}-{base_letter}{row})"
            analysis_ws[f'{comp_col_letter}{row}'] = diff_formula

    # 表示形式の設定
    number_format_percent = r'0.0%;[Red]-0.0%'

    # C列以降のすべてのデータセル（比較列含む）にパーセント表示を設定
    total_cols = max_col + len(comparison_cols)
    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(2, max_row + 1):
            analysis_ws[f'{col_letter}{row}'].number_format = number_format_percent

    # 列幅の設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True

    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠の固定 (B2)
    analysis_ws.freeze_panes = 'B2'

    debug_log(f"Percentage BS sheet (non-consolidated) created: {analysis_sheet_name}")


def create_percentage_pl_sheet_non_consolidated(workbook, source_sheet_name, debug_log=None):
    """
    損益計算書（単体）から百分率PLシートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "損益計算書"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    analysis_sheet_name = f"{source_sheet_name} 分析_百分率PL"

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 元シートの行数と列数を取得
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    # Find NetSales row (売上高)
    net_sales_row = None
    for row in range(2, min(100, max_row + 1)):
        b_val = source_ws.cell(row, 2).value
        if b_val and 'NetSales' in str(b_val):
            net_sales_row = row
            debug_log(f"Found NetSales row: {row}")
            break

    if net_sales_row is None:
        debug_log(f"Percentage PL sheet (non-consolidated) skipped: NetSales row not found in '{source_sheet_name}'")
        return

    # 1行目: ヘッダー行（元シートを参照）
    header_row = []
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"='{source_sheet_name}'!{col_letter}1"
        header_row.append(formula)
    analysis_ws.append(header_row)

    # 2行目以降: データ行
    for row in range(2, max_row + 1):
        data_row = []
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)

            if col <= 2:
                # A列とB列は元シートを参照
                formula = f"='{source_sheet_name}'!{col_letter}{row}"
                data_row.append(formula)
            else:
                # C列以降: 百分率計算
                formula = (f"=IF(OR(ISBLANK('{source_sheet_name}'!{col_letter}{row}),"
                          f"ISBLANK('{source_sheet_name}'!{col_letter}${net_sales_row})),"
                          f"\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}/'{source_sheet_name}'!{col_letter}${net_sales_row})")
                data_row.append(formula)

        analysis_ws.append(data_row)

    # データ列のどれが有効かを判定（売上高があるか）
    def has_sales_data(col_num):
        """指定した列に売上高データがあるかチェック"""
        if col_num < 3:
            return False
        try:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            value = source_ws[f'{col_letter}{net_sales_row}'].value
            return value is not None and value != ''
        except Exception:
            return False

    # 最古と最新のデータ列を特定
    oldest_col = None
    latest_col = max_col

    for col in range(3, max_col + 1):
        if has_sales_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        debug_log(f"Percentage PL sheet (non-consolidated) created but no valid data columns found")
        return

    # 期間を計算 (kikan = 最新列 - 最古列)
    kikan = latest_col - oldest_col

    # 比較列を追加
    comparison_cols = []

    if kikan >= 10:
        # 10年以上データがある場合
        ten_years_ago_col = latest_col - 10
        if ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 1,
                'latest': latest_col,
                'base': ten_years_ago_col
            })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col and ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': ten_years_ago_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })
    elif kikan >= 5:
        # 5-9年データがある場合
        comparison_cols.append({
            'col': max_col + 1,
            'latest': latest_col,
            'base': oldest_col
        })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': oldest_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })

    # 比較列を追加
    for comp in comparison_cols:
        comp_col = comp['col']
        comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
        latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
        base_letter = openpyxl.utils.get_column_letter(comp['base'])

        # 1行目: 年度表示 (例: "2025-2015")
        year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
        analysis_ws[f'{comp_col_letter}1'] = year_formula

        # 2行目以降: 差分計算
        for row in range(2, max_row + 1):
            diff_formula = f"=IF(OR({latest_letter}{row}=\"\",{base_letter}{row}=\"\"),\"\",{latest_letter}{row}-{base_letter}{row})"
            analysis_ws[f'{comp_col_letter}{row}'] = diff_formula

    # 表示形式の設定
    number_format_percent = r'0.0%;[Red]-0.0%'

    # C列以降のすべてのデータセル（比較列含む）にパーセント表示を設定
    total_cols = max_col + len(comparison_cols)
    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(2, max_row + 1):
            analysis_ws[f'{col_letter}{row}'].number_format = number_format_percent

    # 列幅の設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True

    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠の固定 (B2)
    analysis_ws.freeze_panes = 'B2'

    debug_log(f"Percentage PL sheet (non-consolidated) created: {analysis_sheet_name}")


def create_percentage_ifrs_financial_position_sheet(workbook, source_sheet_name, debug_log=None):
    """
    連結財政状態計算書(IFRS)から百分率シートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "連結財政状態計算書(IFRS)"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    analysis_sheet_name = f"{source_sheet_name}_分析_百分率BS"

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 元シートの行数と列数を取得
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    # Find Total Assets row (資産合計) for IFRS
    # IFRS uses patterns like: jpigp_cor_AssetsIFRS, jpcrp_cor_AssetsIFRS, ifrs_Assets, etc.
    total_assets_row = None
    for row in range(2, min(100, max_row + 1)):
        b_val = source_ws.cell(row, 2).value
        if b_val:
            b_str = str(b_val)
            # Match exact IFRS Assets patterns (total assets only, not subtypes)
            # Must end with "AssetsIFRS" or "_Assets" (not have other words before Assets)
            if b_str in ('jpigp_cor_AssetsIFRS', 'jpcrp_cor_AssetsIFRS', 'ifrs_Assets', 'jppfs_cor_AssetsIFRS', 'jppfs_cor_Assets'):
                total_assets_row = row
                debug_log(f"Found Assets row (IFRS): {row} ({b_val})")
                break
            # Also check for patterns that end with _AssetsIFRS (underscore + AssetsIFRS)
            if b_str.endswith('_AssetsIFRS') and not any(x in b_str for x in [
                'Current', 'Noncurrent', 'Contract', 'Intangible', 'Financial',
                'Deferred', 'Tax', 'Lease', 'Property', 'Investment', 'Held'
            ]):
                total_assets_row = row
                debug_log(f"Found Assets row (IFRS): {row} ({b_val})")
                break

    # If not found with IFRS suffix, try standard Assets pattern
    if total_assets_row is None:
        for row in range(2, min(100, max_row + 1)):
            b_val = source_ws.cell(row, 2).value
            if b_val:
                b_str = str(b_val)
                if b_str in ('jppfs_cor_Assets', 'jppfs_cor_TotalAssets') or \
                   (b_str.endswith('_Assets') and 'Current' not in b_str and 'Noncurrent' not in b_str and
                    'Abstract' not in b_str and 'Lease' not in b_str and 'Property' not in b_str and
                    'Intangible' not in b_str and 'Investments' not in b_str and 'Deferred' not in b_str and
                    'Contract' not in b_str and 'Financial' not in b_str and 'Held' not in b_str):
                    total_assets_row = row
                    debug_log(f"Found Assets row (IFRS, standard pattern): {row} ({b_val})")
                    break

    if total_assets_row is None:
        debug_log(f"Percentage Financial Position sheet (IFRS) skipped: Assets row not found in '{source_sheet_name}'")
        return

    # 1行目: ヘッダー行（元シートを参照）
    header_row = []
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"='{source_sheet_name}'!{col_letter}1"
        header_row.append(formula)
    analysis_ws.append(header_row)

    # 2行目以降: データ行
    for row in range(2, max_row + 1):
        data_row = []
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)

            if col <= 2:
                # A列とB列は元シートを参照
                formula = f"='{source_sheet_name}'!{col_letter}{row}"
                data_row.append(formula)
            else:
                # C列以降: 百分率計算
                formula = (f"=IF(OR('{source_sheet_name}'!{col_letter}${total_assets_row}=\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}=\"\","
                          f"'{source_sheet_name}'!{col_letter}${total_assets_row}=0),"
                          f"\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}/'{source_sheet_name}'!{col_letter}${total_assets_row})")
                data_row.append(formula)

        analysis_ws.append(data_row)

    # データ列のどれが有効かを判定（資産合計があるか）
    def has_assets_data(col_num):
        """指定した列に資産合計データがあるかチェック"""
        if col_num < 3:
            return False
        try:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            value = source_ws[f'{col_letter}{total_assets_row}'].value
            return value is not None and value != ''
        except Exception:
            return False

    # 最古と最新のデータ列を特定
    oldest_col = None
    latest_col = max_col

    for col in range(3, max_col + 1):
        if has_assets_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        debug_log(f"Percentage Financial Position sheet (IFRS) created but no valid data columns found")
        return

    # 期間を計算 (kikan = 最新列 - 最古列)
    kikan = latest_col - oldest_col

    # 比較列を追加
    comparison_cols = []

    if kikan >= 10:
        ten_years_ago_col = latest_col - 10
        if ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 1,
                'latest': latest_col,
                'base': ten_years_ago_col
            })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col and ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': ten_years_ago_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })
    elif kikan >= 5:
        comparison_cols.append({
            'col': max_col + 1,
            'latest': latest_col,
            'base': oldest_col
        })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': oldest_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })

    # 比較列を追加
    for comp in comparison_cols:
        comp_col = comp['col']
        comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
        latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
        base_letter = openpyxl.utils.get_column_letter(comp['base'])

        # 1行目: 年度表示
        year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
        analysis_ws[f'{comp_col_letter}1'] = year_formula

        # 2行目以降: 差分計算
        for row in range(2, max_row + 1):
            diff_formula = f"=IF(OR({latest_letter}{row}=\"\",{base_letter}{row}=\"\"),\"\",{latest_letter}{row}-{base_letter}{row})"
            analysis_ws[f'{comp_col_letter}{row}'] = diff_formula

    # 表示形式の設定
    number_format_percent = r'0.0%;[Red]-0.0%'

    # C列以降のすべてのデータセル（比較列含む）にパーセント表示を設定
    total_cols = max_col + len(comparison_cols)
    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(2, max_row + 1):
            analysis_ws[f'{col_letter}{row}'].number_format = number_format_percent

    # 列幅の設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True

    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠の固定 (B2)
    analysis_ws.freeze_panes = 'B2'

    debug_log(f"Percentage Financial Position sheet (IFRS) created: {analysis_sheet_name}")


def create_percentage_ifrs_income_sheet(workbook, source_sheet_name, debug_log=None):
    """
    連結損益計算書(IFRS)から百分率PLシートを生成

    Args:
        workbook: openpyxlワークブック
        source_sheet_name: 元シート名（例: "連結損益計算書(IFRS)"）
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    if source_sheet_name not in workbook.sheetnames:
        return

    source_ws = workbook[source_sheet_name]
    analysis_sheet_name = f"{source_sheet_name} 分析_百分率PL"

    # 既存の分析シートがあれば削除
    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    # 新しいシートを作成
    analysis_ws = workbook.create_sheet(analysis_sheet_name)

    # 元シートの行数と列数を取得
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    # Find Revenue/Sales row (売上高) for IFRS
    # IFRS uses "Revenue" instead of "NetSales"
    net_sales_row = None
    for row in range(2, min(100, max_row + 1)):
        b_val = source_ws.cell(row, 2).value
        if b_val and ('Revenue' in str(b_val) or 'NetSales' in str(b_val) or '売上' in str(b_val)):
            net_sales_row = row
            debug_log(f"Found Revenue/Sales row (IFRS): {row}")
            break

    if net_sales_row is None:
        debug_log(f"Percentage Income sheet (IFRS) skipped: Revenue/Sales row not found in '{source_sheet_name}'")
        return

    # 1行目: ヘッダー行（元シートを参照）
    header_row = []
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"='{source_sheet_name}'!{col_letter}1"
        header_row.append(formula)
    analysis_ws.append(header_row)

    # 2行目以降: データ行
    for row in range(2, max_row + 1):
        data_row = []
        for col in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)

            if col <= 2:
                # A列とB列は元シートを参照
                formula = f"='{source_sheet_name}'!{col_letter}{row}"
                data_row.append(formula)
            else:
                # C列以降: 百分率計算
                formula = (f"=IF(OR(ISBLANK('{source_sheet_name}'!{col_letter}{row}),"
                          f"ISBLANK('{source_sheet_name}'!{col_letter}${net_sales_row})),"
                          f"\"\","
                          f"'{source_sheet_name}'!{col_letter}{row}/'{source_sheet_name}'!{col_letter}${net_sales_row})")
                data_row.append(formula)

        analysis_ws.append(data_row)

    # データ列のどれが有効かを判定（売上高があるか）
    def has_sales_data(col_num):
        """指定した列に売上高データがあるかチェック"""
        if col_num < 3:
            return False
        try:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            value = source_ws[f'{col_letter}{net_sales_row}'].value
            return value is not None and value != ''
        except Exception:
            return False

    # 最古と最新のデータ列を特定
    oldest_col = None
    latest_col = max_col

    for col in range(3, max_col + 1):
        if has_sales_data(col):
            oldest_col = col
            break

    if oldest_col is None:
        debug_log(f"Percentage Income sheet (IFRS) created but no valid data columns found")
        return

    # 期間を計算 (kikan = 最新列 - 最古列)
    kikan = latest_col - oldest_col

    # 比較列を追加
    comparison_cols = []

    if kikan >= 10:
        ten_years_ago_col = latest_col - 10
        if ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 1,
                'latest': latest_col,
                'base': ten_years_ago_col
            })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col and ten_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': ten_years_ago_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })
    elif kikan >= 5:
        comparison_cols.append({
            'col': max_col + 1,
            'latest': latest_col,
            'base': oldest_col
        })

        five_years_ago_col = latest_col - 5
        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 2,
                'latest': five_years_ago_col,
                'base': oldest_col
            })

        if five_years_ago_col >= oldest_col:
            comparison_cols.append({
                'col': max_col + 3,
                'latest': latest_col,
                'base': five_years_ago_col
            })

    # 比較列を追加
    for comp in comparison_cols:
        comp_col = comp['col']
        comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
        latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
        base_letter = openpyxl.utils.get_column_letter(comp['base'])

        # 1行目: 年度表示
        year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
        analysis_ws[f'{comp_col_letter}1'] = year_formula

        # 2行目以降: 差分計算
        for row in range(2, max_row + 1):
            diff_formula = f"=IF(OR({latest_letter}{row}=\"\",{base_letter}{row}=\"\"),\"\",{latest_letter}{row}-{base_letter}{row})"
            analysis_ws[f'{comp_col_letter}{row}'] = diff_formula

    # 表示形式の設定
    number_format_percent = r'0.0%;[Red]-0.0%'

    # C列以降のすべてのデータセル（比較列含む）にパーセント表示を設定
    total_cols = max_col + len(comparison_cols)
    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(2, max_row + 1):
            analysis_ws[f'{col_letter}{row}'].number_format = number_format_percent

    # 列幅の設定
    analysis_ws.column_dimensions['A'].width = 28
    analysis_ws.column_dimensions['B'].hidden = True

    for col in range(3, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        analysis_ws.column_dimensions[col_letter].width = 12

    # ウィンドウ枠の固定 (B2)
    analysis_ws.freeze_panes = 'B2'

    debug_log(f"Percentage Income sheet (IFRS) created: {analysis_sheet_name}")


def add_financial_analysis_sheets(workbook, debug_log=None):
    """
    財務分析シートを追加する（メイン関数）

    Args:
        workbook: openpyxlワークブック
        debug_log: デバッグログ関数（オプション）
    """
    # デバッグログ関数がない場合はダミー関数を使用
    if debug_log is None:
        def debug_log(msg):
            pass

    # 主要な経営指標等の推移（連結）シートを検索してROE分析シートを生成
    for sheet_name in workbook.sheetnames:
        if '主要な経営指標等の推移' in sheet_name and '連結' in sheet_name and '_' not in sheet_name:
            # "_"が含まれない = オリジナルシート（分析シートではない）
            try:
                create_roe_analysis_sheet(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create ROE analysis sheet for '{sheet_name}': {e}")
                # ROE分析シート生成に失敗してもメイン処理は継続

    # 連結財政状態計算書(IFRS)シートを検索して百分率BSシートを生成
    for sheet_name in workbook.sheetnames:
        if '財政状態計算書' in sheet_name and 'IFRS' in sheet_name and '_' not in sheet_name:
            try:
                create_percentage_ifrs_financial_position_sheet(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create percentage Financial Position sheet (IFRS) for '{sheet_name}': {e}")

    # 連結貸借対照表シートを検索して百分率BSシートを生成
    for sheet_name in workbook.sheetnames:
        if '連結貸借対照表' in sheet_name and '日本基準' in sheet_name and '_' not in sheet_name:
            try:
                create_percentage_bs_sheet(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create percentage BS sheet for '{sheet_name}': {e}")

    # 連結損益計算書(IFRS)シートを検索して百分率PLシートを生成
    for sheet_name in workbook.sheetnames:
        if '損益計算書' in sheet_name and 'IFRS' in sheet_name and '包括' not in sheet_name and '_' not in sheet_name:
            try:
                create_percentage_ifrs_income_sheet(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create percentage Income sheet (IFRS) for '{sheet_name}': {e}")

    # 連結損益計算書シートを検索して百分率PLシートを生成
    for sheet_name in workbook.sheetnames:
        if '連結損益計算書' in sheet_name and '日本基準' in sheet_name and '_' not in sheet_name:
            try:
                create_percentage_pl_sheet(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create percentage PL sheet for '{sheet_name}': {e}")

    # 主要な経営指標等の推移（単体）シートを検索してROE分析シートを生成
    for sheet_name in workbook.sheetnames:
        if '主要な経営指標等の推移' in sheet_name and '単体' in sheet_name and '_' not in sheet_name:
            try:
                create_roe_analysis_sheet_non_consolidated(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create ROE analysis sheet for '{sheet_name}': {e}")

    # 貸借対照表（単体）シートを検索して百分率BSシートを生成
    for sheet_name in workbook.sheetnames:
        if sheet_name == '貸借対照表':
            try:
                create_percentage_bs_sheet_non_consolidated(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create percentage BS sheet (non-consolidated) for '{sheet_name}': {e}")

    # 損益計算書（単体）シートを検索して百分率PLシートを生成
    for sheet_name in workbook.sheetnames:
        if sheet_name == '損益計算書':
            try:
                create_percentage_pl_sheet_non_consolidated(workbook, sheet_name, debug_log)
            except Exception as e:
                debug_log(f"Warning: Failed to create percentage PL sheet (non-consolidated) for '{sheet_name}': {e}")
