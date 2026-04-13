"""
Diversity Analysis Module for XBRL to Excel Conversion

ダイバーシティ関連指標（管理職女性割合・男性育児休業取得率・賃金差異）の
シートを生成するモジュール
"""

import re
import unicodedata
import html
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _extract_subsidiary_names_from_ixbrl(ixbrl_files):
    """iXBRL HTMLファイルを解析して連番RowNMember→会社名のマッピングを返す。

    連結子会社ダイバーシティ要素（SUBSIDIARY_ELEMENTS）と同じ行に存在する
    会社名テキストを各RowN別に抽出する。

    Returns:
        dict: {row_num (int): company_name (str)}
    """
    row_names = {}

    # 対象要素のローカル名（namespace prefix 除去後）
    target_local_names = {el.replace('jpcrp_cor_', 'jpcrp_cor:') for el, _ in SUBSIDIARY_ELEMENTS}

    for filepath in (ixbrl_files or []):
        try:
            with open(filepath, encoding='utf-8', errors='ignore') as f:
                content = f.read()
        except OSError:
            continue

        # ダイバーシティ関連要素かつRowNMember contextのタグを探す
        pattern = re.compile(
            r'contextRef="[^"]*Row(\d+)Member[^"]*"[^>]*name="('
            + '|'.join(re.escape(n) for n in target_local_names)
            + r')"',
        )
        for m in pattern.finditer(content):
            row_num = int(m.group(1))
            if row_num in row_names:
                continue
            pos = m.start()
            # 直近N行の <tr> を、現在の行に近い順から検索する
            # 非常に広範囲なテーブルレイアウトに対応するため、検索窓をさらに拡大 (30000)
            chunk = content[max(0, pos - 30000):pos]
            tr_matches = list(re.finditer(r'<tr[ >]', chunk))
            
            found_name = False
            for i in range(1, min(len(tr_matches), 5) + 1):
                start_tr = tr_matches[-i].start()
                # 次の <tr> があればそこまで、なければ chunk の終わりまで
                end_tr = tr_matches[-i+1].start() if i > 1 else len(chunk)
                row_slice = chunk[start_tr:end_tr]
                
                tds = re.findall(r'<td[^>]*>(.*?)</td>', row_slice, re.DOTALL)
                for td in tds:
                    # HTML実体参照（&#160;等）をデコードしてからタグを除去
                    text = html.unescape(td)
                    text = re.sub('<[^>]+>', ' ', text)
                    text = re.sub(r'\s+', '', text).strip()
                    # 社名の後ろに付いている注釈（※4等）を除去
                    text = re.sub(r'[※\*＊].*$', '', text).strip()
                    # 数字・記号・注釈マーカー・連番ラベルのみの文字列は社名として除外
                    if (text and 
                        not re.match(r'^[※\d\.,\(\)\-\%（）\*＊注]+$', text) and 
                        not re.search(r'連番[：:]?\d+件目', text) and
                        len(text) > 1):
                        row_names[row_num] = text
                        found_name = True
                        break
                if found_name:
                    break

    return row_names


# 提出会社の指標（dim_label = "全体" or NonConsolidatedMember）
REPORTING_COMPANY_ELEMENTS = [
    (
        'jpcrp_cor_RatioOfFemaleEmployeesInManagerialPositionsMetricsOfReportingCompany',
        '管理職に占める女性労働者の割合、提出会社の指標',
    ),
    (
        'jpcrp_cor_RegularEmployeesCalculatedBasedOnProvisionsOfActOnPromotionOfWomensActiveEngagementInProfessionalLifeRatioOfMaleEmployeesTakingChildcareLeaveMetricsOfReportingCompany',
        '正規雇用労働者、女性の職業生活における活躍の推進に関する法律の規定に基づき算出、男性労働者の育児休業取得率、提出会社の指標',
    ),
    (
        'jpcrp_cor_NonRegularEmployeesCalculatedBasedOnProvisionsOfActOnPromotionOfWomensActiveEngagementInProfessionalLifeRatioOfMaleEmployeesTakingChildcareLeaveMetricsOfReportingCompany',
        '非正規雇用労働者、女性の職業生活における活躍の推進に関する法律の規定に基づき算出、男性労働者の育児休業取得率、提出会社の指標',
    ),
    (
        'jpcrp_cor_AllEmployeesDifferencesInWagesBetweenMaleAndFemaleEmployeesMetricsOfReportingCompany',
        '全労働者、労働者の男女の賃金の差異、提出会社の指標',
    ),
    (
        'jpcrp_cor_RegularEmployeesDifferencesInWagesBetweenMaleAndFemaleEmployeesMetricsOfReportingCompany',
        '正規雇用労働者、労働者の男女の賃金の差異、提出会社の指標',
    ),
    (
        'jpcrp_cor_NonRegularEmployeesDifferencesInWagesBetweenMaleAndFemaleEmployeesMetricsOfReportingCompany',
        '非正規雇用労働者、労働者の男女の賃金の差異、提出会社の指標',
    ),
]

# 連結子会社の指標（dim_label = "Sequential Numbers：RowN"）
SUBSIDIARY_ELEMENTS = [
    (
        'jpcrp_cor_RegularEmployeesRatioOfMaleEmployeesTakingChildcareLeaveMetricsOfConsolidatedSubsidiaries',
        '正規雇用労働者、男性労働者の育児休業取得率、連結子会社の指標',
    ),
    (
        'jpcrp_cor_NonRegularEmployeesRatioOfMaleEmployeesTakingChildcareLeaveMetricsOfConsolidatedSubsidiaries',
        '非正規雇用労働者、男性労働者の育児休業取得率、連結子会社の指標',
    ),
    (
        'jpcrp_cor_AllEmployeesDifferencesInWagesBetweenMaleAndFemaleEmployeesMetricsOfConsolidatedSubsidiaries',
        '全労働者、労働者の男女の賃金の差異、連結子会社の指標',
    ),
    (
        'jpcrp_cor_RegularEmployeesDifferencesInWagesBetweenMaleAndFemaleEmployeesMetricsOfConsolidatedSubsidiaries',
        '正規雇用労働者、労働者の男女の賃金の差異、連結子会社の指標',
    ),
    (
        'jpcrp_cor_NonRegularEmployeesDifferencesInWagesBetweenMaleAndFemaleEmployeesMetricsOfConsolidatedSubsidiaries',
        '非正規雇用労働者、労働者の男女の賃金の差異、連結子会社の指標',
    ),
]

# 会社名エレメント
SUBSIDIARY_NAME_EL = 'jpcrp_cor_ConsolidatedSubsidiariesMetricsOfConsolidatedSubsidiaries'


def _format_period(period):
    """期間文字列 (YYYY-MM-DD) を表示用フォーマット (YYYY/MM) に変換"""
    if isinstance(period, str) and len(period) >= 7 and '-' in period:
        parts = period.split('-')
        if len(parts) >= 2:
            return f"{parts[0]}/{parts[1]}"
    return str(period)


def _is_row_dim(dim_label):
    """連番ディメンション（'連番：N件目' または 'Sequential Numbers：RowN'）か判定"""
    return '連番：' in dim_label or 'Row' in dim_label


def _row_sort_key(dim_label):
    """連番ディメンションから行番号を抽出してソートキーを返す"""
    m = re.search(r'\d+', dim_label)
    return int(m.group(0)) if m else 9999


def _parse_value(val_str):
    """XBRL値文字列を float に変換。変換失敗時は None。"""
    if val_str is None:
        return None
    s = unicodedata.normalize('NFKC', str(val_str)).replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


def add_diversity_sheet(workbook, global_element_period_values, debug_log=None, ixbrl_files=None, subsidiary_row_names=None):
    """
    ダイバーシティシートを生成してワークブックに追加する。

    シート構成:
        A列: 会社名（提出会社 = "単体"、連結子会社 = 会社名）
        B列: 項目
        C列: 年度
        D列: 値（%）

    Args:
        workbook: openpyxlワークブック
        global_element_period_values: {element: {(fact_std, dim_label, period): value}}
        debug_log: デバッグログ関数（省略可）
    """
    if debug_log is None:
        def debug_log(msg):
            pass

    # =========================================================================
    # 1. 提出会社データ収集
    #    dim_label = "全体" または "単体" 優先
    # =========================================================================
    reporting_items = []  # [(label, {period: value}), ...]
    for el_name, label in REPORTING_COMPANY_ELEMENTS:
        vals = global_element_period_values.get(el_name, {})
        if not vals:
            continue
        period_value = {}
        for (fact_std, dim_label, period), raw_val in vals.items():
            if not period:
                continue
            v = _parse_value(raw_val)
            if v is None:
                continue
            if period not in period_value:
                period_value[period] = (v, dim_label)
            else:
                _, existing_dim = period_value[period]
                if dim_label in ('全体', '単体') and existing_dim not in ('全体', '単体'):
                    period_value[period] = (v, dim_label)
        if period_value:
            reporting_items.append((label, {p: vd[0] for p, vd in period_value.items()}))
            debug_log(f"[Diversity] 提出会社 '{label}': {len(period_value)} 年度")

    # =========================================================================
    # 2. 連結子会社データ収集
    #    会社名マップ: {(dim_label, period): company_name}
    # =========================================================================
    sub_name_map = {}
    for (fact_std, dim_label, period), raw_val in global_element_period_values.get(SUBSIDIARY_NAME_EL, {}).items():
        if not period or not raw_val:
            continue
        company_name = str(raw_val).split('\n')[0].strip()
        if company_name:
            sub_name_map[(dim_label, period)] = company_name

    # XBRL要素で会社名が取得できない場合のフォールバック用マップ {row_num (int): company_name (str)}
    # subsidiary_row_names が渡された場合はそれを優先、次にiXBRL HTMLから抽出
    if subsidiary_row_names:
        ixbrl_row_names = subsidiary_row_names
    elif ixbrl_files:
        ixbrl_row_names = _extract_subsidiary_names_from_ixbrl(ixbrl_files)
    else:
        ixbrl_row_names = {}

    # 連結子会社項目: {el_name: {(dim_label, period): value}}
    # dim_label = "Sequential Numbers：RowN"
    subsidiary_items = []  # [(label, {(dim_label, period): value}), ...]
    for el_name, label in SUBSIDIARY_ELEMENTS:
        vals = global_element_period_values.get(el_name, {})
        if not vals:
            continue
        row_period_val = {}
        for (fact_std, dim_label, period), raw_val in vals.items():
            if not period or not _is_row_dim(dim_label):
                continue
            v = _parse_value(raw_val)
            if v is None:
                continue
            row_period_val[(dim_label, period)] = v
        if row_period_val:
            subsidiary_items.append((label, row_period_val))
            debug_log(f"[Diversity] 連結子会社 '{label}': {len(row_period_val)} エントリ")

    if not reporting_items and not subsidiary_items:
        debug_log("[Diversity] ダイバーシティデータなし、シート作成をスキップ")
        return

    # =========================================================================
    # 3. シート作成
    # =========================================================================
    sheet_name = 'ダイバーシティ'
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # --- スタイル定義 ---
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    reporting_fill = PatternFill(fill_type='solid', fgColor='D6E4F0')  # 提出会社（薄青）
    subsidiary_fill = PatternFill(fill_type='solid', fgColor='E2EFDA')  # 連結子会社（薄緑）
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # --- ヘッダー行 ---
    headers = ['会社名', '項目', '年度', '値（%）']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    current_row = 2

    def write_row(company, label, period_str, val, fill):
        nonlocal current_row
        a = ws.cell(row=current_row, column=1, value=company)
        a.fill = fill
        a.font = normal_font
        a.alignment = left_align

        b = ws.cell(row=current_row, column=2, value=label)
        b.fill = fill
        b.font = normal_font
        b.alignment = left_align

        c = ws.cell(row=current_row, column=3, value=period_str)
        c.font = normal_font
        c.alignment = center_align

        d = ws.cell(row=current_row, column=4, value=val)
        d.font = normal_font
        d.alignment = right_align
        d.number_format = '0.0%'

        current_row += 1

    # =========================================================================
    # 4. 提出会社セクション（会社名 = "単体"）
    # =========================================================================
    for label, period_val in reporting_items:
        sorted_periods = sorted(period_val.keys())
        for period in sorted_periods:
            write_row('単体', label, _format_period(period), period_val[period],
                      reporting_fill)
        current_row += 1  # 項目間空行

    # =========================================================================
    # 5. 連結子会社セクション（会社ごとに展開）
    # =========================================================================
    for label, row_period_val in subsidiary_items:
        # RowN ごとにグループ化
        rows_dict = {}  # {dim_label: {period: value}}
        for (dim_label, period), v in row_period_val.items():
            if dim_label not in rows_dict:
                rows_dict[dim_label] = {}
            rows_dict[dim_label][period] = v

        # RowN 順にソート
        sorted_dims = sorted(rows_dict.keys(), key=_row_sort_key)
        for dim_label in sorted_dims:
            period_val = rows_dict[dim_label]
            sorted_periods = sorted(period_val.keys())

            # 会社名を取得（最初の期間から）
            company_name = None
            for period in sorted_periods:
                company_name = sub_name_map.get((dim_label, period))
                if company_name:
                    break
            if not company_name:
                # iXBRL HTMLから抽出した会社名でフォールバック
                # 抽出された社名があればそれを優先。なければラベルを使用。
                # ただし「連番：...」のようなラベルは社名として適切でないため除外。
                row_num = _row_sort_key(dim_label)
                company_name = ixbrl_row_names.get(row_num)
                if not company_name or re.search(r'連番[：:]\d+件目', company_name):
                    if dim_label and not re.search(r'連番[：:]\d+件目', dim_label):
                        company_name = dim_label
                    else:
                        company_name = company_name or ""

            for period in sorted_periods:
                write_row(company_name, label, _format_period(period), period_val[period],
                          subsidiary_fill)
            current_row += 1  # 会社間空行

        current_row += 1  # 項目間空行（連結子会社）

    # =========================================================================
    # 6. 列幅・枠固定
    # =========================================================================
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.freeze_panes = 'A2'

    debug_log(f"[Diversity] シート '{sheet_name}' 作成完了")


# =============================================================================
# 人的資本シート
# =============================================================================

_HC_PREFIX = 'OfMetricsRelatedToPolicyOnDevelopmentOfHumanResourcesAndInternalEnvironmentAndTargetsAndPerformanceUsingSuchMetrics'

HC_METRICS_EL    = f'jpcrp_cor_MetricsDescription{_HC_PREFIX}'
HC_TARGETS_EL    = f'jpcrp_cor_TargetsDescription{_HC_PREFIX}'
HC_UNIT_EL       = f'jpcrp_cor_MetricsUnitDescription{_HC_PREFIX}'
HC_PERFORMANCE_EL = f'jpcrp_cor_PerformanceDescription{_HC_PREFIX}'


def add_human_capital_sheet(
    workbook, 
    global_element_period_values, 
    debug_log=None, 
    filer_name=None
):
    """
    人的資本シートを生成してワークブックに追加する。

    列構成: 会社名 | 指標名 | 年度 | 目標数値 | 実績数値 | 単位

    各行は (連番RowN, period) の組み合わせ。
    指標名は MetricsDescription から取得。
    会社名は filer_name 引数（省略時は global_element_period_values から取得）。

    Args:
        workbook: openpyxlワークブック
        global_element_period_values: {element: {(fact_std, dim_label, period): value}}
        debug_log: デバッグログ関数（省略可）
        filer_name: 提出会社名（省略時は CompanyNameCoverPage 等から自動取得）
    """
    if debug_log is None:
        def debug_log(msg):
            pass

    # 提出会社名を解決
    if not filer_name:
        for suffix in ('CompanyNameCoverPage', 'EntityNameCompanyName', 'EntityNameEntityReportingName'):
            for el_name, vals in global_element_period_values.items():
                if el_name.endswith(suffix) and vals:
                    filer_name = str(next(iter(vals.values()))).split('\n')[0].strip()
                    break
            if filer_name:
                break
    if not filer_name:
        filer_name = '提出会社'

    # -------------------------------------------------------------------------
    # データ収集: dim_label (連番RowN) × period → 各フィールド値
    # -------------------------------------------------------------------------
    def collect_row_period(el_name):
        """{(dim_label, period): value} を返す（連番ディメンションのみ）"""
        result = {}
        for (fact_std, dim_label, period), raw_val in global_element_period_values.get(el_name, {}).items():
            if not period or not _is_row_dim(dim_label):
                continue
            val = str(raw_val).strip() if raw_val is not None else ''
            if val:
                result[(dim_label, period)] = val
        return result

    metrics_map     = collect_row_period(HC_METRICS_EL)
    targets_map     = collect_row_period(HC_TARGETS_EL)
    unit_map        = collect_row_period(HC_UNIT_EL)
    performance_map = collect_row_period(HC_PERFORMANCE_EL)

    # 出現した (dim_label, period) の全組み合わせを収集
    all_keys = set(metrics_map) | set(targets_map) | set(performance_map)
    if not all_keys:
        debug_log("[HumanCapital] データなし、シート作成をスキップ")
        return

    # -------------------------------------------------------------------------
    # 指標名ベースに再編成
    # (dim_label, period) → metric_name を解決し、
    # {metric_name: {period: {target, unit, perf}}} に変換する
    # -------------------------------------------------------------------------

    # 全期間を収集し最新期間を特定
    all_periods = {period for (_, period) in all_keys}
    latest_period = max(all_periods) if all_periods else None

    # 最新期間での dim_label → metric_name マッピングを取得（登場順を確定するため）
    # dim_label の Row番号順に並べて、metric_name の登場順インデックスを決定
    latest_dim_to_metric = {}  # {dim_label: metric_name}
    for (dim_label, period), val in metrics_map.items():
        if period == latest_period:
            name = val.split('\n')[0].strip()
            if name:
                latest_dim_to_metric[dim_label] = name

    # 最新期間での登場順（Row番号順）
    latest_metric_order = {}  # {metric_name: index}
    for dim_label in sorted(latest_dim_to_metric.keys(), key=_row_sort_key):
        name = latest_dim_to_metric[dim_label]
        if name not in latest_metric_order:
            latest_metric_order[name] = len(latest_metric_order)

    # (dim_label, period) ごとに指標名を解決してデータを集約
    # {metric_name: {period: {'target': ..., 'unit': ..., 'perf': ...}}}
    metric_data = {}  # type: dict

    all_dim_labels = {dim_label for (dim_label, _) in all_keys}
    for dim_label in all_dim_labels:
        periods_for_dim = {period for (dl, period) in all_keys if dl == dim_label}

        # fallback: いずれかの期間に存在する指標名（最古優先）
        fallback_name = dim_label
        for period in sorted(periods_for_dim):
            n = metrics_map.get((dim_label, period), '').split('\n')[0].strip()
            if n:
                fallback_name = n
                break

        for period in periods_for_dim:
            metric_name = metrics_map.get((dim_label, period), '').split('\n')[0].strip() or fallback_name
            if metric_name not in metric_data:
                metric_data[metric_name] = {}
            if period not in metric_data[metric_name]:
                metric_data[metric_name][period] = {
                    'target': targets_map.get((dim_label, period), ''),
                    'unit':   unit_map.get((dim_label, period), ''),
                    'perf':   performance_map.get((dim_label, period), ''),
                }

    # 指標名を最新年度の登場順でソート（最新年度にない指標は末尾）
    def metric_sort_key(name):
        return latest_metric_order.get(name, len(latest_metric_order))

    sorted_metrics = sorted(metric_data.keys(), key=metric_sort_key)
    debug_log(f"[HumanCapital] {len(sorted_metrics)} 指標 × 複数年度")

    # -------------------------------------------------------------------------
    # シート作成
    # -------------------------------------------------------------------------
    sheet_name = '人的資本'
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    row_fill    = PatternFill(fill_type='solid', fgColor='EBF3FB')
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')
    right_align  = Alignment(horizontal='right',  vertical='center')

    # ヘッダー
    headers = ['会社名', '指標名', '年度', '目標数値', '実績数値', '単位']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    current_row = 2

    for metric_name in sorted_metrics:
        period_data = metric_data[metric_name]
        sorted_periods = sorted(period_data.keys())

        for period in sorted_periods:
            entry      = period_data[period]
            target_val = entry['target']
            unit_val   = entry['unit']
            perf_val   = entry['perf']

            # 数値に変換試行
            def to_num(s):
                try:
                    return float(unicodedata.normalize('NFKC', s).replace(',', ''))
                except (ValueError, AttributeError):
                    return s if s else None

            is_pct = (unit_val.strip() == '%')

            def apply_num(raw_float):
                if not isinstance(raw_float, float):
                    return raw_float, None
                if is_pct:
                    return raw_float / 100, '0.0%'
                else:
                    return raw_float, '#,##0.0'

            target_num, target_fmt = apply_num(to_num(target_val))
            perf_num,   perf_fmt   = apply_num(to_num(perf_val))

            cells_data = [
                (filer_name,             left_align,   None),
                (metric_name,            left_align,   None),
                (_format_period(period), center_align, None),
                (target_num,             right_align,  target_fmt),
                (perf_num,               right_align,  perf_fmt),
                (unit_val,               center_align, None),
            ]
            for col_idx, (val, align, fmt) in enumerate(cells_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = normal_font
                cell.fill = row_fill
                cell.alignment = align
                if fmt and isinstance(val, float):
                    cell.number_format = fmt

            current_row += 1

        current_row += 1  # 指標間に空行

    # 列幅
    ws.column_dimensions['A'].width = 25  # 会社名
    ws.column_dimensions['B'].width = 40  # 指標名
    ws.column_dimensions['C'].width = 12  # 年度
    ws.column_dimensions['D'].width = 14  # 目標数値
    ws.column_dimensions['E'].width = 14  # 実績数値
    ws.column_dimensions['F'].width = 8   # 単位
    ws.freeze_panes = 'A2'

    debug_log(f"[HumanCapital] シート '{sheet_name}' 作成完了")
def add_officers_gender_ratio_sheet(workbook, global_element_period_values, debug_log=None):
    """
    「役員の男女比率」シートを生成して追加する。
    """
    if debug_log is None:
        def debug_log(msg): pass

    sheet_name = '役員の男女比率'
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # --- 定義 ---
    elements = [
        ('役員のうち男性の人数', 'jpcrp_cor_NumberOfMaleDirectorsAndOtherOfficers', '#,##0'),
        ('役員のうち女性の人数', 'jpcrp_cor_NumberOfFemaleDirectorsAndOtherOfficers', '#,##0'),
        ('役員数合計', 'TOTAL', '#,##0'),
        ('役員の女性比率', 'jpcrp_cor_RatioOfFemaleDirectorsAndOtherOfficers', '0.0%'),
    ]

    # --- スタイル ---
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ヘッダー
    ws.cell(row=1, column=1, value='区分').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border

    for i, (label, _, _) in enumerate(elements, start=2):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # データ抽出
    # data_store[period][element_name] = value
    data_store = {}
    periods_seen = set()

    for label, el_name, fmt in elements:
        if el_name == 'TOTAL': continue
        vals = global_element_period_values.get(el_name, {})
        if not vals:
            continue
            
        period_value_map = {}
        for (std, dim, period), raw_val in vals.items():
            if not period:
                continue
            v = _parse_value(raw_val)
            if v is None:
                continue
            
            # 優先順位: 全体/単体 > その他
            if period not in period_value_map:
                period_value_map[period] = (v, dim)
            else:
                _, existing_dim = period_value_map[period]
                if dim in ('全体', '単体') and existing_dim not in ('全体', '単体'):
                    period_value_map[period] = (v, dim)

        for p, v_data in period_value_map.items():
            if p not in data_store: data_store[p] = {}
            data_store[p][el_name] = v_data[0]
            periods_seen.add(p)
        
        if period_value_map:
            debug_log(f"[Officers] {el_name} 抽出: {len(period_value_map)} 年度分")

    # 年度軸の作成 (2015-2025)
    years = range(2015, 2026)
    current_row = 2
    for y in years:
        period_to_use = None
        matches = [p for p in periods_seen if p.startswith(str(y))]
        if matches:
            # 提出日(FilingDate)や決算期末など複数の日付がある可能性があるが、最新のものを採用
            period_to_use = sorted(matches)[-1]
        
        # 表示名 (YYYY/M/D形式)
        # FilingDateInstantなどは決算期末(3/31)ではない場合が多いがそのまま表示
        from NumberOfShareholders import _format_date # 既存のフォーマッタを再利用
        display_date = _format_date(period_to_use) if period_to_use else f"{y}/3/31"
        
        cell = ws.cell(row=current_row, column=1, value=display_date)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

        for i, (_, el_name, fmt) in enumerate(elements, start=2):
            val = None
            if el_name == 'TOTAL':
                v_m = data_store.get(period_to_use, {}).get('jpcrp_cor_NumberOfMaleDirectorsAndOtherOfficers')
                v_f = data_store.get(period_to_use, {}).get('jpcrp_cor_NumberOfFemaleDirectorsAndOtherOfficers')
                if v_m is not None or v_f is not None:
                    val = (v_m or 0) + (v_f or 0)
            else:
                val = data_store.get(period_to_use, {}).get(el_name)

            cell = ws.cell(row=current_row, column=i, value=val)
            cell.font = normal_font
            cell.alignment = right_align
            cell.number_format = fmt
            cell.border = thin_border
        current_row += 1

    ws.column_dimensions['A'].width = 15
    for i in range(2, 6):
        ws.column_dimensions[chr(64 + i)].width = 25
    ws.freeze_panes = 'B2'
    debug_log(f"[Officers] '{sheet_name}' 出力完了")
