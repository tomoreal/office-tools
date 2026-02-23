import csv
import openpyxl
import os

input_file = '/home/tomo/work_office/download20260223_1843.csv'
output_file = '/home/tomo/work_office/download20260223_1843_横展開.xlsx'

dict_data = {}
dict_years = set()
dict_items_order = {}
dict_item_names = {} # current_type -> { unique_key: original_name }

current_year = ""
first_year = ""
current_type = ""
prev_item = ""
hierarchy_stack = []

# 厳密な表名称のリスト（アンカーのリセットを防ぐため）
TARGET_SHEET_NAMES = [
    "連結貸借対照表",
    "連結損益計算書",
    "連結包括利益計算書",
    "連結損益（及び包括利益）計算書",
    "連結キャッシュ・フロー計算書",
    "連結株主資本等変動計算書"
]

with open(input_file, encoding='cp932', errors='replace') as f:
    reader = csv.reader(f)
    for row in reader:
        if not row or len(row) < 1:
            continue
            
        raw_col_0 = row[0]
        col_0 = raw_col_0.strip()
        col_1 = row[1].strip() if len(row) > 1 else ""
        col_2 = row[2].strip() if len(row) > 2 else ""
        col_3 = row[3].strip() if len(row) > 3 else ""
        
        # 不要なヘッダーデータを強制除外（勘定科目に混入させない）
        if "現在" in col_0 and ("/" in col_0 or "年" in col_0):
            # 年度の判定
            current_year = col_0.replace("現在", "").strip()
            dict_years.add(current_year)
            if first_year == "":
                first_year = current_year
            continue
            
        if col_0 == "表名称":
            if col_1 in TARGET_SHEET_NAMES:
                current_type = col_1
                prev_item = ""
                hierarchy_stack = []
                if current_type not in dict_data:
                    dict_data[current_type] = {}
                    dict_items_order[current_type] = []
                    dict_item_names[current_type] = {}
            continue
            
        if col_0 in ["企業名", "証券ｺｰﾄﾞ", "（百万円）"] or ("/" in col_0 and "-" in col_0):
            continue
            
        # 表の開始行がA列にある場合の厳密判定
        if col_0 != "" and col_1 != "" and col_2 == "" and col_3 == "":
            if col_0 in TARGET_SHEET_NAMES:
                current_type = col_0
                prev_item = ""
                hierarchy_stack = []
                if current_type not in dict_data:
                    dict_data[current_type] = {}
                    dict_items_order[current_type] = []
                    dict_item_names[current_type] = {}
                continue
            
        # 勘定科目と金額の判定
        if col_0 != "" and current_type != "":
            item_name = raw_col_0.rstrip()
            amount = ""
            if col_2 != "" and (col_2.replace("-","").replace(",","").isdigit() or col_2 == "-"):
                amount = col_2
            elif col_3 != "" and (col_3.replace("-","").replace(",","").isdigit() or col_3 == "-"):
                amount = col_3
            
            # インデントレベルの計算 (全角・半角スペース、タブの混在に対応)
            stripped_len = len(raw_col_0.lstrip(' \t　'))
            indent_level = len(raw_col_0) - stripped_len
            
            # 階層スタックの更新
            while hierarchy_stack and hierarchy_stack[-1][0] >= indent_level:
                hierarchy_stack.pop()
            hierarchy_stack.append((indent_level, col_0))
            
            # 階層を持った一意なキーを作成 (例: "資産の部::流動資産::その他")
            unique_key = "::".join([x[1] for x in hierarchy_stack])
            
            # アイテム名の保存 (表示用に元の名前を保持)
            dict_item_names[current_type][unique_key] = item_name
            
            # アイテムオーダーに登録 (アンカー挿入ロジック・1年目保護版)
            if unique_key not in dict_items_order[current_type]:
                # 1年目（ベースライン作成期）は無条件で末尾に追加し、重複乱れを防ぐ
                if current_year != first_year and prev_item in dict_items_order[current_type]:
                    idx = dict_items_order[current_type].index(prev_item)
                    dict_items_order[current_type].insert(idx + 1, unique_key)
                else:
                    dict_items_order[current_type].append(unique_key)
            
            # 次の行のためのアンカー更新
            prev_item = unique_key
            
            if unique_key not in dict_data[current_type]:
                dict_data[current_type][unique_key] = {}
                
            if current_year != "" and amount != "":
                dict_data[current_type][unique_key][current_year] = amount

sorted_years = sorted(list(dict_years))

# Excelへ出力
wb = openpyxl.Workbook()
default_sheet = wb.active

for type_name in dict_data.keys():
    # 勘定科目が1件も抽出されていない空のシート（例：『連結損益（及び包括利益）計算書』の見出し指定のみ等）は出力しない
    items = dict_items_order[type_name]
    if len(items) == 0:
        continue
        
    safe_title = type_name.replace("・", "").replace(" ", "").replace("/", "")[:31]
    if safe_title == "":
        continue
    ws = wb.create_sheet(title=safe_title)
    
    ws.cell(row=1, column=1, value="勘定科目")
    for i, year in enumerate(sorted_years):
        ws.cell(row=1, column=i+2, value=year)
        
    for row_idx, unique_key in enumerate(items, start=2):
        # 画面に表示する名前は元の名前 (item_name) とする
        display_name = dict_item_names[type_name].get(unique_key, unique_key)
        ws.cell(row=row_idx, column=1, value=display_name)
        
        for col_idx, year in enumerate(sorted_years, start=2):
            val = dict_data[type_name][unique_key].get(year, "")
            try:
                if val != "":
                    numeric_val = float(val.replace(",", ""))
                    ws.cell(row=row_idx, column=col_idx, value=numeric_val)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)
            except ValueError:
                ws.cell(row=row_idx, column=col_idx, value=val)

if len(wb.sheetnames) > 1:
    wb.remove(default_sheet)

wb.save(output_file)
print(f"変換完了 (階層キーによる重複解消版): {output_file}")
