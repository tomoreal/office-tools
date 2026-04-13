"""
Cross-shareholdings (Policy-holding stocks) Analysis Module - V8 (Shareholders' Equity)
政策保有株式シート。分母を「連結純資産」から「連結株主資本」へ変更し、
見出しおよび参照タグ（ShareholdersEquity）を更新。
"""

import re
import unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _parse_value(val_str):
    """XBRL値文字列を float に変換。"""
    if val_str is None or val_str == '':
        return None
    s = unicodedata.normalize('NFKC', str(val_str)).replace(',', '').strip()
    s = s.replace('－', '-').replace('−', '-')
    if s == '-' or s == '':
        return None
    try:
        if '.' in s:
            return float(s)
        return int(s)
    except ValueError:
        return None


def _format_date(period_str):
    """YYYY-MM-DD -> YYYY/M/D"""
    if not period_str or not isinstance(period_str, str):
        return period_str
    parts = period_str.split('-')
    if len(parts) == 3:
        try:
            return f"{int(parts[0])}/{int(parts[1])}/{int(parts[2])}"
        except ValueError:
            pass
    return period_str


def _get_bs_info(wb):
    """
    連結株主資本（または同等の持分）のセル位置を検索する。
    """
    target_sheets = [
        '連結貸借対照表(日本基準)', '連結貸借対照表', 
        '連結財政状態計算書', '連結財政状態計算書(IFRS)',
        'ConsolidatedBalanceSheet', 'ConsolidatedStatementOfFinancialPosition'
    ]
    # 株主資本(ShareholdersEquity)を最優先とする。IFRSの場合は親会社所有者持分。
    target_tags = [
        'jppfs_cor_ShareholdersEquity', 
        'jpigp_cor_EquityAttributableToOwnersOfParent',
        'jppfs_cor_NetAssets', # Fallback
        'jpigp_cor_TotalEquity' # Fallback
    ]
    
    for sn in target_sheets:
        if sn in wb.sheetnames:
            ws = wb[sn]
            found_row = None
            for r in range(1, 500):
                if ws.cell(row=r, column=2).value in target_tags:
                    found_row = r
                    break
            
            if found_row:
                year_col_map = {}
                for c in range(3, 20):
                    header_val = str(ws.cell(row=1, column=c).value or "")
                    match = re.search(r'(\d{4})', header_val)
                    if match:
                        year_col_map[match.group(1)] = c
                return sn, found_row, year_col_map
                
    return None, None, {}


def add_cross_shareholdings_sheet(workbook, global_element_period_values, debug_log=None):
    """
    「政策保有株式」シートを生成して追加する。
    """
    if debug_log is None:
        def debug_log(msg): pass

    sheet_name = '政策保有株式'
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # --- 定義 ---
    metrics_def = [
        ('銘柄数', 'NumberOfIssues', '#,##0'),
        ('貸借対照表計上額', 'CarryingAmount', '#,##0'),
        ('増加銘柄数', 'NumberOfIssuesWhoseNumberOfSharesIncreased', '#,##0'),
        ('増加取得価額', 'TotalAcquisitionCostForIncreasedShares', '#,##0'),
        ('減少銘柄数', 'NumberOfIssuesWhoseNumberOfSharesDecreased', '#,##0'),
        ('減少売却価額', 'TotalSaleAmountForDecreasedShares', '#,##0'),
    ]

    groups = [
        ('非上場株式', 'SharesNotListedInvestmentSharesHeldForPurposesOtherThanPureInvestmentReportingCompany'),
        ('非上場株式以外の株式', 'SharesOtherThanThoseNotListedInvestmentSharesHeldForPurposesOtherThanPureInvestmentReportingCompany'),
    ]

    # --- スタイル ---
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    section_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 見出しを「連結株主資本」系に変更
    header_labels = ['株式区分', '年度'] + [m[0] for m in metrics_def] + ['連結株主資本', '連結株主資本比率']
    
    def write_header(row):
        for col_idx, text in enumerate(header_labels, start=1):
            cell = ws.cell(row=row, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    all_years = range(2015, 2026)
    
    # 株式統計抽出
    data_store = {}
    periods_seen = set()
    for g_label, g_suffix in groups:
        for m_label, m_el_prefix, m_fmt in metrics_def:
            full_el = f"jpcrp_cor_{m_el_prefix}{g_suffix}"
            data_store[full_el] = {}
            vals = global_element_period_values.get(full_el, {})
            for (fact_std, dim, period), raw_val in vals.items():
                if period:
                    v = _parse_value(raw_val)
                    if v is not None:
                        data_store[full_el][period] = v
                        periods_seen.add(period)

    # 連結株主資本抽出 (Fallback)
    se_tag_candidates = [
        'jppfs_cor_ShareholdersEquity', 
        'jpigp_cor_EquityAttributableToOwnersOfParent',
        'jppfs_cor_NetAssets',
        'jpigp_cor_TotalEquity'
    ]
    se_by_period = {}
    for tag in se_tag_candidates:
        if tag in global_element_period_values:
            vals = global_element_period_values[tag]
            found = False
            for (fact_std, dim, period), raw_val in vals.items():
                if not period: continue
                dim_s = str(dim) if dim else ""
                if not dim or dim == "" or "Consolidated" in dim_s or "提出会社" in dim_s:
                    v = _parse_value(raw_val)
                    if v is not None:
                        if period not in se_by_period:
                            se_by_period[period] = v
                            periods_seen.add(period)
                            found = True
            if found: break

    # BSシート情報取得
    bs_sheet, bs_row, year_col_map = _get_bs_info(workbook)
    
    data_row_map = {}
    current_row = 1
    BS_COL_LETTER = "D"
    SE_COL_LETTER = "I"

    def write_table_rows(target_groups, is_total=False):
        nonlocal current_row
        start_group_row = current_row
        
        for y in all_years:
            y_str = str(y)
            period_to_use = None
            matches = [p for p in periods_seen if p.startswith(y_str)]
            if matches: period_to_use = sorted(matches)[-1]
            display_date = _format_date(period_to_use) if period_to_use else f"{y}/3/31"
            
            label = '合計' if is_total else target_groups[0][0]
            if not is_total: data_row_map[(label, y_str)] = current_row
            
            ws.cell(row=current_row, column=1, value=label).font = section_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=2, value=display_date).border = thin_border
            ws.cell(row=current_row, column=2).alignment = center_align

            # 株式統計
            for col_idx, (m_l, m_pref, m_f) in enumerate(metrics_def, start=3):
                if is_total:
                    r1 = data_row_map.get(('非上場株式', y_str))
                    r2 = data_row_map.get(('非上場株式以外の株式', y_str))
                    col_l = get_column_letter(col_idx)
                    val = f'=IF(AND({col_l}{r1}="", {col_l}{r2}=""), "", SUM({col_l}{r1}, {col_l}{r2}))' if r1 and r2 else ""
                else:
                    full_el = f"jpcrp_cor_{m_pref}{target_groups[0][1]}"
                    val = data_store[full_el].get(period_to_use) if period_to_use else None
                
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.number_format = m_f
                cell.border = thin_border
                cell.alignment = right_align

            # I: 連結株主資本 (数式参照 or 値)
            bs_col_idx = year_col_map.get(y_str)
            if bs_sheet and bs_row and bs_col_idx:
                col_ref = get_column_letter(bs_col_idx)
                formula = f"=IF('{bs_sheet}'!{col_ref}{bs_row}=\"\", \"\", '{bs_sheet}'!{col_ref}{bs_row})"
                cell_se = ws.cell(row=current_row, column=9, value=formula)
            else:
                se_val = se_by_period.get(period_to_use) if period_to_use else None
                cell_se = ws.cell(row=current_row, column=9, value=se_val)
            
            cell_se.number_format = '#,##0'
            cell_se.border = thin_border
            cell_se.alignment = right_align

            # J: 連結株主資本比率
            formula_r = f'=IF(OR({BS_COL_LETTER}{current_row}="", {SE_COL_LETTER}{current_row}="", {SE_COL_LETTER}{current_row}=0), "", {BS_COL_LETTER}{current_row}/{SE_COL_LETTER}{current_row})'
            ws.cell(row=current_row, column=10, value=formula_r).number_format = '0.00%'
            ws.cell(row=current_row, column=10).border = thin_border
            ws.cell(row=current_row, column=10).alignment = right_align
            current_row += 1
            
        ws.merge_cells(start_row=start_group_row, end_row=current_row-1, start_column=1, end_column=1)

    # 各表の出力
    write_header(current_row); current_row += 1
    write_table_rows([groups[0]]); current_row += 2
    write_header(current_row); current_row += 1
    write_table_rows([groups[1]]); current_row += 2
    ws.cell(row=current_row, column=1, value='【合計 (非上場株式 ＋ 非上場株式以外)】').font = Font(bold=True, size=11)
    current_row += 1; write_header(current_row); current_row += 1
    write_table_rows(groups, is_total=True)

    # 後処理
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for i in range(3, 11):
        ws.column_dimensions[get_column_letter(i)].width = 22
    
    ws.freeze_panes = 'C2'
    debug_log(f"[CrossShare] '{sheet_name}' 株主資本へ変更完了")
