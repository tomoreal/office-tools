#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EDINET Taxonomy Dictionary Auto-Update Script

This script automatically downloads the latest EDINET taxonomy from FSA
and regenerates the edinet_taxonomy_dict.py file.

Features:
    - Automatic remote update detection using ETag/Last-Modified headers
    - Efficient HEAD request to check for updates without downloading
    - Conditional GET (If-None-Match/If-Modified-Since) for 304 optimization
    - SHA256 hash verification for downloaded files
    - Metadata tracking to avoid unnecessary downloads

Update Detection Strategy (Hybrid Approach):
    1. HEAD request to get remote ETag/Last-Modified
    2. Compare with local metadata (.edinet_taxonomy.meta)
    3. Download only if remote has changed
    4. Save old file hash BEFORE download
    5. Compare old hash vs new hash AFTER download
    6. Regenerate dictionary only if hash changed

Usage:
    python3 update_edinet_taxonomy.py [--force] [--debug]

Options:
    --force    Force update even if the file hasn't changed
               (skips remote check and hash verification)
    --debug    Enable debug-level logging (verbose output)

Exit Codes:
    0 - Success (updated or no update needed)
    1 - Failure (download error, generation error, etc.)
"""

import os
import sys
import urllib.request
import hashlib
import openpyxl
import json
import logging
from datetime import datetime
from contextlib import contextmanager

# Import fcntl for file locking (Unix/Linux/Mac)
# Fallback to no-op lock on Windows
try:
    import fcntl
    HAS_FCNTL = True
except ImportError:
    HAS_FCNTL = False

# Logging configuration
LOG_FILE = "update_edinet_taxonomy.log"

def setup_logging(debug=False):
    """Setup logging configuration with file and console output"""
    log_level = logging.DEBUG if debug else logging.INFO

    # Create formatters
    file_formatter = logging.Formatter(
        '%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_formatter = logging.Formatter('%(message)s')

    # File handler (always INFO or DEBUG)
    file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
    file_handler.setLevel(log_level)
    file_handler.setFormatter(file_formatter)

    # Console handler (INFO or DEBUG)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(log_level)
    console_handler.setFormatter(console_formatter)

    # Configure root logger
    logger = logging.getLogger()
    logger.setLevel(log_level)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logging.getLogger(__name__)

# EDINET Taxonomy URL
EDINET_TAXONOMY_URL = "https://disclosure2dl.edinet-fsa.go.jp/guide/static/disclosure/download/ESE140115.xlsx"
TAXONOMY_FILE = "edinet_taxonomy_elements.xlsx"
OUTPUT_FILE = "edinet_taxonomy_dict.py"
HASH_FILE = ".edinet_taxonomy.hash"
METADATA_FILE = ".edinet_taxonomy.meta"  # Stores ETag and Last-Modified
LOCK_FILE = ".edinet_taxonomy_update.lock"  # Process-level lock for update operations

# Logger will be initialized in main()
logger = None

@contextmanager
def file_lock(lock_path):
    """
    Cross-process file lock using fcntl (Unix/Linux/Mac).

    This prevents race conditions when multiple processes try to update
    the taxonomy dictionary simultaneously (e.g., cron jobs, manual runs).

    Args:
        lock_path: Path to lock file

    Yields:
        None

    Note:
        On Windows (where fcntl is unavailable), this becomes a no-op.
        Windows users should avoid running multiple update processes simultaneously.
    """
    lock_file = None
    try:
        if HAS_FCNTL:
            # Create lock file if it doesn't exist
            lock_file = open(lock_path, 'w')

            # Acquire exclusive lock (blocks until available)
            # LOCK_EX: exclusive lock
            # This will wait if another process holds the lock
            try:
                logger.debug(f"Acquiring file lock: {lock_path}")
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
                logger.debug(f"File lock acquired: {lock_path}")
            except IOError as e:
                logger.warning(f"Could not acquire file lock: {e}")
                # Continue anyway (non-fatal)

            yield
        else:
            # Windows: no file lock available
            # User must ensure no concurrent update processes
            logger.debug("fcntl not available, skipping file lock (ensure no concurrent updates on Windows)")
            yield
    finally:
        if lock_file:
            try:
                # Release lock and close file
                if HAS_FCNTL:
                    fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
                    logger.debug(f"File lock released: {lock_path}")
                lock_file.close()
            except:
                pass

def check_remote_update():
    """
    Check if remote file has been updated using ETag or Last-Modified headers.
    Returns:
        tuple: (needs_update: bool, metadata: dict)
    """
    logger.info("Checking for remote updates...")

    try:
        # Send HEAD request to get metadata without downloading
        req = urllib.request.Request(EDINET_TAXONOMY_URL, method='HEAD')
        with urllib.request.urlopen(req, timeout=10) as response:
            remote_etag = response.headers.get('ETag')
            remote_last_modified = response.headers.get('Last-Modified')
            remote_content_length = response.headers.get('Content-Length')

            remote_metadata = {
                'etag': remote_etag,
                'last_modified': remote_last_modified,
                'content_length': remote_content_length,
                'checked_at': datetime.now().isoformat()
            }

            logger.debug(f"  Remote ETag: {remote_etag or 'N/A'}")
            logger.debug(f"  Remote Last-Modified: {remote_last_modified or 'N/A'}")
            logger.debug(f"  Remote Size: {remote_content_length or 'N/A'} bytes")

            # Load local metadata if exists
            if os.path.exists(METADATA_FILE):
                try:
                    with open(METADATA_FILE, 'r') as f:
                        local_metadata = json.load(f)

                    # Compare ETag (most reliable)
                    if remote_etag and local_metadata.get('etag'):
                        if remote_etag == local_metadata['etag']:
                            logger.info("  ✓ ETag matches - no remote update")
                            return False, remote_metadata
                        else:
                            logger.info("  ✗ ETag changed - remote update detected")
                            return True, remote_metadata

                    # Fallback to Last-Modified
                    if remote_last_modified and local_metadata.get('last_modified'):
                        if remote_last_modified == local_metadata['last_modified']:
                            logger.info("  ✓ Last-Modified matches - no remote update")
                            return False, remote_metadata
                        else:
                            logger.info("  ✗ Last-Modified changed - remote update detected")
                            return True, remote_metadata

                    # Fallback to Content-Length
                    if remote_content_length and local_metadata.get('content_length'):
                        if remote_content_length != local_metadata['content_length']:
                            logger.info("  ✗ File size changed - remote update detected")
                            return True, remote_metadata
                        else:
                            logger.warning("  ⚠ No ETag/Last-Modified, but size unchanged - assuming no update")
                            return False, remote_metadata

                except json.JSONDecodeError:
                    logger.warning("  ⚠ Local metadata corrupted, assuming update needed")
                    return True, remote_metadata
            else:
                logger.warning("  ⚠ No local metadata found - first run or forced update")
                return True, remote_metadata

            # If we reach here, metadata exists but couldn't determine - assume update needed
            logger.warning("  ⚠ Could not determine update status - assuming update needed")
            return True, remote_metadata

    except Exception as e:
        logger.warning(f"  ⚠ Remote check failed ({e}), will proceed with download")
        return True, {}

def download_taxonomy(use_conditional_request=False, metadata=None):
    """
    Download EDINET taxonomy file with optional conditional request.
    Uses atomic rename to prevent partial file corruption.

    Args:
        use_conditional_request: If True, use If-None-Match or If-Modified-Since headers
        metadata: Previous metadata dict with ETag/Last-Modified

    Returns:
        tuple: (success: bool, was_modified: bool)
    """
    logger.info(f"Downloading EDINET taxonomy from: {EDINET_TAXONOMY_URL}")

    # Use temporary file for atomic download
    import tempfile
    temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix='edinet_taxonomy_')

    try:
        req = urllib.request.Request(EDINET_TAXONOMY_URL)

        # Add conditional headers if available
        if use_conditional_request and metadata:
            if metadata.get('etag'):
                req.add_header('If-None-Match', metadata['etag'])
                logger.debug(f"  Using If-None-Match: {metadata['etag']}")
            elif metadata.get('last_modified'):
                req.add_header('If-Modified-Since', metadata['last_modified'])
                logger.debug(f"  Using If-Modified-Since: {metadata['last_modified']}")

        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                # Download to temporary file first (atomic operation)
                with os.fdopen(temp_fd, 'wb') as f:
                    f.write(response.read())

                # Verify file was written successfully
                if not os.path.exists(temp_path) or os.path.getsize(temp_path) == 0:
                    raise IOError("Downloaded file is empty or missing")

                # Atomic rename: only replace original file if download succeeded
                # This prevents partial/corrupted files
                backup_path = None
                if os.path.exists(TAXONOMY_FILE):
                    # Backup old file before replacing (optional safety measure)
                    backup_path = TAXONOMY_FILE + '.bak'
                    if os.path.exists(backup_path):
                        os.remove(backup_path)
                    os.rename(TAXONOMY_FILE, backup_path)

                # Atomic rename from temp to final location
                os.rename(temp_path, TAXONOMY_FILE)

                # Remove backup if everything succeeded
                if backup_path and os.path.exists(backup_path):
                    os.remove(backup_path)

                logger.info(f"✓ Downloaded: {TAXONOMY_FILE}")

                # Save new metadata
                new_metadata = {
                    'etag': response.headers.get('ETag'),
                    'last_modified': response.headers.get('Last-Modified'),
                    'content_length': response.headers.get('Content-Length'),
                    'downloaded_at': datetime.now().isoformat()
                }
                save_metadata(new_metadata)

                return True, True

        except urllib.error.HTTPError as e:
            if e.code == 304:
                # 304 Not Modified - no need to download
                logger.info("✓ Remote file unchanged (304 Not Modified)")

                # Update metadata even on 304 to track last check time
                # The metadata passed in contains the latest ETag/Last-Modified from HEAD request
                if metadata:
                    updated_metadata = metadata.copy()
                    updated_metadata['checked_at'] = datetime.now().isoformat()
                    save_metadata(updated_metadata)

                # Clean up temp file
                try:
                    os.close(temp_fd)
                except:
                    pass
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return True, False
            else:
                raise

    except Exception as e:
        logger.error(f"✗ Download failed: {e}")
        # Clean up temp file on error
        try:
            os.close(temp_fd)
        except:
            pass
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass
        return False, False

def save_metadata(metadata):
    """Save remote file metadata (ETag, Last-Modified, etc.)"""
    try:
        with open(METADATA_FILE, 'w') as f:
            json.dump(metadata, f, indent=2)
        logger.debug(f"✓ Saved remote metadata")
    except Exception as e:
        logger.warning(f"⚠ Could not save metadata: {e}")

def load_metadata():
    """Load saved remote file metadata"""
    if os.path.exists(METADATA_FILE):
        try:
            with open(METADATA_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            return None
    return None

def calculate_file_hash(filename):
    """Calculate SHA256 hash of file"""
    if not os.path.exists(filename):
        return None

    sha256_hash = hashlib.sha256()
    with open(filename, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()

def check_if_file_changed_after_download(old_hash):
    """
    Check if downloaded file differs from the previous version.

    Args:
        old_hash: Hash of the file before download (or None if no previous file)

    Returns:
        bool: True if file has changed (or is new), False if unchanged
    """
    if not os.path.exists(TAXONOMY_FILE):
        logger.error("✗ Downloaded file not found (unexpected error)")
        return False

    new_hash = calculate_file_hash(TAXONOMY_FILE)

    if old_hash is None:
        logger.info("✓ New file downloaded (no previous version)")
        return True

    if new_hash == old_hash:
        logger.info("✓ Downloaded file is identical to previous version (hash match)")
        return False

    logger.info(f"✓ File content changed")
    logger.debug(f"  Old hash: {old_hash[:16]}...")
    logger.debug(f"  New hash: {new_hash[:16]}...")
    return True

def get_current_file_hash():
    """Get hash of current taxonomy file (before download)"""
    if not os.path.exists(TAXONOMY_FILE):
        return None
    return calculate_file_hash(TAXONOMY_FILE)

def save_hash():
    """Save current hash of taxonomy file"""
    current_hash = calculate_file_hash(TAXONOMY_FILE)
    with open(HASH_FILE, 'w') as f:
        f.write(current_hash)
    logger.debug(f"✓ Saved taxonomy hash: {current_hash[:16]}...")

def get_column_index_map(ws, header_row=2):
    """
    Create column name to index mapping from header row.
    This makes the code resilient to column reordering.

    Args:
        ws: openpyxl worksheet
        header_row: Row number containing headers (1-indexed)

    Returns:
        dict: {column_name: index}
    """
    headers = [cell.value for cell in ws[header_row]]
    idx_map = {name: i for i, name in enumerate(headers) if name}
    return idx_map

def generate_dictionary():
    """Generate dictionary from EDINET taxonomy (all industry sheets)"""
    logger.info(f"Generating dictionary from: {TAXONOMY_FILE}")

    try:
        wb = openpyxl.load_workbook(TAXONOMY_FILE, data_only=True)

        # Skip metadata sheets (not taxonomy data)
        SKIP_SHEETS = ['目次', '勘定科目リストについて']

        # Extract from ALL industry sheets (not just '一般商工業')
        # This ensures we capture industry-specific elements (banking, insurance, etc.)
        edinet_dict = {}
        sheets_processed = []
        namespace_stats = {}  # Track namespace usage for transparency

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            sheet_count = 0

            # Build column index map from header (row 2)
            # This makes code resilient to column reordering
            idx_map = get_column_index_map(ws, header_row=2)

            # Validate required columns exist
            required_columns = ['要素名', '名前空間プレフィックス', '標準ラベル（日本語）']
            missing_columns = [col for col in required_columns if col not in idx_map]
            if missing_columns:
                logger.warning(f"  ⚠ Skipping sheet '{sheet_name}': Missing columns {missing_columns}")
                continue

            # Namespace filtering: Use BLACKLIST instead of whitelist
            # This allows IFRS, extensions, and future taxonomies
            #
            # Design rationale:
            # - EDINET taxonomy may add new namespaces in the future (jpigp_cor, ifrs_full, etc.)
            # - Whitelist approach would require code changes for each taxonomy update
            # - Blacklist approach automatically accepts valid namespaces from EDINET
            # - Company-specific extensions (e.g., jpcrp030000-asr_E01225-000) won't match
            #   during XBRL parsing anyway (element names differ)
            NAMESPACE_BLACKLIST = {
                '名前空間プレフィックス',  # Header itself (not actual data)
                None,                      # Empty namespace
                '',                        # Empty string
                # Add more here if needed (e.g., internal test namespaces)
            }

            for row in ws.iter_rows(min_row=3, values_only=True):
                # Use header-based indexing instead of hard-coded positions
                element_name = row[idx_map['要素名']]
                namespace = row[idx_map['名前空間プレフィックス']]
                jp_label = row[idx_map['標準ラベル（日本語）']]

                # Apply blacklist filter (more permissive than whitelist)
                # This captures: jppfs_cor, jpigp_cor, ifrs_full, jpcrp_cor, extensions, etc.
                if not element_name or not jp_label or namespace in NAMESPACE_BLACKLIST:
                    continue

                # Skip if already exists (first occurrence wins - usually from '一般商工業')
                if element_name in edinet_dict:
                    continue

                label = str(jp_label)
                # Shorten labels with "or loss" notation
                if '又は' in label and ('損失' in label or '損' in label) and '（△）' in label:
                    parts = label.split('又は')
                    if len(parts) == 2:
                        label = parts[0].strip()
                edinet_dict[element_name] = label
                sheet_count += 1

                # Track namespace usage
                namespace_stats[namespace] = namespace_stats.get(namespace, 0) + 1

            if sheet_count > 0:
                sheets_processed.append(f"{sheet_name}({sheet_count})")

        logger.info(f"✓ Extracted {len(edinet_dict)} items from EDINET taxonomy")
        logger.info(f"  Processed {len(sheets_processed)} sheets:")
        # Show detailed sheet statistics
        for i, sheet_info in enumerate(sheets_processed):
            if i < 10:  # Show first 10 sheets
                logger.info(f"    - {sheet_info}")
        if len(sheets_processed) > 10:
            logger.info(f"    ... and {len(sheets_processed) - 10} more sheets")

        # Show namespace statistics (for transparency)
        if namespace_stats:
            logger.debug(f"  Namespaces found:")
            for ns, count in sorted(namespace_stats.items(), key=lambda x: x[1], reverse=True):
                logger.debug(f"    - {ns}: {count} elements")

        # Custom mappings (IFRS variants, abbreviations, etc.)
        custom_mappings = {
            # Generic abbreviations
            'Notes': '注記',
            'Inventory': '棚卸資産',

            # Financial statement names
            'ConsolidatedBalanceSheet': '連結貸借対照表',
            'ConsolidatedStatementOfIncome': '連結損益計算書',
            'ConsolidatedStatementOfCashFlows': '連結キャッシュ・フロー計算書',
            'ConsolidatedStatementOfChangesInEquity': '連結株主資本等変動計算書',
            'ConsolidatedStatementOfFinancialPosition': '連結財政状態計算書',
            'ConsolidatedStatementOfProfitOrLoss': '連結損益計算書',
            'ConsolidatedStatementOfFinancialPositionIFRS': '連結財政状態計算書',
            'ConsolidatedStatementOfProfitOrLossIFRS': '連結損益計算書',
            'ConsolidatedStatementOfCashFlowsIFRS': '連結キャッシュ・フロー計算書',
            'ConsolidatedStatementOfChangesInEquityIFRS': '連結株主資本等変動計算書',
            'ConsolidatedStatementOfComprehensiveIncomeIFRS': '連結包括利益計算書',

            # IFRS variants / Abbreviations
            'NetSalesIFRS': '売上収益',
            'RevenueIFRS': '売上収益',
            'CostOfSalesIFRS': '売上原価',
            'GrossProfitIFRS': '売上総利益',
            'SellingGeneralAndAdministrativeExpensesIFRS': '販売費及び一般管理費',
            'OtherOperatingIncome': 'その他の営業収益',
            'OtherOperatingExpenses': 'その他の営業費用',
            'OtherOperatingExpense': 'その他の営業費用',
            'OtherIncomeIFRS': 'その他の収益',
            'OtherExpensesIFRS': 'その他の費用',
            'OtherOperatingIncomeIFRS': 'その他の営業収益',
            'OtherOperatingExpensesIFRS': 'その他の営業費用',
            'ShareOfProfitLossOfInvestmentsAccountedForUsingEquityMethodIFRS': '持分法による投資利益',
            'OperatingProfitLossIFRS': '営業利益',
            'FinanceIncomeIFRS': '金融収益',
            'FinanceCostsIFRS': '金融費用',
            'ProfitLossBeforeTaxIFRS': '税引前当期利益',
            'IncomeTaxExpenseIFRS': '法人所得税費用',
            'ProfitLossIFRS': '当期利益',
            'ProfitLossAttributableToOwnersOfParentIFRS': '親会社の所有者に帰属する当期利益',
            'ProfitLossAttributableToNonControllingInterestsIFRS': '非支配持分',
            'BasicEarningsPerShareIFRS': '基本的１株当たり当期利益（円）',

            # J-GAAP variants / Abbreviations
            'OperatingProfit': '営業利益',
            'FinanceIncome': '金融収益',
            'FinancialIncome': '金融収益',
            'FinanceCosts': '金融費用',
            'FinancialExpenses': '金融費用',
            'FinanceExpenses': '金融費用',
            'ProfitBeforeTax': '税引前利益',
            'IncomeTaxExpense': '法人所得税費用',
            'Profit': '当期利益',
            'NetIncome': '当期利益',
            'ProfitLossAttributableToAbstract': '当期利益の帰属',
            'ProfitAttributableToOwnersOfParent': '親会社株主に帰属する当期純利益',
            'ProfitAttributableToNoncontrollingInterests': '非支配持分',
            'ProfitLossAttributableToNoncontrollingInterests': '非支配持分',
            'BasicEarningsPerShare': '基本的１株当たり当期純利益（円）',
            'BasicEarningsLossPerShare': '基本的１株当たり当期純利益（円）',
            'SellingGeneralAndAdministrativeExpense': '販売費及び一般管理費',
            'ShareOfProfitLossOfAssociatesAndJointVenturesAccountedForUsingEquityMethod': '持分法による投資利益',
        }

        # Merge dictionaries with priority control
        # IMPORTANT: EDINET official definitions take priority over custom mappings
        # Custom mappings are only used when EDINET doesn't define the element
        final_dict = {}

        # Step 1: Add all EDINET definitions (highest priority)
        final_dict.update(edinet_dict)

        # Step 2: Add custom mappings only if NOT already defined by EDINET
        # This prevents custom mappings from overwriting official definitions
        custom_added = 0
        custom_skipped = 0
        for key, value in custom_mappings.items():
            if key not in final_dict:
                final_dict[key] = value
                custom_added += 1
            else:
                # Key already exists in EDINET - skip to preserve official definition
                custom_skipped += 1

        logger.info(f"✓ Total items: {len(final_dict)} (EDINET: {len(edinet_dict)}, Custom: {custom_added})")
        if custom_skipped > 0:
            logger.warning(f"  ⚠ Skipped {custom_skipped} custom mappings (already defined in EDINET)")

        return final_dict, len(edinet_dict), len(custom_mappings)

    except Exception as e:
        logger.error(f"✗ Dictionary generation failed: {e}")
        import traceback
        traceback.print_exc()
        return None, 0, 0

def write_dictionary_file(final_dict, edinet_count, custom_count):
    """Write dictionary to Python file"""
    logger.info(f"Writing dictionary to: {OUTPUT_FILE}")

    try:
        with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
            f.write('#!/usr/bin/env python3\n')
            f.write('# -*- coding: utf-8 -*-\n')
            f.write('"""\n')
            f.write('EDINET Taxonomy Dictionary\n')
            f.write('\n')
            f.write('Auto-generated from EDINET Official Taxonomy:\n')
            f.write(f'{EDINET_TAXONOMY_URL}\n')
            f.write('\n')
            f.write(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
            f.write(f'Total: {len(final_dict):,} items\n')
            f.write(f'- EDINET Official Taxonomy: {edinet_count:,} items\n')
            f.write(f'- Custom Mappings (IFRS variants, etc.): {custom_count} items\n')
            f.write('"""\n')
            f.write('\n')
            f.write('# EDINET Taxonomy common dictionary\n')
            f.write('# Maps element names to Japanese account labels\n')
            f.write('common_dict = {\n')

            # Write entries in insertion order (preserves Excel's semantic order)
            # This maintains the original accounting structure from EDINET taxonomy
            # (Balance Sheet: Assets -> Liabilities -> Equity, etc.)
            for key in final_dict.keys():
                value = final_dict[key]
                # Escape single quotes
                value_escaped = value.replace("'", "\\'")
                f.write(f"    '{key}': '{value_escaped}',\n")

            f.write('}\n')

        logger.info(f"✓ Dictionary written to: {OUTPUT_FILE}")
        return True

    except Exception as e:
        logger.error(f"✗ File write failed: {e}")
        return False

def main():
    """Main function"""
    global logger

    # Parse command-line arguments
    force = '--force' in sys.argv
    debug = '--debug' in sys.argv

    # Setup logging (must be done first)
    logger = setup_logging(debug=debug)

    logger.info("=" * 80)
    logger.info("EDINET Taxonomy Dictionary Auto-Update")
    logger.info("=" * 80)
    logger.info("")

    # Acquire file lock to prevent concurrent update processes
    # This protects against race conditions when:
    # - Multiple cron jobs run simultaneously
    # - Manual update runs while cron job is active
    # - convert_xbrl_to_excel.py reads files being updated
    with file_lock(LOCK_FILE):
        return _main_locked(force, debug)

def _main_locked(force, debug):
    """Main function implementation (runs under file lock)"""
    global logger

    # Step 1: Check for remote updates (unless file doesn't exist)
    needs_download = False
    remote_metadata = None

    if not os.path.exists(TAXONOMY_FILE):
        logger.info("Local taxonomy file not found - download required")
        needs_download = True
    elif force:
        logger.info("Force update requested - skipping remote check")
        needs_download = True
    else:
        # Check if remote file has been updated
        needs_remote_update, remote_metadata = check_remote_update()
        if needs_remote_update:
            logger.info("Remote update detected - download required")
            needs_download = True
        else:
            logger.info("Remote file unchanged - no download needed")
            # Save metadata even when no update needed to track last check time
            if remote_metadata:
                save_metadata(remote_metadata)

    # Step 2: Download if needed
    file_changed = False
    if needs_download:
        # CRITICAL: Save hash of old file BEFORE download
        old_hash = get_current_file_hash()
        logger.debug(f"Old file hash: {old_hash[:16] + '...' if old_hash else 'N/A (no previous file)'}")

        # Load existing metadata for conditional request (304 optimization)
        existing_metadata = load_metadata()
        success, was_modified = download_taxonomy(
            use_conditional_request=not force,
            metadata=existing_metadata
        )

        if not success:
            logger.error("\n✗ Update failed: Could not download taxonomy file")
            return 1

        if not was_modified and not force:
            logger.info("\n✓ No update needed (304 Not Modified)")
            return 0

        # Step 3: Verify if downloaded file differs from old file (hash comparison)
        if force:
            logger.info("\nForce mode: Skipping hash verification")
            file_changed = True  # Force regeneration
        else:
            logger.info("\nVerifying downloaded file...")
            file_changed = check_if_file_changed_after_download(old_hash)

            if not file_changed:
                logger.info("\n✓ No update needed (downloaded file is identical to previous version)")
                return 0

    # If we didn't download, check if we even have a file
    elif not os.path.exists(TAXONOMY_FILE):
        logger.error("\n✗ No taxonomy file found and no download performed")
        return 1
    else:
        # No download needed, file exists, assume it's already processed
        logger.info("\n✓ No update needed (remote unchanged, local file exists)")
        return 0

    # Step 4: Generate dictionary (only if file changed or force)
    logger.info("")
    logger.info("Generating dictionary from updated taxonomy file...")
    final_dict, edinet_count, custom_count = generate_dictionary()
    if final_dict is None:
        logger.error("\n✗ Update failed: Could not generate dictionary")
        return 1

    # Step 5: Write dictionary file
    if not write_dictionary_file(final_dict, edinet_count, custom_count):
        logger.error("\n✗ Update failed: Could not write dictionary file")
        return 1

    # Step 6: Save hash of new file
    save_hash()

    logger.info("")
    logger.info("=" * 80)
    logger.info("✓ Dictionary update completed successfully!")
    logger.info("=" * 80)

    return 0

if __name__ == '__main__':
    sys.exit(main())
