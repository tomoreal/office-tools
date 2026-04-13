"""
Number of Shareholders Analysis Module for XBRL to Excel Conversion - V3 (Diversity Analysis Method)

株式の所有者別状況（株主数、所有株式数、割合）のシートを生成するモジュール
diversity_analysis.py のデータ抽出手法を再利用して、コンテキストの優先順位を制御します。
"""

import re
import unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _parse_value(val_str):
    """XBRL値文字列を float に変換。変換失敗時は None。 (diversity_analysis.py準拠)"""
    if val_str is None or val_str == '':
        return None
    # Normalize NFKC to handle commas and full-width characters
    s = unicodedata.normalize('NFKC', str(val_str)).replace(',', '').strip()
    # Normalize full-width minus/hyphen
    s = s.replace('－', '-').replace('−', '-')
    if s == '-' or s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _format_date(period_str):
    """YYYY-MM-DD -> YYYY/M/D (ユーザー指定形式)"""
    if not period_str or not isinstance(period_str, str):
        return period_str
    parts = period_str.split('-')
    if len(parts) == 3:
        try:
            return f"{int(parts[0])}/{int(parts[1])}/{int(parts[2])}"
        except ValueError:
            pass
    return period_str


def _get_best_value_for_element(global_element_period_values, el_name, debug_log=None):
    """
    指定された要素について、各年度(period)で最も適切なコンテキスト(dim_label)の値を取得する。
    diversity_analysis.py の提出会社データ収集ロジックを一般化。
    """
    vals = global_element_period_values.get(el_name, {})
    if not vals:
        return {}
        
    period_value_map = {} # {period: (value, dim_label)}
    for (fact_std, dim_label, period), raw_val in vals.items():
        if not period:
            continue
        v = _parse_value(raw_val)
        if v is None:
            continue
            
        dim_l = str(dim_label).lower()
        # 優先順位: 1. 普通株式 2. 全体/単体 3. その他
        if period not in period_value_map:
            period_value_map[period] = (v, dim_label)
        else:
            _, existing_dim = period_value_map[period]
            existing_dim_l = str(existing_dim).lower()
            
            # 現在のdim_labelが「普通株式」または「Ordinary」を含む場合、最優先
            is_current_ordinary = ('普通株式' in dim_label or 'ordinary' in dim_l)
            is_existing_ordinary = ('普通株式' in existing_dim or 'ordinary' in existing_dim_l)
            
            if is_current_ordinary and not is_existing_ordinary:
                period_value_map[period] = (v, dim_label)
            elif (is_current_ordinary == is_existing_ordinary): # 両方Ordinary、または両方非Ordinary
                # 次に「全体」「単体」を優先
                if dim_label in ('全体', '単体') and existing_dim not in ('全体', '単体'):
                    period_value_map[period] = (v, dim_label)
                # その他、同一条件なら何もしない（最初の方を優先）

    return {p: val[0] for p, val in period_value_map.items()}


def add_number_of_shareholders_sheet(workbook, global_element_period_values, debug_log=None):
    """
    「株式の所有者別状況」シートを生成して追加する。
    """
    if debug_log is None:
        def debug_log(msg): pass

    sheet_name = '株式の所有者別状況'
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # --- カテゴリー定義 ---
    categories = [
        ('政府及び地方公共団体', 'NationalAndLocalGovernments'),
        ('金融機関', 'FinancialInstitutions'),
        ('金融商品取引業者', 'FinancialServiceProviders'),
        ('その他の法人', 'OtherCorporations'),
        ('外国法人等', 'ForeignInvestors', 'Foreigners'), # Total
        ('個人以外', 'ForeignInvestorsOtherThanIndividuals', 'ForeignersOtherThanIndividuals'),
        ('個人', 'ForeignIndividualInvestors', 'ForeignIndividuals'),
        ('個人その他', 'IndividualsAndOthers'),
        ('計', 'Total'),
    ]

    # --- セクション定義 ---
    groups = [
        ('株主数（人）', 'NumberOfShareholders', '#,##0'),
        ('所有株式数（単元）', 'NumberOfSharesHeldNumberOfUnits', '#,##0'),
        ('所有株式数の割合（％）', 'PercentageOfShareholdings', '0.00%'),
    ]

    # --- スタイル準備 ---
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    section_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
    section_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ヘッダー描画
    ws.cell(row=1, column=1, value='区分').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border

    for col_idx, (cat_name, *_) in enumerate(categories, start=2):
        cell = ws.cell(row=1, column=col_idx, value=cat_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 全期間を収集 (2015-2025を基本軸とする)
    years = range(2015, 2026)
    
    # 実際にあるデータのアグリゲーション
    # extracted_data[el_prefix][suffix][period] = value
    data_store = {}
    periods_seen = set()

    for group_label, el_prefix, num_fmt in groups:
        data_store[el_prefix] = {}
        for col_idx, (cat_name, *el_suffixes) in enumerate(categories, start=2):
            suffix = el_suffixes[0]
            if el_prefix == 'PercentageOfShareholdings' and len(el_suffixes) > 1:
                suffix = el_suffixes[1]
            
            el_name = f"jpcrp_cor_{el_prefix}{suffix}"
            best_vals = _get_best_value_for_element(global_element_period_values, el_name, debug_log)
            data_store[el_prefix][suffix] = best_vals
            periods_seen.update(best_vals.keys())
            if best_vals:
                debug_log(f"[Shareholder] {el_name} 抽出: {len(best_vals)} 年度分")

        # 外国法人等の合算フォールバック
        for p in periods_seen:
            # ForeignInvestors (Total) がない場合
            total_suffix = 'ForeignInvestors' if el_prefix != 'PercentageOfShareholdings' else 'Foreigners'
            if p not in data_store[el_prefix].get(total_suffix, {}):
                v1_suffix = 'ForeignInvestorsOtherThanIndividuals' if el_prefix != 'PercentageOfShareholdings' else 'ForeignersOtherThanIndividuals'
                v2_suffix = 'ForeignIndividualInvestors' if el_prefix != 'PercentageOfShareholdings' else 'ForeignIndividuals'
                v1 = data_store[el_prefix].get(v1_suffix, {}).get(p)
                v2 = data_store[el_prefix].get(v2_suffix, {}).get(p)
                if v1 is not None or v2 is not None:
                    if total_suffix not in data_store[el_prefix]: data_store[el_prefix][total_suffix] = {}
                    data_store[el_prefix][total_suffix][p] = (v1 or 0) + (v2 or 0)

    # 行の作成
    current_row = 2
    for group_label, el_prefix, num_fmt in groups:
        # セクションヘッダー
        cell = ws.cell(row=current_row, column=1, value=group_label)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(1, len(categories) + 2):
            ws.cell(row=current_row, column=c).fill = section_fill
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1
        
        # 年度行
        for y in years:
            # 表示年度に対応する実際の期間(YYYY-MM-DD)を探す
            period_to_use = None
            matches = [p for p in periods_seen if p.startswith(str(y))]
            if matches:
                period_to_use = sorted(matches)[-1] # 当年の決算期末
            
            display_date = _format_date(period_to_use) if period_to_use else f"{y}/3/31"
            ws.cell(row=current_row, column=1, value=display_date).font = normal_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border

            for col_idx, (cat_name, *el_suffixes) in enumerate(categories, start=2):
                suffix = el_suffixes[0]
                if el_prefix == 'PercentageOfShareholdings' and len(el_suffixes) > 1:
                    suffix = el_suffixes[1]
                
                val = None
                if period_to_use:
                    val = data_store[el_prefix].get(suffix, {}).get(period_to_use)

                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = normal_font
                cell.alignment = right_align
                cell.number_format = num_fmt
                cell.border = thin_border
            current_row += 1
        current_row += 1

    # 後処理
    ws.column_dimensions['A'].width = 15
    for i in range(2, 11):
        ws.column_dimensions[chr(64 + i)].width = 20
    ws.freeze_panes = 'B2'
    debug_log(f"[NumberOfShareholders] '{sheet_name}' 出力完了")
