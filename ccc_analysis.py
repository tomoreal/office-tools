import openpyxl.utils

def add_ccc_analysis_sheets(workbook, debug_log=None):
    """
    CCC分析シートを生成する
    """
    if debug_log is None:
        def debug_log(msg):
            pass

    # 判定対象の基準ペアリスト (貸借対照表, 損益計算書)
    standards_to_try = [
        ('連結貸借対照表(日本基準)', '連結損益計算書(日本基準)'),
        ('連結財政状態計算書(IFRS)', '連結損益計算書(IFRS)'),
        ('連結貸借対照表(IFRS)', '連結損益計算書(IFRS)'),
        ('連結貸借対照表(US GAAP)', '連結損益計算書(US GAAP)'),
        ('連結貸借対照表(JMIS)', '連結損益計算書(JMIS)'),
        # 単体
        ('貸借対照表', '損益計算書'),
    ]

    processed_any = False
    for bs_sheet_name, pl_sheet_name in standards_to_try:
        if bs_sheet_name in workbook.sheetnames and pl_sheet_name in workbook.sheetnames:
            _perform_ccc_analysis(workbook, bs_sheet_name, pl_sheet_name, debug_log)
            processed_any = True

    if not processed_any:
        debug_log("CCC analysis skipped: No matching Balance Sheet/Profit and Loss sheets found.")

def _perform_ccc_analysis(workbook, bs_sheet_name, pl_sheet_name, debug_log):
    """
    具体的なCCC分析処理（内部関数）
    """

    bs_ws = workbook[bs_sheet_name]
    pl_ws = workbook[pl_sheet_name]

    # CCC分析シート名 (BSシート名にサフィックスを追加)
    analysis_sheet_name = f"{bs_sheet_name}_分析_CCC"
    
    if len(analysis_sheet_name) > 31:
        # シート名が31文字を超える場合の処理
        # 「分析_CCC」を優先して残し、前半をカット
        suffix = "_分析_CCC"
        allowed_base_len = 31 - len(suffix)
        analysis_sheet_name = f"{bs_sheet_name[:allowed_base_len]}{suffix}"

    if analysis_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[analysis_sheet_name])

    ccc_ws = workbook.create_sheet(analysis_sheet_name)

    max_col = bs_ws.max_column

    # 1. 1行目 ヘッダーの作成
    header_row = ["分類", "勘定科目", "項目（英名）"]
    for col in range(3, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        header_row.append(f"='{bs_sheet_name}'!{col_letter}1")
    ccc_ws.append(header_row)

    # 2. 売上高の抽出 (PL)
    sales_keywords_jp = ['NetSales', 'Sales']
    sales_keywords_ifrs = ['Revenue']
    sales_keywords = sales_keywords_jp + sales_keywords_ifrs
    sales_row_nums = []
    for row in range(2, pl_ws.max_row + 1):
        b_val = pl_ws.cell(row, 2).value
        # PLのB列は英語名
        if b_val:
            s_val = str(b_val)
            if 'Abstract' in s_val or 'Heading' in s_val or 'LineItems' in s_val:
                continue
            if any(kw in s_val for kw in sales_keywords):
                sales_row_nums.append(row)
            
    if not sales_row_nums:
        debug_log(f"CCC analysis skipped: NetSales/Revenue not found in {pl_sheet_name}")
        return

    # 優先順位付け: IFRS(Revenue)を優先する (年度の新しいほうを採用)
    def sales_priority(row_num):
        b_val = str(pl_ws.cell(row_num, 2).value or "")
        if any(kw in b_val for kw in sales_keywords_ifrs):
            return 0  # High priority
        return 1  # Lower priority
    
    sales_row_nums.sort(key=sales_priority)
    debug_log(f"CCC Sales rows (prioritized): {sales_row_nums}")

    # PLシートの年度(1行目の値)と列のマッピングを作成
    pl_date_to_col = {}
    for col in range(3, pl_ws.max_column + 1):
        dt = pl_ws.cell(1, col).value
        if dt:
            pl_date_to_col[str(dt).strip()] = openpyxl.utils.get_column_letter(col)

    # 売上高行を追加 (対応する年度の列を参照)
    # 複数行ある場合は優先順位に従って「年度の新しいほう（IFRS等）」を採用する (Deduplication)
    label = f"=TRIM('{pl_sheet_name}'!A{sales_row_nums[0]})"
    names = " / ".join([f"'{pl_sheet_name}'!B{r}" for r in sales_row_nums])
    row_data = ["売上高", label, f"={names}" if len(sales_row_nums) > 1 else f"='{pl_sheet_name}'!B{sales_row_nums[0]}"]
    
    for col in range(3, max_col + 1):
        bs_dt = bs_ws.cell(1, col).value
        bs_date_key = str(bs_dt).strip() if bs_dt is not None else None
        
        pl_col_letter = pl_date_to_col.get(bs_date_key) if bs_date_key else None
        if pl_col_letter:
            if len(sales_row_nums) > 1:
                # 優先順位に従った数式を作成: IF(収益<>"", 収益, 売上高)
                # 3つ以上ある場合はネスト
                formula = f"'{pl_sheet_name}'!{pl_col_letter}{sales_row_nums[-1]}"
                for r in reversed(sales_row_nums[:-1]):
                    curr_cell = f"'{pl_sheet_name}'!{pl_col_letter}{r}"
                    formula = f"IF({curr_cell}<>\"\",{curr_cell},{formula})"
                row_data.append(f"={formula}")
            else:
                row_data.append(f"='{pl_sheet_name}'!{pl_col_letter}{sales_row_nums[0]}")
        else:
            row_data.append("")
    ccc_ws.append(row_data)

    ccc_sales_data_row = ccc_ws.max_row  # 売上高の行番号（2のはず）

    # 3. 棚卸資産、売上債権、仕入債務の抽出 (BS)
    inventory_kws_jp = ['MerchandiseAndFinishedGoods', 'WorkInProcess', 'RawMaterialsAndSupplies', 'Inventories', 'Goods', 'SemiFinishedGoods', 'Merchandise', 'FinishedGoods']
    inventory_kws_ifrs = [] # InventoriesはJP側でもカバー
    inventory_kws = inventory_kws_jp + inventory_kws_ifrs

    receivable_kws_jp = ['NotesAndAccountsReceivableTrade', 'AccountsReceivableTrade', 'NotesReceivableTrade', 'ElectronicallyRecordedMonetaryClaimsOperating', 'ContractAssets']
    receivable_kws_ifrs = ['TradeAndOtherReceivables', 'TradeReceivables', 'OtherReceivables', 'ContractAssets']
    receivable_kws = receivable_kws_jp + receivable_kws_ifrs

    payable_kws_jp = ['NotesAndAccountsPayableTrade', 'AccountsPayableTrade', 'NotesPayableTrade', 'ElectronicallyRecordedObligationsOperating']
    payable_kws_ifrs = ['TradeAndOtherPayables', 'TradePayables', 'OtherPayables']
    payable_kws = payable_kws_jp + payable_kws_ifrs

    def get_category(eng_name):
        if 'Abstract' in eng_name:
            return None
        if any(kw in eng_name for kw in inventory_kws): return "棚卸資産"
        if any(kw in eng_name for kw in receivable_kws): return "売上債権"
        if any(kw in eng_name for kw in payable_kws): return "仕入債務"
        return None

    # カテゴリごとに行をまとめる
    inventory_rows = []
    receivable_rows = []
    payable_rows = []

    for row in range(2, bs_ws.max_row + 1):
        b_val = bs_ws.cell(row, 2).value
        if b_val:
            cat = get_category(str(b_val))
            if cat == "棚卸資産":
                inventory_rows.append(row)
            elif cat == "売上債権":
                receivable_rows.append(row)
            elif cat == "仕入債務":
                payable_rows.append(row)

    categories_to_add = [
        ("棚卸資産", inventory_rows),
        ("売上債権", receivable_rows),
        ("仕入債務", payable_rows)
    ]

    # CCC計算用に行を記憶
    category_start_end_rows = {} # { category: (start_row, end_row) }

    for cat_name, rows in categories_to_add:
        if not rows:
            continue
        start_r = ccc_ws.max_row + 1
        for bs_row in rows:
            row_data = [cat_name, f"=TRIM('{bs_sheet_name}'!A{bs_row})", f"='{bs_sheet_name}'!B{bs_row}"]
            for col in range(3, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                row_data.append(f"='{bs_sheet_name}'!{col_letter}{bs_row}")
            ccc_ws.append(row_data)
        category_start_end_rows[cat_name] = (start_r, ccc_ws.max_row)

    # 空行
    ccc_ws.append([''] * max_col)

    # 4. 合計行の追加
    sum_rows = {}
    for cat_name in ["棚卸資産", "売上債権", "仕入債務"]:
        current_r = ccc_ws.max_row + 1
        row_data = ["", cat_name, ""] # 分類空白、勘定科目にカテゴリ名
        for col in range(3, max_col + 1):
            new_col_letter = openpyxl.utils.get_column_letter(col + 1)
            row_data.append(f"=SUMIFS({new_col_letter}:{new_col_letter}, $A:$A, $B{current_r})")
        ccc_ws.append(row_data)
        sum_rows[cat_name] = current_r

    # 空行
    ccc_ws.append([''] * max_col)

    # 5. 回転期間・CCC行の追加
    # 売上債権回転期間、仕入債務回転期間、棚卸資産回転期間、CCC
    calc_rows = {}
    for calc_name in ["棚卸資産回転期間", "売上債権回転期間", "仕入債務回転期間", "CCC"]:
        row_data = ["", calc_name, ""]
        current_r = ccc_ws.max_row + 1
        
        target_sum_row = sum_rows.get(calc_name.replace("回転期間", ""), None)
        
        for col in range(3, max_col + 1):
            new_col_letter = openpyxl.utils.get_column_letter(col + 1)
            if calc_name == "CCC":
                # 売上債権回転期間 - 仕入債務回転期間 + 棚卸資産回転期間
                r_receiv = calc_rows.get("売上債権回転期間")
                r_payable = calc_rows.get("仕入債務回転期間")
                r_inv = calc_rows.get("棚卸資産回転期間")
                if r_receiv and r_payable and r_inv:
                    row_data.append(f"={new_col_letter}{r_receiv}-{new_col_letter}{r_payable}+{new_col_letter}{r_inv}")
                else:
                    row_data.append('')
            else:
                # 分類 / 売上高 * 365
                if target_sum_row and sales_row_nums:
                    row_data.append(f"=IF({new_col_letter}{ccc_sales_data_row}=0,\"\",({new_col_letter}{target_sum_row}/{new_col_letter}{ccc_sales_data_row})*365)")
                else:
                    row_data.append('')
                    
        ccc_ws.append(row_data)
        calc_rows[calc_name] = current_r

    # 空行 確保
    for _ in range(6):
        ccc_ws.append([''] * max_col)

    # 6. 書式設定
    # 最後に一括で適用するため定義のみ
    # Excel内部では[Red]を使用します。日本語環境でもExcelが自動で[赤]として表示します。
    number_format_integer = '#,##0;[Red]-#,##0'
    
    # 7. 比較列の追加 (2025-2015, etc.) - Percent BSのロジックを流用
    oldest_col = None
    latest_col = max_col
    comparison_cols = []

    def has_data(col_num):
        try:
            bs_dt = bs_ws.cell(1, col_num).value
            bs_date_key = str(bs_dt).strip() if bs_dt is not None else None
            pl_col_letter = pl_date_to_col.get(bs_date_key) if bs_date_key else None
            
            if not pl_col_letter:
                return False
                
            for r in sales_row_nums:
                val = pl_ws[f'{pl_col_letter}{r}'].value
                if val is not None and not (isinstance(val, str) and val.strip() in ('', '-', '0')) and not (isinstance(val, (int, float)) and val == 0):
                    return True
            return False
        except Exception:
            return False

    for col in range(3, max_col + 1):
        if has_data(col):
            oldest_col = col
            break

    if oldest_col is not None:
        kikan = latest_col - oldest_col
        
        new_latest_col = latest_col + 1
        new_oldest_col = oldest_col + 1
        start_diff_col = max_col + 2

        if kikan >= 10:
            ten_years_ago_col = new_latest_col - 10
            if ten_years_ago_col >= new_oldest_col:
                comparison_cols.append({'col': start_diff_col, 'latest': new_latest_col, 'base': ten_years_ago_col})
            
            five_years_ago_col = new_latest_col - 5
            if five_years_ago_col >= new_oldest_col and ten_years_ago_col >= new_oldest_col:
                comparison_cols.append({'col': start_diff_col + 1, 'latest': five_years_ago_col, 'base': ten_years_ago_col})
                
            if five_years_ago_col >= new_oldest_col:
                comparison_cols.append({'col': start_diff_col + 2, 'latest': new_latest_col, 'base': five_years_ago_col})
        elif kikan >= 5:
            comparison_cols.append({'col': start_diff_col, 'latest': new_latest_col, 'base': new_oldest_col})
            five_years_ago_col = new_latest_col - 5
            if five_years_ago_col >= new_oldest_col:
                comparison_cols.append({'col': start_diff_col + 1, 'latest': five_years_ago_col, 'base': new_oldest_col})
                comparison_cols.append({'col': start_diff_col + 2, 'latest': new_latest_col, 'base': five_years_ago_col})
        elif kikan > 0:
            comparison_cols.append({'col': start_diff_col, 'latest': new_latest_col, 'base': new_oldest_col})

        for comp in comparison_cols:
            comp_col = comp['col']
            comp_col_letter = openpyxl.utils.get_column_letter(comp_col)
            latest_letter = openpyxl.utils.get_column_letter(comp['latest'])
            base_letter = openpyxl.utils.get_column_letter(comp['base'])

            # Header
            year_formula = f"=YEAR({latest_letter}1) & \"-\" & YEAR({base_letter}1)"
            ccc_ws[f'{comp_col_letter}1'] = year_formula

            # Difference rows (only for Turnover and CCC rows)
            for calc_name, calc_row in calc_rows.items():
                diff_formula = f"=IF(OR({latest_letter}{calc_row}=\"\",{base_letter}{calc_row}=\"\"),\"\",{latest_letter}{calc_row}-{base_letter}{calc_row})"
                ccc_ws[f'{comp_col_letter}{calc_row}'] = diff_formula

    # 8. 見た目の調整
    ccc_ws.column_dimensions['A'].width = 12
    ccc_ws.column_dimensions['B'].width = 30
    ccc_ws.column_dimensions['C'].hidden = True
    
    total_cols = max_col + 1 + (len(comparison_cols) if oldest_col is not None else 0)
    for col in range(4, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        ccc_ws.column_dimensions[col_letter].width = 15
        
        # すべての数値列のセルに書式を適用 (ヘッダー以外)
        for row in range(2, ccc_ws.max_row + 1):
            ccc_ws[f'{col_letter}{row}'].number_format = number_format_integer

    ccc_ws.freeze_panes = 'D2'
    debug_log(f"CCC analysis sheet created: {analysis_sheet_name}")
